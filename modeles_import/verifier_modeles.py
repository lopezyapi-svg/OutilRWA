# -*- coding: utf-8 -*-
"""Vérifie que les quatre modèles générés passent bien les contrôles d'import.

Le contrôle du risque de crédit appelle directement le validateur du backend.
Les trois autres rejouent en Python la logique des dialogues d'import Flutter
(détection de feuille, correspondance des colonnes et des libellés, règles de
rejet ligne à ligne).

    python modeles_import/verifier_modeles.py
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

DOSSIER = Path(__file__).resolve().parent
RACINE = DOSSIER.parent
sys.path.insert(0, str(DOSSIER))
sys.path.insert(0, str(RACINE / "backend"))

from openpyxl import load_workbook  # noqa: E402

from _referentiels import (  # noqa: E402
    BIC_POSTES, ENTETES_ACTIONS, ENTETES_OBLIGATIONS, FP_POSTES,
    LIGNES_METIER, TYPES_EVENEMENT,
)
from generer_modeles import FICHIERS  # noqa: E402

anomalies: list[str] = []


def signaler(condition: bool, message: str) -> bool:
    if condition:
        print(f"   OK   {message}")
    else:
        print(f"   ECHEC {message}")
        anomalies.append(message)
    return condition


# ── Normalisation identique aux dialogues Flutter ───────────────────────────

def _norm(valeur: str) -> str:
    resultat = valeur.lower()
    for source, cible in (("àâä", "a"), ("éèêë", "e"), ("îï", "i"),
                          ("ôö", "o"), ("ùûü", "u")):
        for caractere in source:
            resultat = resultat.replace(caractere, cible)
    resultat = re.sub(r"[^a-z0-9]", "_", resultat)
    resultat = re.sub(r"_+", "_", resultat)
    return resultat.strip("_")


def _normalize_header(valeur: str) -> str:
    """Équivalent de `_normalizeHeader` (import marché)."""
    return re.sub(r"\s+", " ", (valeur or "").strip())


def _cellules(feuille, ligne_index: int) -> list[str]:
    valeurs = next(
        feuille.iter_rows(min_row=ligne_index, max_row=ligne_index, values_only=True),
        (),
    )
    return ["" if valeur is None else str(valeur).strip() for valeur in valeurs]


# ── 1. Risque de crédit : validateur réel du backend ────────────────────────

def verifier_credit() -> None:
    print("\n[1] Risque de credit")
    from app.validators.excel_import_validator import inspect_workbook_structure
    from database.services.excel_import_service import excel_import_service

    octets = FICHIERS["credit"].read_bytes()
    classeur = load_workbook(FICHIERS["credit"], data_only=True, read_only=True)
    try:
        inspection = inspect_workbook_structure(classeur)
    finally:
        classeur.close()

    signaler(inspection["valid"], "structure acceptee par le validateur du backend")
    for erreur in inspection["errors"][:10]:
        print(f"        - {erreur}")

    rapport = excel_import_service.inspect_uploaded_workbook(
        octets, FICHIERS["credit"].name
    )
    signaler(rapport["valid"], "inspection complete (service d'import) valide")
    lignes = rapport["rows_read_by_sheet"]
    print(f"        lignes lues : {lignes}")
    signaler(
        lignes.get("Template données") == 1000,
        f"1 000 expositions lues (obtenu {lignes.get('Template données')})",
    )

    # Rejeu du calcul prudentiel, comme à l'import réel.
    feuilles = load_workbook(FICHIERS["credit"], data_only=True, read_only=True)
    try:
        analyse = excel_import_service._parse_workbook(
            feuilles,
            profiler=type(
                "P", (), {"measure": lambda self, _n: __import__("contextlib").nullcontext()}
            )(),
        )
    finally:
        feuilles.close()

    signaler(
        analyse.rejected_rows == 0,
        f"aucune ligne rejetee au calcul (obtenu {analyse.rejected_rows})",
    )
    for erreur in analyse.errors[:5]:
        print(f"        - {erreur}")
    signaler(
        len(analyse.exposure_records) == 1000,
        f"1 000 expositions calculees (obtenu {len(analyse.exposure_records)})",
    )


# ── 2. Risque de marché : logique de `market_data_import_store.dart` ────────

def verifier_marche() -> None:
    print("\n[2] Risque de marche")
    classeur = load_workbook(FICHIERS["marche"], data_only=True, read_only=True)
    try:
        for feuille_nom, entetes_attendus, minimum in (
            ("Saisir donnée", ENTETES_OBLIGATIONS, 700),
            ("Actions", ENTETES_ACTIONS, 300),
        ):
            present = feuille_nom in classeur.sheetnames
            if not signaler(present, f"feuille « {feuille_nom} » presente"):
                continue
            feuille = classeur[feuille_nom]
            entetes = [_normalize_header(valeur) for valeur in _cellules(feuille, 1)]
            manquants = [
                entete for entete in entetes_attendus
                if _normalize_header(entete) not in entetes
            ]
            signaler(
                not manquants,
                f"« {feuille_nom} » : toutes les colonnes requises presentes"
                + (f" (manquantes : {manquants})" if manquants else ""),
            )
            lignes = sum(
                1
                for valeurs in feuille.iter_rows(min_row=2, values_only=True)
                if any(valeur not in (None, "") for valeur in valeurs)
            )
            signaler(
                lignes == minimum,
                f"« {feuille_nom} » : {minimum} lignes de donnees (obtenu {lignes})",
            )
    finally:
        classeur.close()


# ── 3. Risque opérationnel : pertes + BIC ───────────────────────────────────

_ALIAS_PERTES = {
    "date_occurrence": ["date_occurrence", "date occurrence", "date d occurrence",
                        "date d'occurrence", "date", "date_occ", "date_incident"],
    "description": ["description", "desc", "libelle", "libellé", "objet"],
    "ligne_metier": ["ligne_metier", "ligne de metier", "ligne de métier",
                     "ligne metier", "ligne métier", "metier", "métier", "business_line"],
    "type_evenement": ["type_evenement", "type d evenement", "type d'evenement",
                       "type d'événement", "type evenement", "type", "type_evt"],
    "cause_racine": ["cause_racine", "cause racine", "cause", "cause_rac", "root_cause"],
    "perte_brute": ["perte_brute", "perte brute", "perte brute (fcfa)", "montant brut",
                    "brut", "perte", "montant", "gross_loss"],
    "perte_recuperee": ["perte_recuperee", "perte recuperee", "perte récupérée",
                        "perte récupérée (fcfa)", "recuperee", "recouv", "récupéré", "recovery"],
    "statut": ["statut", "etat", "état", "status"],
}
_OBLIGATOIRES_PERTES = ["date_occurrence", "description", "ligne_metier",
                        "type_evenement", "perte_brute"]


def _champ_perte(entete: str):
    normalise = _norm(entete)
    for champ, alias in _ALIAS_PERTES.items():
        if any(_norm(valeur) == normalise for valeur in alias):
            return champ
    return None


def verifier_operationnel() -> None:
    print("\n[3] Risque operationnel")
    classeur = load_workbook(FICHIERS["operationnel"], data_only=True, read_only=True)
    try:
        # ── a) import des pertes ───────────────────────────────────────────
        candidates = [nom for nom in classeur.sheetnames if "incident" in nom.lower()]
        if not signaler(bool(candidates), "feuille « Incidents » detectee"):
            return
        feuille = classeur[candidates[0]]

        meilleure = {}
        ligne_entete = 1
        for index in range(1, min(7, feuille.max_row + 1)):
            candidate = {}
            for position, entete in enumerate(_cellules(feuille, index)):
                if not entete:
                    continue
                champ = _champ_perte(entete)
                if champ:
                    candidate[position] = champ
            if len(candidate) > len(meilleure):
                meilleure, ligne_entete = candidate, index

        manquants = [
            champ for champ in _OBLIGATOIRES_PERTES
            if champ not in meilleure.values()
        ]
        signaler(
            not manquants,
            "colonnes obligatoires des pertes reconnues"
            + (f" (manquantes : {manquants})" if manquants else ""),
        )

        index_par_champ = {champ: position for position, champ in meilleure.items()}
        invalides, total = [], 0
        for numero, valeurs in enumerate(
            feuille.iter_rows(min_row=ligne_entete + 1, values_only=True),
            start=ligne_entete + 1,
        ):
            if all(valeur in (None, "") for valeur in valeurs):
                continue
            total += 1
            erreurs = []

            def lire(champ):
                position = index_par_champ.get(champ)
                if position is None or position >= len(valeurs):
                    return ""
                return "" if valeurs[position] is None else str(valeurs[position]).strip()

            if not re.match(r"^\d{4}-\d{2}-\d{2}", lire("date_occurrence")):
                erreurs.append("date")
            if not lire("description"):
                erreurs.append("description")
            if lire("ligne_metier") not in LIGNES_METIER:
                erreurs.append("ligne_metier")
            if lire("type_evenement") not in TYPES_EVENEMENT:
                erreurs.append("type_evenement")
            try:
                if float(lire("perte_brute").replace(",", ".")) <= 0:
                    erreurs.append("perte_brute")
            except ValueError:
                erreurs.append("perte_brute")
            if erreurs:
                invalides.append((numero, erreurs))

        signaler(total == 1000, f"1 000 incidents lus (obtenu {total})")
        signaler(
            not invalides,
            f"aucun incident rejete (obtenu {len(invalides)} ligne(s) en erreur)",
        )
        for numero, erreurs in invalides[:5]:
            print(f"        - ligne {numero} : {', '.join(erreurs)}")

        # ── b) import BIC / CCR3 ───────────────────────────────────────────
        annees = []
        for nom in classeur.sheetnames:
            trouve = re.search(r"(19|20)\d{2}", nom.strip())
            if nom.strip().isdigit() and 1900 <= int(nom.strip()) <= 2100:
                annees.append(int(nom.strip()))
            elif trouve:
                annees.append(int(trouve.group(0)))
        signaler(
            sorted(annees) == [2023, 2024, 2025],
            f"onglets d'exercice detectes : {sorted(annees)}",
        )

        libelles_normalises = {_norm(poste): poste for poste in BIC_POSTES}
        for annee in sorted(annees):
            feuille_annee = classeur[str(annee)]
            entete_index = None
            poste_col = valeur_col = None
            for index in range(1, 7):
                cellules = _cellules(feuille_annee, index)
                p = v = None
                for position, brut in enumerate(cellules):
                    if not brut:
                        continue
                    normalise = _norm(brut)
                    if p is None and "poste" in normalise:
                        p = position
                    elif v is None and any(
                        normalise == alias or alias in normalise
                        for alias in ("valeur", "montant", "value", "val", "fcfa")
                    ):
                        v = position
                if p is not None and v is not None:
                    entete_index, poste_col, valeur_col = index, p, v
                    break
            if not signaler(
                entete_index is not None,
                f"exercice {annee} : colonnes Poste / Valeur detectees",
            ):
                continue
            trouves = set()
            for valeurs in feuille_annee.iter_rows(
                min_row=entete_index + 1, values_only=True
            ):
                if poste_col >= len(valeurs) or valeurs[poste_col] is None:
                    continue
                libelle = str(valeurs[poste_col]).strip()
                if _norm(libelle) in libelles_normalises:
                    trouves.add(libelles_normalises[_norm(libelle)])
            manquants = [poste for poste in BIC_POSTES if poste not in trouves]
            signaler(
                not manquants,
                f"exercice {annee} : 14 postes reconnus"
                + (f" (manquants : {manquants})" if manquants else ""),
            )
    finally:
        classeur.close()


# ── 4. Fonds propres ────────────────────────────────────────────────────────

def verifier_fonds_propres() -> None:
    print("\n[4] Fonds propres")
    classeur = load_workbook(FICHIERS["fonds_propres"], data_only=True, read_only=True)
    try:
        candidates = [
            nom for nom in classeur.sheetnames
            if "fonds" in nom.lower() or "propres" in nom.lower()
        ]
        if not signaler(bool(candidates), "feuille « Fonds propres » detectee"):
            return
        feuille = classeur[candidates[0]]

        entete_index = poste_col = valeur_col = None
        for index in range(1, 7):
            p = v = None
            for position, brut in enumerate(_cellules(feuille, index)):
                if not brut:
                    continue
                normalise = _norm(brut)
                if p is None and "poste" in normalise:
                    p = position
                elif v is None and any(
                    normalise == alias or alias in normalise
                    for alias in ("valeur", "montant", "value", "val", "fcfa")
                ):
                    v = position
            if p is not None and v is not None:
                entete_index, poste_col, valeur_col = index, p, v
                break

        if not signaler(entete_index is not None, "colonnes Poste / Valeur detectees"):
            return

        libelles = [poste for _groupe, poste in FP_POSTES]
        normalises = {_norm(poste): position for position, poste in enumerate(libelles)}
        valeurs_lues, non_reconnus = {}, []
        for valeurs in feuille.iter_rows(min_row=entete_index + 1, values_only=True):
            if poste_col >= len(valeurs) or valeurs[poste_col] is None:
                continue
            libelle = str(valeurs[poste_col]).strip()
            if not libelle:
                continue
            index_champ = normalises.get(_norm(libelle))
            if index_champ is None:
                # Repli par proximité, comme le dialogue Flutter.
                normalise = _norm(libelle)
                longueur = min(8, len(normalise))
                for position, poste in enumerate(libelles):
                    reference = _norm(poste)
                    if (
                        reference.find(normalise[:longueur]) >= 0
                        or normalise.find(reference[: min(8, len(reference))]) >= 0
                    ):
                        index_champ = position
                        break
            if index_champ is None:
                non_reconnus.append(libelle)
                continue
            brut = valeurs[valeur_col] if valeur_col < len(valeurs) else 0
            valeurs_lues[index_champ] = float(brut or 0)

        manquants = [
            libelles[position] for position in range(len(libelles))
            if position not in valeurs_lues
        ]
        signaler(
            not manquants,
            "les 11 postes sont lus" + (f" (manquants : {manquants})" if manquants else ""),
        )
        signaler(
            not non_reconnus,
            "aucun libelle non reconnu"
            + (f" (ignores : {non_reconnus})" if non_reconnus else ""),
        )

        cet1 = sum(valeurs_lues.get(i, 0) for i in range(4)) - valeurs_lues.get(4, 0)
        at1 = valeurs_lues.get(5, 0) + valeurs_lues.get(6, 0) - valeurs_lues.get(7, 0)
        tier2 = valeurs_lues.get(8, 0) + valeurs_lues.get(9, 0) - valeurs_lues.get(10, 0)
        total = cet1 + at1 + tier2
        print(
            f"        CET1 {cet1 / 1e9:,.2f} Md | AT1 {at1 / 1e9:,.2f} Md | "
            f"Tier 2 {tier2 / 1e9:,.2f} Md | total {total / 1e9:,.2f} Md"
        )
        signaler(cet1 > 0 and at1 > 0 and tier2 > 0, "les trois compartiments sont positifs")
    finally:
        classeur.close()


def main() -> int:
    print("=" * 78)
    print("Verification des modeles d'import")
    print("=" * 78)
    for fichier in FICHIERS.values():
        if not fichier.is_file():
            print(f"Fichier manquant : {fichier}")
            return 1

    verifier_credit()
    verifier_marche()
    verifier_operationnel()
    verifier_fonds_propres()

    print("\n" + "=" * 78)
    if anomalies:
        print(f"{len(anomalies)} anomalie(s) :")
        for anomalie in anomalies:
            print(f"  - {anomalie}")
        return 1
    print("Les quatre modeles passent tous les controles d'import.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
