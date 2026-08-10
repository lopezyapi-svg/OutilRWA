# -*- coding: utf-8 -*-
"""Modèle d'import « Risque de crédit » — 1 000 expositions + CRM associées.

Structure produite (strictement conforme à IMPORT_SHEET_SPECS côté backend) :
  • Template données   — 1 ligne = 1 exposition (10 colonnes obligatoires + 30 optionnelles)
  • CRM_non_financee   — 1 ligne = 1 garantie personnelle
  • CRM_financée       — 1 ligne = 1 sûreté financière
  • Notice             — feuille d'aide, ignorée à l'import
"""

from __future__ import annotations

import random
from datetime import date, timedelta

from openpyxl import Workbook

from _referentiels import (
    BANQUE, CAS_INSTITUTION_BANCAIRE, CAS_SOUVERAIN_AUCUN, CAS_SOUVERAIN_BCEAO,
    CAS_SOUVERAIN_ORG, CAS_SOUVERAIN_UEMOA, CATEGORIES, CATEGORIES_GARANT,
    DATE_ANALYSE, MATURITES_CRM, NIVEAUX_RISQUE_HB, NOTATION_PAYS,
    NOTATIONS_CRM_FINANCEE, PAYS_HORS_UEMOA, PAYS_UEMOA,
    PONDERATIONS_AVANT_DEFAUT, TAUX_XOF, TYPES_AUTRES_ACTIFS,
    TYPES_EMETTEUR_CRM,
)
from _styles import FMT_MONTANT, ecrire_entetes, ecrire_lignes, feuille_notice

# ── Colonnes exactes de chaque feuille (obligatoires puis optionnelles) ─────

COLONNES_OBLIGATOIRES_TEMPLATE = [
    "Date d'analyse", "ID_Exposition", "Contrepartie",
    "Notation_externe_contrepartie", "Pays_contrepartie",
    "Notation_externe_pays", "Catégorie d'exposition",
    "Montant_exposition_but_au_bilan", "Devise", "Type_CRM",
]

COLONNES_OPTIONNELLES_TEMPLATE = [
    "Date d'octroi", "Date d'échéance", "PRÊT TOTAL",
    "Montant d'exposition au HB", "Niveau de risque HB", "Statut",
    "Provisions", "Jours_impayes", "Commentaire",
    "Cas_particulier_souverain", "Souverain_ponderation_pref_nulle",
    "Souverain_OCE_etabli", "Souverain_note_OCE",
    "Organisme_public_cas_UEMOA_FCFA", "Organisme_public_activite_non_publique",
    "BMD_cas_haute_qualite", "BMD_cas_UEMOA_FCFA",
    "BMD_criteres_UEMOA_respectes", "BMD_institution_listee_FCFA",
    "Cas_institution_bancaire", "Type_autre_actif",
    "Clientele_detail_criteres_respectes", "Immobilier_residentiel_eligible",
    "Immobilier_commercial_eligible", "Ponderation_initiale_avant_defaut",
    "Defaut_pret_immo_residentiel", "Defaut_provision_min_20pct",
    "Entreprise_depasse_seuil_degradation_BCEAO",
    "Entreprise_procedure_prudentielle",
    "Entreprise_investissement_hors_loi_bancaire",
]

COLONNES_TEMPLATE = COLONNES_OBLIGATOIRES_TEMPLATE + COLONNES_OPTIONNELLES_TEMPLATE

COLONNES_CRM_NON_FIN = [
    "ID_Exposition", "Nom du garant", "Catégorie du garant", "Note_garant",
    "Pays_garant", "Note_pays_garant", "Part couverte",
]

COLONNES_CRM_FIN_OBLIGATOIRES = [
    "ID_Exposition", "Valeur_Collatéral", "Type_emetteur", "Notation",
    "Bloc", "Maturite",
]
COLONNES_CRM_FIN_OPTIONNELLES = [
    "Devise_Collatéral", "Type_Collatéral",
    "Obligation_convertible_indice_principal", "Decote_OPCVM_max",
]
COLONNES_CRM_FIN = COLONNES_CRM_FIN_OBLIGATOIRES + COLONNES_CRM_FIN_OPTIONNELLES

# ── Répartition du portefeuille (1 000 lignes) ──────────────────────────────
# Calquée sur la structure d'une banque universelle de l'UMOA : forte
# granularité sur le détail, concentration des encours sur les entreprises et
# le souverain.
REPARTITION = [
    ("a", 30), ("b", 22), ("c", 8), ("d", 40), ("e", 210), ("f", 455),
    ("g", 85), ("h", 40), ("i", 60), ("j", 5), ("k", 45),
]

DATE_REF = date(2026, 6, 30)

# ── Dictionnaires de noms fictifs ───────────────────────────────────────────
_ENT_PREFIXES = [
    "Société", "Compagnie", "Groupe", "Entreprise", "Établissements",
    "Comptoir", "Manufacture", "Coopérative",
]
_ENT_ACTIVITES = [
    "Agro-Industrielle", "de Distribution", "Cotonnière", "de Transformation",
    "des Oléagineux", "de Négoce", "Cimentière", "de Transport", "Portuaire",
    "d'Assemblage", "Textile", "Pharmaceutique", "de Travaux Publics",
    "Avicole", "Halieutique", "Minière", "Énergétique", "Immobilière",
    "de Logistique", "Sucrière", "Brassicole", "de Téléphonie",
    "d'Ingénierie", "Hôtelière", "Aquacole", "de Câblage",
]
_ENT_SUFFIXES = ["SA", "SARL", "SAS", "SA Holding", "Group SA"]
_ENT_MARQUES = [
    "Sahel", "Atlantis", "Kolda", "Bandama", "Zambé", "Nokoué", "Faso",
    "Bagoé", "Wouri", "Comoé", "Djoliba", "Sanaga", "Sirba", "Mono",
    "Korhogo", "Tiassalé", "Kaolack", "Ouémé", "Nazinon", "Bandiagara",
]

_PRENOMS = [
    "Kouamé", "Aminata", "Ibrahim", "Fatoumata", "Yao", "Awa", "Seydou",
    "Mariam", "Adama", "Rokia", "Boubacar", "Salimata", "Koffi", "Nafissatou",
    "Moussa", "Aïcha", "Souleymane", "Bintou", "Mamadou", "Djeneba",
    "Abdoulaye", "Kadiatou", "Alassane", "Oumou", "Cheikh", "Hawa",
    "Zakaria", "Ramata", "Issouf", "Assitan",
]
_NOMS = [
    "Traoré", "Diallo", "Koné", "Ouattara", "Bamba", "Sanogo", "Cissé",
    "Kouassi", "N'Guessan", "Diop", "Ndiaye", "Fall", "Sow", "Barry",
    "Camara", "Touré", "Sylla", "Bakayoko", "Dembélé", "Zongo", "Compaoré",
    "Sawadogo", "Ouédraogo", "Gnassingbé", "Adjovi", "Hounkpatin", "Keita",
    "Sidibé", "Coulibaly", "Doumbia",
]

_BANQUES_CORRESPONDANTES = [
    "Banque Régionale du Littoral", "Union Bancaire du Sahel",
    "Banque de Développement de l'Ouest", "Crédit Interbancaire Ivoirien",
    "Banque Commerciale du Delta", "Société Bancaire du Fleuve",
    "Banque d'Investissement de la Lagune", "Caisse de Refinancement Régionale",
    "Banque Panafricaine de Dépôts", "Établissement Financier du Golfe",
]

_ORGANISMES_PUBLICS = [
    "Agence Nationale de l'Électrification Rurale",
    "Office National de l'Assainissement",
    "District Autonome du Littoral",
    "Régie Autonome des Transports Urbains",
    "Agence de Gestion des Routes",
    "Office National de l'Eau Potable",
    "Conseil Régional de la Vallée",
    "Agence Nationale du Logement Social",
    "Port Autonome de la Baie",
    "Office des Marchés Agricoles",
    "Agence de Développement Rural",
    "Centre Hospitalier Universitaire Régional",
]

_BMD = [
    ("Banque Ouest-Africaine de Développement", "AAA"),
    ("Banque Africaine de Développement", "AAA"),
    ("Société Financière Internationale", "AAA"),
    ("Banque Islamique de Développement", "AA"),
    ("Banque Européenne d'Investissement", "AAA"),
    ("Banque Internationale pour la Reconstruction", "AAA"),
    ("Banque de Développement des États d'Afrique Centrale", "A"),
    ("Fonds Africain de Garantie", "A-"),
]

_OBJETS_CREDIT_DETAIL = [
    "Crédit à la consommation", "Prêt personnel", "Découvert autorisé",
    "Crédit scolarité", "Crédit véhicule", "Crédit équipement ménager",
    "Avance sur salaire", "Microcrédit commerçant", "Crédit campagne agricole",
    "Facilité de caisse particulier",
]


def _montant(rng: random.Random, mini: float, maxi: float, arrondi: int = 100000) -> float:
    """Montant tiré selon une loi log-uniforme puis arrondi (queue longue
    réaliste : beaucoup de petits encours, quelques très gros)."""
    import math

    tirage = math.exp(rng.uniform(math.log(mini), math.log(maxi)))
    return float(int(round(tirage / arrondi)) * arrondi)


def _date_iso(jour: date) -> str:
    return jour.isoformat()


def _dates_credit(rng: random.Random, duree_mois_min: int, duree_mois_max: int):
    """Couple (date d'octroi, date d'échéance) encadrant la date d'analyse."""
    anciennete = rng.randint(30, 2200)
    octroi = DATE_REF - timedelta(days=anciennete)
    duree = rng.randint(duree_mois_min, duree_mois_max)
    echeance = octroi + timedelta(days=int(duree * 30.44))
    if echeance <= DATE_REF:
        echeance = DATE_REF + timedelta(days=rng.randint(20, 900))
    return _date_iso(octroi), _date_iso(echeance)


def _nom_entreprise(rng: random.Random) -> str:
    if rng.random() < 0.45:
        return (
            f"{rng.choice(_ENT_PREFIXES)} {rng.choice(_ENT_MARQUES)} "
            f"{rng.choice(_ENT_ACTIVITES)} {rng.choice(_ENT_SUFFIXES)}"
        )
    return f"{rng.choice(_ENT_MARQUES)}-{rng.choice(_ENT_MARQUES)} {rng.choice(_ENT_SUFFIXES)}"


def _nom_particulier(rng: random.Random) -> str:
    civilite = rng.choice(["M.", "Mme", "M.", "Mme", "Mlle"])
    return f"{civilite} {rng.choice(_NOMS).upper()} {rng.choice(_PRENOMS)}"


def _pays(rng: random.Random, part_uemoa: float = 0.88) -> str:
    return (
        rng.choice(PAYS_UEMOA)
        if rng.random() < part_uemoa
        else rng.choice(PAYS_HORS_UEMOA)
    )


def _devise(rng: random.Random, part_xof: float = 0.9) -> str:
    if rng.random() < part_xof:
        return "XOF"
    return "EUR" if rng.random() < 0.6 else "USD"


def _notation_entreprise(rng: random.Random) -> str:
    return rng.choices(
        ["A", "A-", "BBB+", "BBB", "BBB-", "BB+", "BB", "BB-", "B+", "B", "B-", "Non noté"],
        weights=[1, 2, 3, 5, 6, 8, 10, 10, 8, 6, 3, 38],
    )[0]


# ── Construction d'une exposition par catégorie ─────────────────────────────

def _ligne_souverain(rng, identifiant, index):
    pays = PAYS_UEMOA[index % len(PAYS_UEMOA)] if rng.random() < 0.85 else rng.choice(PAYS_HORS_UEMOA)
    uemoa = pays in PAYS_UEMOA
    devise = "XOF" if uemoa else _devise(rng, 0.2)
    montant = _montant(rng, 900_000_000, 12_000_000_000, 1_000_000)
    octroi, echeance = _dates_credit(rng, 6, 120)
    if uemoa and devise == "XOF":
        cas = CAS_SOUVERAIN_BCEAO if rng.random() < 0.12 else CAS_SOUVERAIN_UEMOA
    elif rng.random() < 0.25:
        cas = rng.choice(CAS_SOUVERAIN_ORG)
    else:
        cas = CAS_SOUVERAIN_AUCUN
    libelle = (
        f"Titre public — État {pays}" if uemoa
        else f"Créance souveraine — {pays}"
    )
    ligne = {
        "Contrepartie": libelle,
        "Notation_externe_contrepartie": NOTATION_PAYS[pays],
        "Pays_contrepartie": pays,
        "Notation_externe_pays": NOTATION_PAYS[pays],
        "Catégorie d'exposition": CATEGORIES["a"],
        "Montant_exposition_but_au_bilan": montant,
        "Devise": devise,
        "Date d'octroi": octroi,
        "Date d'échéance": echeance,
        "PRÊT TOTAL": montant,
        "Cas_particulier_souverain": cas,
        "Souverain_ponderation_pref_nulle": "Oui" if cas != CAS_SOUVERAIN_AUCUN else "Non",
        "Commentaire": "Portefeuille de titres publics — refinancement BCEAO éligible"
        if uemoa else "Exposition souveraine hors UMOA",
    }
    if cas == CAS_SOUVERAIN_AUCUN and NOTATION_PAYS[pays] == "Non noté":
        ligne["Souverain_OCE_etabli"] = "Oui"
        ligne["Souverain_note_OCE"] = str(rng.randint(3, 6))
    return ligne


def _ligne_organisme_public(rng, identifiant, index):
    pays = _pays(rng, 0.95)
    montant = _montant(rng, 180_000_000, 4_500_000_000, 100_000)
    octroi, echeance = _dates_credit(rng, 24, 144)
    uemoa_fcfa = pays in PAYS_UEMOA
    return {
        "Contrepartie": _ORGANISMES_PUBLICS[index % len(_ORGANISMES_PUBLICS)]
        + f" — {pays}",
        "Notation_externe_contrepartie": rng.choice(["BBB", "BBB-", "BB+", "Non noté", "Non noté"]),
        "Pays_contrepartie": pays,
        "Notation_externe_pays": NOTATION_PAYS[pays],
        "Catégorie d'exposition": CATEGORIES["b"],
        "Montant_exposition_but_au_bilan": montant,
        "Devise": "XOF" if uemoa_fcfa else _devise(rng, 0.3),
        "Date d'octroi": octroi,
        "Date d'échéance": echeance,
        "PRÊT TOTAL": montant,
        "Organisme_public_cas_UEMOA_FCFA": "Oui" if uemoa_fcfa else "Non",
        "Organisme_public_activite_non_publique": "Oui" if rng.random() < 0.18 else "Non",
        "Commentaire": "Financement d'infrastructure publique",
    }


def _ligne_bmd(rng, identifiant, index):
    nom, notation = _BMD[index % len(_BMD)]
    montant = _montant(rng, 600_000_000, 7_000_000_000, 1_000_000)
    octroi, echeance = _dates_credit(rng, 36, 180)
    fcfa = rng.random() < 0.6
    return {
        "Contrepartie": nom,
        "Notation_externe_contrepartie": notation,
        "Pays_contrepartie": "Cote d'Ivoire" if fcfa else "France",
        "Notation_externe_pays": NOTATION_PAYS["Cote d'Ivoire" if fcfa else "France"],
        "Catégorie d'exposition": CATEGORIES["c"],
        "Montant_exposition_but_au_bilan": montant,
        "Devise": "XOF" if fcfa else "EUR",
        "Date d'octroi": octroi,
        "Date d'échéance": echeance,
        "PRÊT TOTAL": montant,
        "BMD_cas_haute_qualite": "Oui" if notation in ("AAA", "AA") else "Non",
        "BMD_institution_listee_FCFA": "Oui" if fcfa else "Non",
        "BMD_cas_UEMOA_FCFA": "Oui" if fcfa else "Non",
        "BMD_criteres_UEMOA_respectes": "Oui" if fcfa else "Non",
        "Commentaire": "Ligne de refinancement multilatérale",
    }


def _ligne_institution_financiere(rng, identifiant, index):
    pays = _pays(rng, 0.75)
    nom = _BANQUES_CORRESPONDANTES[index % len(_BANQUES_CORRESPONDANTES)]
    montant = _montant(rng, 90_000_000, 9_000_000_000, 100_000)
    octroi, echeance = _dates_credit(rng, 1, 60)
    notation = rng.choices(
        ["A", "A-", "BBB+", "BBB", "BBB-", "BB+", "BB", "Non noté"],
        weights=[3, 5, 8, 10, 10, 8, 6, 25],
    )[0]
    ligne = {
        "Contrepartie": f"{nom} — {pays}",
        "Notation_externe_contrepartie": notation,
        "Pays_contrepartie": pays,
        "Notation_externe_pays": NOTATION_PAYS[pays],
        "Catégorie d'exposition": CATEGORIES["d"],
        "Montant_exposition_but_au_bilan": montant,
        "Devise": _devise(rng, 0.72),
        "Date d'octroi": octroi,
        "Date d'échéance": echeance,
        "PRÊT TOTAL": montant,
        "Cas_institution_bancaire": rng.choices(
            list(CAS_INSTITUTION_BANCAIRE), weights=[45, 8, 47]
        )[0],
        "Commentaire": rng.choice(
            ["Placement interbancaire", "Ligne de correspondant bancaire",
             "Prêt interbancaire à terme", "Compte nostro"]
        ),
    }
    if rng.random() < 0.25:
        ligne["Montant d'exposition au HB"] = round(montant * rng.uniform(0.05, 0.4), -5)
        ligne["Niveau de risque HB"] = rng.choice(NIVEAUX_RISQUE_HB[:4])
    return ligne


def _ligne_entreprise(rng, identifiant, index):
    pays = _pays(rng, 0.9)
    montant = _montant(rng, 25_000_000, 8_000_000_000, 100_000)
    octroi, echeance = _dates_credit(rng, 6, 120)
    pret_total = round(montant * rng.uniform(1.0, 1.35), -5)
    ligne = {
        "Contrepartie": _nom_entreprise(rng),
        "Notation_externe_contrepartie": _notation_entreprise(rng),
        "Pays_contrepartie": pays,
        "Notation_externe_pays": NOTATION_PAYS[pays],
        "Catégorie d'exposition": CATEGORIES["e"],
        "Montant_exposition_but_au_bilan": montant,
        "Devise": _devise(rng, 0.86),
        "Date d'octroi": octroi,
        "Date d'échéance": echeance,
        "PRÊT TOTAL": pret_total,
        "Entreprise_depasse_seuil_degradation_BCEAO": "Oui" if rng.random() < 0.06 else "Non",
        "Entreprise_procedure_prudentielle": "Oui" if rng.random() < 0.03 else "Non",
        "Entreprise_investissement_hors_loi_bancaire": "Oui" if rng.random() < 0.04 else "Non",
        "Commentaire": rng.choice(
            ["Crédit d'investissement", "Crédit de campagne", "Ligne d'escompte",
             "Découvert d'exploitation", "Crédit-bail adossé",
             "Financement de stock", "Préfinancement export"]
        ),
    }
    if rng.random() < 0.28:
        ligne["Montant d'exposition au HB"] = round(montant * rng.uniform(0.08, 0.6), -5)
        ligne["Niveau de risque HB"] = rng.choices(
            list(NIVEAUX_RISQUE_HB), weights=[15, 20, 35, 22, 8]
        )[0]
    if rng.random() < 0.09:
        ligne["Jours_impayes"] = rng.randint(1, 85)
        ligne["Provisions"] = round(montant * rng.uniform(0.01, 0.09), -5)
    return ligne


def _ligne_detail(rng, identifiant, index):
    pays = _pays(rng, 0.98)
    montant = _montant(rng, 350_000, 145_000_000, 10_000)
    octroi, echeance = _dates_credit(rng, 6, 84)
    ligne = {
        "Contrepartie": _nom_particulier(rng),
        "Notation_externe_contrepartie": "Non noté",
        "Pays_contrepartie": pays,
        "Notation_externe_pays": NOTATION_PAYS[pays],
        "Catégorie d'exposition": CATEGORIES["f"],
        "Montant_exposition_but_au_bilan": montant,
        "Devise": "XOF",
        "Date d'octroi": octroi,
        "Date d'échéance": echeance,
        "PRÊT TOTAL": round(montant * rng.uniform(1.0, 1.2), -4),
        "Clientele_detail_criteres_respectes": "Oui" if rng.random() < 0.94 else "Non",
        "Commentaire": rng.choice(_OBJETS_CREDIT_DETAIL),
    }
    if rng.random() < 0.07:
        ligne["Montant d'exposition au HB"] = round(montant * rng.uniform(0.1, 0.5), -4)
        ligne["Niveau de risque HB"] = rng.choice(NIVEAUX_RISQUE_HB[:3])
    if rng.random() < 0.11:
        ligne["Jours_impayes"] = rng.randint(1, 89)
        ligne["Provisions"] = round(montant * rng.uniform(0.01, 0.12), -4)
    return ligne


def _ligne_immo_residentiel(rng, identifiant, index):
    pays = _pays(rng, 0.99)
    montant = _montant(rng, 6_000_000, 165_000_000, 10_000)
    octroi, echeance = _dates_credit(rng, 84, 300)
    return {
        "Contrepartie": _nom_particulier(rng),
        "Notation_externe_contrepartie": "Non noté",
        "Pays_contrepartie": pays,
        "Notation_externe_pays": NOTATION_PAYS[pays],
        "Catégorie d'exposition": CATEGORIES["g"],
        "Montant_exposition_but_au_bilan": montant,
        "Devise": "XOF",
        "Date d'octroi": octroi,
        "Date d'échéance": echeance,
        "PRÊT TOTAL": round(montant * rng.uniform(1.0, 1.15), -4),
        "Immobilier_residentiel_eligible": "Oui" if rng.random() < 0.88 else "Non",
        "Commentaire": "Crédit acquisition logement — hypothèque de 1er rang",
    }


def _ligne_immo_commercial(rng, identifiant, index):
    pays = _pays(rng, 0.96)
    montant = _montant(rng, 45_000_000, 2_400_000_000, 100_000)
    octroi, echeance = _dates_credit(rng, 60, 240)
    return {
        "Contrepartie": _nom_entreprise(rng),
        "Notation_externe_contrepartie": _notation_entreprise(rng),
        "Pays_contrepartie": pays,
        "Notation_externe_pays": NOTATION_PAYS[pays],
        "Catégorie d'exposition": CATEGORIES["h"],
        "Montant_exposition_but_au_bilan": montant,
        "Devise": _devise(rng, 0.93),
        "Date d'octroi": octroi,
        "Date d'échéance": echeance,
        "PRÊT TOTAL": round(montant * rng.uniform(1.0, 1.2), -5),
        "Immobilier_commercial_eligible": "Oui" if rng.random() < 0.75 else "Non",
        "Entreprise_depasse_seuil_degradation_BCEAO": "Non",
        "Entreprise_procedure_prudentielle": "Non",
        "Commentaire": rng.choice(
            ["Acquisition d'entrepôt logistique", "Immeuble de bureaux locatif",
             "Centre commercial — financement long", "Plateforme industrielle"]
        ),
    }


def _ligne_souffrance(rng, identifiant, index):
    pays = _pays(rng, 0.97)
    montant = _montant(rng, 9_000_000, 4_500_000_000, 10_000)
    octroi, echeance = _dates_credit(rng, 12, 120)
    taux_provision = rng.uniform(0.05, 0.85)
    provisions = round(montant * taux_provision, -4)
    est_immo = rng.random() < 0.18
    return {
        "Contrepartie": _nom_entreprise(rng) if rng.random() < 0.4 else _nom_particulier(rng),
        "Notation_externe_contrepartie": rng.choice(["< B-", "B-", "Non noté", "Non noté"]),
        "Pays_contrepartie": pays,
        "Notation_externe_pays": NOTATION_PAYS[pays],
        "Catégorie d'exposition": CATEGORIES["i"],
        "Montant_exposition_but_au_bilan": montant,
        "Devise": "XOF",
        "Date d'octroi": octroi,
        "Date d'échéance": echeance,
        "PRÊT TOTAL": montant,
        "Provisions": provisions,
        "Jours_impayes": rng.choice([92, 105, 121, 148, 180, 215, 270, 365, 420, 540]),
        "Ponderation_initiale_avant_defaut": rng.choice(PONDERATIONS_AVANT_DEFAUT),
        "Defaut_pret_immo_residentiel": "Oui" if est_immo else "Non",
        "Defaut_provision_min_20pct": "Oui" if taux_provision >= 0.20 else "Non",
        "Commentaire": rng.choice(
            ["Créance déclassée — recouvrement amiable en cours",
             "Contentieux — mise en demeure notifiée",
             "Dossier transmis au service juridique",
             "Rééchelonnement refusé — provisionnement renforcé"]
        ),
    }


def _ligne_risque_eleve(rng, identifiant, index):
    pays = _pays(rng, 0.9)
    montant = _montant(rng, 60_000_000, 1_600_000_000, 100_000)
    octroi, echeance = _dates_credit(rng, 12, 84)
    return {
        "Contrepartie": _nom_entreprise(rng),
        "Notation_externe_contrepartie": rng.choice(["< B-", "B-", "Non noté"]),
        "Pays_contrepartie": pays,
        "Notation_externe_pays": NOTATION_PAYS[pays],
        "Catégorie d'exposition": CATEGORIES["j"],
        "Montant_exposition_but_au_bilan": montant,
        "Devise": _devise(rng, 0.8),
        "Date d'octroi": octroi,
        "Date d'échéance": echeance,
        "PRÊT TOTAL": montant,
        "Commentaire": "Secteur signalé à risque élevé par le régulateur",
    }


# Un bilan bancaire est dominé par l'encaisse et les immobilisations : les
# types d'actifs les plus fortement pondérés (participations significatives,
# établissements non conformes) restent marginaux.
_POIDS_AUTRES_ACTIFS = (22, 8, 8, 7, 22, 15, 4, 4, 5, 3, 2)


def _ligne_autre_actif(rng, identifiant, index):
    type_actif = rng.choices(list(TYPES_AUTRES_ACTIFS), weights=_POIDS_AUTRES_ACTIFS)[0]
    montant = _montant(rng, 15_000_000, 9_000_000_000, 100_000)
    return {
        "Contrepartie": f"Poste de bilan — {type_actif[:60]}",
        "Notation_externe_contrepartie": "Non noté",
        "Pays_contrepartie": "Cote d'Ivoire",
        "Notation_externe_pays": NOTATION_PAYS["Cote d'Ivoire"],
        "Catégorie d'exposition": CATEGORIES["k"],
        "Montant_exposition_but_au_bilan": montant,
        "Devise": _devise(rng, 0.94),
        "PRÊT TOTAL": montant,
        "Type_autre_actif": type_actif,
        "Commentaire": "Élément d'actif non ventilé en portefeuille de crédit",
    }


_FABRIQUES = {
    "a": _ligne_souverain,
    "b": _ligne_organisme_public,
    "c": _ligne_bmd,
    "d": _ligne_institution_financiere,
    "e": _ligne_entreprise,
    "f": _ligne_detail,
    "g": _ligne_immo_residentiel,
    "h": _ligne_immo_commercial,
    "i": _ligne_souffrance,
    "j": _ligne_risque_eleve,
    "k": _ligne_autre_actif,
}

# Catégories sur lesquelles une CRM est plausible.
_CRM_ELIGIBLE = {"b", "d", "e", "f", "g", "h", "i", "j"}

# Garants proposés selon le profil de l'exposition couverte : un fonds de
# garantie de proximité pour le détail, une contre-garantie multilatérale ou
# bancaire pour les gros dossiers d'entreprise.
_GARANTS_DETAIL = [
    ("Mutuelle de Garantie des Artisans", "Clientèle de détail", "Non noté"),
    ("Fonds de Garantie des Crédits aux PME", "Organismes pub. hors Adm c", "BBB"),
    ("Société de Cautionnement Mutuel de l'Ouest", "Entreprises", "BB"),
    ("Fonds Régional de Garantie de l'Habitat", "Organismes pub. hors Adm c", "BB+"),
]
_GARANTS_ENTREPRISE = [
    ("Société de Cautionnement Mutuel de l'Ouest", "Entreprises", "BB"),
    ("Compagnie d'Assurance-Crédit Régionale", "Entreprises", "BBB-"),
    ("Groupe Sahel Holding SA", "Entreprises", "BB-"),
    ("Union Bancaire du Sahel", "Institutions financières", "BBB-"),
    ("Banque Régionale du Littoral", "Institutions financières", "BB+"),
]
_GARANTS_GRANDS_COMPTES = [
    ("Banque Ouest-Africaine de Développement", "Expositions sur les BMD", "AAA"),
    ("Fonds de Garantie des Dépôts de l'UMOA", "Souverains", "AAA"),
    ("État — Fonds de Garantie Agricole", "Souverains", "BB"),
    ("Fonds Africain de Garantie", "Expositions sur les BMD", "A-"),
    ("Union Bancaire du Sahel", "Institutions financières", "BBB-"),
]

_BLOCS_CRM = [
    "Dépôt à terme nanti", "Compte-espèces bloqué", "Nantissement de titres publics",
    "Gage sur bons du Trésor", "Nantissement de parts d'OPCVM",
    "Nantissement d'actions cotées", "Or physique déposé en coffre",
    "Titres bancaires nantis",
]


def _crm_financee(rng, exposition_id, montant, devise):
    type_collateral = rng.choices(
        [
            "Liquidités dans la même devise",
            "Titre de dette souverain",
            "Titre non noté émis par un État de l UMOA",
            "Titre de dette émis par un autre émetteur",
            "Action de l indice BRVM 10",
            "OPCVM / FI",
            "Or",
            "Titre bancaire non noté",
        ],
        weights=[34, 22, 12, 9, 8, 6, 4, 5],
    )[0]
    est_souverain = "souverain" in type_collateral or "UMOA" in type_collateral
    # La sûreté est adossée à l'exposition : sa valeur est exprimée dans la
    # même devise, celle que reprend la colonne Devise_Collatéral.
    valeur = round(montant * rng.uniform(0.25, 1.15), -2)
    ligne = {
        "ID_Exposition": exposition_id,
        "Valeur_Collatéral": valeur,
        "Type_emetteur": TYPES_EMETTEUR_CRM[0] if est_souverain else TYPES_EMETTEUR_CRM[1],
        "Notation": rng.choices(
            list(NOTATIONS_CRM_FINANCEE),
            weights=[3, 2, 3, 4, 5, 6, 6, 7, 8, 8, 7, 6, 5, 3, 27],
        )[0],
        "Bloc": rng.choice(_BLOCS_CRM),
        "Maturite": rng.choices(list(MATURITES_CRM), weights=[35, 28, 18, 13, 6])[0],
        "Devise_Collatéral": devise,
        "Type_Collatéral": type_collateral,
    }
    if type_collateral == "OPCVM / FI":
        ligne["Decote_OPCVM_max"] = rng.choice([0.15, 0.20, 0.30])
    if type_collateral == "Obligation convertible en action":
        ligne["Obligation_convertible_indice_principal"] = "Oui"
    return ligne


def _crm_non_financee(rng, exposition_id, montant, montant_xof, code_categorie):
    # Le choix du garant s'apprécie sur la contre-valeur FCFA, pour rester
    # cohérent quelle que soit la devise de l'exposition couverte.
    if code_categorie in ("f", "g") or montant_xof < 150_000_000:
        vivier = _GARANTS_DETAIL
    elif montant_xof >= 1_500_000_000:
        vivier = _GARANTS_GRANDS_COMPTES
    else:
        vivier = _GARANTS_ENTREPRISE
    nom, categorie, note = rng.choice(vivier)
    pays = rng.choice(PAYS_UEMOA)
    return {
        "ID_Exposition": exposition_id,
        "Nom du garant": nom,
        "Catégorie du garant": categorie if categorie in CATEGORIES_GARANT else "Entreprises",
        "Note_garant": note,
        "Pays_garant": pays,
        "Note_pays_garant": NOTATION_PAYS[pays],
        "Part couverte": round(montant * rng.uniform(0.20, 0.95), -2),
    }


# Champs monétaires de la feuille de saisie : ils sont tous exprimés dans la
# devise de la ligne (colonne « Devise »), jamais en contre-valeur FCFA —
# l'outil applique lui-même la conversion à l'affichage et à l'agrégation.
_CHAMPS_MONETAIRES = (
    "Montant_exposition_but_au_bilan",
    "PRÊT TOTAL",
    "Montant d'exposition au HB",
    "Provisions",
)


def _convertir_en_devise(ligne: dict, echelle: float) -> None:
    """Applique le facteur d'échelle puis exprime les montants dans la devise
    de l'exposition (les fabriques les produisent en FCFA)."""
    taux = TAUX_XOF.get(ligne["Devise"], 1.0)
    arrondi = 10_000 if taux == 1.0 else 100
    for champ in _CHAMPS_MONETAIRES:
        valeur = ligne.get(champ)
        if valeur is None:
            continue
        converti = valeur * echelle / taux
        ligne[champ] = float(max(arrondi, round(converti / arrondi) * arrondi))


def construire_donnees(graine: int = 20260630, echelle: float = 1.0):
    """Génère les 1 000 expositions et leurs CRM. Retourne (template, non_fin, fin).

    `echelle` multiplie tous les encours : le RWA de crédit lui étant
    proportionnel, elle sert à caler le poids du risque de crédit dans le RWA
    total sans toucher à la structure du portefeuille.
    """
    rng = random.Random(graine)
    template, crm_non_fin, crm_fin = [], [], []
    compteur = 0

    plan = []
    for code, nombre in REPARTITION:
        plan.extend([code] * nombre)
    rng.shuffle(plan)

    index_par_code = {code: 0 for code, _ in REPARTITION}

    for code in plan:
        compteur += 1
        identifiant = f"EXP-2026-{compteur:05d}"
        index = index_par_code[code]
        index_par_code[code] += 1

        ligne = _FABRIQUES[code](rng, identifiant, index)
        ligne["Date d'analyse"] = DATE_ANALYSE
        ligne["ID_Exposition"] = identifiant
        _convertir_en_devise(ligne, echelle)

        montant = ligne["Montant_exposition_but_au_bilan"]
        devise = ligne["Devise"]
        montant_xof = montant * TAUX_XOF.get(devise, 1.0)

        # Statut de GESTION (libre) — à ne pas confondre avec le statut
        # prudentiel, que l'outil dérive lui-même des jours d'impayés.
        jours = ligne.get("Jours_impayes") or 0
        if code == "i":
            ligne["Statut"] = rng.choice(["Contentieux", "En recouvrement", "Restructurée"])
        elif jours > 0:
            ligne["Statut"] = "En recouvrement"
        else:
            ligne["Statut"] = "Active"

        # Technique de réduction du risque de crédit.
        type_crm = "Aucune"
        if code in _CRM_ELIGIBLE:
            tirage = rng.random()
            if tirage < 0.13:
                type_crm = "CRM financee"
                crm_fin.append(_crm_financee(rng, identifiant, montant, devise))
            elif tirage < 0.24:
                type_crm = "CRM non financee"
                crm_non_fin.append(
                    _crm_non_financee(rng, identifiant, montant, montant_xof, code)
                )
        ligne["Type_CRM"] = type_crm

        template.append(ligne)

    return template, crm_non_fin, crm_fin


def construire_classeur(chemin, template, crm_non_fin, crm_fin):
    wb = Workbook()
    wb.remove(wb.active)

    formats_template = {
        "Montant_exposition_but_au_bilan": FMT_MONTANT,
        "PRÊT TOTAL": FMT_MONTANT,
        "Montant d'exposition au HB": FMT_MONTANT,
        "Provisions": FMT_MONTANT,
    }

    ws = wb.create_sheet("Template données")
    ecrire_entetes(ws, COLONNES_TEMPLATE, ligne=1,
                   obligatoires=COLONNES_OBLIGATOIRES_TEMPLATE)
    ecrire_lignes(ws, template, COLONNES_TEMPLATE, 2, formats_template)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(COLONNES_TEMPLATE)).column_letter}{len(template) + 1}"

    ws_nf = wb.create_sheet("CRM_non_financee")
    ecrire_entetes(ws_nf, COLONNES_CRM_NON_FIN, ligne=1)
    ecrire_lignes(ws_nf, crm_non_fin, COLONNES_CRM_NON_FIN, 2,
                  {"Part couverte": FMT_MONTANT})
    ws_nf.freeze_panes = "B2"

    ws_f = wb.create_sheet("CRM_financée")
    ecrire_entetes(ws_f, COLONNES_CRM_FIN, ligne=1,
                   obligatoires=COLONNES_CRM_FIN_OBLIGATOIRES)
    ecrire_lignes(ws_f, crm_fin, COLONNES_CRM_FIN, 2,
                  {"Valeur_Collatéral": FMT_MONTANT})
    ws_f.freeze_panes = "B2"

    feuille_notice(
        wb,
        f"Modèle d'import — Risque de crédit — {BANQUE}",
        [
            ("Date d'arrêté", f"Toutes les expositions sont observées au {DATE_ANALYSE}."),
            ("Template données",
             f"{len(template)} expositions. Les 10 premières colonnes (fond bleu) sont "
             "obligatoires, les 30 suivantes (fond gris) ne concernent que certaines "
             "catégories et peuvent rester vides."),
            ("CRM_non_financee",
             f"{len(crm_non_fin)} garanties personnelles. Chaque ID_Exposition y figurant "
             "porte Type_CRM = « CRM non financee » dans Template données."),
            ("CRM_financée",
             f"{len(crm_fin)} sûretés financières. Chaque ID_Exposition y figurant porte "
             "Type_CRM = « CRM financee » dans Template données."),
            ("Colonnes calculées",
             "Aucune pondération, EAD, RWA ni capital n'est saisi : l'outil recalcule "
             "tout à l'import avec le moteur prudentiel de la saisie manuelle."),
            ("Mode d'import",
             "« Remplacer la base » écrase les expositions existantes ; « Compléter la "
             "base » ajoute les nouvelles et met à jour les ID déjà connus."),
            ("Notice",
             "Cette feuille est purement informative : l'import ne lit que les trois "
             "feuilles ci-dessus."),
        ],
    )

    wb.save(chemin)
    return chemin
