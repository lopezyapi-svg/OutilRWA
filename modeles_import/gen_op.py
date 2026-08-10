# -*- coding: utf-8 -*-
"""Modèle d'import « Risque opérationnel » — pertes + indicateur d'activité.

Un seul classeur alimente les deux imports du module :
  • Incidents  — 1 000 pertes opérationnelles (import « Pertes »)
  • 2023 / 2024 / 2025 — un onglet par exercice, colonnes Poste / Valeur
    (import « BIC / CCR3 »)
  • Notice — feuille d'aide, ignorée par les deux imports

Les deux dialogues d'import ignorent les feuilles qui ne les concernent pas :
celui des pertes ne lit que la feuille contenant « incident », celui du BIC ne
lit que les onglets dont le nom porte une année.
"""

from __future__ import annotations

import random
from datetime import date, timedelta

from openpyxl import Workbook

from _referentiels import BANQUE, BIC_POSTES
from _styles import (
    BLUE_LIGHT, BORDER, FMT_MONTANT, NAVY_LIGHT, ecrire_entetes, ecrire_lignes,
    feuille_notice, fill, center, left, titre_bandeau,
)
from openpyxl.styles import Font

COLONNES_INCIDENTS = [
    "date_occurrence", "description", "ligne_metier", "type_evenement",
    "cause_racine", "perte_brute", "perte_recuperee", "statut",
]
COLONNES_INCIDENTS_OBLIGATOIRES = [
    "date_occurrence", "description", "ligne_metier", "type_evenement",
    "perte_brute",
]

NB_INCIDENTS = 1000
DEBUT = date(2023, 1, 2)
FIN = date(2026, 6, 30)

ANNEES_BIC = (2023, 2024, 2025)

# ── Catalogue d'incidents : (ligne métier, type, cause, gabarit, échelle) ────
# L'échelle pilote l'ordre de grandeur de la perte brute (loi log-normale) :
# beaucoup d'incidents modestes, quelques sinistres majeurs.
_CATALOGUE = [
    ("Banque de détail", "Externe", "Fraude externe",
     "Fraude carte bancaire — {n} transactions non autorisées sur automate", 6.0, 0.9),
    ("Banque de détail", "Externe", "Fraude externe",
     "Retrait frauduleux au guichet — usurpation d'identité agence {ville}", 6.2, 0.8),
    ("Banque de détail", "Personnel", "Erreur humaine",
     "Erreur de caisse — écart de fin de journée agence {ville}", 5.4, 0.7),
    ("Banque de détail", "Processus", "Processus inadéquat",
     "Double débit clientèle — anomalie de traitement des prélèvements", 5.8, 0.8),
    ("Banque de détail", "Externe", "Événement externe",
     "Vol par effraction — distributeur automatique agence {ville}", 6.6, 0.7),
    ("Banque de détail", "Système", "Défaillance système",
     "Indisponibilité du réseau monétique {n} heures", 6.3, 0.8),
    ("Banque commerciale", "Interne", "Fraude interne",
     "Détournement par un chargé de clientèle — comptes dormants", 7.1, 1.0),
    ("Banque commerciale", "Processus", "Processus inadéquat",
     "Garantie non enregistrée — perte de sûreté sur dossier crédit", 7.0, 0.9),
    ("Banque commerciale", "Juridique", "Processus inadéquat",
     "Condamnation — défaut d'information précontractuelle", 7.3, 0.9),
    ("Banque commerciale", "Personnel", "Erreur humaine",
     "Décaissement erroné — mauvais compte bénéficiaire", 6.8, 0.9),
    ("Paiements et règlements", "Système", "Défaillance système",
     "Interruption du système de compensation — {n} opérations en attente", 6.9, 0.9),
    ("Paiements et règlements", "Processus", "Erreur humaine",
     "Virement international erroné — code SWIFT incorrect", 6.7, 0.9),
    ("Paiements et règlements", "Externe", "Fraude externe",
     "Fraude au virement — usurpation d'identité fournisseur", 7.2, 1.0),
    ("Paiements et règlements", "Système", "Défaillance système",
     "Rejets massifs de prélèvements — incompatibilité de format", 6.1, 0.8),
    ("Activités de marché", "Processus", "Erreur humaine",
     "Erreur de saisie d'ordre — quantité exécutée erronée", 7.0, 1.0),
    ("Activités de marché", "Système", "Défaillance système",
     "Défaut de valorisation — flux de cours indisponible {n} jours", 6.6, 0.9),
    ("Activités de marché", "Processus", "Processus inadéquat",
     "Dépassement de limite non détecté — position dénouée à perte", 7.4, 1.0),
    ("Financement d'entreprise", "Juridique", "Processus inadéquat",
     "Litige client — devoir de conseil sur montage structuré", 7.5, 1.0),
    ("Financement d'entreprise", "Processus", "Processus inadéquat",
     "Documentation de sûreté incomplète — recours limité", 7.2, 0.9),
    ("Financement d'entreprise", "Interne", "Fraude interne",
     "Contournement du circuit de délégation — engagement non autorisé", 7.6, 1.0),
    ("Fonctions d'agent", "Processus", "Erreur humaine",
     "Retard de règlement-livraison — pénalité contractuelle", 6.2, 0.8),
    ("Fonctions d'agent", "Processus", "Processus inadéquat",
     "Erreur de conservation de titres — réconciliation tardive", 6.4, 0.8),
    ("Gestion d'actifs", "Processus", "Erreur humaine",
     "Erreur de calcul de valeur liquidative — indemnisation porteurs", 6.8, 0.9),
    ("Gestion d'actifs", "Juridique", "Processus inadéquat",
     "Non-respect du ratio de dispersion d'un OPCVM", 6.5, 0.8),
    ("Courtage de détail", "Processus", "Erreur humaine",
     "Ordre client exécuté hors instruction — rachat de position", 6.3, 0.9),
    ("Courtage de détail", "Externe", "Fraude externe",
     "Compromission d'accès client — ordres frauduleux", 6.6, 0.9),
    ("Banque de détail", "Personnel", "Événement externe",
     "Accident du travail — arrêt d'activité agence {ville}", 5.9, 0.7),
    ("Banque commerciale", "Externe", "Événement externe",
     "Dégât des eaux — remise en état agence {ville}", 6.4, 0.8),
    ("Banque de détail", "Système", "Défaillance système",
     "Panne du core banking — {n} heures d'indisponibilité", 6.7, 0.9),
    ("Banque commerciale", "Interne", "Fraude interne",
     "Falsification de pièces justificatives — dossier de crédit", 7.2, 0.9),
    ("Paiements et règlements", "Juridique", "Processus inadéquat",
     "Sanction pour retard de déclaration réglementaire", 6.5, 0.8),
    ("Activités de marché", "Externe", "Événement externe",
     "Défaut de contrepartie sur opération de pension livrée", 7.7, 1.0),
]

_VILLES = [
    "Abidjan Plateau", "Yopougon", "Cocody", "Bouaké", "San-Pédro", "Korhogo",
    "Dakar Centre", "Thiès", "Saint-Louis", "Cotonou", "Porto-Novo",
    "Ouagadougou", "Bobo-Dioulasso", "Bamako", "Ségou", "Lomé", "Kara",
    "Niamey", "Maradi", "Bissau",
]

_STATUTS_PAR_ANCIENNETE = [
    (0, 60, ["Ouvert", "En cours", "En cours"]),
    (60, 240, ["En cours", "Résolu", "Résolu"]),
    (240, 10_000, ["Résolu", "Clôturé", "Clôturé", "Clôturé"]),
]


def construire_incidents(graine: int = 20260630):
    import math

    rng = random.Random(graine)
    total_jours = (FIN - DEBUT).days
    lignes = []

    for numero in range(1, NB_INCIDENTS + 1):
        ligne_metier, type_evt, cause, gabarit, echelle, sigma = (
            _CATALOGUE[rng.randrange(len(_CATALOGUE))]
        )
        # Volumétrie croissante : la collecte des incidents se densifie au fil
        # des exercices, comme dans un dispositif qui monte en maturité.
        position = rng.random() ** 0.72
        jour = DEBUT + timedelta(days=int(position * total_jours))

        description = gabarit.format(
            n=rng.randint(2, 48), ville=rng.choice(_VILLES)
        )

        perte_brute = math.exp(rng.gauss(echelle * 2.302585, sigma))
        perte_brute = float(int(round(perte_brute / 1000)) * 1000)
        if perte_brute < 25_000:
            perte_brute = 25_000.0

        # Récupération (assurance, recouvrement) sur une minorité de dossiers.
        if rng.random() < 0.38:
            perte_recuperee = float(
                int(round(perte_brute * rng.uniform(0.05, 0.85) / 1000)) * 1000
            )
        else:
            perte_recuperee = 0.0

        anciennete = (FIN - jour).days
        for borne_min, borne_max, statuts in _STATUTS_PAR_ANCIENNETE:
            if borne_min <= anciennete < borne_max:
                statut = rng.choice(statuts)
                break
        else:
            statut = "Clôturé"

        lignes.append({
            "date_occurrence": jour.isoformat(),
            "description": description,
            "ligne_metier": ligne_metier,
            "type_evenement": type_evt,
            "cause_racine": cause,
            "perte_brute": perte_brute,
            "perte_recuperee": perte_recuperee,
            "statut": statut,
        })

    lignes.sort(key=lambda item: item["date_occurrence"])
    return lignes


# ── Postes BIC / CCR3 par exercice (FCFA) ───────────────────────────────────
# Trajectoire d'une banque universelle de l'UMOA d'environ 1 200 Md de total
# de bilan, en croissance régulière.
BIC_PAR_ANNEE = {
    2023: {
        "Intérêts perçus": 79_400_000_000,
        "Intérêts versés": 29_800_000_000,
        "Dividendes perçus": 940_000_000,
        "Trésorerie & Banques centrales": 128_500_000_000,
        "Créances sur Étab. de crédit": 76_200_000_000,
        "Créances clientèle (brut)": 812_000_000_000,
        "Provisions sur créances": 52_400_000_000,
        "Autres produits d'exploitation": 5_600_000_000,
        "Autres charges d'exploitation": 2_700_000_000,
        "Commissions perçues": 17_900_000_000,
        "Commissions versées": 4_600_000_000,
        "Résultat net Ptf négociation": 2_900_000_000,
        "Résultat net Ptf bancaire": 1_700_000_000,
    },
    2024: {
        "Intérêts perçus": 85_900_000_000,
        "Intérêts versés": 32_100_000_000,
        "Dividendes perçus": 1_080_000_000,
        "Trésorerie & Banques centrales": 137_300_000_000,
        "Créances sur Étab. de crédit": 82_600_000_000,
        "Créances clientèle (brut)": 858_000_000_000,
        "Provisions sur créances": 57_100_000_000,
        "Autres produits d'exploitation": 6_100_000_000,
        "Autres charges d'exploitation": 2_950_000_000,
        "Commissions perçues": 19_600_000_000,
        "Commissions versées": 5_050_000_000,
        "Résultat net Ptf négociation": 3_400_000_000,
        "Résultat net Ptf bancaire": 1_950_000_000,
    },
    2025: {
        "Intérêts perçus": 92_300_000_000,
        "Intérêts versés": 34_400_000_000,
        "Dividendes perçus": 1_210_000_000,
        "Trésorerie & Banques centrales": 145_100_000_000,
        "Créances sur Étab. de crédit": 88_400_000_000,
        "Créances clientèle (brut)": 902_000_000_000,
        "Provisions sur créances": 61_300_000_000,
        "Autres produits d'exploitation": 6_550_000_000,
        "Autres charges d'exploitation": 3_120_000_000,
        "Commissions perçues": 21_200_000_000,
        "Commissions versées": 5_420_000_000,
        "Résultat net Ptf négociation": 3_820_000_000,
        "Résultat net Ptf bancaire": 2_240_000_000,
    },
}


def _pnb(postes: dict) -> float:
    """PNB reconstitué à partir des postes saisis (marge d'intérêt, commissions
    nettes, résultats de portefeuille et dividendes)."""
    return (
        postes["Intérêts perçus"] - postes["Intérêts versés"]
        + postes["Commissions perçues"] - postes["Commissions versées"]
        + postes["Résultat net Ptf négociation"]
        + postes["Résultat net Ptf bancaire"]
        + postes["Dividendes perçus"]
        + postes["Autres produits d'exploitation"]
        - postes["Autres charges d'exploitation"]
    )


def calculer_bic(parametres=None) -> dict:
    """Reproduit `calcul_bic` du backend (CRR3, ILM = 1) sur BIC_PAR_ANNEE."""
    seuil_ildc = 0.0225
    seuil1 = 655_957_000_000.0
    seuil2 = 19_678_710_000_000.0
    c1, c2, c3 = 0.12, 0.15, 0.18
    multiplicateur = 12.5

    def moyenne(poste: str) -> float:
        return sum(BIC_PAR_ANNEE[annee][poste] for annee in ANNEES_BIC) / 3

    ic = moyenne("Intérêts perçus") - moyenne("Intérêts versés")
    ac = (
        moyenne("Trésorerie & Banques centrales")
        + moyenne("Créances sur Étab. de crédit")
        + (moyenne("Créances clientèle (brut)") - moyenne("Provisions sur créances"))
    )
    ildc = min(abs(ic), ac * seuil_ildc) + moyenne("Dividendes perçus")
    sc = (
        max(moyenne("Autres produits d'exploitation"), moyenne("Autres charges d'exploitation"))
        + max(moyenne("Commissions perçues"), moyenne("Commissions versées"))
    )
    fc = abs(moyenne("Résultat net Ptf négociation")) + abs(moyenne("Résultat net Ptf bancaire"))
    bi = ildc + sc + fc

    if bi <= seuil1:
        bic = bi * c1
        tranche = 1
    elif bi <= seuil2:
        bic = seuil1 * c1 + (bi - seuil1) * c2
        tranche = 2
    else:
        bic = seuil1 * c1 + (seuil2 - seuil1) * c2 + (bi - seuil2) * c3
        tranche = 3

    return {
        "ildc": ildc, "sc": sc, "fc": fc, "bi": bi, "bic": bic,
        "tranche": tranche, "ofr_crr3": bic, "rea_crr3": bic * multiplicateur,
    }


def _feuille_bic(wb, annee: int) -> None:
    ws = wb.create_sheet(str(annee))
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 22
    titre_bandeau(ws, f"Exercice {annee} — Indicateur d'activité BIC / CCR3", 2)

    ws.row_dimensions[2].height = 20
    for col_index, label in ((1, "Poste"), (2, "Valeur (FCFA)")):
        cell = ws.cell(row=2, column=col_index, value=label)
        cell.font = Font(bold=True, size=10, color=NAVY_LIGHT)
        cell.fill = fill(BLUE_LIGHT)
        cell.border = BORDER
        cell.alignment = center()

    postes = BIC_PAR_ANNEE[annee]
    valeurs = dict(postes)
    valeurs["PNB (BIA — si non calculé automatiquement)"] = _pnb(postes)

    for offset, poste in enumerate(BIC_POSTES):
        ligne = 3 + offset
        c_label = ws.cell(row=ligne, column=1, value=poste)
        c_label.font = Font(size=10)
        c_label.border = BORDER
        c_label.alignment = left(wrap=True)
        c_valeur = ws.cell(row=ligne, column=2, value=float(valeurs[poste]))
        c_valeur.font = Font(size=10)
        c_valeur.border = BORDER
        c_valeur.alignment = center()
        c_valeur.number_format = FMT_MONTANT
    ws.freeze_panes = "A3"


def construire_classeur(chemin, incidents):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Incidents")
    ecrire_entetes(ws, COLONNES_INCIDENTS, ligne=1,
                   obligatoires=COLONNES_INCIDENTS_OBLIGATOIRES)
    ws.column_dimensions["B"].width = 62
    ecrire_lignes(ws, incidents, COLONNES_INCIDENTS, 2,
                  {"perte_brute": FMT_MONTANT, "perte_recuperee": FMT_MONTANT})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(incidents) + 1}"

    for annee in ANNEES_BIC:
        _feuille_bic(wb, annee)

    resultat = calculer_bic()
    perte_brute_totale = sum(item["perte_brute"] for item in incidents)
    perte_nette_totale = sum(
        item["perte_brute"] - item["perte_recuperee"] for item in incidents
    )

    feuille_notice(
        wb,
        f"Modèle d'import — Risque opérationnel — {BANQUE}",
        [
            ("Deux imports, un seul fichier",
             "La feuille « Incidents » alimente l'import des pertes ; les onglets "
             "2023, 2024 et 2025 alimentent l'import BIC / CCR3. Chaque dialogue "
             "ignore les feuilles qui ne le concernent pas."),
            ("Incidents",
             f"{len(incidents)} pertes du {DEBUT.isoformat()} au {FIN.isoformat()} — "
             f"perte brute cumulée {perte_brute_totale / 1e9:,.2f} Md FCFA, perte "
             f"nette {perte_nette_totale / 1e9:,.2f} Md FCFA."),
            ("Colonnes obligatoires",
             "date_occurrence, description, ligne_metier, type_evenement et "
             "perte_brute (strictement positive). cause_racine, perte_recuperee "
             "et statut sont optionnels."),
            ("Onglets d'exercice",
             "Le nom de l'onglet porte l'année : pour ajouter un exercice, "
             "dupliquer un onglet et le renommer. Un onglet sans année "
             "reconnaissable est ignoré."),
            ("Indicateur d'activité",
             f"BI = {resultat['bi'] / 1e9:,.2f} Md FCFA (tranche {resultat['tranche']}) "
             f"→ BIC = {resultat['bic'] / 1e9:,.2f} Md FCFA."),
            ("Notice",
             "Feuille informative : elle ne porte pas d'année et n'est donc lue "
             "par aucun des deux imports."),
        ],
    )

    wb.save(chemin)
    return chemin, resultat
