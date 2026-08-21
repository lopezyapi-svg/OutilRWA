# -*- coding: utf-8 -*-
"""Mise en forme commune aux quatre modèles d'import (registre institutionnel)."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "0F1B3D"
NAVY_LIGHT = "1D4ED8"
BLUE_LIGHT = "DBEAFE"
GREY_HEADER = "F1F5F9"
GREY_ROW = "F8FAFC"
BORDER_CLR = "CBD5E1"

_thin = Side(style="thin", color=BORDER_CLR)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

FMT_MONTANT = "# ##0"
FMT_MONTANT_2 = "# ##0.00"
FMT_TAUX = "0.000"
FMT_DATE = "yyyy-mm-dd"


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def center(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def left(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def titre_bandeau(ws, texte: str, derniere_colonne: int, hauteur: int = 30) -> None:
    """Bandeau de titre fusionné sur la ligne 1."""
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=max(1, derniere_colonne)
    )
    cell = ws.cell(row=1, column=1, value=texte)
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = fill(NAVY)
    cell.alignment = center()
    ws.row_dimensions[1].height = hauteur


def ecrire_entetes(ws, entetes, ligne: int = 1, obligatoires=None) -> None:
    """Écrit la ligne d'en-tête. Les colonnes obligatoires sont en bleu,
    les optionnelles en gris - même code couleur que le modèle de l'outil."""
    obligatoires = set(obligatoires or entetes)
    ws.row_dimensions[ligne].height = 30
    for index, entete in enumerate(entetes, start=1):
        est_obligatoire = entete in obligatoires
        cell = ws.cell(row=ligne, column=index, value=entete)
        cell.font = Font(
            bold=True, size=9.5, color=NAVY_LIGHT if est_obligatoire else "475569"
        )
        cell.fill = fill(BLUE_LIGHT if est_obligatoire else GREY_HEADER)
        cell.border = BORDER
        cell.alignment = center(wrap=True)
        largeur = max(13, min(34, len(str(entete)) + 3))
        ws.column_dimensions[get_column_letter(index)].width = largeur


def ecrire_lignes(ws, lignes, entetes, premiere_ligne: int, formats=None) -> None:
    """Écrit les lignes de données (liste de dicts) et applique les formats."""
    formats = formats or {}
    for offset, valeurs in enumerate(lignes):
        ligne = premiere_ligne + offset
        for index, entete in enumerate(entetes, start=1):
            valeur = valeurs.get(entete)
            cell = ws.cell(row=ligne, column=index, value=valeur)
            cell.font = Font(size=9.5)
            cell.border = BORDER
            cell.alignment = left() if isinstance(valeur, str) else center()
            fmt = formats.get(entete)
            if fmt and isinstance(valeur, (int, float)):
                cell.number_format = fmt


def feuille_notice(wb, titre: str, sections, largeur_label: int = 28,
                   largeur_texte: int = 105):
    """Feuille d'aide (non lue par l'import) décrivant le contenu du classeur."""
    ws = wb.create_sheet("Notice")
    ws.column_dimensions["A"].width = largeur_label
    ws.column_dimensions["B"].width = largeur_texte
    titre_bandeau(ws, titre, 2, hauteur=32)
    for offset, (label, texte) in enumerate(sections):
        ligne = 2 + offset
        c_label = ws.cell(row=ligne, column=1, value=label)
        c_label.font = Font(bold=True, size=10, color=NAVY_LIGHT)
        c_label.fill = fill(BLUE_LIGHT)
        c_label.border = BORDER
        c_label.alignment = left(wrap=True)
        c_texte = ws.cell(row=ligne, column=2, value=texte)
        c_texte.font = Font(size=10)
        c_texte.border = BORDER
        c_texte.alignment = left(wrap=True)
        ws.row_dimensions[ligne].height = max(18, 15 * (1 + len(texte) // 95))
    return ws
