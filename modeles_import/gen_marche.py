# -*- coding: utf-8 -*-
"""Modèle d'import « Risque de marché » - portefeuille titres complet.

Structure produite (noms de feuilles imposés par l'import de l'outil) :
  • Saisir donnée - portefeuille obligataire (30 colonnes obligatoires)
  • Actions       - portefeuille actions (15 colonnes obligatoires)
  • Notice        - feuille d'aide, ignorée à l'import

Les deux feuilles étant présentes, l'écran « Importer données de marché »
reconnaît automatiquement le périmètre « Portefeuille complet ».
"""

from __future__ import annotations

import random
from datetime import date, timedelta

from openpyxl import Workbook

from _referentiels import (
    BANQUE, DATE_ANALYSE, ENTETES_ACTIONS, ENTETES_OBLIGATIONS, TAUX_XOF,
)
from _styles import (
    FMT_MONTANT, FMT_MONTANT_2, FMT_TAUX, ecrire_entetes, ecrire_lignes,
    feuille_notice,
)

FEUILLE_OBLIGATIONS = "Saisir donnée"
FEUILLE_ACTIONS = "Actions"

NB_OBLIGATIONS = 700
NB_ACTIONS = 300

DATE_REF = date(2026, 6, 30)

# ── Émetteurs obligataires ──────────────────────────────────────────────────
# (libellé, pays, zone, code ISO3, notation S&P/Moody's/Fitch, souverain ?)
_EMETTEURS_SOUVERAINS = [
    ("État de Côte d'Ivoire", "Cote d'Ivoire", "UEMOA", "CIV", ("BB", "Ba2", "BB-")),
    ("État du Sénégal", "Senegal", "UEMOA", "SEN", ("B+", "B1", "B+")),
    ("État du Bénin", "Benin", "UEMOA", "BEN", ("BB-", "Ba3", "BB-")),
    ("État du Burkina Faso", "Burkina Faso", "UEMOA", "BFA", ("Non noté", "Non noté", "Non noté")),
    ("État du Mali", "Mali", "UEMOA", "MLI", ("Non noté", "Non noté", "Non noté")),
    ("État du Niger", "Niger", "UEMOA", "NER", ("Non noté", "Non noté", "Non noté")),
    ("État du Togo", "Togo", "UEMOA", "TGO", ("B", "B3", "B")),
    ("État de Guinée-Bissau", "Guinee-Bissau", "UEMOA", "GNB", ("Non noté", "Non noté", "Non noté")),
]

_EMETTEURS_MULTILATERAUX = [
    ("Banque Ouest-Africaine de Développement", "Togo", "UEMOA", "TGO", ("BBB", "Baa1", "BBB")),
    ("Banque Africaine de Développement", "Cote d'Ivoire", "Afrique", "CIV", ("AAA", "Aaa", "AAA")),
    ("Banque Islamique de Développement", "Arabie saoudite", "Hors zone", "SAU", ("AAA", "Aaa", "AAA")),
]

_EMETTEURS_CORPORATE = [
    ("Compagnie Sahel Énergie SA", "Cote d'Ivoire", "UEMOA", "CIV", ("BB-", "Ba3", "BB-")),
    ("Groupe Bandama Agro-Industriel SA", "Cote d'Ivoire", "UEMOA", "CIV", ("B+", "B1", "B+")),
    ("Société Nokoué Cimentière SA", "Benin", "UEMOA", "BEN", ("Non noté", "Non noté", "Non noté")),
    ("Union Bancaire du Sahel", "Burkina Faso", "UEMOA", "BFA", ("BB", "Ba2", "BB")),
    ("Banque Régionale du Littoral", "Senegal", "UEMOA", "SEN", ("BB-", "Ba3", "BB-")),
    ("Comoé Télécom SA", "Cote d'Ivoire", "UEMOA", "CIV", ("BBB-", "Baa3", "BBB-")),
    ("Faso Logistique et Transit SA", "Burkina Faso", "UEMOA", "BFA", ("Non noté", "Non noté", "Non noté")),
    ("Djoliba Immobilier SA", "Mali", "UEMOA", "MLI", ("B", "B2", "B")),
    ("Wouri Distribution SA", "Cameroun", "CEMAC", "CMR", ("B-", "B3", "B-")),
    ("Atlantis Shipping Holding SA", "France", "Europe", "FRA", ("BBB", "Baa2", "BBB")),
]

_TYPES_SOUVERAIN = [
    ("Bon du Trésor", "BT", 0.25, 1.0, True),
    ("Bon Assimilable du Trésor", "BAT", 0.5, 2.0, True),
    ("Obligation du Trésor", "OT", 3.0, 10.0, False),
    ("Obligation Assimilable du Trésor", "OAT", 5.0, 15.0, False),
    ("Titre Public Émis par Adjudication", "TPA", 3.0, 7.0, False),
]

_TYPES_CORPORATE = [
    ("Obligation d'entreprise", "OBC", 3.0, 10.0, False),
    ("Emprunt obligataire coté BRVM", "EOB", 5.0, 12.0, False),
    ("Obligation privée indexée", "OPI", 3.0, 8.0, False),
    ("Billet de trésorerie", "BTR", 0.25, 1.0, True),
]

# Les émetteurs multilatéraux ont leurs propres libellés d'instrument : un
# intitulé contenant « Trésor » les ferait classer à tort en titre souverain
# par la reconnaissance de catégorie de l'outil.
_TYPES_MULTILATERAL = [
    ("Emprunt obligataire multilatéral", "EOM", 5.0, 15.0, False),
    ("Obligation de développement", "ODV", 3.0, 12.0, False),
    ("Sukuk de développement", "SDV", 3.0, 10.0, False),
]

# Part du portefeuille classée en négociation : c'est elle qui détermine
# l'assiette des exigences taux et actions (le portefeuille bancaire en est
# exclu). La banque modélisée est très active sur le marché secondaire
# régional, d'où un portefeuille de transaction volontairement étoffé.
PART_TRADING_OBLIGATIONS = 0.40
PART_TRADING_ACTIONS = 0.55


def _intentions_obligations(part_trading: float):
    reste = 1.0 - part_trading
    return [
        ("HTM - Détenu jusqu'à échéance", reste * 0.72),
        ("AFS - Disponible à la vente", reste * 0.28),
        ("Trading - Portefeuille de négociation", part_trading),
    ]

_MODES_PLACEMENT = [
    "Adjudication BCEAO", "Syndication BRVM", "Placement privé",
    "Marché secondaire", "Appel public à l'épargne",
]

_CLASSIFICATIONS = [
    "Titre de créance négociable", "Titre d'investissement",
    "Titre de transaction", "Titre de placement",
]

_PROFILS = [
    ("In fine", 0.72), ("Amortissement constant", 0.16),
    ("Annuités constantes", 0.08), ("Zéro coupon", 0.04),
]

_FREQUENCES = [
    ("Annuelle", 1), ("Semestrielle", 2), ("Trimestrielle", 4),
]

_NOTATIONS_INTERNES = ["1 - Excellent", "2 - Bon", "3 - Correct", "4 - Sous surveillance", "5 - Sensible"]

# ── Actions ─────────────────────────────────────────────────────────────────
_SOCIETES_ACTIONS = [
    ("Sahel Télécom SA", "Télécommunications", "Cote d'Ivoire", "BRVM"),
    ("Bandama Agro SA", "Agro-alimentaire", "Cote d'Ivoire", "BRVM"),
    ("Comoé Distribution SA", "Distribution", "Cote d'Ivoire", "BRVM"),
    ("Union Bancaire du Sahel", "Banque", "Burkina Faso", "BRVM"),
    ("Banque Régionale du Littoral", "Banque", "Senegal", "BRVM"),
    ("Nokoué Ciments SA", "Matériaux de construction", "Benin", "BRVM"),
    ("Faso Énergie SA", "Énergie", "Burkina Faso", "BRVM"),
    ("Djoliba Assurances SA", "Assurance", "Mali", "BRVM"),
    ("Mono Industries SA", "Industrie", "Togo", "BRVM"),
    ("Kolda Palm Oil SA", "Agro-alimentaire", "Senegal", "BRVM"),
    ("Zambé Transport SA", "Transport", "Cote d'Ivoire", "BRVM"),
    ("Sirba Mining SA", "Mines", "Burkina Faso", "BRVM"),
    ("Ouémé Hôtellerie SA", "Tourisme", "Benin", "BRVM"),
    ("Korhogo Coton SA", "Textile", "Cote d'Ivoire", "BRVM"),
    ("Bagoé Pharma SA", "Santé", "Mali", "BRVM"),
    ("Atlantis Maritime SA", "Transport maritime", "France", "Euronext Paris"),
    ("Wouri Brasseries SA", "Boissons", "Cameroun", "BVMAC"),
    ("Nazinon Immobilier SA", "Immobilier", "Burkina Faso", "BRVM"),
    ("Tiassalé Cacao SA", "Agro-alimentaire", "Cote d'Ivoire", "BRVM"),
    ("Sanaga Utilities SA", "Services aux collectivités", "Cameroun", "BVMAC"),
]

_TYPES_ACTION = [
    ("Action ordinaire", 0.78), ("Action de préférence", 0.10),
    ("Part d'OPCVM actions", 0.08), ("Certificat d'investissement", 0.04),
]


def _tirage_pondere(rng, options):
    valeurs = [option[0] for option in options]
    poids = [option[1] for option in options]
    return rng.choices(valeurs, weights=poids)[0]


def _pire_notation(sp: str, moodys: str, fitch: str) -> str:
    """Retient la notation la plus sévère parmi les trois agences."""
    echelle = [
        "AAA", "AA+", "AA", "AA-", "A+", "A", "A-", "BBB+", "BBB", "BBB-",
        "BB+", "BB", "BB-", "B+", "B", "B-", "< B-",
    ]
    equivalences = {
        "Aaa": "AAA", "Aa1": "AA+", "Aa2": "AA", "Aa3": "AA-",
        "A1": "A+", "A2": "A", "A3": "A-",
        "Baa1": "BBB+", "Baa2": "BBB", "Baa3": "BBB-",
        "Ba1": "BB+", "Ba2": "BB", "Ba3": "BB-",
        "B1": "B+", "B2": "B", "B3": "B-",
    }
    notes = []
    for brute in (sp, moodys, fitch):
        note = equivalences.get(brute, brute)
        if note in echelle:
            notes.append(echelle.index(note))
    if not notes:
        return "Non noté"
    return echelle[max(notes)]


def construire_obligations(
    graine: int = 20260630,
    echelle: float = 1.0,
    part_trading: float = PART_TRADING_OBLIGATIONS,
):
    rng = random.Random(graine)
    intentions = _intentions_obligations(part_trading)
    lignes = []
    for numero in range(1, NB_OBLIGATIONS + 1):
        tirage = rng.random()
        if tirage < 0.68:
            emetteur, pays, zone, iso3, notations = rng.choice(_EMETTEURS_SOUVERAINS)
            type_libelle, code, mini, maxi, court = rng.choices(
                _TYPES_SOUVERAIN, weights=[24, 16, 26, 24, 10]
            )[0]
        elif tirage < 0.78:
            emetteur, pays, zone, iso3, notations = rng.choice(_EMETTEURS_MULTILATERAUX)
            type_libelle, code, mini, maxi, court = rng.choice(_TYPES_MULTILATERAL)
        else:
            emetteur, pays, zone, iso3, notations = rng.choice(_EMETTEURS_CORPORATE)
            type_libelle, code, mini, maxi, court = rng.choices(
                _TYPES_CORPORATE, weights=[40, 30, 15, 15]
            )[0]

        sp, moodys, fitch = notations
        duree_ans = rng.uniform(mini, maxi)
        emission = DATE_REF - timedelta(days=int(rng.uniform(30, min(duree_ans, 8.0) * 365)))
        echeance = emission + timedelta(days=int(duree_ans * 365.25))
        if echeance <= DATE_REF + timedelta(days=15):
            echeance = DATE_REF + timedelta(days=rng.randint(20, 400))

        maturite_mois = max(1, int(round((echeance - emission).days / 30.44)))
        maturite_residuelle = max(0, int(round((echeance - DATE_REF).days / 30.44)))

        est_souverain_uemoa = zone == "UEMOA" and "État" in emetteur
        # La position ouverte en devises d'une banque de l'UMOA reste limitée :
        # les lignes hors zone franc sont peu nombreuses et de taille modeste.
        devise = "XOF" if (zone == "UEMOA" or rng.random() < 0.45) else rng.choice(["EUR", "USD"])

        nominal = 10_000.0 if devise == "XOF" else 1_000.0
        if devise == "XOF":
            quantite = float(rng.choice([2_000, 5_000, 10_000, 20_000, 50_000, 100_000, 250_000]))
        else:
            quantite = float(rng.choice([300, 600, 1_200, 2_500]))
        quantite = float(max(1, round(quantite * echelle)))
        capital_initial = nominal * quantite

        if court:
            coupon = 0.0
            profil = "Zéro coupon"
            frequence_libelle, _ = _FREQUENCES[0]
            prix_emission = round(rng.uniform(94.0, 99.2), 3)
        else:
            base = 5.6 if est_souverain_uemoa else 6.9
            coupon = round(rng.gauss(base, 0.7), 3)
            coupon = float(min(max(coupon, 3.0), 9.5))
            profil = _tirage_pondere(rng, _PROFILS)
            frequence_libelle, _ = rng.choices(_FREQUENCES, weights=[45, 40, 15])[0]
            prix_emission = round(rng.uniform(98.0, 100.5), 3)

        prix_remboursement = round(rng.choice([100.0, 100.0, 100.0, 101.0, 102.0]), 3)
        prime_emission = round((prix_emission - 100.0) / 100.0 * nominal, 2)
        prime_remboursement = round((prix_remboursement - 100.0) / 100.0 * nominal, 2)

        intention = _tirage_pondere(rng, intentions)
        pire = _pire_notation(sp, moodys, fitch)

        lignes.append({
            "ID Titre": f"{code}-{iso3}-{emission.year}-{numero:04d}",
            "Date d'analyse": DATE_ANALYSE,
            "Pays émetteur": pays,
            "Zone": zone,
            "Type d'instrument": type_libelle,
            "Code type d'instrument": code,
            "Emetteur": emetteur,
            "Mode de Placement": rng.choice(_MODES_PLACEMENT),
            "Notation externe_S&P": sp,
            "Notation externe_Moody's": moodys,
            "Notation externe_Fitch": fitch,
            "La pire notation externe": pire,
            "Notation Interne": rng.choice(_NOTATIONS_INTERNES),
            "Intention comptable": intention,
            "Classification des titres": (
                "Titre de transaction" if intention.startswith("Trading")
                else "Titre d'investissement" if intention.startswith("HTM")
                else rng.choice(_CLASSIFICATIONS)
            ),
            "Date d'émission": emission.isoformat(),
            "Devise": devise,
            "Valeur nominale unitaire": nominal,
            "quantités": quantite,
            "Capital initial": capital_initial,
            "Prix d'émission": prix_emission,
            "Prime d'émission": prime_emission,
            "Prix de remboursement": prix_remboursement,
            "Prime de remboursement": prime_remboursement,
            "Date d'échéance": echeance.isoformat(),
            "Coupon (%)": coupon,
            "Profil d'amortissement": profil,
            "Fréquence de paiement des intérêts": frequence_libelle,
            "Maturité (mois)": maturite_mois,
            "Maturité résiduelle (mois)": maturite_residuelle,
        })
    return lignes


def construire_actions(
    graine: int = 20260701,
    echelle: float = 1.0,
    part_trading: float = PART_TRADING_ACTIONS,
):
    rng = random.Random(graine)
    # Le caractère « liquide et bien diversifié » (qui ramène le risque
    # spécifique de 8 % à 4 %) s'apprécie par ligne de titres, pas ligne à
    # ligne : il est donc fixé une fois pour toutes par société.
    liquides = {
        societe: (index % 3 == 0)
        for index, (societe, *_reste) in enumerate(_SOCIETES_ACTIONS)
    }
    lignes = []
    for numero in range(1, NB_ACTIONS + 1):
        societe, secteur, pays, bourse = rng.choice(_SOCIETES_ACTIONS)
        devise = "XOF" if bourse == "BRVM" else ("EUR" if bourse == "Euronext Paris" else "XAF")
        if devise == "XAF":  # non géré par le convertisseur : ramené au XOF
            devise = "XOF"
        type_instrument = _tirage_pondere(rng, _TYPES_ACTION)
        # Une action ne peut pas être « détenue jusqu'à l'échéance » : hors
        # négociation, elle est classée en titres disponibles à la vente ou en
        # titres de participation, tous deux au portefeuille bancaire.
        intention = _tirage_pondere(rng, [
            ("Trading - Portefeuille de négociation", part_trading),
            ("AFS - Disponible à la vente", (1.0 - part_trading) * 0.74),
            ("Titres de participation - portefeuille bancaire", (1.0 - part_trading) * 0.26),
        ])

        prix_acquisition = round(rng.uniform(1_200, 24_000) if devise == "XOF"
                                 else rng.uniform(8, 220), 2)
        variation = rng.gauss(0.04, 0.19)
        cours_actuel = round(max(prix_acquisition * (1 + variation), prix_acquisition * 0.25), 2)
        quantite = float(rng.choice([500, 1_000, 1_800, 3_000, 5_000, 9_000, 18_000, 35_000]))
        quantite = float(max(1, round(quantite * echelle)))

        lignes.append({
            "ID Instrument": f"ACT-{societe[:3].upper()}-{numero:04d}",
            "Type d'instrument": type_instrument,
            "Émetteur / Société": societe,
            "Secteur": secteur,
            "Pays / marché": pays,
            "Bourse": bourse,
            "Devise": devise,
            "Intention comptable": intention,
            "Quantité": quantite,
            "Prix d'acquisition": prix_acquisition,
            "Cours actuel": cours_actuel,
            "Bêta": round(rng.uniform(0.45, 1.65), 2),
            "Volatilité annualisée (%)": round(rng.uniform(9.0, 42.0), 2),
            "Rendement dividende (%)": round(max(0.0, rng.gauss(5.4, 2.4)), 2),
            "Liquide et diversifié (Oui/Non)": "Oui" if liquides[societe] else "Non",
        })
    return lignes


def construire_classeur(chemin, obligations, actions):
    wb = Workbook()
    wb.remove(wb.active)

    formats_obligations = {
        "Valeur nominale unitaire": FMT_MONTANT,
        "quantités": FMT_MONTANT,
        "Capital initial": FMT_MONTANT,
        "Prix d'émission": FMT_MONTANT_2,
        "Prime d'émission": FMT_MONTANT_2,
        "Prix de remboursement": FMT_MONTANT_2,
        "Prime de remboursement": FMT_MONTANT_2,
        "Coupon (%)": FMT_TAUX,
    }
    ws = wb.create_sheet(FEUILLE_OBLIGATIONS)
    ecrire_entetes(ws, ENTETES_OBLIGATIONS, ligne=1)
    ecrire_lignes(ws, obligations, ENTETES_OBLIGATIONS, 2, formats_obligations)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = (
        f"A1:{ws.cell(row=1, column=len(ENTETES_OBLIGATIONS)).column_letter}"
        f"{len(obligations) + 1}"
    )

    formats_actions = {
        "Quantité": FMT_MONTANT,
        "Prix d'acquisition": FMT_MONTANT_2,
        "Cours actuel": FMT_MONTANT_2,
        "Bêta": "0.00",
        "Volatilité annualisée (%)": "0.00",
        "Rendement dividende (%)": "0.00",
    }
    ws_a = wb.create_sheet(FEUILLE_ACTIONS)
    ecrire_entetes(ws_a, ENTETES_ACTIONS, ligne=1)
    ecrire_lignes(ws_a, actions, ENTETES_ACTIONS, 2, formats_actions)
    ws_a.freeze_panes = "B2"
    ws_a.auto_filter.ref = (
        f"A1:{ws_a.cell(row=1, column=len(ENTETES_ACTIONS)).column_letter}"
        f"{len(actions) + 1}"
    )

    valeur_obligations = sum(
        ligne["Capital initial"] * TAUX_XOF.get(ligne["Devise"], 1.0)
        for ligne in obligations
    )
    valeur_actions = sum(
        ligne["Quantité"] * ligne["Cours actuel"] * TAUX_XOF.get(ligne["Devise"], 1.0)
        for ligne in actions
    )

    feuille_notice(
        wb,
        f"Modèle d'import - Risque de marché - {BANQUE}",
        [
            ("Date d'arrêté", f"Portefeuille observé au {DATE_ANALYSE}."),
            ("Saisir donnée",
             f"{len(obligations)} lignes obligataires - encours nominal "
             f"{valeur_obligations / 1e9:,.1f} Md FCFA. Le nom de la feuille est "
             "imposé par l'outil : ne pas le renommer."),
            ("Actions",
             f"{len(actions)} lignes actions - valorisation "
             f"{valeur_actions / 1e9:,.1f} Md FCFA."),
            ("Périmètre reconnu",
             "Les deux feuilles étant présentes, l'écran « Importer données de "
             "marché » bascule automatiquement sur « Portefeuille complet »."),
            ("Intention comptable",
             "Elle détermine le périmètre prudentiel : seules les lignes de "
             "négociation (Trading) entrent dans l'exigence taux et actions. Les "
             "lignes HTM et AFS relèvent du portefeuille bancaire. Le risque de "
             "change porte, lui, sur l'ensemble des positions."),
            ("Devises",
             "XOF, EUR et USD sont convertis en francs CFA par le convertisseur "
             "de l'outil. Toute autre devise serait traitée au taux 1."),
            ("Notice",
             "Feuille informative : l'import ne lit que « Saisir donnée » et "
             "« Actions »."),
        ],
    )

    wb.save(chemin)
    return chemin, valeur_obligations, valeur_actions
