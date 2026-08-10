# -*- coding: utf-8 -*-
"""Modèle d'import « Fonds propres réglementaires » — CET1 / AT1 / Tier 2.

Une seule photo (pas de dimension exercice) : 11 postes, une ligne par poste.
Les montants sont calibrés sur les RWA réellement produits par les trois
autres modèles, de façon à obtenir un ratio de solvabilité plausible pour une
banque de l'UMOA correctement capitalisée.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

from _referentiels import BANQUE, FP_POSTES
from _styles import (
    BLUE_LIGHT, BORDER, FMT_MONTANT, GREY_ROW, NAVY_LIGHT, center,
    feuille_notice, fill, left, titre_bandeau,
)

FEUILLE = "Fonds propres"

# Structure de fonds propres visée, en part du total (banque universelle de
# l'UMOA : capitalisation très majoritairement CET1).
PART_CET1 = 0.815
PART_AT1 = 0.055
PART_T2 = 0.130

# Ratio de solvabilité global visé. Le minimum réglementaire est de 9 % ;
# porté à 11,5 % avec le coussin de conservation de 2,5 %. La cible retenue
# laisse donc une marge de gestion d'environ 1,5 point au-dessus de l'exigence
# coussin compris, ce qui correspond au niveau d'une banque de l'UMOA
# correctement capitalisée.
RATIO_CIBLE = 0.130


def calibrer(rwa_total: float, ratio_cible: float = RATIO_CIBLE) -> dict:
    """Décompose les fonds propres visés en 11 postes cohérents entre eux."""
    total_vise = rwa_total * ratio_cible

    cet1 = total_vise * PART_CET1
    at1 = total_vise * PART_AT1
    tier2 = total_vise * PART_T2

    def arrondir(valeur: float) -> float:
        return float(round(valeur / 1_000_000) * 1_000_000)

    # ── CET1 : capital + réserves + report + résultat − déductions ──────────
    deductions_cet1 = arrondir(cet1 * 0.085)   # incorporels, impôts différés
    brut_cet1 = cet1 + deductions_cet1
    capital_ordinaire = arrondir(brut_cet1 * 0.52)
    reserves = arrondir(brut_cet1 * 0.27)
    resultats_report = arrondir(brut_cet1 * 0.13)
    resultat_eligible = arrondir(
        brut_cet1 - capital_ordinaire - reserves - resultats_report
    )

    # ── AT1 : instruments + primes − déductions ────────────────────────────
    deductions_at1 = arrondir(at1 * 0.04)
    brut_at1 = at1 + deductions_at1
    instruments_at1 = arrondir(brut_at1 * 0.86)
    primes_emission_at1 = arrondir(brut_at1 - instruments_at1)

    # ── Tier 2 : dettes subordonnées + provisions générales − déductions ────
    deductions_t2 = arrondir(tier2 * 0.03)
    brut_t2 = tier2 + deductions_t2
    dettes_subordonnees_t2 = arrondir(brut_t2 * 0.78)
    provisions_generales_t2 = arrondir(brut_t2 - dettes_subordonnees_t2)

    valeurs = {
        "Capital ordinaire": capital_ordinaire,
        "Réserves": reserves,
        "Résultats en report": resultats_report,
        "Résultat éligible": resultat_eligible,
        "Réduction prudentielle (CET1)": deductions_cet1,
        "Instruments additionnels (AT1)": instruments_at1,
        "Primes d'émission (AT1)": primes_emission_at1,
        "Réduction prudentielle (AT1)": deductions_at1,
        "Dettes subordonnées (Tier 2)": dettes_subordonnees_t2,
        "Provisions générales (Tier 2)": provisions_generales_t2,
        "Réduction prudentielle (Tier 2)": deductions_t2,
    }

    cet1_effectif = (
        valeurs["Capital ordinaire"] + valeurs["Réserves"]
        + valeurs["Résultats en report"] + valeurs["Résultat éligible"]
        - valeurs["Réduction prudentielle (CET1)"]
    )
    at1_effectif = (
        valeurs["Instruments additionnels (AT1)"]
        + valeurs["Primes d'émission (AT1)"]
        - valeurs["Réduction prudentielle (AT1)"]
    )
    t2_effectif = (
        valeurs["Dettes subordonnées (Tier 2)"]
        + valeurs["Provisions générales (Tier 2)"]
        - valeurs["Réduction prudentielle (Tier 2)"]
    )
    total_effectif = cet1_effectif + at1_effectif + t2_effectif

    return {
        "valeurs": valeurs,
        "cet1": cet1_effectif,
        "at1": at1_effectif,
        "tier1": cet1_effectif + at1_effectif,
        "tier2": t2_effectif,
        "total": total_effectif,
        "ratio": total_effectif / rwa_total if rwa_total > 0 else 0.0,
        "ratio_cet1": cet1_effectif / rwa_total if rwa_total > 0 else 0.0,
        "ratio_tier1": (cet1_effectif + at1_effectif) / rwa_total if rwa_total > 0 else 0.0,
    }


def construire_classeur(chemin, calibration: dict, synthese: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = FEUILLE

    titre_bandeau(
        ws,
        "Modèle d'import — Fonds propres réglementaires (CET1 / AT1 / Tier 2)",
        3,
        hauteur=32,
    )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 22
    ws.row_dimensions[2].height = 22
    for col_index, label in ((1, "Groupe"), (2, "Poste"), (3, "Valeur (FCFA)")):
        cell = ws.cell(row=2, column=col_index, value=label)
        cell.font = Font(bold=True, size=10, color=NAVY_LIGHT)
        cell.fill = fill(BLUE_LIGHT)
        cell.border = BORDER
        cell.alignment = center()

    valeurs = calibration["valeurs"]
    for offset, (groupe, poste) in enumerate(FP_POSTES):
        ligne = 3 + offset
        remplissage = fill("FFFFFF") if ligne % 2 == 1 else fill(GREY_ROW)

        c_groupe = ws.cell(row=ligne, column=1, value=groupe)
        c_groupe.font = Font(size=10, bold=True, color=NAVY_LIGHT)
        c_groupe.fill = remplissage
        c_groupe.border = BORDER
        c_groupe.alignment = center()

        c_poste = ws.cell(row=ligne, column=2, value=poste)
        c_poste.font = Font(size=10)
        c_poste.fill = remplissage
        c_poste.border = BORDER
        c_poste.alignment = left(wrap=True)

        c_valeur = ws.cell(row=ligne, column=3, value=float(valeurs[poste]))
        c_valeur.font = Font(size=10)
        c_valeur.fill = remplissage
        c_valeur.border = BORDER
        c_valeur.alignment = center()
        c_valeur.number_format = FMT_MONTANT

    ws.freeze_panes = "A3"

    # Rappel des agrégats, volontairement placé hors des colonnes « Poste » et
    # « Valeur » : l'import lit toutes les lignes sous l'en-tête et signalerait
    # sinon « Total CET1 » comme un poste non reconnu.
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 22
    rappels = [
        ("Total CET1", calibration["cet1"]),
        ("Total AT1", calibration["at1"]),
        ("Total Tier 1", calibration["tier1"]),
        ("Total Tier 2", calibration["tier2"]),
        ("Fonds propres globaux", calibration["total"]),
    ]
    entete_rappel = ws.cell(row=2, column=5, value="Rappel (non importé)")
    entete_rappel.font = Font(bold=True, size=10, color=NAVY_LIGHT)
    entete_rappel.fill = fill(GREY_ROW)
    entete_rappel.alignment = center()
    for offset, (label, valeur) in enumerate(rappels):
        ligne = 3 + offset
        c_label = ws.cell(row=ligne, column=5, value=label)
        c_label.font = Font(size=10, bold=True, color=NAVY_LIGHT)
        c_label.alignment = left()
        c_valeur = ws.cell(row=ligne, column=6, value=float(valeur))
        c_valeur.font = Font(size=10, bold=True)
        c_valeur.number_format = FMT_MONTANT
        c_valeur.alignment = center()

    milliards = lambda montant: f"{montant / 1e9:,.2f} Md FCFA".replace(",", " ")

    feuille_notice(
        wb,
        f"Modèle d'import — Fonds propres réglementaires — {BANQUE}",
        [
            ("Format",
             "Une ligne par poste, trois colonnes : Groupe, Poste, Valeur. "
             "Les 11 libellés de la colonne « Poste » doivent rester intacts : "
             "c'est sur eux que l'import fait la correspondance."),
            ("Portée de l'import",
             "L'import remplace intégralement les fonds propres enregistrés — "
             "il n'y a pas de dimension exercice, c'est une photo unique."),
            ("Calibrage",
             "Les montants sont ajustés sur les RWA produits par les trois "
             "autres modèles de ce dossier, pour un ratio de solvabilité global "
             f"de {calibration['ratio'] * 100:.2f} % (minimum réglementaire 9 %)."),
            ("RWA crédit", milliards(synthese["rwa_credit"])),
            ("RWA marché", milliards(synthese["rwa_marche"])),
            ("RWA opérationnel", milliards(synthese["rwa_operationnel"])),
            ("RWA total", milliards(synthese["rwa_total"])),
            ("Fonds propres globaux", milliards(calibration["total"])),
            ("Ratio CET1", f"{calibration['ratio_cet1'] * 100:.2f} %"),
            ("Ratio Tier 1", f"{calibration['ratio_tier1'] * 100:.2f} %"),
            ("Ratio de solvabilité global", f"{calibration['ratio'] * 100:.2f} %"),
            ("Lignes de rappel",
             "Les totaux affichés sous le tableau sont indicatifs : l'import ne "
             "retient que les lignes dont le libellé correspond à l'un des 11 "
             "postes."),
        ],
    )

    wb.save(chemin)
    return chemin
