# -*- coding: utf-8 -*-
"""Charge les quatre modèles dans la base et affiche le tableau de bord obtenu.

Par défaut le script travaille sur une **copie** de la base (bac à sable) : il
ne modifie rien et sert à vérifier ce que l'application affichera. Avec
`--base-reelle`, il écrit dans la base de l'outil après en avoir fait une
sauvegarde horodatée.

    python modeles_import/charger_dans_la_base.py                # simulation
    python modeles_import/charger_dans_la_base.py --base-reelle  # écriture

Le risque de marché est le seul poste que ce script ne peut pas produire par
le chemin normal : son exigence de fonds propres est calculée côté application
puis publiée par l'écran « Importer données de marché ». Le script écrit donc
l'estimation de `calc_marche.py`, que l'application recalculera — et écrasera —
au premier import du modèle marché.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

DOSSIER = Path(__file__).resolve().parent
RACINE = DOSSIER.parent
sys.path.insert(0, str(DOSSIER))
sys.path.insert(0, str(RACINE / "backend"))

BASE_REELLE = RACINE / "backend" / "data" / "rwa_data.db"


def _preparer_bac_a_sable() -> Path:
    """Redirige la couche d'accès aux données vers une copie de la base."""
    import app.core.runtime_paths as runtime_paths

    bac = DOSSIER / "_bac_a_sable"
    (bac / "data").mkdir(parents=True, exist_ok=True)
    copie = bac / "data" / "rwa_data.db"
    shutil.copy2(BASE_REELLE, copie)
    for suffixe in ("-wal", "-shm"):
        source = BASE_REELLE.with_name(BASE_REELLE.name + suffixe)
        if source.exists():
            shutil.copy2(source, copie.with_name(copie.name + suffixe))

    runtime_paths.backend_source_root = lambda: bac
    return copie


def _sauvegarder_base_reelle() -> Path:
    horodatage = datetime.now().strftime("%Y%m%d_%H%M%S")
    destination = BASE_REELLE.with_name(f"rwa_data.avant_modeles_{horodatage}.db")
    shutil.copy2(BASE_REELLE, destination)
    return destination


def _md(montant: float) -> str:
    return f"{montant / 1e9:>10,.2f} Md".replace(",", " ")


def main() -> int:
    analyseur = argparse.ArgumentParser(description=__doc__)
    analyseur.add_argument(
        "--base-reelle",
        action="store_true",
        help="écrit dans la base de l'outil (sauvegarde automatique) au lieu d'une copie",
    )
    options = analyseur.parse_args()

    if not BASE_REELLE.is_file():
        print(f"Base introuvable : {BASE_REELLE}")
        return 1

    if options.base_reelle:
        sauvegarde = _sauvegarder_base_reelle()
        print(f"Sauvegarde de la base : {sauvegarde.name}")
        cible = BASE_REELLE
    else:
        cible = _preparer_bac_a_sable()
        print("Mode simulation : la base de l'outil n'est pas modifiee.")
    print(f"Base ciblee : {cible}")

    # Les imports doivent venir APRES la redirection eventuelle du chemin.
    from app.dashboard.models import FondsPropresUpdate
    from app.dashboard.services import get_dashboard_snapshot, update_fonds_propres
    from app.market.services import MARKET_CAPITAL_REQUIREMENT_KEY
    from app.risque_operationnel.models import OpRiskInputUpdate
    from app.risque_operationnel.services import upsert_op_risk_input
    from database.connection import database_manager
    from database.services.excel_import_service import excel_import_service

    import calc_marche
    import gen_marche
    import gen_op
    from generer_modeles import FICHIERS
    from openpyxl import load_workbook

    # ── 1. Expositions (chemin d'import reel du backend) ───────────────────
    fichier_credit = FICHIERS["credit"]
    rapport = excel_import_service.import_uploaded_workbook(
        fichier_credit.read_bytes(), fichier_credit.name, mode="replace"
    )
    print(
        f"\n[1] Credit      : {rapport['imported_rows']} expositions importees, "
        f"{rapport['rejected_rows']} rejetee(s)"
    )

    # ── 2. Postes BIC / CCR3 ───────────────────────────────────────────────
    correspondances = {
        "Intérêts perçus": "interets_percus",
        "Intérêts versés": "interets_verses",
        "Dividendes perçus": "dividendes_percus",
        "Trésorerie & Banques centrales": "tresorerie_et_banques_centrales",
        "Créances sur Étab. de crédit": "creances_etablissements_credit",
        "Créances clientèle (brut)": "creances_clientele",
        "Provisions sur créances": "provisions",
        "Autres produits d'exploitation": "autres_produits_exploitation",
        "Autres charges d'exploitation": "autres_charges_exploitation",
        "Commissions perçues": "commissions_percues",
        "Commissions versées": "commissions_versees",
        "Résultat net Ptf négociation": "resultat_portefeuille_negociation",
        "Résultat net Ptf bancaire": "resultat_portefeuille_bancaire",
    }
    for annee in gen_op.ANNEES_BIC:
        postes = gen_op.BIC_PAR_ANNEE[annee]
        charge = {
            champ: float(postes[libelle]) for libelle, champ in correspondances.items()
        }
        upsert_op_risk_input(annee, OpRiskInputUpdate(**charge))
    resultat_bic = gen_op.calculer_bic()
    print(
        f"[2] Operationnel: exercices {list(gen_op.ANNEES_BIC)} enregistres, "
        f"BIC {resultat_bic['bic'] / 1e9:,.2f} Md"
    )

    # ── 3. Exigence de fonds propres marche (estimation, cf. en-tete) ──────
    obligations, actions = _relire_marche(load_workbook, gen_marche, FICHIERS["marche"])
    marche = calc_marche.exigence_totale(obligations, actions)
    with database_manager.transaction() as connexion:
        connexion.execute(
            """
            INSERT INTO metadonnees_app(cle, valeur) VALUES(?, ?)
            ON CONFLICT(cle) DO UPDATE SET valeur = excluded.valeur
            """,
            (
                MARKET_CAPITAL_REQUIREMENT_KEY,
                json.dumps(
                    {
                        "capital_requis": marche["capital_requis"],
                        "rwa_marche": marche["rwa_marche"],
                    }
                ),
            ),
        )
    print(
        f"[3] Marche      : capital requis {marche['capital_requis'] / 1e9:,.2f} Md, "
        f"RWA {marche['rwa_marche'] / 1e9:,.2f} Md (estimation)"
    )

    # ── 4. Fonds propres ───────────────────────────────────────────────────
    valeurs = _lire_fonds_propres(load_workbook, FICHIERS["fonds_propres"])
    update_fonds_propres(FondsPropresUpdate(**valeurs))
    print("[4] Fonds propres: 11 postes enregistres")

    # ── Restitution : ce que l'application affichera ───────────────────────
    instantane = get_dashboard_snapshot()
    metriques = {m.key: m.value for m in instantane.metrics}
    rwa_total = metriques["rwa"]

    print("\n" + "=" * 66)
    print("Tableau de bord obtenu")
    print("=" * 66)
    for cle, libelle in (
        ("rwa_credit", "RWA Credit"),
        ("rwa_market", "RWA Marche"),
        ("rwa_op", "RWA Operationnel"),
    ):
        montant = metriques[cle]
        part = montant / rwa_total * 100 if rwa_total else 0.0
        print(f"  {libelle:<18}{_md(montant)} FCFA{part:>8.1f} %")
    print(f"  {'RWA total':<18}{_md(rwa_total)} FCFA{100.0:>8.1f} %")
    print(f"\n  {'Fonds propres':<18}{_md(metriques['capital'])} FCFA")
    print(f"  {'Ratio CET1':<18}{metriques['cet1_ratio'] * 100:>10.2f} %")
    print(f"  {'Ratio Tier 1':<18}{metriques['tier1_ratio'] * 100:>10.2f} %")
    print(f"  {'Ratio solvabilite':<18}{metriques['solvabilite'] * 100:>10.2f} %")
    print(f"  {'Encours brut':<18}{_md(metriques['encours'])} FCFA")

    if not options.base_reelle:
        print(
            "\nSimulation terminee : relancez avec --base-reelle pour appliquer, "
            "ou importez les quatre fichiers depuis l'application."
        )
    return 0


def _relire_marche(load_workbook, gen_marche, chemin):
    """Relit le classeur marché produit, pour estimer l'exigence sur les
    données réellement livrées plutôt que sur un nouveau tirage."""
    classeur = load_workbook(chemin, data_only=True, read_only=True)
    try:
        obligations = _lignes(classeur, gen_marche.FEUILLE_OBLIGATIONS)
        actions = _lignes(classeur, gen_marche.FEUILLE_ACTIONS)
    finally:
        classeur.close()
    return obligations, actions


def _lignes(classeur, nom_feuille):
    feuille = classeur[nom_feuille]
    iterateur = feuille.iter_rows(values_only=True)
    entetes = [str(valeur).strip() if valeur is not None else "" for valeur in next(iterateur)]
    lignes = []
    for valeurs in iterateur:
        if all(valeur in (None, "") for valeur in valeurs):
            continue
        lignes.append({
            entete: (valeurs[index] if index < len(valeurs) else None)
            for index, entete in enumerate(entetes)
            if entete
        })
    return lignes


def _lire_fonds_propres(load_workbook, chemin):
    champs = [
        "capital_ordinaire", "reserves", "resultats_report", "resultat_eligible",
        "deductions_prud_cet1", "instruments_at1", "primes_emission_at1",
        "deductions_prud_at1", "dettes_subordonnees_t2", "provisions_generales_t2",
        "deductions_prud_t2",
    ]
    classeur = load_workbook(chemin, data_only=True, read_only=True)
    try:
        feuille = classeur["Fonds propres"]
        valeurs = [
            float(ligne[2] or 0)
            for ligne in feuille.iter_rows(min_row=3, max_row=13, values_only=True)
        ]
    finally:
        classeur.close()
    return dict(zip(champs, valeurs))


if __name__ == "__main__":
    raise SystemExit(main())
