from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

def build_fonds_propres_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fonds propres"
    
    # Removing freeze_panes
    # ws.freeze_panes = "A3"
    
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = "Modèle d'import — Fonds Propres Réglementaires"
    
    # Also commenting out row heights to be safe
    # ws.row_dimensions[2].height = 24

    ws["A2"] = "Groupe"
    ws["B2"] = "Poste (ne pas modifier)"
    ws["C2"] = "Valeur (XOF)"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

if __name__ == "__main__":
    with open("test_python.xlsx", "wb") as f:
        f.write(build_fonds_propres_import_template())
