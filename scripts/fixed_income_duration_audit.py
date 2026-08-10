#!/usr/bin/env python3
"""Audit et recalcul professionnel des durations obligataires.

Le script lit un classeur Excel de portefeuille obligataire, reconstruit les
cash flows, recalcule prix, YTM, duration de Macaulay, duration modifiee,
convexite, poids de portefeuille et controles de coherence, puis genere un
classeur final et un rapport methodologique.

Exemple:
    python scripts/fixed_income_duration_audit.py ^
        --input "Base_GO_modifie (1).xlsx" ^
        --output "Base_GO_CORRIGE_FINAL.xlsx" ^
        --report "rapport_methodologique_duration_obligataire.md"
"""

from __future__ import annotations

import argparse
import math
import re
import tempfile
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from scipy.optimize import brentq as scipy_brentq
except Exception:  # pragma: no cover - fallback si scipy n'est pas installe.
    scipy_brentq = None


DEFAULT_INPUT = "Base_GO_modifie (1).xlsx"
DEFAULT_OUTPUT = "Base_GO_CORRIGE_FINAL.xlsx"
DEFAULT_REPORT = "rapport_methodologique_duration_obligataire.md"

THEME = {
    "navy": "13203A",
    "blue": "2563EB",
    "cyan": "06B6D4",
    "green": "10B981",
    "amber": "F59E0B",
    "red": "EF4444",
    "violet": "7C3AED",
    "border": "DDE7F5",
    "soft": "F8FAFC",
    "white": "FFFFFF",
}


ALIASES: dict[str, list[str]] = {
    "instrument_id": [
        "isin",
        "code isin",
        "code titre",
        "code instrument",
        "instrument code",
        "security id",
        "identifiant",
        "reference",
        "ref",
        "id",
    ],
    "issuer": ["emetteur", "issuer", "contrepartie", "debiteur", "nom emetteur"],
    "nominal": [
        "nominal",
        "valeur nominale",
        "face value",
        "par value",
        "principal",
        "capital restant du",
        "capital restant dû",
        "crd",
        "encours nominal",
        "outstanding nominal",
        "notional",
    ],
    "quantity": ["quantite", "quantité", "nombre titres", "nb titres", "quantity"],
    "market_value": [
        "market value",
        "valeur de marche",
        "valeur de marché",
        "valeur actuelle",
        "valeur actualisee",
        "valeur actualisée",
        "present value",
        "pv",
        "exposition",
        "exposure",
        "fair value",
    ],
    "clean_price": [
        "clean price",
        "prix clean",
        "prix sec",
        "cours clean",
        "cours pied de coupon",
        "prix hors coupon couru",
        "price clean",
    ],
    "dirty_price": [
        "dirty price",
        "prix dirty",
        "prix sale",
        "prix plein coupon",
        "prix coupon couru inclus",
        "full price",
        "gross price",
    ],
    "accrued_interest": [
        "coupon couru",
        "interet couru",
        "intérêt couru",
        "accrued interest",
        "accrued",
    ],
    "coupon_rate": [
        "coupon",
        "taux coupon",
        "coupon rate",
        "taux facial",
        "taux nominal",
        "nominal rate",
        "interest rate",
    ],
    "ytm": [
        "ytm",
        "yield to maturity",
        "rendement actuariel",
        "taux actuariel",
        "yield",
        "rendement",
        "tra",
    ],
    "issue_date": [
        "date emission",
        "date d emission",
        "date d'emission",
        "issue date",
        "settlement issue",
    ],
    "settlement_date": [
        "date analyse",
        "date d analyse",
        "date d'analyse",
        "date valeur",
        "valuation date",
        "settlement date",
        "as of date",
    ],
    "maturity_date": [
        "maturite",
        "maturité",
        "date maturite",
        "date maturité",
        "date echeance",
        "date échéance",
        "echeance",
        "échéance",
        "maturity",
        "maturity date",
    ],
    "previous_coupon_date": [
        "date precedent coupon",
        "date précédent coupon",
        "previous coupon",
        "previous coupon date",
        "last coupon date",
    ],
    "next_coupon_date": [
        "date prochain coupon",
        "next coupon",
        "next coupon date",
        "prochain coupon",
    ],
    "frequency": [
        "frequence",
        "fréquence",
        "frequence coupon",
        "fréquence coupon",
        "periodicite",
        "périodicité",
        "coupon frequency",
        "frequency",
        "freq",
    ],
    "day_count": [
        "day count",
        "daycount",
        "base",
        "convention",
        "convention jour",
        "base de calcul",
        "basis",
    ],
    "amortization_type": [
        "type amortissement",
        "amortissement",
        "amortization type",
        "amortisation type",
    ],
    "amortization_rate": [
        "taux amortissement",
        "amortization rate",
        "amortisation rate",
    ],
    "amortization_amount": [
        "montant amortissement",
        "amortization amount",
        "principal amortization",
        "principal repayment",
    ],
    "existing_macaulay": [
        "duration macaulay",
        "duration de macaulay",
        "macaulay duration",
        "dmac",
    ],
    "existing_modified": [
        "duration modifiee",
        "duration modifiée",
        "modified duration",
        "dmod",
        "duration mod",
    ],
    "existing_convexity": ["convexite", "convexité", "convexity", "cvx"],
    "existing_weight": [
        "poids",
        "ponderation",
        "pondération",
        "weight",
        "portfolio weight",
    ],
}

CANONICAL_OUTPUT_SHEETS = [
    "DATA_ORIGINALE",
    "CASHFLOWS",
    "PRICING",
    "DURATION_DETAIL",
    "PORTFOLIO_METRICS",
    "CONTROLES",
    "DIAGNOSTIC_ERREURS",
]


@dataclass
class Diagnostic:
    instrument_id: str
    source_sheet: str
    source_row: int | None
    error: str
    financial_impact: str
    correction: str
    mathematical_justification: str
    severity: str = "WARNING"


@dataclass
class Control:
    instrument_id: str
    control: str
    status: str
    observed_value: Any
    threshold: str
    comment: str


@dataclass
class BondInput:
    source_sheet: str
    source_row: int
    instrument_id: str
    issuer: str
    nominal: float | None
    quantity: float | None
    market_value_input: float | None
    clean_price_input: float | None
    dirty_price_input: float | None
    accrued_interest_input: float | None
    coupon_rate: float | None
    ytm_input: float | None
    issue_date: date | None
    settlement_date: date
    maturity_date: date | None
    previous_coupon_date: date | None
    next_coupon_date: date | None
    frequency: int | None
    frequency_label: str
    day_count: str | None
    amortization_type: str
    amortization_rate: float | None
    amortization_amount: float | None
    existing_macaulay: float | None
    existing_modified: float | None
    existing_convexity: float | None
    existing_weight: float | None
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass
class Cashflow:
    instrument_id: str
    source_row: int
    payment_date: date
    years: float
    period_index: float
    coupon_amount: float
    principal_amount: float
    cashflow: float
    nominal_remaining_before: float
    discount_factor: float | None = None
    present_value: float | None = None
    flow_type: str = "Coupon"


@dataclass
class BondResult:
    instrument_id: str
    source_sheet: str
    source_row: int
    issuer: str
    status: str
    nominal: float | None
    coupon_rate: float | None
    ytm_input: float | None
    ytm_corrected: float | None
    clean_price_input: float | None
    dirty_price_input: float | None
    accrued_interest: float | None
    clean_price_corrected: float | None
    dirty_price_corrected: float | None
    dirty_value_corrected: float | None
    market_value: float | None
    frequency: int | None
    frequency_label: str
    day_count: str | None
    settlement_date: date
    maturity_date: date | None
    maturity_years: float | None
    macaulay_duration: float | None
    modified_duration: float | None
    convexity: float | None
    price_up_100bps: float | None
    price_down_100bps: float | None
    repriced_return_up_100bps: float | None
    repriced_return_down_100bps: float | None
    duration_approx_up_100bps: float | None
    duration_convexity_approx_up_100bps: float | None
    portfolio_weight: float | None = None
    duration_contribution: float | None = None
    convexity_contribution: float | None = None
    notes: str = ""


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower().replace("\n", " ").replace("\r", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return str(value).strip()


def clean_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        resolved = float(value)
        return resolved if math.isfinite(resolved) else None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "-", "n/a"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.replace("\u00a0", " ").replace(" ", "")
    text = text.replace("(", "").replace(")", "")
    text = text.replace("%", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", ".", "-", "-."}:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if negative:
        number = -number
    return number if math.isfinite(number) else None


def clean_rate(value: Any) -> float | None:
    raw = clean_number(value)
    if raw is None:
        return None
    text = str(value)
    if "%" in text or abs(raw) > 1.5:
        raw /= 100.0
    return raw if math.isfinite(raw) else None


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text_value = str(value).strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text_value):
        parsed = pd.to_datetime(text_value, errors="coerce", format="%Y-%m-%d")
    else:
        parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.date()


def detect_columns(columns: Iterable[Any]) -> dict[str, str]:
    normalized = {normalize_header(col): col for col in columns}
    mapped: dict[str, str] = {}
    for canonical, aliases in ALIASES.items():
        candidates = [normalize_header(alias) for alias in aliases]
        for candidate in candidates:
            if candidate in normalized:
                mapped[canonical] = normalized[candidate]
                break
        if canonical in mapped:
            continue
        for norm_col, original in normalized.items():
            if any(candidate and candidate in norm_col for candidate in candidates):
                mapped[canonical] = original
                break
    return mapped


def sheet_bond_score(df: pd.DataFrame) -> int:
    mapped = detect_columns(df.columns)
    required = {
        "nominal",
        "coupon_rate",
        "maturity_date",
        "frequency",
        "day_count",
        "ytm",
        "clean_price",
        "dirty_price",
        "market_value",
    }
    score = sum(1 for key in required if key in mapped)
    score += 2 if "maturity_date" in mapped else 0
    score += 2 if ("ytm" in mapped or "clean_price" in mapped or "dirty_price" in mapped) else 0
    return score


def cell_value(row: pd.Series, colmap: dict[str, str], key: str) -> Any:
    column = colmap.get(key)
    if column is None:
        return None
    return row.get(column)


def standardize_frequency(value: Any, coupon_rate: float | None = None) -> tuple[int | None, str]:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        if coupon_rate is not None and abs(coupon_rate) < 1e-12:
            return 0, "Zero Coupon"
        return None, "Non renseignee"
    if isinstance(value, (int, float, np.integer, np.floating)):
        number = int(round(float(value)))
        if number in {0, 1, 2, 4, 12}:
            labels = {0: "Zero Coupon", 1: "Annual", 2: "Semi-Annual", 4: "Quarterly", 12: "Monthly"}
            return number, labels[number]
    text = normalize_header(value)
    if any(token in text for token in ["zero", "zc", "sans coupon"]):
        return 0, "Zero Coupon"
    if any(token in text for token in ["annual", "annuel", "annuelle", "yearly"]):
        return 1, "Annual"
    if any(token in text for token in ["semi", "semes", "bi annual", "biannual"]):
        return 2, "Semi-Annual"
    if any(token in text for token in ["quarter", "trimes", "quarterly"]):
        return 4, "Quarterly"
    if any(token in text for token in ["month", "mens", "monthly"]):
        return 12, "Monthly"
    number = clean_number(value)
    if number is not None and int(round(number)) in {0, 1, 2, 4, 12}:
        return int(round(number)), str(int(round(number)))
    return None, str(value)


def standardize_day_count(value: Any) -> str | None:
    text = normalize_header(value)
    if not text:
        return None
    compact = text.replace(" ", "")
    if compact in {"act360", "actual360", "a360"} or "act 360" in text:
        return "ACT/360"
    if compact in {"act365", "actual365", "a365"} or "act 365" in text:
        return "ACT/365"
    if compact in {"30360", "30e360", "bondbasis"} or "30 360" in text:
        return "30/360"
    if compact in {"actact", "actualactual"} or "act act" in text:
        return "ACT/ACT"
    return None


def is_leap_year(year: int) -> bool:
    return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)


def year_fraction(start: date, end: date, convention: str) -> float:
    if end < start:
        return -year_fraction(end, start, convention)
    convention = convention.upper()
    if convention == "ACT/360":
        return (end - start).days / 360.0
    if convention == "ACT/365":
        return (end - start).days / 365.0
    if convention == "ACT/ACT":
        if start == end:
            return 0.0
        total = 0.0
        cursor = start
        while cursor < end:
            next_year = date(cursor.year + 1, 1, 1)
            stop = min(end, next_year)
            total += (stop - cursor).days / (366.0 if is_leap_year(cursor.year) else 365.0)
            cursor = stop
        return total
    if convention == "30/360":
        d1 = min(start.day, 30)
        d2 = end.day
        if d1 == 30:
            d2 = min(d2, 30)
        return ((end.year - start.year) * 360 + (end.month - start.month) * 30 + (d2 - d1)) / 360.0
    raise ValueError(f"Convention day count non supportee: {convention}")


def add_months(input_date: date, months: int) -> date:
    ts = pd.Timestamp(input_date) + pd.DateOffset(months=months)
    return ts.date()


def subtract_months(input_date: date, months: int) -> date:
    ts = pd.Timestamp(input_date) - pd.DateOffset(months=months)
    return ts.date()


def generate_coupon_schedule(settlement_date: date, maturity_date: date, frequency: int) -> list[date]:
    if settlement_date >= maturity_date:
        return []
    if frequency == 0:
        return [maturity_date]
    months = int(12 / frequency)
    payment_dates: list[date] = []
    cursor = maturity_date
    guard = 0
    while cursor > settlement_date and guard < 1200:
        payment_dates.append(cursor)
        cursor = subtract_months(cursor, months)
        guard += 1
    return sorted(set(payment_dates))


def infer_previous_coupon_date(
    settlement_date: date,
    next_coupon_date: date,
    frequency: int,
    issue_date: date | None,
) -> date:
    if frequency <= 0:
        return settlement_date
    months = int(12 / frequency)
    previous = subtract_months(next_coupon_date, months)
    if issue_date is not None and previous < issue_date <= settlement_date:
        return issue_date
    return previous


def generate_cashflows(bond: BondInput) -> tuple[list[Cashflow], float, list[Diagnostic]]:
    """Reconstruit les cash flows futurs et le coupon couru d'une obligation."""
    diagnostics: list[Diagnostic] = []
    if bond.nominal is None or bond.nominal <= 0:
        diagnostics.append(
            Diagnostic(
                bond.instrument_id,
                bond.source_sheet,
                bond.source_row,
                "Nominal absent ou invalide",
                "Impossible de convertir les prix en valeur monetaire et de projeter les coupons.",
                "Ligne exclue du recalcul quantitatif jusqu'a correction du nominal.",
                "Les cash flows obligataires sont proportionnels au nominal restant du.",
                "BLOCKED",
            )
        )
        return [], 0.0, diagnostics
    if bond.maturity_date is None:
        diagnostics.append(
            Diagnostic(
                bond.instrument_id,
                bond.source_sheet,
                bond.source_row,
                "Date de maturite absente",
                "La maturite residuelle et les dates de paiement ne peuvent pas etre reconstruites.",
                "Ligne exclue du recalcul quantitatif jusqu'a renseignement de l'echeance.",
                "La duration est une moyenne temporelle des cash flows actualises.",
                "BLOCKED",
            )
        )
        return [], 0.0, diagnostics
    if bond.day_count is None:
        diagnostics.append(
            Diagnostic(
                bond.instrument_id,
                bond.source_sheet,
                bond.source_row,
                "Convention day count absente ou non reconnue",
                "Le temps actuariel et le coupon couru seraient approximatifs.",
                "Aucun calcul de duration n'est produit sans convention explicite.",
                "Le temps t de chaque cash flow doit respecter la convention contractuelle.",
                "BLOCKED",
            )
        )
        return [], 0.0, diagnostics
    if bond.frequency is None:
        diagnostics.append(
            Diagnostic(
                bond.instrument_id,
                bond.source_sheet,
                bond.source_row,
                "Frequence de coupon absente ou non reconnue",
                "Le calendrier des coupons et la formule Dmod = Dmac/(1+y/m) sont indetermines.",
                "Aucun calcul de duration n'est produit sans periodicite explicite.",
                "La periodicite m intervient dans l'actualisation et la duration modifiee.",
                "BLOCKED",
            )
        )
        return [], 0.0, diagnostics
    if bond.settlement_date >= bond.maturity_date:
        diagnostics.append(
            Diagnostic(
                bond.instrument_id,
                bond.source_sheet,
                bond.source_row,
                "Obligation maturee a la date d'analyse",
                "Aucun flux futur positif ne subsiste.",
                "Ligne exclue des sensibilites futures.",
                "La duration des flux futurs est nulle lorsque la maturite est depassee.",
                "WARNING",
            )
        )
        return [], 0.0, diagnostics

    coupon_rate = bond.coupon_rate or 0.0
    frequency = bond.frequency
    schedule = generate_coupon_schedule(bond.settlement_date, bond.maturity_date, frequency)
    if not schedule:
        return [], 0.0, diagnostics

    accrued_interest = 0.0
    if frequency > 0 and coupon_rate != 0:
        next_coupon = bond.next_coupon_date or schedule[0]
        previous_coupon = bond.previous_coupon_date or infer_previous_coupon_date(
            bond.settlement_date,
            next_coupon,
            frequency,
            bond.issue_date,
        )
        if previous_coupon < bond.settlement_date < next_coupon:
            elapsed = year_fraction(previous_coupon, bond.settlement_date, bond.day_count)
            full = year_fraction(previous_coupon, next_coupon, bond.day_count)
            if full > 0:
                accrued_interest = bond.nominal * coupon_rate / frequency * (elapsed / full)
        elif bond.accrued_interest_input is not None:
            accrued_interest = bond.accrued_interest_input

    remaining = bond.nominal
    cashflows: list[Cashflow] = []
    amort_type = normalize_header(bond.amortization_type)
    explicit_amortizing = bool(amort_type and not any(token in amort_type for token in ["bullet", "in fine", "final"]))
    linear_amortizing = any(token in amort_type for token in ["linear", "lineaire", "constant"])
    amort_amount = bond.amortization_amount if bond.amortization_amount and bond.amortization_amount > 0 else None
    amort_rate = bond.amortization_rate if bond.amortization_rate and bond.amortization_rate > 0 else None
    if amort_rate and amort_rate > 1:
        amort_rate /= 100.0

    for index, payment_date in enumerate(schedule, start=1):
        nominal_before = remaining
        coupon_amount = 0.0 if frequency == 0 else nominal_before * coupon_rate / frequency
        is_final = payment_date == schedule[-1]
        principal_amount = 0.0
        if is_final:
            principal_amount = remaining
        elif amort_amount is not None:
            principal_amount = min(remaining, amort_amount)
        elif amort_rate is not None and explicit_amortizing:
            principal_amount = min(remaining, bond.nominal * amort_rate / max(frequency, 1))
        elif linear_amortizing:
            remaining_periods = len(schedule) - index + 1
            principal_amount = remaining / remaining_periods

        remaining = max(0.0, remaining - principal_amount)
        cashflow_amount = coupon_amount + principal_amount
        flow_type = "Coupon + Principal" if principal_amount > 0 and coupon_amount > 0 else (
            "Principal" if principal_amount > 0 else "Coupon"
        )
        years = year_fraction(bond.settlement_date, payment_date, bond.day_count)
        period_index = years * max(frequency, 1)
        cashflows.append(
            Cashflow(
                instrument_id=bond.instrument_id,
                source_row=bond.source_row,
                payment_date=payment_date,
                years=years,
                period_index=period_index,
                coupon_amount=coupon_amount,
                principal_amount=principal_amount,
                cashflow=cashflow_amount,
                nominal_remaining_before=nominal_before,
                flow_type=flow_type,
            )
        )

    return cashflows, accrued_interest, diagnostics


def compute_dirty_price(cashflows: list[Cashflow], ytm: float, frequency: int | None) -> float:
    """Prix dirty en valeur monetaire par actualisation actuarielle discrete."""
    if not cashflows:
        return float("nan")
    m = max(frequency or 1, 1)
    base = 1.0 + ytm / m
    if base <= 0:
        return float("nan")
    return float(sum(cf.cashflow / (base ** (m * cf.years)) for cf in cashflows))


def _bisect_root(function, lower: float, upper: float, tolerance: float = 1e-10, max_iter: int = 200) -> float:
    f_lower = function(lower)
    f_upper = function(upper)
    if not math.isfinite(f_lower) or not math.isfinite(f_upper) or f_lower * f_upper > 0:
        raise ValueError("Root is not bracketed")
    for _ in range(max_iter):
        middle = (lower + upper) / 2.0
        f_middle = function(middle)
        if abs(f_middle) < tolerance or (upper - lower) / 2.0 < tolerance:
            return middle
        if f_lower * f_middle <= 0:
            upper = middle
            f_upper = f_middle
        else:
            lower = middle
            f_lower = f_middle
    return (lower + upper) / 2.0


def compute_ytm(cashflows: list[Cashflow], target_dirty_price: float, frequency: int | None) -> float:
    """Resout le YTM coherent avec le prix dirty via scipy.brentq si disponible."""
    if not cashflows or target_dirty_price <= 0:
        return float("nan")

    def objective(rate: float) -> float:
        return compute_dirty_price(cashflows, rate, frequency) - target_dirty_price

    grid = np.concatenate(
        [
            np.linspace(-0.95, 0.50, 120),
            np.linspace(0.55, 5.00, 120),
        ]
    )
    previous_rate = None
    previous_value = None
    for rate in grid:
        value = objective(float(rate))
        if not math.isfinite(value):
            continue
        if previous_value is not None and previous_value * value <= 0:
            lower = float(previous_rate)
            upper = float(rate)
            if scipy_brentq is not None:
                return float(scipy_brentq(objective, lower, upper, xtol=1e-12, rtol=1e-12, maxiter=200))
            return float(_bisect_root(objective, lower, upper))
        previous_rate = float(rate)
        previous_value = value
    return float("nan")


def macaulay_duration(cashflows: list[Cashflow], ytm: float, frequency: int | None) -> float:
    price = compute_dirty_price(cashflows, ytm, frequency)
    if not price or not math.isfinite(price) or price <= 0:
        return float("nan")
    m = max(frequency or 1, 1)
    base = 1.0 + ytm / m
    numerator = sum(cf.years * cf.cashflow / (base ** (m * cf.years)) for cf in cashflows)
    return float(numerator / price)


def modified_duration(macaulay: float, ytm: float, frequency: int | None) -> float:
    if not math.isfinite(macaulay):
        return float("nan")
    m = max(frequency or 1, 1)
    denominator = 1.0 + ytm / m
    if denominator <= 0:
        return float("nan")
    return float(macaulay / denominator)


def convexity(cashflows: list[Cashflow], ytm: float, frequency: int | None) -> float:
    price = compute_dirty_price(cashflows, ytm, frequency)
    if not price or not math.isfinite(price) or price <= 0:
        return float("nan")
    m = max(frequency or 1, 1)
    base = 1.0 + ytm / m
    numerator = 0.0
    for cf in cashflows:
        periods = m * cf.years
        numerator += cf.cashflow * periods * (periods + 1.0) / (base ** (periods + 2.0))
    return float(numerator / (price * (m**2)))


def price_to_value(price: float | None, nominal: float | None) -> float | None:
    if price is None or nominal is None or nominal <= 0:
        return None
    if abs(price) <= 300:
        return nominal * price / 100.0
    return price


def value_to_price(value: float | None, nominal: float | None) -> float | None:
    if value is None or nominal is None or nominal <= 0:
        return None
    return value / nominal * 100.0


def safe_ratio(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or denominator == 0:
        return None
    result = numerator / denominator
    return result if math.isfinite(result) else None


def build_bond_inputs(
    bond_df: pd.DataFrame,
    analysis_date: date,
    diagnostics: list[Diagnostic],
) -> list[BondInput]:
    colmap = detect_columns(bond_df.columns)
    bonds: list[BondInput] = []
    for idx, row in bond_df.iterrows():
        source_sheet = normalize_text(row.get("_source_sheet")) or "DATA"
        source_row_raw = clean_number(row.get("_source_row"))
        source_row = int(source_row_raw or idx + 2)
        coupon_rate = clean_rate(cell_value(row, colmap, "coupon_rate"))
        frequency, frequency_label = standardize_frequency(cell_value(row, colmap, "frequency"), coupon_rate)
        settlement = parse_date(cell_value(row, colmap, "settlement_date"))
        if settlement is None:
            settlement = analysis_date
            diagnostics.append(
                Diagnostic(
                    instrument_id=normalize_text(cell_value(row, colmap, "instrument_id")) or f"Ligne {source_row}",
                    source_sheet=source_sheet,
                    source_row=source_row,
                    error="Date d'analyse absente",
                    financial_impact="La date de valorisation pilote maturite residuelle, coupon couru et facteurs d'actualisation.",
                    correction=f"Date d'analyse parametree utilisee: {analysis_date.isoformat()}.",
                    mathematical_justification="Toutes les echeances t sont mesurees depuis la date d'analyse.",
                    severity="WARNING",
                )
            )
        instrument_id = normalize_text(cell_value(row, colmap, "instrument_id")) or f"{source_sheet}!R{source_row}"
        issuer = normalize_text(cell_value(row, colmap, "issuer")) or "Non renseigne"
        day_count = standardize_day_count(cell_value(row, colmap, "day_count"))
        bond = BondInput(
            source_sheet=source_sheet,
            source_row=source_row,
            instrument_id=instrument_id,
            issuer=issuer,
            nominal=clean_number(cell_value(row, colmap, "nominal")),
            quantity=clean_number(cell_value(row, colmap, "quantity")),
            market_value_input=clean_number(cell_value(row, colmap, "market_value")),
            clean_price_input=clean_number(cell_value(row, colmap, "clean_price")),
            dirty_price_input=clean_number(cell_value(row, colmap, "dirty_price")),
            accrued_interest_input=clean_number(cell_value(row, colmap, "accrued_interest")),
            coupon_rate=coupon_rate,
            ytm_input=clean_rate(cell_value(row, colmap, "ytm")),
            issue_date=parse_date(cell_value(row, colmap, "issue_date")),
            settlement_date=settlement,
            maturity_date=parse_date(cell_value(row, colmap, "maturity_date")),
            previous_coupon_date=parse_date(cell_value(row, colmap, "previous_coupon_date")),
            next_coupon_date=parse_date(cell_value(row, colmap, "next_coupon_date")),
            frequency=frequency,
            frequency_label=frequency_label,
            day_count=day_count,
            amortization_type=normalize_text(cell_value(row, colmap, "amortization_type")) or "Bullet",
            amortization_rate=clean_rate(cell_value(row, colmap, "amortization_rate")),
            amortization_amount=clean_number(cell_value(row, colmap, "amortization_amount")),
            existing_macaulay=clean_number(cell_value(row, colmap, "existing_macaulay")),
            existing_modified=clean_number(cell_value(row, colmap, "existing_modified")),
            existing_convexity=clean_number(cell_value(row, colmap, "existing_convexity")),
            existing_weight=clean_rate(cell_value(row, colmap, "existing_weight")),
            raw=row.to_dict(),
        )
        bonds.append(bond)
    return bonds


def compute_bond_result(
    bond: BondInput,
    diagnostics: list[Diagnostic],
    controls: list[Control],
) -> tuple[BondResult, list[Cashflow]]:
    cashflows, accrued_interest, cashflow_diagnostics = generate_cashflows(bond)
    diagnostics.extend(cashflow_diagnostics)
    maturity_years = (
        year_fraction(bond.settlement_date, bond.maturity_date, bond.day_count)
        if bond.maturity_date and bond.day_count
        else None
    )
    empty_result = BondResult(
        instrument_id=bond.instrument_id,
        source_sheet=bond.source_sheet,
        source_row=bond.source_row,
        issuer=bond.issuer,
        status="Non calcule",
        nominal=bond.nominal,
        coupon_rate=bond.coupon_rate,
        ytm_input=bond.ytm_input,
        ytm_corrected=None,
        clean_price_input=bond.clean_price_input,
        dirty_price_input=bond.dirty_price_input,
        accrued_interest=accrued_interest,
        clean_price_corrected=None,
        dirty_price_corrected=None,
        dirty_value_corrected=None,
        market_value=bond.market_value_input,
        frequency=bond.frequency,
        frequency_label=bond.frequency_label,
        day_count=bond.day_count,
        settlement_date=bond.settlement_date,
        maturity_date=bond.maturity_date,
        maturity_years=maturity_years,
        macaulay_duration=None,
        modified_duration=None,
        convexity=None,
        price_up_100bps=None,
        price_down_100bps=None,
        repriced_return_up_100bps=None,
        repriced_return_down_100bps=None,
        duration_approx_up_100bps=None,
        duration_convexity_approx_up_100bps=None,
        notes="Calcul bloque par donnees critiques manquantes.",
    )
    if not cashflows:
        controls.append(
            Control(
                bond.instrument_id,
                "Cash flows futurs disponibles",
                "BLOCKED",
                0,
                "> 0",
                "Aucun flux futur exploitable.",
            )
        )
        return empty_result, []

    dirty_value_input = price_to_value(bond.dirty_price_input, bond.nominal)
    clean_value_input = price_to_value(bond.clean_price_input, bond.nominal)
    target_dirty_value = None
    price_source = "Non disponible"
    if dirty_value_input is not None:
        target_dirty_value = dirty_value_input
        price_source = "Dirty Price"
    elif clean_value_input is not None:
        target_dirty_value = clean_value_input + accrued_interest
        price_source = "Clean Price + Coupon Couru"
    elif bond.market_value_input is not None and bond.market_value_input > 0:
        target_dirty_value = bond.market_value_input
        price_source = "Market Value"

    corrected_ytm = None
    notes = []
    if target_dirty_value is not None:
        corrected_ytm = compute_ytm(cashflows, target_dirty_value, bond.frequency)
        if corrected_ytm is not None and math.isfinite(corrected_ytm):
            notes.append(f"YTM resolu depuis {price_source}.")
        elif bond.ytm_input is not None:
            corrected_ytm = bond.ytm_input
            notes.append("YTM source utilise car la resolution depuis prix n'a pas converge.")
            diagnostics.append(
                Diagnostic(
                    bond.instrument_id,
                    bond.source_sheet,
                    bond.source_row,
                    "Resolution YTM non convergente",
                    "Le couple prix/cash flows ne permet pas de racine actuarielle stable.",
                    "YTM source conserve avec diagnostic.",
                    "Le YTM est la racine de P = somme CF_t/(1+y/m)^(m*t).",
                    "WARNING",
                )
            )
    elif bond.ytm_input is not None:
        corrected_ytm = bond.ytm_input
        notes.append("Prix reconstruit depuis le YTM source.")
    else:
        diagnostics.append(
            Diagnostic(
                bond.instrument_id,
                bond.source_sheet,
                bond.source_row,
                "Prix et YTM absents",
                "Impossible de valoriser les cash flows et de calculer une duration coherent prix-rendement.",
                "Renseigner au minimum un prix clean/dirty, une valeur de marche ou un YTM.",
                "La duration de Macaulay utilise les flux actualises par le rendement actuariel.",
                "BLOCKED",
            )
        )
        controls.append(
            Control(
                bond.instrument_id,
                "Prix ou YTM disponible",
                "BLOCKED",
                "Absent",
                "Prix ou YTM requis",
                "Aucune mesure de sensibilite produite.",
            )
        )
        return empty_result, cashflows

    if corrected_ytm is None or not math.isfinite(corrected_ytm):
        return empty_result, cashflows

    dirty_value_corrected = compute_dirty_price(cashflows, corrected_ytm, bond.frequency)
    if target_dirty_value is not None and math.isfinite(corrected_ytm):
        dirty_value_corrected = target_dirty_value
    clean_value_corrected = dirty_value_corrected - accrued_interest
    clean_price_corrected = value_to_price(clean_value_corrected, bond.nominal)
    dirty_price_corrected = value_to_price(dirty_value_corrected, bond.nominal)
    market_value = dirty_value_corrected

    m_duration = macaulay_duration(cashflows, corrected_ytm, bond.frequency)
    mod_duration = modified_duration(m_duration, corrected_ytm, bond.frequency)
    cvx = convexity(cashflows, corrected_ytm, bond.frequency)
    price_up = compute_dirty_price(cashflows, corrected_ytm + 0.01, bond.frequency)
    price_down = compute_dirty_price(cashflows, corrected_ytm - 0.01, bond.frequency)
    repriced_up = safe_ratio(price_up - dirty_value_corrected, dirty_value_corrected)
    repriced_down = safe_ratio(price_down - dirty_value_corrected, dirty_value_corrected)
    duration_approx_up = -mod_duration * 0.01 if math.isfinite(mod_duration) else None
    duration_convexity_approx_up = (
        -mod_duration * 0.01 + 0.5 * cvx * (0.01**2)
        if math.isfinite(mod_duration) and math.isfinite(cvx)
        else None
    )

    for cf in cashflows:
        base = 1.0 + corrected_ytm / max(bond.frequency or 1, 1)
        cf.discount_factor = 1.0 / (base ** (max(bond.frequency or 1, 1) * cf.years))
        cf.present_value = cf.cashflow * cf.discount_factor

    if bond.ytm_input is not None and abs(bond.ytm_input - corrected_ytm) > 1e-4:
        diagnostics.append(
            Diagnostic(
                bond.instrument_id,
                bond.source_sheet,
                bond.source_row,
                "YTM source incoherent avec le prix",
                f"Ecart YTM de {(corrected_ytm - bond.ytm_input) * 10000:.2f} bps.",
                "YTM recalcule par resolution actuarielle sur le prix dirty coherent.",
                "Le rendement est la racine de P_dirty = somme CF_t/(1+y/m)^(m*t).",
                "ERROR",
            )
        )
    if bond.existing_macaulay is not None and math.isfinite(m_duration) and abs(bond.existing_macaulay - m_duration) > 1e-3:
        diagnostics.append(
            Diagnostic(
                bond.instrument_id,
                bond.source_sheet,
                bond.source_row,
                "Duration de Macaulay source differente du recalcul",
                f"Ecart de {m_duration - bond.existing_macaulay:.6f} annee(s).",
                "Duration de Macaulay reconstruite comme moyenne des echeances ponderees par PV des flux.",
                "Dmac = somme(t x PV(CF_t)) / P_dirty.",
                "ERROR",
            )
        )
    if bond.existing_modified is not None and math.isfinite(mod_duration) and abs(bond.existing_modified - mod_duration) > 1e-3:
        diagnostics.append(
            Diagnostic(
                bond.instrument_id,
                bond.source_sheet,
                bond.source_row,
                "Duration modifiee source differente du recalcul",
                f"Ecart de {mod_duration - bond.existing_modified:.6f} annee(s).",
                "Duration modifiee recalculee depuis Dmac et YTM periodique.",
                "Dmod = Dmac / (1 + y/m).",
                "ERROR",
            )
        )
    if bond.existing_convexity is not None and math.isfinite(cvx) and abs(bond.existing_convexity - cvx) > 1e-2:
        diagnostics.append(
            Diagnostic(
                bond.instrument_id,
                bond.source_sheet,
                bond.source_row,
                "Convexite source differente du recalcul",
                f"Ecart de {cvx - bond.existing_convexity:.6f}.",
                "Convexite recalculee au second ordre sur les cash flows actualises.",
                "C = somme(CF x k x (k+1)/(1+y/m)^(k+2))/(P x m^2).",
                "WARNING",
            )
        )

    controls.extend(build_bond_controls(bond, m_duration, mod_duration, cvx, repriced_up, repriced_down))

    return (
        BondResult(
            instrument_id=bond.instrument_id,
            source_sheet=bond.source_sheet,
            source_row=bond.source_row,
            issuer=bond.issuer,
            status="Calcule",
            nominal=bond.nominal,
            coupon_rate=bond.coupon_rate,
            ytm_input=bond.ytm_input,
            ytm_corrected=corrected_ytm,
            clean_price_input=bond.clean_price_input,
            dirty_price_input=bond.dirty_price_input,
            accrued_interest=accrued_interest,
            clean_price_corrected=clean_price_corrected,
            dirty_price_corrected=dirty_price_corrected,
            dirty_value_corrected=dirty_value_corrected,
            market_value=market_value,
            frequency=bond.frequency,
            frequency_label=bond.frequency_label,
            day_count=bond.day_count,
            settlement_date=bond.settlement_date,
            maturity_date=bond.maturity_date,
            maturity_years=maturity_years,
            macaulay_duration=m_duration,
            modified_duration=mod_duration,
            convexity=cvx,
            price_up_100bps=price_up,
            price_down_100bps=price_down,
            repriced_return_up_100bps=repriced_up,
            repriced_return_down_100bps=repriced_down,
            duration_approx_up_100bps=duration_approx_up,
            duration_convexity_approx_up_100bps=duration_convexity_approx_up,
            notes=" ".join(notes),
        ),
        cashflows,
    )


def build_bond_controls(
    bond: BondInput,
    macaulay: float,
    modified: float,
    cvx: float,
    repriced_up: float | None,
    repriced_down: float | None,
) -> list[Control]:
    controls: list[Control] = []
    if math.isfinite(macaulay) and math.isfinite(modified):
        controls.append(
            Control(
                bond.instrument_id,
                "Duration modifiee < Duration Macaulay",
                "PASS" if modified <= macaulay + 1e-9 else "FAIL",
                f"Dmod={modified:.8f}; Dmac={macaulay:.8f}",
                "Dmod <= Dmac",
                "Controle de coherence du facteur 1+y/m.",
            )
        )
    if bond.frequency == 0 and bond.maturity_date and bond.day_count and math.isfinite(macaulay):
        maturity = year_fraction(bond.settlement_date, bond.maturity_date, bond.day_count)
        controls.append(
            Control(
                bond.instrument_id,
                "Zero coupon => Duration = Maturite",
                "PASS" if abs(macaulay - maturity) < 1e-6 else "FAIL",
                f"Dmac={macaulay:.8f}; Mat={maturity:.8f}",
                "Ecart < 1e-6",
                "Un zero coupon a un flux unique a maturite.",
            )
        )
    if repriced_up is not None:
        controls.append(
            Control(
                bond.instrument_id,
                "Prix baisse si taux +100 bps",
                "PASS" if repriced_up < 0 else "FAIL",
                f"{repriced_up:.6%}",
                "< 0",
                "Relation prix/taux decroissante.",
            )
        )
    if repriced_down is not None:
        controls.append(
            Control(
                bond.instrument_id,
                "Prix monte si taux -100 bps",
                "PASS" if repriced_down > 0 else "FAIL",
                f"{repriced_down:.6%}",
                "> 0",
                "Relation prix/taux decroissante.",
            )
        )
    if math.isfinite(cvx):
        controls.append(
            Control(
                bond.instrument_id,
                "Convexite positive",
                "PASS" if cvx >= 0 else "FAIL",
                f"{cvx:.8f}",
                ">= 0",
                "Convexite standard positive pour obligations option-free.",
            )
        )
    return controls


def portfolio_duration(results: list[BondResult]) -> dict[str, float]:
    valid = [
        result
        for result in results
        if result.market_value is not None
        and result.market_value > 0
        and result.macaulay_duration is not None
        and result.modified_duration is not None
        and result.convexity is not None
    ]
    total_market_value = sum(result.market_value or 0.0 for result in valid)
    if total_market_value <= 0:
        return {
            "total_market_value": 0.0,
            "macaulay_duration": float("nan"),
            "modified_duration": float("nan"),
            "convexity": float("nan"),
            "weighted_ytm": float("nan"),
            "weighted_coupon": float("nan"),
        }
    macaulay = sum((result.market_value or 0.0) / total_market_value * (result.macaulay_duration or 0.0) for result in valid)
    modified = sum((result.market_value or 0.0) / total_market_value * (result.modified_duration or 0.0) for result in valid)
    cvx = sum((result.market_value or 0.0) / total_market_value * (result.convexity or 0.0) for result in valid)
    weighted_ytm = sum((result.market_value or 0.0) / total_market_value * (result.ytm_corrected or 0.0) for result in valid)
    weighted_coupon = sum((result.market_value or 0.0) / total_market_value * (result.coupon_rate or 0.0) for result in valid)
    return {
        "total_market_value": total_market_value,
        "macaulay_duration": macaulay,
        "modified_duration": modified,
        "convexity": cvx,
        "weighted_ytm": weighted_ytm,
        "weighted_coupon": weighted_coupon,
    }


def update_portfolio_weights(
    bonds: list[BondInput],
    results: list[BondResult],
    diagnostics: list[Diagnostic],
) -> dict[str, float]:
    metrics = portfolio_duration(results)
    total = metrics["total_market_value"]
    bond_by_key = {(bond.source_sheet, bond.source_row, bond.instrument_id): bond for bond in bonds}
    for result in results:
        if total > 0 and result.market_value is not None:
            result.portfolio_weight = result.market_value / total
            result.duration_contribution = (
                result.portfolio_weight * result.modified_duration
                if result.modified_duration is not None
                else None
            )
            result.convexity_contribution = (
                result.portfolio_weight * result.convexity
                if result.convexity is not None
                else None
            )
        bond = bond_by_key.get((result.source_sheet, result.source_row, result.instrument_id))
        if (
            bond
            and bond.existing_weight is not None
            and result.portfolio_weight is not None
            and abs(bond.existing_weight - result.portfolio_weight) > 1e-4
        ):
            diagnostics.append(
                Diagnostic(
                    result.instrument_id,
                    result.source_sheet,
                    result.source_row,
                    "Ponderation portefeuille source incoherente",
                    f"Ecart de {(result.portfolio_weight - bond.existing_weight) * 100:.4f} point(s) de pourcentage.",
                    "Poids recalcule sur Market Value reelle.",
                    "w_i = MV_i / somme(MV_i), pas nominal_i / somme(nominal_i).",
                    "ERROR",
                )
            )
    return metrics


def workbook_formula_inventory(input_path: Path) -> tuple[list[Diagnostic], list[dict[str, Any]]]:
    diagnostics: list[Diagnostic] = []
    inventory: list[dict[str, Any]] = []
    try:
        workbook = load_workbook(input_path, data_only=False, read_only=False)
    except Exception:
        return diagnostics, inventory
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    inventory.append(
                        {
                            "Sheet": sheet.title,
                            "Cell": cell.coordinate,
                            "Formula": value,
                        }
                    )
                    if "#REF!" in value.upper():
                        diagnostics.append(
                            Diagnostic(
                                "Classeur",
                                sheet.title,
                                cell.row,
                                f"Reference cassee dans {cell.coordinate}",
                                "Une formule avec #REF! rend les indicateurs dependants non fiables.",
                                "Reconstruction complete des calculs dans le fichier corrige.",
                                "Une reference Excel cassee invalide la chaine de calcul.",
                                "ERROR",
                            )
                        )
    workbook.close()
    return diagnostics, inventory


def read_excel_sheets(input_path: Path) -> dict[str, pd.DataFrame]:
    sheets: dict[str, pd.DataFrame] = {}
    with pd.ExcelFile(input_path) as excel:
        for sheet_name in excel.sheet_names:
            sheets[sheet_name] = pd.read_excel(excel, sheet_name=sheet_name, dtype=object)
    return sheets


def combine_original_data(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for sheet_name, df in sheets.items():
        frame = df.copy()
        frame.insert(0, "_source_sheet", sheet_name)
        frame.insert(1, "_source_row", np.arange(2, len(frame) + 2))
        frames.append(frame)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True, sort=False)


def select_bond_data(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    scored = [(sheet_name, sheet_bond_score(df), df) for sheet_name, df in sheets.items()]
    candidates = [(name, score, df) for name, score, df in scored if score >= 5 and not df.empty]
    if not candidates:
        best = max(scored, key=lambda item: item[1], default=(None, 0, pd.DataFrame()))
        if best[1] < 3:
            raise ValueError(
                "Aucune feuille obligataire exploitable detectee. Verifier les en-tetes: nominal, coupon, maturite, "
                "frequence, day count, prix ou YTM."
            )
        candidates = [best]

    frames: list[pd.DataFrame] = []
    for sheet_name, _, df in candidates:
        frame = df.copy()
        frame.insert(0, "_source_sheet", sheet_name)
        frame.insert(1, "_source_row", np.arange(2, len(frame) + 2))
        frames.append(frame)
    return pd.concat(frames, ignore_index=True, sort=False)


def diagnostics_to_df(diagnostics: list[Diagnostic]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Instrument": item.instrument_id,
                "Feuille source": item.source_sheet,
                "Ligne source": item.source_row,
                "Severite": item.severity,
                "Erreur detectee": item.error,
                "Impact financier": item.financial_impact,
                "Correction appliquee": item.correction,
                "Justification mathematique": item.mathematical_justification,
            }
            for item in diagnostics
        ]
    )


def controls_to_df(controls: list[Control]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Instrument": item.instrument_id,
                "Controle": item.control,
                "Statut": item.status,
                "Valeur observee": item.observed_value,
                "Seuil": item.threshold,
                "Commentaire": item.comment,
            }
            for item in controls
        ]
    )


def cashflows_to_df(cashflows: list[Cashflow]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Instrument": cf.instrument_id,
                "Ligne source": cf.source_row,
                "Date paiement": cf.payment_date,
                "Temps annees": cf.years,
                "Index periode": cf.period_index,
                "Type flux": cf.flow_type,
                "Nominal avant flux": cf.nominal_remaining_before,
                "Coupon": cf.coupon_amount,
                "Principal": cf.principal_amount,
                "Cash Flow": cf.cashflow,
                "Facteur actualisation": cf.discount_factor,
                "Valeur actuelle": cf.present_value,
            }
            for cf in cashflows
        ]
    )


def pricing_to_df(results: list[BondResult]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Instrument": result.instrument_id,
                "Emetteur": result.issuer,
                "Statut": result.status,
                "Nominal": result.nominal,
                "Coupon": result.coupon_rate,
                "YTM source": result.ytm_input,
                "YTM corrige": result.ytm_corrected,
                "Clean Price source": result.clean_price_input,
                "Dirty Price source": result.dirty_price_input,
                "Coupon couru": result.accrued_interest,
                "Clean Price corrige": result.clean_price_corrected,
                "Dirty Price corrige": result.dirty_price_corrected,
                "Dirty Value corrigee": result.dirty_value_corrected,
                "Market Value": result.market_value,
                "Poids portefeuille": result.portfolio_weight,
                "Frequence": result.frequency_label,
                "Day Count": result.day_count,
                "Date analyse": result.settlement_date,
                "Maturite": result.maturity_date,
                "Notes": result.notes,
            }
            for result in results
        ]
    )


def duration_to_df(results: list[BondResult]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Instrument": result.instrument_id,
                "Ligne source": result.source_row,
                "Maturite residuelle": result.maturity_years,
                "Duration Macaulay": result.macaulay_duration,
                "Duration Modifiee": result.modified_duration,
                "Convexite": result.convexity,
                "Contribution Duration": result.duration_contribution,
                "Contribution Convexite": result.convexity_contribution,
                "Prix +100bps": result.price_up_100bps,
                "Prix -100bps": result.price_down_100bps,
                "Variation revalorisee +100bps": result.repriced_return_up_100bps,
                "Variation revalorisee -100bps": result.repriced_return_down_100bps,
                "Approx duration +100bps": result.duration_approx_up_100bps,
                "Approx duration+convexite +100bps": result.duration_convexity_approx_up_100bps,
            }
            for result in results
        ]
    )


def portfolio_metrics_to_df(metrics: dict[str, float], results: list[BondResult]) -> pd.DataFrame:
    valid_count = sum(1 for result in results if result.status == "Calcule")
    rows = [
        (
            "Nombre obligations calculees",
            valid_count,
            "=COUNTIF(PRICING!C:C,\"Calcule\")",
            "Titres disposant de donnees critiques suffisantes.",
        ),
        (
            "Market Value portefeuille",
            metrics["total_market_value"],
            "=SUM(PRICING!N:N)",
            "Somme des Market Values reelles, base de ponderation.",
        ),
        (
            "YTM portefeuille",
            metrics["weighted_ytm"],
            "=SUMPRODUCT(PRICING!O:O,PRICING!G:G)",
            "YTM pondere par Market Value.",
        ),
        (
            "Coupon moyen pondere",
            metrics["weighted_coupon"],
            "=SUMPRODUCT(PRICING!O:O,PRICING!E:E)",
            "Coupon facial pondere par Market Value.",
        ),
        (
            "Duration Macaulay portefeuille",
            metrics["macaulay_duration"],
            "=SUMPRODUCT(PRICING!O:O,DURATION_DETAIL!C:C)",
            "Somme des Dmac individuelles ponderees par Market Value.",
        ),
        (
            "Duration Modifiee portefeuille",
            metrics["modified_duration"],
            "=SUMPRODUCT(PRICING!O:O,DURATION_DETAIL!D:D)",
            "Somme des Dmod individuelles ponderees par Market Value.",
        ),
        (
            "Convexite portefeuille",
            metrics["convexity"],
            "=SUMPRODUCT(PRICING!O:O,DURATION_DETAIL!E:E)",
            "Somme des convexites individuelles ponderees par Market Value.",
        ),
    ]
    return pd.DataFrame(rows, columns=["Metrique", "Valeur", "Formule Excel", "Methodologie"])


def build_portfolio_controls(results: list[BondResult], metrics: dict[str, float]) -> list[Control]:
    controls: list[Control] = []
    weights = [result.portfolio_weight for result in results if result.portfolio_weight is not None]
    weight_sum = sum(weights)
    controls.append(
        Control(
            "Portefeuille",
            "Somme des poids",
            "PASS" if abs(weight_sum - 1.0) < 1e-8 or not weights else "FAIL",
            f"{weight_sum:.10f}",
            "1.0000000000",
            "Les poids doivent etre bases sur la Market Value totale.",
        )
    )
    controls.append(
        Control(
            "Portefeuille",
            "Duration portefeuille calculable",
            "PASS" if math.isfinite(metrics.get("modified_duration", float("nan"))) else "BLOCKED",
            metrics.get("modified_duration"),
            "Nombre fini",
            "Aggregation des sensibilites individuelles.",
        )
    )
    return controls


def sanitize_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    frame = df.copy()
    for column in frame.columns:
        frame[column] = frame[column].map(lambda value: None if isinstance(value, float) and not math.isfinite(value) else value)
    return frame


def write_output_workbook(
    output_path: Path,
    data_original: pd.DataFrame,
    cashflows: pd.DataFrame,
    pricing: pd.DataFrame,
    duration_detail: pd.DataFrame,
    portfolio_metrics: pd.DataFrame,
    controls: pd.DataFrame,
    diagnostics: pd.DataFrame,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        sanitize_for_excel(data_original).to_excel(writer, sheet_name=CANONICAL_OUTPUT_SHEETS[0], index=False)
        sanitize_for_excel(cashflows).to_excel(writer, sheet_name=CANONICAL_OUTPUT_SHEETS[1], index=False)
        sanitize_for_excel(pricing).to_excel(writer, sheet_name=CANONICAL_OUTPUT_SHEETS[2], index=False)
        sanitize_for_excel(duration_detail).to_excel(writer, sheet_name=CANONICAL_OUTPUT_SHEETS[3], index=False)
        sanitize_for_excel(portfolio_metrics).to_excel(writer, sheet_name=CANONICAL_OUTPUT_SHEETS[4], index=False)
        sanitize_for_excel(controls).to_excel(writer, sheet_name=CANONICAL_OUTPUT_SHEETS[5], index=False)
        sanitize_for_excel(diagnostics).to_excel(writer, sheet_name=CANONICAL_OUTPUT_SHEETS[6], index=False)

    workbook = load_workbook(output_path)
    apply_workbook_style(workbook)
    apply_metric_formulas(workbook)
    add_portfolio_chart(workbook)
    workbook.save(output_path)
    workbook.close()


def apply_workbook_style(workbook) -> None:
    thin_border = Side(style="thin", color=THEME["border"])
    header_fill = PatternFill("solid", fgColor=THEME["navy"])
    soft_fill = PatternFill("solid", fgColor=THEME["soft"])
    header_font = Font(color=THEME["white"], bold=True)
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        max_row = sheet.max_row
        max_col = sheet.max_column
        if max_row == 0 or max_col == 0:
            continue
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="medium", color=THEME["blue"]))
        for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = Border(bottom=thin_border)
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if cell.row % 2 == 0:
                    cell.fill = soft_fill
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            values = [sheet.cell(row=row, column=col_idx).value for row in range(1, min(max_row, 80) + 1)]
            width = min(max(max(len(str(value)) if value is not None else 0 for value in values) + 2, 12), 42)
            sheet.column_dimensions[letter].width = width
        if max_row >= 2 and max_col >= 1:
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            table_name = re.sub(r"[^A-Za-z0-9_]", "_", f"tbl_{sheet.title}")[:31]
            table = Table(displayName=table_name, ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False,
            )
            try:
                sheet.add_table(table)
            except ValueError:
                pass
        status_columns = [
            cell.column
            for cell in sheet[1]
            if normalize_header(cell.value) in {"statut", "severite"}
        ]
        for col_idx in status_columns:
            col_letter = get_column_letter(col_idx)
            range_ref = f"{col_letter}2:{col_letter}{max_row}"
            sheet.conditional_formatting.add(
                range_ref,
                CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor="D1FAE5")),
            )
            sheet.conditional_formatting.add(
                range_ref,
                CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill("solid", fgColor="FEE2E2")),
            )
            sheet.conditional_formatting.add(
                range_ref,
                CellIsRule(operator="equal", formula=['"ERROR"'], fill=PatternFill("solid", fgColor="FEE2E2")),
            )
            sheet.conditional_formatting.add(
                range_ref,
                CellIsRule(operator="equal", formula=['"WARNING"'], fill=PatternFill("solid", fgColor="FEF3C7")),
            )
            sheet.conditional_formatting.add(
                range_ref,
                CellIsRule(operator="equal", formula=['"BLOCKED"'], fill=PatternFill("solid", fgColor="E5E7EB")),
            )
        for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                header = normalize_header(sheet.cell(row=1, column=cell.column).value)
                if any(token in header for token in ["ytm", "coupon", "poids", "variation", "return"]):
                    cell.number_format = "0.0000%"
                elif any(token in header for token in ["duration", "convexite", "maturite", "temps"]):
                    cell.number_format = "0.000000"
                elif any(token in header for token in ["price", "prix"]):
                    cell.number_format = "0.000000"
                elif any(token in header for token in ["value", "valeur", "nominal", "coupon couru", "cash flow", "principal"]):
                    cell.number_format = '#,##0.00'
                elif isinstance(cell.value, date):
                    cell.number_format = "yyyy-mm-dd"


def apply_metric_formulas(workbook) -> None:
    if "PORTFOLIO_METRICS" not in workbook.sheetnames:
        return
    sheet = workbook["PORTFOLIO_METRICS"]
    for row in range(2, sheet.max_row + 1):
        formula = sheet.cell(row=row, column=3).value
        if isinstance(formula, str) and formula.startswith("="):
            sheet.cell(row=row, column=2).value = formula
    sheet.column_dimensions["D"].width = 72
    for row in range(2, sheet.max_row + 1):
        metric = normalize_header(sheet.cell(row=row, column=1).value)
        if any(token in metric for token in ["ytm", "coupon"]):
            sheet.cell(row=row, column=2).number_format = "0.0000%"
        elif any(token in metric for token in ["duration", "convexite"]):
            sheet.cell(row=row, column=2).number_format = "0.000000"
        elif "market value" in metric:
            sheet.cell(row=row, column=2).number_format = '#,##0.00'


def add_portfolio_chart(workbook) -> None:
    if "PORTFOLIO_METRICS" not in workbook.sheetnames:
        return
    sheet = workbook["PORTFOLIO_METRICS"]
    if sheet.max_row < 7:
        return
    chart = BarChart()
    chart.title = "Sensibilite portefeuille"
    chart.y_axis.title = "Annees / Convexite"
    chart.x_axis.title = "Indicateur"
    data = Reference(sheet, min_col=2, min_row=5, max_row=7)
    categories = Reference(sheet, min_col=1, min_row=5, max_row=7)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 18
    sheet.add_chart(chart, "F2")


def write_methodology_report(
    report_path: Path,
    input_path: Path,
    output_path: Path,
    metrics: dict[str, float],
    diagnostics: list[Diagnostic],
    controls: list[Control],
    formula_inventory: list[dict[str, Any]],
    analysis_date: date,
) -> None:
    severity_counts = pd.Series([item.severity for item in diagnostics]).value_counts().to_dict() if diagnostics else {}
    control_counts = pd.Series([item.status for item in controls]).value_counts().to_dict() if controls else {}
    lines = [
        "# Rapport methodologique - Audit des durations obligataires",
        "",
        f"Date d'analyse: {analysis_date.isoformat()}",
        f"Fichier source: `{input_path}`",
        f"Fichier corrige: `{output_path}`",
        "",
        "## Synthese portefeuille",
        "",
        f"- Market Value portefeuille: {metrics.get('total_market_value', float('nan')):,.2f}",
        f"- YTM portefeuille: {metrics.get('weighted_ytm', float('nan')):.6%}",
        f"- Coupon moyen pondere: {metrics.get('weighted_coupon', float('nan')):.6%}",
        f"- Duration Macaulay portefeuille: {metrics.get('macaulay_duration', float('nan')):.6f}",
        f"- Duration modifiee portefeuille: {metrics.get('modified_duration', float('nan')):.6f}",
        f"- Convexite portefeuille: {metrics.get('convexity', float('nan')):.6f}",
        "",
        "## Methodologie appliquee",
        "",
        "Les cash flows sont reconstruits titre par titre a partir du nominal, de la frequence de coupon, "
        "de la maturite, du coupon facial, de la convention day count et du type d'amortissement lorsqu'il est renseigne.",
        "",
        "Le prix dirty coherent est obtenu par actualisation actuarielle discrete:",
        "",
        "`P_dirty = somme_t CF_t / (1 + y/m)^(m x t)`",
        "",
        "La duration de Macaulay est calculee comme centre de gravite des flux actualises:",
        "",
        "`D_Mac = somme_t t x PV(CF_t) / P_dirty`",
        "",
        "La duration modifiee est calculee strictement selon:",
        "",
        "`D_Mod = D_Mac / (1 + y/m)`",
        "",
        "La convexite est calculee au second ordre selon la convention de composition discrete:",
        "",
        "`C = somme_t CF_t x k x (k+1) / (1+y/m)^(k+2) / (P_dirty x m^2)`, avec `k = m x t`.",
        "",
        "Les poids portefeuille sont bases exclusivement sur la Market Value corrigee, jamais sur le nominal brut.",
        "",
        "## Conventions gerees",
        "",
        "- Frequences: Annual, Semi-Annual, Quarterly, Monthly, Zero Coupon.",
        "- Day count: ACT/360, ACT/365, 30/360, ACT/ACT.",
        "- Prix: clean, dirty, valeur de marche, coupon couru.",
        "- Sensibilite: stress tests +100 bps et -100 bps, approximation duration seule et duration + convexite.",
        "",
        "## Diagnostic des erreurs",
        "",
        f"Repartition severites: {severity_counts or 'Aucune anomalie detectee.'}",
        "",
    ]
    if diagnostics:
        for item in diagnostics[:80]:
            lines.extend(
                [
                    f"### {item.severity} - {item.instrument_id}",
                    f"- Erreur detectee: {item.error}",
                    f"- Impact financier: {item.financial_impact}",
                    f"- Correction appliquee: {item.correction}",
                    f"- Justification mathematique: {item.mathematical_justification}",
                    "",
                ]
            )
        if len(diagnostics) > 80:
            lines.append(f"_Diagnostics supplementaires disponibles dans le classeur: {len(diagnostics) - 80} lignes._")
            lines.append("")
    lines.extend(
        [
            "## Controles de coherence",
            "",
            f"Repartition statuts: {control_counts or 'Aucun controle disponible.'}",
            "",
            "Les controles incluent notamment: Dmod <= Dmac, zero coupon egal maturite, prix en baisse sous choc +100 bps, "
            "prix en hausse sous choc -100 bps, convexite positive et somme des poids portefeuille.",
            "",
            "## Audit des formules source",
            "",
            f"Nombre de formules inventoriees dans le classeur source: {len(formula_inventory)}.",
            "Les references cassees detectees sont remontees dans `DIAGNOSTIC_ERREURS`.",
            "",
            "## Limites",
            "",
            "Aucune convention critique n'est supposee silencieusement. Les lignes sans nominal, maturite, frequence, "
            "day count, prix ou YTM exploitable sont marquees `BLOCKED` et doivent etre completees avant validation finale.",
            "",
        ]
    )
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(lines), encoding="utf-8")


def run_audit(input_path: Path, output_path: Path, report_path: Path, analysis_date: date) -> dict[str, Any]:
    if not input_path.exists():
        raise FileNotFoundError(f"Fichier source introuvable: {input_path}")
    diagnostics, formula_inventory = workbook_formula_inventory(input_path)
    sheets = read_excel_sheets(input_path)
    original = combine_original_data(sheets)
    bond_df = select_bond_data(sheets)
    bonds = build_bond_inputs(bond_df, analysis_date, diagnostics)
    controls: list[Control] = []
    results: list[BondResult] = []
    all_cashflows: list[Cashflow] = []
    for bond in bonds:
        result, cashflows = compute_bond_result(bond, diagnostics, controls)
        results.append(result)
        all_cashflows.extend(cashflows)
    metrics = update_portfolio_weights(bonds, results, diagnostics)
    controls.extend(build_portfolio_controls(results, metrics))
    write_output_workbook(
        output_path=output_path,
        data_original=original,
        cashflows=cashflows_to_df(all_cashflows),
        pricing=pricing_to_df(results),
        duration_detail=duration_to_df(results),
        portfolio_metrics=portfolio_metrics_to_df(metrics, results),
        controls=controls_to_df(controls),
        diagnostics=diagnostics_to_df(diagnostics),
    )
    write_methodology_report(
        report_path,
        input_path,
        output_path,
        metrics,
        diagnostics,
        controls,
        formula_inventory,
        analysis_date,
    )
    return {
        "input": str(input_path),
        "output": str(output_path),
        "report": str(report_path),
        "bond_count": len(bonds),
        "calculated_count": sum(1 for result in results if result.status == "Calcule"),
        "diagnostic_count": len(diagnostics),
        "control_count": len(controls),
        "metrics": metrics,
    }


def create_demo_input(path: Path) -> None:
    data = pd.DataFrame(
        [
            {
                "ISIN": "DEMO-FIX-5Y",
                "Emetteur": "Demo Treasury",
                "Nominal": 1_000_000,
                "Coupon": 0.045,
                "YTM": 0.052,
                "Date analyse": "2026-05-31",
                "Maturite": "2031-05-31",
                "Frequence coupon": "Annual",
                "Day Count": "ACT/365",
                "Prix clean": 96.95,
                "Duration Modifiee": 1.0,
            },
            {
                "ISIN": "DEMO-ZC-3Y",
                "Emetteur": "Demo Zero",
                "Nominal": 500_000,
                "Coupon": 0.0,
                "YTM": 0.041,
                "Date analyse": "2026-05-31",
                "Maturite": "2029-05-31",
                "Frequence coupon": "Zero Coupon",
                "Day Count": "ACT/365",
                "Prix dirty": 88.64,
            },
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name="Saisir donnee", index=False)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Audit professionnel des durations obligataires.")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Classeur source Excel.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Classeur Excel corrige a produire.")
    parser.add_argument("--report", default=DEFAULT_REPORT, help="Rapport methodologique Markdown a produire.")
    parser.add_argument(
        "--analysis-date",
        default=date.today().isoformat(),
        help="Date d'analyse utilisee lorsque la source ne fournit pas de date de valorisation.",
    )
    parser.add_argument(
        "--self-test",
        action="store_true",
        help="Execute un test fonctionnel sur un mini-portefeuille synthetique dans un dossier temporaire.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    analysis_date = parse_date(args.analysis_date)
    if analysis_date is None:
        raise ValueError("--analysis-date doit etre une date ISO ou reconnue par pandas.")
    if args.self_test:
        with tempfile.TemporaryDirectory(prefix="duration_audit_") as tmp:
            tmp_path = Path(tmp)
            demo_input = tmp_path / "demo_input.xlsx"
            demo_output = tmp_path / DEFAULT_OUTPUT
            demo_report = tmp_path / DEFAULT_REPORT
            create_demo_input(demo_input)
            summary = run_audit(demo_input, demo_output, demo_report, analysis_date)
            print("SELF_TEST_OK")
            print(f"calculated_count={summary['calculated_count']}")
            print(f"diagnostic_count={summary['diagnostic_count']}")
            print(f"output_exists={demo_output.exists()}")
            print(f"report_exists={demo_report.exists()}")
        return 0

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    report_path = Path(args.report).expanduser().resolve()
    summary = run_audit(input_path, output_path, report_path, analysis_date)
    print("AUDIT_DURATION_OBLIGATAIRE_TERMINE")
    print(f"Fichier corrige: {summary['output']}")
    print(f"Rapport: {summary['report']}")
    print(f"Obligations detectees: {summary['bond_count']}")
    print(f"Obligations calculees: {summary['calculated_count']}")
    print(f"Diagnostics: {summary['diagnostic_count']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
