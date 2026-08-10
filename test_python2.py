from io import BytesIO
from openpyxl import Workbook

def build_fonds_propres_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fonds propres"
    
    ws["A2"] = "Groupe"
    ws["B2"] = "Poste (ne pas modifier)"
    ws["C2"] = "Valeur (XOF)"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

if __name__ == "__main__":
    with open("test_python2.xlsx", "wb") as f:
        f.write(build_fonds_propres_import_template())
