"""Récupération en ligne de la courbe des taux UEMOA (UMOA-Titres).

Source : page « Courbes de taux » d'UMOA-Titres, qui publie un fichier Excel
par date (une feuille par pays de l'UEMOA). Le parsing est 100 %
déterministe (la source est déjà tabulaire) : aucune extraction par IA n'est
utilisée dans ce chemin, conformément à l'exigence prudentielle.

Structure d'une feuille pays (vérifiée sur les fichiers publiés) :
- titre en haut : « <Pays> - COURBE DES TAUX <JJ/MM/AAAA> » ;
- tableau à partir de l'en-tête « Maturité » : colonnes Maturité,
  Zero Coupon, Taux Après Lissage ;
- maturités textuelles : « 3 mois », « 6 mois », …, « 1 an », « 2 ans »… ;
- taux exprimés en fraction (0,0648 = 6,48 %).

On retient le « Taux Après Lissage » (courbe de référence lissée) et on le
convertit en pourcentage pour la couche de données VaR (historique_taux :
date, maturite_annees, taux_pct).
"""

from __future__ import annotations

import io
import re
import unicodedata
import urllib.request
from datetime import date

import openpyxl

PAGE_COURBES = "https://www.umoatitres.org/en/ressources-2/courbe-des-taux/"

_ENTETE_HTTP = {"User-Agent": "Mozilla/5.0 (compatible; OutilRWA/1.0)"}
_DELAI = 30

# Mois français pour lire les dates des noms de fichiers.
_MOIS_FR = {
    "janvier": 1, "fevrier": 2, "mars": 3, "avril": 4, "mai": 5, "juin": 6,
    "juillet": 7, "aout": 8, "septembre": 9, "octobre": 10,
    "novembre": 11, "decembre": 12,
}

PAYS_UEMOA = (
    "Bénin", "Burkina", "Côte d'Ivoire", "Guinée-Bissau",
    "Mali", "Niger", "Sénégal", "Togo",
)


class ErreurCourbeUmoa(RuntimeError):
    """Échec de récupération ou de lecture de la courbe UMOA-Titres."""


def _sans_accents(texte: str) -> str:
    plie = unicodedata.normalize("NFD", texte.lower())
    return "".join(c for c in plie if unicodedata.category(c) != "Mn").strip()


def _lire_url(url: str) -> bytes:
    requete = urllib.request.Request(url, headers=_ENTETE_HTTP)
    with urllib.request.urlopen(requete, timeout=_DELAI) as reponse:
        return reponse.read()


def _date_depuis_nom(url: str) -> date | None:
    """Extrait la date d'un nom de fichier de courbe (deux conventions)."""

    # 1. JJ.MM.AAAA (ex. COURBES-DE-TAUX-au.14.10.2024.xlsx)
    m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", url)
    if m:
        jour, mois, annee = (int(g) for g in m.groups())
        try:
            return date(annee, mois, jour)
        except ValueError:
            return None
    # 2. « J-mois-AAAA » ou « J mois AAAA » (ex. Courbe-de-taux-2-septembre-2024)
    m = re.search(r"(\d{1,2})[-_\s]+([a-zA-Zéûôàè]+)[-_\s]+(\d{4})", url)
    if m:
        jour = int(m.group(1))
        mois = _MOIS_FR.get(_sans_accents(m.group(2)))
        annee = int(m.group(3))
        if mois:
            try:
                return date(annee, mois, jour)
            except ValueError:
                return None
    return None


def lien_courbe_la_plus_recente() -> tuple[str, date]:
    """Retourne (url, date) du fichier de courbe le plus récent publié."""

    html = _lire_url(PAGE_COURBES).decode("utf-8", "ignore")
    liens = re.findall(
        r"https://[^\"'\s]+[Cc]ourbe[s]?-?[- ]?[Dd]e-?[- ]?[Tt]aux[^\"'\s]+\.xlsx",
        html,
    )
    liens = list(dict.fromkeys(liens))
    if not liens:
        raise ErreurCourbeUmoa(
            "Aucun fichier de courbe (.xlsx) trouvé sur la page UMOA-Titres. "
            "La structure du site a peut-être changé."
        )
    dates = [(url, _date_depuis_nom(url)) for url in liens]
    dates = [(url, jour) for url, jour in dates if jour is not None]
    if not dates:
        raise ErreurCourbeUmoa("Impossible de dater les fichiers de courbe publiés.")
    url, jour = max(dates, key=lambda couple: couple[1])
    return url, jour


def _maturite_en_annees(texte: object) -> float | None:
    brut = _sans_accents(str(texte))
    m = re.match(r"(\d+(?:[.,]\d+)?)\s*mois", brut)
    if m:
        return float(m.group(1).replace(",", ".")) / 12.0
    m = re.match(r"(\d+(?:[.,]\d+)?)\s*an", brut)
    if m:
        return float(m.group(1).replace(",", "."))
    return None


def _feuille_du_pays(classeur: openpyxl.Workbook, pays: str):
    cible = _sans_accents(pays)
    for nom in classeur.sheetnames:
        if _sans_accents(nom) == cible:
            return classeur[nom]
    # Tolérance : correspondance par préfixe (« cote d ivoire » vs « cote d'ivoire »).
    cible_compact = cible.replace("'", "").replace(" ", "")
    for nom in classeur.sheetnames:
        if _sans_accents(nom).replace("'", "").replace(" ", "") == cible_compact:
            return classeur[nom]
    return None


def _points_feuille(feuille) -> tuple[list[tuple[float, float]], date | None]:
    """Extrait les (maturité en années, taux %) d'une feuille pays."""

    entete = None
    for ligne in feuille.iter_rows():
        for cellule in ligne:
            if cellule.value and _sans_accents(str(cellule.value)) == "maturite":
                entete = (cellule.row, cellule.column)
                break
        if entete:
            break
    if entete is None:
        return [], None

    ligne_entete, col_maturite = entete
    col_taux = col_maturite + 2  # colonne « Taux Après Lissage »

    date_courbe = None
    from datetime import datetime as _dt
    for cellule in feuille["A"]:
        if isinstance(cellule.value, _dt):
            date_courbe = cellule.value.date()

    points: list[tuple[float, float]] = []
    rang = ligne_entete + 1
    while True:
        maturite = feuille.cell(rang, col_maturite).value
        taux = feuille.cell(rang, col_taux).value
        if maturite is None:
            break
        annees = _maturite_en_annees(maturite)
        if annees is not None and taux is not None:
            points.append((round(annees, 4), round(float(taux) * 100.0, 4)))
        rang += 1
    return points, date_courbe


def recuperer_courbe(pays: str) -> dict:
    """Télécharge la courbe la plus récente et renvoie les points du pays.

    Retourne {pays, date, source_url, points: [{maturite_annees, taux_pct}]}.
    Lève ErreurCourbeUmoa en cas d'échec réseau/lecture ou de pays inconnu.
    """

    try:
        url, date_fichier = lien_courbe_la_plus_recente()
        contenu = _lire_url(url)
    except ErreurCourbeUmoa:
        raise
    except Exception as exc:  # réseau, timeout, etc.
        raise ErreurCourbeUmoa(
            f"Récupération de la courbe UMOA-Titres impossible : {exc}"
        ) from exc

    try:
        classeur = openpyxl.load_workbook(io.BytesIO(contenu), data_only=True)
    except Exception as exc:
        raise ErreurCourbeUmoa(
            f"Fichier de courbe illisible : {exc}"
        ) from exc

    feuille = _feuille_du_pays(classeur, pays)
    if feuille is None:
        disponibles = ", ".join(classeur.sheetnames)
        raise ErreurCourbeUmoa(
            f"Pays '{pays}' absent du fichier. Feuilles disponibles : {disponibles}."
        )

    points, date_feuille = _points_feuille(feuille)
    if not points:
        raise ErreurCourbeUmoa(
            f"Aucun point de courbe lisible pour le pays '{pays}'."
        )

    jour = (date_feuille or date_fichier).isoformat()
    return {
        "pays": pays,
        "date": jour,
        "source_url": url,
        "points": [
            {"maturite_annees": annees, "taux_pct": taux}
            for annees, taux in points
        ],
    }
