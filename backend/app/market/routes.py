"""Routes marche pour l'actualisation des courbes de taux."""

from __future__ import annotations

import base64
import json
import re
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from typing import Any

from fastapi import APIRouter, HTTPException, status, UploadFile, File
from fastapi.responses import Response

from app.core.config import settings
import pandas as pd
from app.market.services import MARKET_CAPITAL_REQUIREMENT_KEY
from database.connection import database_manager, utcnow_iso
import os
from pathlib import Path

router = APIRouter(prefix="/market", tags=["Market"])

_MARKET_FX_POSITIONS_KEY = "market_fx_positions_v1"

_MARKET_PORTFOLIOS_PAYLOAD_KEY = "market_portfolios_payload_v1"
_MARKET_PORTFOLIOS_SAVED_AT_KEY = "market_portfolios_saved_at"
_BEAC_PAGE_URL = "https://www.beac.int/economie-stats/statistiques-titres-publics/"
_BEAC_COUNTRY_PDF_URLS = {
    "Cameroun": (
        "https://www.beac.int/wp-content/uploads/2016/10/"
        "Courbe-des-taux-de-rendement-des-titres-publics-Camerounais-mars-26.pdf"
    ),
    "Congo": (
        "https://www.beac.int/wp-content/uploads/2016/10/"
        "Courbe-des-taux-de-rendement-des-titres-publics-Congolais-mars-26.pdf"
    ),
    "Gabon": (
        "https://www.beac.int/wp-content/uploads/2016/10/"
        "Courbe-des-taux-de-rendement-des-titres-publics-Gabonais-mars-26.pdf"
    ),
}
_BEAC_COUNTRY_URL_TOKENS = {
    "Cameroun": ("cameroun", "camerounais"),
    "Congo": ("congo", "congolais"),
    "Gabon": ("gabon", "gabonais"),
}
_CEMAC_MARCH_2026_EXACT_CURVES = {
    "Cameroun": [
        ("3 mois", 0.25, 6.31),
        ("6 mois", 0.5, 7.16),
        ("1 an", 1, 8.32),
        ("1,5 ans", 1.5, 9.00),
        ("2 ans", 2, 9.41),
        ("3 ans", 3, 9.79),
        ("3,5 ans", 3.5, 9.88),
        ("4 ans", 4, 9.92),
        ("5 ans", 5, 9.97),
        ("6 ans", 6, 9.98),
        ("7 ans", 7, 9.99),
        ("8 ans", 8, 9.99),
        ("9 ans", 9, 9.98),
        ("10 ans", 10, 9.98),
        ("11 ans", 11, 9.98),
        ("12 ans", 12, 9.98),
        ("13 ans", 13, 9.98),
        ("14 ans", 14, 9.98),
        ("15 ans", 15, 9.98),
    ],
    "Congo": [
        ("3 mois", 0.25, 6.64),
        ("6 mois", 0.5, 7.67),
        ("1 an", 1, 9.09),
        ("1,5 ans", 1.5, 9.96),
        ("2 ans", 2, 10.49),
        ("3 ans", 3, 11.03),
        ("3,5 ans", 3.5, 11.17),
        ("4 ans", 4, 11.25),
        ("5 ans", 5, 11.35),
        ("6 ans", 6, 11.40),
        ("7 ans", 7, 11.43),
        ("8 ans", 8, 11.45),
        ("9 ans", 9, 11.46),
        ("10 ans", 10, 11.47),
    ],
    "Gabon": [
        ("3 mois", 0.25, 6.12),
        ("6 mois", 0.5, 7.18),
        ("1 an", 1, 8.21),
        ("1,5 ans", 1.5, 8.38),
        ("2 ans", 2, 8.13),
        ("3 ans", 3, 7.19),
        ("3,5 ans", 3.5, 6.70),
        ("4 ans", 4, 6.24),
        ("5 ans", 5, 5.47),
        ("6 ans", 6, 4.88),
        ("7 ans", 7, 4.42),
        ("8 ans", 8, 4.07),
        ("9 ans", 9, 3.79),
        ("10 ans", 10, 3.57),
    ],
}
_GEMINI_ENDPOINT_TEMPLATE = (
    "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
)
_CEMAC_COUNTRIES = {
    "cameroun": "Cameroun",
    "republique centrafricaine": "République Centrafricaine",
    "centrafrique": "République Centrafricaine",
    "congo": "Congo",
    "gabon": "Gabon",
    "guinee equatoriale": "Guinée équatoriale",
    "tchad": "Tchad",
}
_HTTP_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/125.0 Safari/537.36"
    ),
}


@router.get("/portfolios")
def get_market_portfolios() -> dict[str, Any]:
    """Retourne le portefeuille marché persistant depuis SQLite local."""

    with database_manager.read_connection() as connection:
        row = connection.execute(
            "SELECT valeur AS value FROM metadonnees_app WHERE cle = ?",
            (_MARKET_PORTFOLIOS_PAYLOAD_KEY,),
        ).fetchone()

    if row is None or not str(row["value"] or "").strip():
        return {
            "storage_backend": "sqlite",
            "empty": True,
            "payload": None,
        }

    try:
        payload = json.loads(str(row["value"]))
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "code": "MARKET_PORTFOLIO_SQL_PAYLOAD_INVALID",
                "message": "Le portefeuille marché stocké en SQLite est illisible.",
            },
        ) from exc

    return {
        "storage_backend": "sqlite",
        "empty": False,
        "payload": payload,
    }

def _dates_en_iso(df: "pd.DataFrame", colonnes: list[str]) -> "pd.DataFrame":
    """Normalise les colonnes de dates au format AAAA-MM-JJ attendu par la
    couche de données VaR (les cellules Excel arrivent en Timestamp pandas)."""

    for colonne in colonnes:
        if colonne in df.columns:
            df[colonne] = pd.to_datetime(df[colonne], errors="coerce").dt.strftime(
                "%Y-%m-%d"
            )
    return df


def _separer_positions_et_historique(
    df: "pd.DataFrame",
    id_col: str,
    positions_cols: list[str],
    date_col: str,
    prix_col: str,
    hist_cols: list[str],
) -> tuple["pd.DataFrame", "pd.DataFrame"]:
    """Sépare un onglet au format long (une ligne = un titre à une date) en
    (positions dédupliquées, historique des prix datés).

    Les caractéristiques du titre sont répétées sur chaque ligne : seule la
    première occurrence de chaque identifiant est conservée pour les
    positions. Les lignes sans Date ni Prix (position sans historique
    encore fourni) n'alimentent pas l'historique."""

    df = df.dropna(subset=[id_col])
    positions = (
        df[[c for c in positions_cols if c in df.columns]]
        .drop_duplicates(subset=[id_col], keep="first")
        .reset_index(drop=True)
    )
    if date_col in df.columns and prix_col in df.columns:
        historique = df.dropna(subset=[date_col, prix_col])[hist_cols].reset_index(drop=True)
    else:
        historique = pd.DataFrame(columns=hist_cols)
    return positions, historique


@router.post("/upload-var-history")
async def upload_var_history(file: UploadFile = File(...)) -> dict[str, Any]:
    """Parse le fichier Excel (2 onglets : Obligations, Actions) et écrit
    les CSV lus par le module de calcul VaR (backend/data/*.csv).

    Chaque onglet est au format long : une ligne = un titre à une date. Les
    caractéristiques statiques du titre alimentent les positions (première
    occurrence de l'ID) ; les couples (Date, Prix) alimentent l'historique.
    Si un portefeuille a déjà été importé via « Importer données de marché »,
    ses positions restent prioritaires (voir portefeuille_data.py) : ce
    fichier ne sert alors qu'à fournir l'historique de prix, à condition de
    couvrir les mêmes identifiants (ISIN / ticker) que ce portefeuille.
    """
    try:
        content = await file.read()
        import io

        from app.var_marche import portefeuille_data

        excel_file = io.BytesIO(content)

        df_bonds = pd.read_excel(excel_file, sheet_name="Obligations")
        df_bonds = df_bonds.rename(columns={
            "ID Titre": "isin", "Emetteur": "emetteur", "Devise": "devise",
            "Valeur nominale unitaire": "valeur_nominale", "Coupon (%)": "taux_coupon_pct",
            "Fréquence de paiement des intérêts": "frequence_coupon", "Date d'émission": "date_emission",
            "Date d'échéance": "date_echeance", "quantités": "quantite",
            "Date": "date", "Prix de marché (%)": "prix_marche_pct",
        })
        df_bonds = _dates_en_iso(df_bonds, ["date_emission", "date_echeance", "date"])
        if "frequence_coupon" in df_bonds.columns:
            # La colonne accepte un libellé (Annuelle, Semestrielle…) comme
            # dans le formulaire d'import du portefeuille marché ; convertie
            # ici vers le code numérique attendu à la lecture (1, 2, 4, 12).
            df_bonds["frequence_coupon"] = df_bonds["frequence_coupon"].apply(
                lambda v: portefeuille_data.normaliser_frequence_coupon("" if pd.isna(v) else str(v))
            )
        df_bonds_positions, df_hist_bonds = _separer_positions_et_historique(
            df_bonds,
            id_col="isin",
            positions_cols=[
                "isin", "emetteur", "devise", "valeur_nominale", "taux_coupon_pct",
                "frequence_coupon", "date_emission", "date_echeance", "quantite",
            ],
            date_col="date",
            prix_col="prix_marche_pct",
            hist_cols=["date", "isin", "prix_marche_pct"],
        )

        df_equities = pd.read_excel(excel_file, sheet_name="Actions")
        df_equities = df_equities.rename(columns={
            "ID Instrument": "ticker", "Émetteur / Société": "libelle",
            "Secteur": "secteur", "Quantité": "quantite",
            "Date": "date", "Cours de clôture": "cours_cloture",
        })
        df_equities = _dates_en_iso(df_equities, ["date"])
        df_equities_positions, df_hist_equities = _separer_positions_et_historique(
            df_equities,
            id_col="ticker",
            positions_cols=["ticker", "libelle", "secteur", "quantite"],
            date_col="date",
            prix_col="cours_cloture",
            hist_cols=["date", "ticker", "cours_cloture"],
        )
        racine = portefeuille_data.repertoire_data()
        racine.mkdir(parents=True, exist_ok=True)

        _dataframe_vers_csv(
            df_bonds_positions,
            racine / "positions_obligations.csv",
            colonnes_source=[
                "isin", "emetteur", "devise", "valeur_nominale", "taux_coupon_pct",
                "frequence_coupon", "date_emission", "date_echeance", "quantite",
            ],
            entetes=[
                "isin", "emetteur", "devise", "valeur_nominale", "taux_coupon_pct",
                "frequence_coupon", "date_emission", "date_echeance", "quantite",
            ],
        )
        _dataframe_vers_csv(
            df_hist_bonds,
            racine / "historique_prix_obligations.csv",
            colonnes_source=["date", "isin", "prix_marche_pct"],
            entetes=["date", "isin", "prix_pct"],
        )
        _dataframe_vers_csv(
            df_equities_positions,
            racine / "positions_actions.csv",
            colonnes_source=["ticker", "libelle", "secteur", "quantite"],
            entetes=["ticker", "libelle", "secteur", "quantite"],
        )
        _dataframe_vers_csv(
            df_hist_equities,
            racine / "historique_cours_actions.csv",
            colonnes_source=["date", "ticker", "cours_cloture"],
            entetes=["date", "ticker", "cours_cloture"],
        )

        portefeuille_data.invalider_cache_series()

        return {
            "status": "ok",
            "message": "Historique importé avec succès.",
            "obligations_positions": int(len(df_bonds_positions)),
            "obligations_historique": int(len(df_hist_bonds)),
            "actions_positions": int(len(df_equities_positions)),
            "actions_historique": int(len(df_hist_equities)),
        }
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": "VAR_HISTORY_UPLOAD_FAILED", "message": str(exc)},
        ) from exc


def _dataframe_vers_csv(
    df: "pd.DataFrame",
    chemin: Path,
    *,
    colonnes_source: list[str],
    entetes: list[str],
) -> None:
    """Écrit un DataFrame en CSV `;`-délimité (format lu par
    portefeuille_data._lire_csv) : décimales en point, cellules vides pour
    les valeurs manquantes, en-têtes explicitement renommés (colonnes_source
    → entetes) pour matcher le nom attendu côté lecture."""

    presentes = [c for c in colonnes_source if c in df.columns]
    entetes_presentes = [entetes[colonnes_source.index(c)] for c in presentes]
    df[presentes].to_csv(
        chemin,
        sep=";",
        index=False,
        header=entetes_presentes,
        encoding="utf-8",
        lineterminator="\n",
    )


@router.get("/var-history-template")
def download_var_history_template() -> Response:
    """Génère et retourne le modèle Excel d'import d'historique VaR."""
    template_bytes = _build_var_history_template()
    return Response(
        content=template_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=modele_import_var.xlsx"
        },
    )


def _build_var_history_template() -> bytes:
    """Classeur Excel de modèle pour l'import de l'historique VaR.

    Deux onglets seulement (Obligations, Actions), au format long : une
    ligne = un titre à une date. Les caractéristiques du titre sont
    répétées sur chaque ligne ; plusieurs lignes pour un même identifiant,
    avec des dates différentes, construisent son historique de prix dans
    le même onglet — pas de correspondance à maintenir entre des onglets
    séparés positions/historique.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    NAVY = "0F1B3D"
    BORDER_CLR = "CBD5E1"
    EXAMPLE_CLR = "FEF3C7"
    thin = Side(style="thin", color=BORDER_CLR)
    thin_b = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ex_fill = PatternFill("solid", fgColor=EXAMPLE_CLR)
    ex_font = Font(name="Calibri", size=10, italic=True, color="6B7280")

    note_font = Font(name="Calibri", size=9, italic=True, color="6B7280")

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_b

    def style_example_rows(ws, start_row, end_row, ncols):
        for row in range(start_row, end_row + 1):
            for col in range(1, ncols + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = ex_fill
                cell.font = ex_font
                cell.border = thin_b

    def add_note(ws, row, text, col=1):
        # Colonne dédiée (hors de la plage de données) par défaut du côté
        # appelant : une note écrite dans la colonne 1 (ID) serait relue par
        # pandas comme une position malformée lors de l'import.
        cell = ws.cell(row=row, column=col)
        cell.value = text
        cell.font = note_font

    def auto_width(ws, ncols, max_width=35):
        for col in range(1, ncols + 1):
            letter = get_column_letter(col)
            best = 10
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
                for cell in row:
                    if cell.value:
                        best = max(best, min(len(str(cell.value)) + 2, max_width))
            ws.column_dimensions[letter].width = best

    # ── Obligations (positions + historique dans le même onglet) ──────────
    ws = wb.active
    ws.title = "Obligations"
    headers_bonds = [
        "ID Titre", "Emetteur", "Devise",
        "Valeur nominale unitaire", "Coupon (%)",
        "Fréquence de paiement des intérêts",
        "Date d'émission", "Date d'échéance", "quantités",
        "Date", "Prix de marché (%)",
    ]
    ws.append(headers_bonds)
    ws.append([
        "OAT-CI-2023-01", "État de Côte d'Ivoire", "XOF",
        10000, 5.75, "Semestrielle",
        "2023-03-15", "2030-03-15", 150000,
        "2025-01-02", 99.8,
    ])
    ws.append([
        "OAT-CI-2023-01", "État de Côte d'Ivoire", "XOF",
        10000, 5.75, "Semestrielle",
        "2023-03-15", "2030-03-15", 150000,
        "2025-01-03", 99.6,
    ])
    ws.append([
        "OAT-CI-2023-01", "État de Côte d'Ivoire", "XOF",
        10000, 5.75, "Semestrielle",
        "2023-03-15", "2030-03-15", 150000,
        "2025-01-06", 100.1,
    ])
    ws.append([
        "OAT-SN-2024-02", "État du Sénégal", "XOF",
        10000, 6.25, "Annuelle",
        "2024-01-10", "2031-01-10", 200000,
        "2025-01-02", 100.3,
    ])
    ws.append([
        "OAT-SN-2024-02", "État du Sénégal", "XOF",
        10000, 6.25, "Annuelle",
        "2024-01-10", "2031-01-10", 200000,
        "2025-01-03", 100.0,
    ])
    ws.append([
        "OAT-SN-2024-02", "État du Sénégal", "XOF",
        10000, 6.25, "Annuelle",
        "2024-01-10", "2031-01-10", 200000,
        "2025-01-06", 99.7,
    ])
    style_header(ws, len(headers_bonds))
    style_example_rows(ws, 2, 7, len(headers_bonds))
    # Notes en colonne 13 (2 colonnes après la dernière donnée, en 11) : une
    # note écrite dans la colonne 1 (ID Titre) serait relue par pandas comme
    # une position malformée lors de l'import.
    notes_col = len(headers_bonds) + 2
    add_note(ws, 1, "Les lignes d'exemple ci-dessous (en jaune) sont à supprimer avant l'import.", col=notes_col)
    add_note(ws, 2, "Une ligne = un titre à une date : répétez l'ID Titre sur plusieurs lignes,", col=notes_col)
    add_note(ws, 3, "avec une Date et un Prix différents, pour construire son historique.", col=notes_col)
    add_note(ws, 4, "Une ligne sans Date ni Prix est acceptée (position sans historique encore fourni).", col=notes_col)
    add_note(ws, 5, "Coupon (%) : en pourcentage annuel (ex: 5.75 pour 5,75 %).", col=notes_col)
    add_note(ws, 6, "Prix de marché : en pourcentage du nominal (ex: 99.5 pour 99,5 %).", col=notes_col)
    add_note(ws, 7, "Fréquence : Annuelle, Semestrielle, Trimestrielle ou Mensuelle.", col=notes_col)
    add_note(ws, 8, "Dates : format AAAA-MM-JJ ou format Excel date.", col=notes_col)
    add_note(ws, 9, "Minimum requis : 250 dates par titre (fenêtre la plus courte proposée par l'écran VaR).", col=notes_col)
    auto_width(ws, len(headers_bonds))
    ws.column_dimensions[get_column_letter(notes_col)].width = 68

    # ── Actions (positions + historique dans le même onglet) ──────────────
    ws2 = wb.create_sheet("Actions")
    headers_equities = [
        "ID Instrument", "Émetteur / Société", "Secteur", "Quantité",
        "Date", "Cours de clôture",
    ]
    ws2.append(headers_equities)
    ws2.append(["SNTS.CI", "Sonatel SA", "Télécommunications", 5000, "2025-01-02", 18400])
    ws2.append(["SNTS.CI", "Sonatel SA", "Télécommunications", 5000, "2025-01-03", 18550])
    ws2.append(["SNTS.CI", "Sonatel SA", "Télécommunications", 5000, "2025-01-06", 18500])
    ws2.append(["SGBC.CI", "Société Générale CI", "Banque", 12000, "2025-01-02", 9700])
    ws2.append(["SGBC.CI", "Société Générale CI", "Banque", 12000, "2025-01-03", 9800])
    ws2.append(["SGBC.CI", "Société Générale CI", "Banque", 12000, "2025-01-06", 9750])
    style_header(ws2, len(headers_equities))
    style_example_rows(ws2, 2, 7, len(headers_equities))
    notes_col2 = len(headers_equities) + 2
    add_note(ws2, 1, "Les lignes d'exemple ci-dessous (en jaune) sont à supprimer avant l'import.", col=notes_col2)
    add_note(ws2, 2, "Une ligne = un titre à une date : répétez l'ID Instrument sur plusieurs lignes,", col=notes_col2)
    add_note(ws2, 3, "avec une Date et un Cours différents, pour construire son historique.", col=notes_col2)
    add_note(ws2, 4, "Une ligne sans Date ni Cours est acceptée (position sans historique encore fourni).", col=notes_col2)
    add_note(ws2, 5, "Cours de clôture : prix unitaire en devise locale (XOF).", col=notes_col2)
    add_note(ws2, 6, "Minimum requis : 250 dates par instrument (fenêtre la plus courte proposée par l'écran VaR).", col=notes_col2)
    auto_width(ws2, len(headers_equities))
    ws2.column_dimensions[get_column_letter(notes_col2)].width = 68

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.put("/portfolios")
def save_market_portfolios(body: dict[str, Any]) -> dict[str, Any]:
    """Persiste le portefeuille marché dans SQLite local."""

    payload = body.get("payload")
    if not isinstance(payload, dict):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={
                "code": "MARKET_PORTFOLIO_PAYLOAD_INVALID",
                "message": "Payload portefeuille marché invalide.",
            },
        )

    encoded_payload = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    saved_at = utcnow_iso()

    with database_manager.transaction() as connection:
        connection.execute(
            """
            INSERT INTO metadonnees_app(cle, valeur)
            VALUES (?, ?)
            ON CONFLICT(cle) DO UPDATE SET valeur = excluded.valeur
            """,
            (_MARKET_PORTFOLIOS_PAYLOAD_KEY, encoded_payload),
        )
        connection.execute(
            """
            INSERT INTO metadonnees_app(cle, valeur)
            VALUES (?, ?)
            ON CONFLICT(cle) DO UPDATE SET valeur = excluded.valeur
            """,
            (_MARKET_PORTFOLIOS_SAVED_AT_KEY, saved_at),
        )
        connection.execute(
            """
            INSERT INTO metadonnees_app(cle, valeur)
            VALUES (?, ?)
            ON CONFLICT(cle) DO UPDATE SET valeur = excluded.valeur
            """,
            ("market_storage_backend", "sqlite"),
        )

    return {
        "status": "ok",
        "storage_backend": "sqlite",
        "saved_at": saved_at,
    }


@router.post("/portfolios")
def save_market_portfolios_compat(body: dict[str, Any]) -> dict[str, Any]:
    """Alias POST pour les clients qui n'exposent pas PUT."""

    return save_market_portfolios(body)


@router.get("/fx-positions")
def get_fx_positions() -> dict[str, Any]:
    """Retourne les positions de change persistees (module Risque de Change)."""

    with database_manager.read_connection() as connection:
        row = connection.execute(
            "SELECT valeur AS value FROM metadonnees_app WHERE cle = ?",
            (_MARKET_FX_POSITIONS_KEY,),
        ).fetchone()

    if row is None or not str(row["value"] or "").strip():
        return {"storage_backend": "sqlite", "empty": True, "payload": []}

    try:
        payload = json.loads(str(row["value"]))
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "code": "MARKET_FX_POSITIONS_SQL_PAYLOAD_INVALID",
                "message": "Les positions de change stockees en SQLite sont illisibles.",
            },
        ) from exc

    if not isinstance(payload, list):
        payload = []

    return {"storage_backend": "sqlite", "empty": not payload, "payload": payload}


@router.put("/fx-positions")
def save_fx_positions(body: dict[str, Any]) -> dict[str, Any]:
    """Persiste les positions de change dans SQLite local."""

    payload = body.get("payload")
    if not isinstance(payload, list):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={
                "code": "MARKET_FX_POSITIONS_PAYLOAD_INVALID",
                "message": "Payload positions de change invalide.",
            },
        )

    encoded_payload = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    saved_at = utcnow_iso()

    with database_manager.transaction() as connection:
        connection.execute(
            """
            INSERT INTO metadonnees_app(cle, valeur)
            VALUES (?, ?)
            ON CONFLICT(cle) DO UPDATE SET valeur = excluded.valeur
            """,
            (_MARKET_FX_POSITIONS_KEY, encoded_payload),
        )

    return {"status": "ok", "storage_backend": "sqlite", "saved_at": saved_at}


@router.post("/fx-positions")
def save_fx_positions_compat(body: dict[str, Any]) -> dict[str, Any]:
    """Alias POST pour les clients qui n'exposent pas PUT."""

    return save_fx_positions(body)


@router.get("/capital-requirement")
def get_capital_requirement() -> dict[str, Any]:
    """Retourne le dernier resultat calcule (RWA + capital requis) du module Marche."""

    with database_manager.read_connection() as connection:
        row = connection.execute(
            "SELECT valeur AS value FROM metadonnees_app WHERE cle = ?",
            (MARKET_CAPITAL_REQUIREMENT_KEY,),
        ).fetchone()

    if row is None or not str(row["value"] or "").strip():
        return {"empty": True, "rwa_marche": 0.0, "capital_requis": 0.0, "computed_at": None}

    try:
        payload = json.loads(str(row["value"]))
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "code": "MARKET_CAPITAL_REQUIREMENT_SQL_PAYLOAD_INVALID",
                "message": "Le capital requis marche stocke en SQLite est illisible.",
            },
        ) from exc

    return {
        "empty": False,
        "rwa_marche": float(payload.get("rwa_marche", 0.0)),
        "capital_requis": float(payload.get("capital_requis", 0.0)),
        "computed_at": payload.get("computed_at"),
    }


@router.put("/capital-requirement")
def save_capital_requirement(body: dict[str, Any]) -> dict[str, Any]:
    """Persiste le RWA/capital requis marche calcule cote client."""

    rwa_marche = body.get("rwa_marche")
    capital_requis = body.get("capital_requis")
    if not isinstance(rwa_marche, (int, float)) or not isinstance(
        capital_requis, (int, float)
    ):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={
                "code": "MARKET_CAPITAL_REQUIREMENT_PAYLOAD_INVALID",
                "message": "rwa_marche et capital_requis doivent etre numeriques.",
            },
        )

    saved_at = utcnow_iso()
    encoded_payload = json.dumps(
        {
            "rwa_marche": float(rwa_marche),
            "capital_requis": float(capital_requis),
            "computed_at": saved_at,
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )

    with database_manager.transaction() as connection:
        connection.execute(
            """
            INSERT INTO metadonnees_app(cle, valeur)
            VALUES (?, ?)
            ON CONFLICT(cle) DO UPDATE SET valeur = excluded.valeur
            """,
            (MARKET_CAPITAL_REQUIREMENT_KEY, encoded_payload),
        )

    return {"status": "ok", "saved_at": saved_at}


@router.post("/yield-curves/cemac/refresh")
def refresh_cemac_yield_curve() -> dict[str, Any]:
    """Actualise les courbes pays CEMAC depuis les PDFs BEAC."""

    if not _resolve_ai_endpoint():
        return _build_static_cemac_extraction(
            warnings=[
                (
                    "Moteur d'extraction non configure: courbes BEAC mars 2026 "
                    "exactes conservees en local."
                )
            ],
        )

    urls = _find_latest_beac_country_pdf_urls()
    extracted_curves: list[Any] = []
    warnings: list[str] = []
    date_source = "Mars 2026"
    for country, source_url in urls.items():
        try:
            pdf_bytes = _download_bytes(source_url)
            ai_payload = _call_ai_extractor(
                pdf_bytes=pdf_bytes,
                source_url=source_url,
            )
            raw_curves = ai_payload.get("courbes") or ai_payload.get("curves")
            if isinstance(raw_curves, list):
                extracted_curves.extend(raw_curves)
            
            # Extract date if available
            ai_date = ai_payload.get("date_source")
            if ai_date and str(ai_date).strip():
                date_source = str(ai_date).strip()

        except HTTPException as exc:
            warnings.append(
                f"{country}: extraction PDF indisponible ({exc.status_code})."
            )

    if not extracted_curves:
        return _build_static_cemac_extraction(
            warnings=[
                *warnings,
                (
                    "Aucune extraction PDF exploitable: courbes BEAC mars 2026 "
                    "exactes conservees en local."
                ),
            ],
        )

    payload = {
        "zone": "CEMAC",
        "source": "BEAC",
        "date_source": date_source,
        "courbes": extracted_curves,
    }
    result = _normalize_cemac_extraction(payload, source_url=_BEAC_PAGE_URL)
    result["warnings"] = [*result.get("warnings", []), *warnings]
    return result


def _find_latest_beac_country_pdf_urls() -> dict[str, str]:
    urls = dict(_BEAC_COUNTRY_PDF_URLS)
    try:
        html = _download_text(_BEAC_PAGE_URL)
    except HTTPException:
        return urls

    candidates = re.findall(
        r"""href=["']([^"']*Courbe[^"']*\.pdf)["']""",
        html,
        flags=re.IGNORECASE,
    )
    resolved_countries = set()
    for raw_href in candidates:
        href = raw_href.replace("&amp;", "&")
        normalized_href = _normalize_text(urllib.parse.unquote(href))
        for country, tokens in _BEAC_COUNTRY_URL_TOKENS.items():
            if country in resolved_countries:
                continue
            if any(token in normalized_href for token in tokens):
                urls[country] = urllib.parse.urljoin(_BEAC_PAGE_URL, href)
                resolved_countries.add(country)
                break
    return urls


def _download_text(url: str) -> str:
    data = _download_bytes(url)
    return data.decode("utf-8", errors="replace")


def _download_bytes(url: str) -> bytes:
    request = urllib.request.Request(url, headers=_HTTP_HEADERS)
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            return response.read()
    except (urllib.error.URLError, TimeoutError) as exc:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail={
                "code": "SOURCE_DOWNLOAD_FAILED",
                "message": f"Source BEAC indisponible: {exc}",
            },
        ) from exc


def _call_ai_extractor(*, pdf_bytes: bytes, source_url: str) -> dict[str, Any]:
    endpoint = _resolve_ai_endpoint()
    if _is_gemini_endpoint(endpoint):
        return _call_gemini_extractor(
            endpoint=endpoint,
            pdf_bytes=pdf_bytes,
            source_url=source_url,
        )

    document_base64 = base64.b64encode(pdf_bytes).decode("ascii")
    payload = {
        "task": "extract_cemac_yield_curves",
        "model": settings.yield_curve_ai_model or None,
        "source": {
            "name": "BEAC",
            "zone": "CEMAC",
            "url": source_url,
            "mime_type": "application/pdf",
        },
        "instructions": (
            "Extraire uniquement les courbes de rendement des titres publics "
            "CEMAC presentes dans le document BEAC. Ne jamais inventer un pays "
            "ou une maturite absente. Retourner un JSON strict avec: zone, "
            "source, date_source, courbes[]. Chaque courbe doit avoir pays et "
            "points[]. Chaque point doit avoir maturite, annees, taux. Si la "
            "source separe taux_brut et taux_lisse, les inclure. Le taux doit "
            "etre en pourcentage annuel, par exemple 6.35 pour 6,35%."
        ),
        "expected_countries": [
            "Cameroun",
            "République Centrafricaine",
            "Congo",
            "Gabon",
            "Guinée équatoriale",
            "Tchad",
        ],
        "document_base64": document_base64,
    }
    body = json.dumps(payload).encode("utf-8")
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    if settings.yield_curve_ai_api_key.strip():
        headers["Authorization"] = f"Bearer {settings.yield_curve_ai_api_key.strip()}"

    request = urllib.request.Request(
        endpoint,
        data=body,
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            raw = response.read().decode("utf-8", errors="replace")
    except (urllib.error.URLError, TimeoutError) as exc:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail={
                "code": "AI_EXTRACTOR_FAILED",
                "message": f"Extraction PDF indisponible: {exc}",
            },
        ) from exc

    try:
        decoded = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail={
                "code": "AI_EXTRACTOR_INVALID_JSON",
                "message": "Le moteur d'extraction n'a pas renvoye un JSON lisible.",
            },
        ) from exc

    return _unwrap_ai_payload(decoded)


def _resolve_ai_endpoint() -> str:
    endpoint = settings.yield_curve_ai_endpoint.strip()
    if endpoint:
        return endpoint
    if not settings.yield_curve_ai_api_key.strip():
        return ""
    model = settings.yield_curve_ai_model.strip() or "gemini-2.5-flash"
    return _GEMINI_ENDPOINT_TEMPLATE.format(model=urllib.parse.quote(model, safe="-_."))


def _is_gemini_endpoint(endpoint: str) -> bool:
    return "generativelanguage.googleapis.com" in endpoint.lower()


def _call_gemini_extractor(
    *,
    endpoint: str,
    pdf_bytes: bytes,
    source_url: str,
) -> dict[str, Any]:
    document_base64 = base64.b64encode(pdf_bytes).decode("ascii")
    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [
                    {
                        "inline_data": {
                            "mime_type": "application/pdf",
                            "data": document_base64,
                        }
                    },
                    {"text": _gemini_extraction_prompt(source_url)},
                ],
            }
        ],
        "generationConfig": {
            "temperature": 0,
            "responseMimeType": "application/json",
            "responseJsonSchema": _cemac_response_schema(),
        },
    }
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "x-goog-api-key": settings.yield_curve_ai_api_key.strip(),
    }
    request = urllib.request.Request(
        endpoint,
        data=json.dumps(payload).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            raw = response.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail={
                "code": "GEMINI_EXTRACTOR_FAILED",
                "message": f"Gemini a refuse l'extraction: {detail[:900]}",
            },
        ) from exc
    except (urllib.error.URLError, TimeoutError) as exc:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail={
                "code": "GEMINI_EXTRACTOR_FAILED",
                "message": f"Gemini indisponible: {exc}",
            },
        ) from exc

    try:
        decoded = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail={
                "code": "GEMINI_EXTRACTOR_INVALID_JSON",
                "message": "Gemini n'a pas renvoye une reponse JSON lisible.",
            },
        ) from exc
    return _unwrap_ai_payload(decoded)


def _gemini_extraction_prompt(source_url: str) -> str:
    return (
        "Tu es un moteur d'extraction de donnees, pas un assistant de chat. "
        "Lis le PDF BEAC fourni et extrais uniquement les courbes de rendement "
        "des titres publics de la zone CEMAC presentes dans le document. "
        "La CEMAC peut inclure Cameroun, Republique Centrafricaine, Congo, "
        "Gabon, Guinee equatoriale et Tchad. Ne jamais inventer une courbe, un "
        "pays, une maturite ou un taux absent du document. Si seules certaines "
        "courbes pays sont visibles, retourne uniquement celles-la. Les taux "
        "doivent etre en pourcentage annuel, par exemple 6.35 pour 6,35 %. "
        "Chaque point doit contenir maturite, annees et taux. Si le document "
        "separe taux brut et taux lisse, retourne aussi taux_brut et "
        "taux_lisse. Identifie également la date (mois et année) de publication "
        "ou de validité des courbes (ex: Juillet 2024) et place la dans le champ "
        "date_source. Source: "
        f"{source_url}"
    )


def _cemac_response_schema() -> dict[str, Any]:
    return {
        "type": "object",
        "properties": {
            "zone": {"type": "string"},
            "source": {"type": "string"},
            "date_source": {"type": "string"},
            "courbes": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "pays": {"type": "string"},
                        "points": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "maturite": {"type": "string"},
                                    "annees": {"type": "number"},
                                    "taux": {"type": "number"},
                                    "taux_brut": {"type": "number"},
                                    "taux_lisse": {"type": "number"},
                                },
                                "required": ["maturite", "annees", "taux"],
                            },
                        },
                    },
                    "required": ["pays", "points"],
                },
            },
        },
        "required": ["zone", "source", "date_source", "courbes"],
    }


def _unwrap_ai_payload(decoded: Any) -> dict[str, Any]:
    if isinstance(decoded, dict) and "choices" in decoded:
        content = (
            decoded.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "")
        )
        decoded = _parse_embedded_json(content)
    elif isinstance(decoded, dict) and "candidates" in decoded:
        parts = (
            decoded.get("candidates", [{}])[0]
            .get("content", {})
            .get("parts", [])
        )
        content = "".join(
            part.get("text", "") for part in parts if isinstance(part, dict)
        )
        decoded = _parse_embedded_json(content)
    elif isinstance(decoded, dict) and isinstance(decoded.get("data"), dict):
        decoded = decoded["data"]
    elif isinstance(decoded, str):
        decoded = _parse_embedded_json(decoded)

    if not isinstance(decoded, dict):
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail={
                "code": "AI_EXTRACTOR_UNSUPPORTED_RESPONSE",
                "message": "Reponse d'extraction non compatible avec le schema attendu.",
            },
        )
    return decoded


def _parse_embedded_json(value: str) -> dict[str, Any]:
    cleaned = value.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
    if match:
        cleaned = match.group(0)
    return json.loads(cleaned)


def _normalize_cemac_extraction(
    payload: dict[str, Any],
    *,
    source_url: str,
) -> dict[str, Any]:
    raw_curves = payload.get("courbes") or payload.get("curves") or []
    curves = []
    warnings = []

    for raw_curve in raw_curves:
        if not isinstance(raw_curve, dict):
            continue
        country = _canonical_country(
            str(
                raw_curve.get("pays")
                or raw_curve.get("country")
                or raw_curve.get("nom")
                or raw_curve.get("name")
                or ""
            )
        )
        if country is None:
            warnings.append("Courbe ignoree: pays CEMAC non reconnu.")
            continue

        points = _normalize_curve_points(raw_curve.get("points") or [])
        if len(points) < 2:
            warnings.append(f"Courbe {country} ignoree: points insuffisants.")
            continue
        curves.append({"country": country, "points": points})

    if not curves:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={
                "code": "AI_EXTRACTOR_EMPTY_CURVES",
                "message": "Aucune courbe pays CEMAC exploitable n'a ete extraite.",
            },
        )

    present = {curve["country"] for curve in curves}
    absent = sorted(set(_CEMAC_COUNTRIES.values()) - present)
    aggregate_points = _aggregate_points(curves)
    return {
        "status": "ok",
        "zone": "CEMAC",
        "source": payload.get("source") or "BEAC",
        "source_url": source_url,
        "source_date_label": _source_date_label(payload),
        "methodology": (
            "Actualisation CEMAC depuis les PDFs BEAC pays, "
            "conservation des courbes extraites et calcul de la courbe zone "
            "par moyenne simple des maturites communes."
        ),
        "curves": curves,
        "aggregate_points": aggregate_points,
        "countries_absent_from_document": absent,
        "warnings": warnings,
        "message": (
            f"CEMAC actualisee: {len(curves)} courbe(s) pays extraite(s)."
        ),
    }


def _build_static_cemac_extraction(
    *,
    warnings: list[str] | None = None,
) -> dict[str, Any]:
    curves = [
        {
            "country": country,
            "points": [
                {
                    "maturity": maturity,
                    "years": years,
                    "rate": rate,
                    "raw_rate": rate,
                    "smoothed_rate": rate,
                }
                for maturity, years, rate in points
            ],
        }
        for country, points in _CEMAC_MARCH_2026_EXACT_CURVES.items()
    ]
    absent = sorted(
        set(_CEMAC_COUNTRIES.values()) - set(_CEMAC_MARCH_2026_EXACT_CURVES)
    )
    return {
        "status": "ok",
        "zone": "CEMAC",
        "source": "BEAC",
        "source_url": _BEAC_PAGE_URL,
        "source_date_label": "Mars 2026",
        "methodology": (
            "Actualisation CEMAC prevue depuis les PDFs "
            "BEAC pays. En absence d'Internet, l'outil conserve ce "
            "jeu local valide BEAC mars 2026; la courbe zone est la moyenne "
            "simple des maturites communes aux pays disponibles."
        ),
        "curves": curves,
        "aggregate_points": _aggregate_points(curves),
        "countries_absent_from_document": absent,
        "warnings": warnings or [],
        "message": "CEMAC chargee depuis les tables BEAC mars 2026.",
    }


def _normalize_curve_points(raw_points: Any) -> list[dict[str, float | str]]:
    if not isinstance(raw_points, list):
        return []

    points = []
    seen = set()
    for raw_point in raw_points:
        if not isinstance(raw_point, dict):
            continue
        label = str(
            raw_point.get("maturite")
            or raw_point.get("maturity")
            or raw_point.get("label")
            or ""
        )
        years = _number_or_none(raw_point.get("annees") or raw_point.get("years"))
        if years is None:
            years = _parse_maturity_years(label)
        rate = _number_or_none(
            raw_point.get("taux_lisse")
            or raw_point.get("taux_lissé")
            or raw_point.get("smoothed_rate")
            or raw_point.get("smoothedRate")
            or raw_point.get("taux")
            or raw_point.get("rate")
        )
        raw_rate = _number_or_none(
            raw_point.get("taux_brut")
            or raw_point.get("raw_rate")
            or raw_point.get("rawRate")
        )
        if years is None or rate is None or years <= 0:
            continue
        if rate <= 1:
            rate *= 100
        if raw_rate is None:
            raw_rate = rate
        elif raw_rate <= 1:
            raw_rate *= 100
        if rate <= 0 or rate > 30:
            continue
        key = round(years, 4)
        if key in seen:
            continue
        seen.add(key)
        points.append(
            {
                "maturity": _format_maturity_label(years, fallback=label),
                "years": round(years, 6),
                "rate": round(rate, 6),
                "raw_rate": round(raw_rate, 6),
                "smoothed_rate": round(rate, 6),
            }
        )

    points.sort(key=lambda point: float(point["years"]))
    return points


def _aggregate_points(curves: list[dict[str, Any]]) -> list[dict[str, float | str]]:
    grouped: dict[float, list[float]] = {}
    for curve in curves:
        for point in curve["points"]:
            key = round(float(point["years"]), 6)
            grouped.setdefault(key, []).append(
                float(point.get("smoothed_rate") or point["rate"])
            )

    aggregate = []
    required_count = len(curves)
    for years, rates in sorted(grouped.items()):
        if required_count > 1 and len(rates) < required_count:
            continue
        aggregate.append(
            {
                "maturity": _format_maturity_label(years),
                "years": years,
                "rate": round(sum(rates) / len(rates), 6),
                "raw_rate": round(sum(rates) / len(rates), 6),
                "smoothed_rate": round(sum(rates) / len(rates), 6),
            }
        )
    return aggregate


def _canonical_country(value: str) -> str | None:
    normalized = _normalize_text(value)
    for token, country in _CEMAC_COUNTRIES.items():
        if token in normalized:
            return country
    return None


def _normalize_text(value: str) -> str:
    without_accents = "".join(
        char
        for char in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(char)
    )
    return re.sub(r"\s+", " ", without_accents.lower()).strip()


def _number_or_none(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    cleaned = str(value).replace("%", "").replace(",", ".").strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_maturity_years(label: str) -> float | None:
    normalized = _normalize_text(label).replace(",", ".")
    match = re.search(r"(\d+(?:\.\d+)?)", normalized)
    if not match:
        return None
    amount = float(match.group(1))
    return amount / 12 if "mois" in normalized else amount


def _format_maturity_label(years: float, fallback: str = "") -> str:
    if fallback.strip():
        return fallback.strip()
    if years < 1:
        months = round(years * 12)
        return f"{months} mois"
    if abs(years - round(years)) < 0.001:
        rounded = round(years)
        return f"{rounded} an" if rounded == 1 else f"{rounded} ans"
    return f"{str(round(years, 1)).replace('.', ',')} ans"


def _source_date_label(payload: dict[str, Any]) -> str:
    for key in ("date_source", "source_date", "sourceDate", "source_date_label"):
        value = payload.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return "Source BEAC actualisee"
