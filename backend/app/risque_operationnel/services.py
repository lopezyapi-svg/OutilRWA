"""Services du module Risque Opérationnel."""

from __future__ import annotations

import uuid
from datetime import date, datetime, timedelta

from database.connection import database_manager, utcnow_iso

from .models import (
    AibCalculResult,
    AsAnneeDetail,
    AsCalculResult,
    AsLigneDetail,
    BetaLigneUpdate,
    BetaLigneView,
    ControleCreate,
    ControleUpdate,
    ControleView,
    DashboardData,
    DecisionCritere,
    DecisionPilotageResult,
    DashboardWidget1,
    DashboardWidget2,
    DashboardWidget3,
    EvolutionPoint,
    HistoriqueView,
    IncidentBulkImportRequest,
    IncidentCreate,
    IncidentImportResult,
    IncidentUpdate,
    IncidentView,
    KriModuleData,
    KriValeurCreate,
    KriValeurView,
    KriView,
    OpRiskBiDetail,
    OpRiskCalculResult,
    OpRiskFcDetail,
    OpRiskIldcDetail,
    OpRiskInput,
    OpRiskInputUpdate,
    OpRiskParametres,
    OpRiskParametresUpdate,
    OpRiskScDetail,
    ParametresAib,
    ParametresAibUpdate,
    ParametresAs,
    ParametresAsUpdate,
    ParametresSeuils,
    ParametresSeuilsUpdate,
    PlanCreate,
    PlanUpdate,
    PlanView,
    PnbAnnuelCreate,
    PnbAnnuelView,
    PnbParLigneCreate,
    PnbParLigneView,
    RepartitionItem,
    RisqueCreate,
    RisqueUpdate,
    RisqueView,
    SyntheseLigne,
    SyntheseResult,
)

_LIGNES_METIER = [
    "Financement d'entreprise",
    "Activités de marché",
    "Banque de détail",
    "Banque commerciale",
    "Paiements et règlements",
    "Fonctions d'agent",
    "Gestion d'actifs",
    "Courtage de détail",
]

# ─── helpers ──────────────────────────────────────────────────────────────────

def _new_id() -> str:
    return str(uuid.uuid4())


def _today() -> str:
    return date.today().isoformat()


def _seq_reference(prefix: str, conn) -> str:
    year = date.today().year
    table_map = {
        "INC": "ro_incidents",
        "CTRL": "ro_controles",
        "ACT": "ro_plans_action",
    }
    table = table_map.get(prefix, "ro_incidents")
    row = conn.execute(
        f"SELECT COUNT(*) as cnt FROM {table} WHERE reference LIKE ?",
        (f"{prefix}-{year}-%",),
    ).fetchone()
    seq = (row["cnt"] if row else 0) + 1
    return f"{prefix}-{year}-{seq:04d}"


def _niveau_label(niveau: float) -> str:
    if niveau <= 4:
        return "Faible"
    if niveau <= 9:
        return "Moyen"
    if niveau <= 16:
        return "Élevé"
    return "Critique"


def _prochaine_date(date_str: str, frequence: str) -> str:
    if not date_str:
        return ""
    try:
        d = date.fromisoformat(date_str)
    except ValueError:
        return ""
    delta_map = {
        "Mensuel": timedelta(days=30),
        "Trimestriel": timedelta(days=91),
        "Semestriel": timedelta(days=182),
        "Annuel": timedelta(days=365),
    }
    delta = delta_map.get(frequence, timedelta(days=30))
    return (d + delta).isoformat()


def _log(conn, menu: str, type_action: str, element: str, ancienne: str = "", nouvelle: str = "") -> None:
    conn.execute(
        """INSERT INTO ro_historique (date_evenement, utilisateur, menu, type_action, element, ancienne_valeur, nouvelle_valeur)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (utcnow_iso(), "Système", menu, type_action, element, ancienne, nouvelle),
    )


# ─── Incidents ────────────────────────────────────────────────────────────────

def _row_to_incident(row) -> IncidentView:
    perte_nette = (row["perte_brute"] or 0) - (row["perte_recuperee"] or 0)
    return IncidentView(
        id=row["id"],
        reference=row["reference"],
        date_occurrence=row["date_occurrence"],
        description=row["description"],
        ligne_metier=row["ligne_metier"],
        type_evenement=row["type_evenement"],
        cause_racine=row["cause_racine"] or "",
        perte_brute=row["perte_brute"] or 0,
        perte_recuperee=row["perte_recuperee"] or 0,
        perte_nette=perte_nette,
        statut=row["statut"],
        significatif=perte_nette > 0,
        cree_le=row["cree_le"],
        modifie_le=row["modifie_le"],
    )


def list_incidents(statut: str | None = None, ligne_metier: str | None = None) -> list[IncidentView]:
    with database_manager.transaction() as conn:
        query = "SELECT * FROM ro_incidents WHERE 1=1"
        params: list = []
        if statut:
            query += " AND statut = ?"
            params.append(statut)
        if ligne_metier:
            query += " AND ligne_metier = ?"
            params.append(ligne_metier)
        query += " ORDER BY date_occurrence DESC"
        rows = conn.execute(query, params).fetchall()
    return [_row_to_incident(r) for r in rows]


def create_incident(data: IncidentCreate) -> IncidentView:
    id_ = _new_id()
    now = utcnow_iso()
    with database_manager.transaction() as conn:
        ref = _seq_reference("INC", conn)
        conn.execute(
            """INSERT INTO ro_incidents (id, reference, date_occurrence, description, ligne_metier,
               type_evenement, cause_racine, perte_brute, perte_recuperee, statut, cree_le, modifie_le)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (id_, ref, data.date_occurrence, data.description, data.ligne_metier,
             data.type_evenement, data.cause_racine, data.perte_brute,
             data.perte_recuperee, data.statut, now, now),
        )
        _log(conn, "Incidents", "CREATE", ref, "", data.description[:80])
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT * FROM ro_incidents WHERE id = ?", (id_,)).fetchone()
    return _row_to_incident(row)


def update_incident(id_: str, data: IncidentUpdate) -> IncidentView:
    now = utcnow_iso()
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        with database_manager.transaction() as conn:
            row = conn.execute("SELECT * FROM ro_incidents WHERE id = ?", (id_,)).fetchone()
        return _row_to_incident(row)
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    with database_manager.transaction() as conn:
        old = conn.execute("SELECT * FROM ro_incidents WHERE id = ?", (id_,)).fetchone()
        conn.execute(
            f"UPDATE ro_incidents SET {set_clause}, modifie_le = ? WHERE id = ?",
            (*updates.values(), now, id_),
        )
        _log(conn, "Incidents", "UPDATE", old["reference"] if old else id_)
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT * FROM ro_incidents WHERE id = ?", (id_,)).fetchone()
    return _row_to_incident(row)


def delete_incident(id_: str) -> None:
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT reference FROM ro_incidents WHERE id = ?", (id_,)).fetchone()
        conn.execute("DELETE FROM ro_incidents WHERE id = ?", (id_,))
        _log(conn, "Incidents", "DELETE", row["reference"] if row else id_)


def build_ro_import_template() -> bytes:
    """Génère un classeur Excel de modèle pour l'import des pertes opérationnelles."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Palette ─────────────────────────────────────────────────────────────
    BLUE_DARK   = "1D4ED8"
    BLUE_LIGHT  = "DBEAFE"
    GREEN_DARK  = "15803D"
    GREEN_LIGHT = "DCFCE7"
    YELLOW_BG   = "FEF9C3"
    RED_LIGHT   = "FEE2E2"
    GREY_HEADER = "F1F5F9"
    BORDER_CLR  = "CBD5E1"

    thin = Side(style="thin", color=BORDER_CLR)
    thin_b = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(bold=True, size=11, color="000000"):
        return Font(bold=bold, size=size, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def left(wrap=False):
        return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

    # ── Feuille 1 : INCIDENTS (saisie) ────────────────────────────────────
    ws = wb.active
    ws.title = "Incidents"
    ws.freeze_panes = "A3"

    # Titre
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Modèle d'import — Pertes Opérationnelles (Risque Opérationnel BCEAO)"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = fill(BLUE_DARK)
    title_cell.alignment = center()

    # En-têtes colonnes
    headers = [
        ("A", "date_occurrence",  "Date d'occurrence",    "AAAA-MM-JJ ou JJ/MM/AAAA",   True,  15),
        ("B", "description",      "Description",           "Libellé de l'incident",       True,  35),
        ("C", "ligne_metier",     "Ligne de métier",       "Voir onglet Lignes_Métier",   True,  22),
        ("D", "type_evenement",   "Type d'événement",      "Voir onglet Types_Événement", True,  18),
        ("E", "cause_racine",     "Cause racine",          "Optionnel",                   False, 20),
        ("F", "perte_brute",      "Perte brute (FCFA)",    "Montant numérique",           True,  18),
        ("G", "perte_recuperee",  "Perte récupérée (FCFA)","Assurance / provisions",      False, 20),
        ("H", "statut",           "Statut",                "Optionnel — défaut: Ouvert",  False, 14),
    ]

    ws.row_dimensions[2].height = 40
    for col_letter, _, label, hint, required, width in headers:
        c = ws[f"{col_letter}2"]
        c.value = f"{'★ ' if required else '○ '}{label}"
        c.font = Font(bold=True, size=10, color=BLUE_DARK if required else "475569")
        c.fill = fill(BLUE_LIGHT if required else GREY_HEADER)
        c.border = thin_b
        c.alignment = center(wrap=True)
        ws.column_dimensions[col_letter].width = width

        # Ligne hint sous l'en-tête (row 3)
        h = ws[f"{col_letter}3"]
        h.value = hint
        h.font = Font(italic=True, size=8, color="64748B")
        h.fill = fill("F8FAFC")
        h.border = thin_b
        h.alignment = center(wrap=True)

    ws.row_dimensions[3].height = 18

    # Lignes d'exemple (une par type d'événement)
    exemples = [
        ("2025-01-15", "Erreur de saisie virement client — doublon déclenché",            "Banque de détail",   "Processus", "Erreur humaine",       450000, 0,       "Résolu"),
        ("2025-02-03", "Tentative de phishing — données compromises employé",              "Banque commerciale", "Interne",   "Fraude interne",       0,      0,       "En cours"),
        ("2025-03-10", "Fraude carte bancaire — 12 transactions non autorisées",           "Banque de détail",   "Externe",   "Fraude externe",       1800000, 600000, "Résolu"),
        ("2025-03-22", "Panne du core banking 6h — transactions en attente",               "Paiements et règlements", "Système", "Défaillance système", 3200000, 0,    "Clôturé"),
        ("2025-04-08", "Litige client — mauvais conseil produit structuré",                "Financement d'entreprise", "Juridique", "Processus inadéquat", 7500000, 2500000, "En cours"),
        ("2025-05-14", "Accident du travail — blessure caissier",                          "Banque de détail",   "Personnel", "Événement externe",    0,      0,       "Clôturé"),
        ("2025-06-01", "Erreur règlement titre — mauvaise quantité exécutée",              "Activités de marché","Processus", "Erreur humaine",       950000, 950000, "Clôturé"),
        ("2025-06-20", "Perte physique fonds — vol agence",                                "Banque de détail",   "Externe",   "Fraude externe",       2100000, 400000, "En cours"),
    ]

    for row_idx, ex in enumerate(exemples, start=4):
        row_fill = fill("FFFFFF") if row_idx % 2 == 0 else fill("F8FAFC")
        vals = list(ex)
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.border = thin_b
            c.fill = row_fill
            c.alignment = left(wrap=(ci == 2))
            c.font = Font(size=10)

    # ── Feuille 2 : Instructions ────────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 58

    ws2.merge_cells("A1:B1")
    ws2["A1"].value = "Instructions — Import des pertes opérationnelles"
    ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = fill(BLUE_DARK)
    ws2["A1"].alignment = center()

    instructions = [
        ("Feuille de saisie",   "Saisir vos incidents dans la feuille « Incidents » à partir de la ligne 4."),
        ("Ligne 2",             "★ = colonne OBLIGATOIRE   ○ = colonne optionnelle (peut rester vide)"),
        ("date_occurrence",     "Format ISO AAAA-MM-JJ (ex : 2025-06-15) ou JJ/MM/AAAA (ex : 15/06/2025)."),
        ("ligne_metier",        "Valeur exacte parmi les 8 options."),
        ("type_evenement",      "Valeur exacte parmi 6 options (Interne, Externe, Processus, Système, Personnel, Juridique)."),
        ("cause_racine",        "Optionnel : Erreur humaine | Défaillance système | Processus inadéquat | Fraude interne | Fraude externe | Événement externe | Non définie."),
        ("perte_brute",         "Montant total de la perte avant récupération, en FCFA (nombre entier ou décimal)."),
        ("perte_recuperee",     "Montant récupéré (assurance, provisions) — 0 si rien. Perte nette = brute − récupérée."),
        ("statut",              "Optionnel : Ouvert | En cours | Résolu | Clôturé. Défaut = Ouvert."),
        ("Lignes vides",        "Les lignes entièrement vides sont ignorées."),
        ("Doublons",            "Aucune vérification de doublon — vérifier avant import en mode remplacement."),
    ]

    for ri, (label, text) in enumerate(instructions, start=2):
        c_label = ws2.cell(row=ri, column=1, value=label)
        c_text  = ws2.cell(row=ri, column=2, value=text)
        row_fill = fill(BLUE_LIGHT) if label and label != "" else fill("FFFFFF")
        c_label.font = Font(bold=bool(label), size=10, color=BLUE_DARK if label else "000000")
        c_label.fill = row_fill
        c_label.border = thin_b
        c_label.alignment = left()
        c_text.font = Font(size=10)
        c_text.fill = fill("FFFFFF")
        c_text.border = thin_b
        c_text.alignment = left(wrap=True)
        ws2.row_dimensions[ri].height = 20

    # Onglet de référence : pas de validation externe ni de feuilles supplémentaires
    ws2.sheet_view.tabSelected = False

    # ── Export en bytes ────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


_LIGNES_METIER = [
    "Financement d'entreprise", "Activités de marché", "Banque de détail",
    "Banque commerciale", "Paiements et règlements", "Fonctions d'agent",
    "Gestion d'actifs", "Courtage de détail",
]
_TYPES_EVENEMENT = ["Interne", "Externe", "Processus", "Système", "Personnel", "Juridique"]
_CAUSES_RACINE = [
    "Erreur humaine", "Défaillance système", "Processus inadéquat",
    "Fraude interne", "Fraude externe", "Événement externe", "Non définie",
]
_STATUTS_INCIDENT = ["Ouvert", "En cours", "Résolu", "Clôturé"]


def bulk_import_incidents(data: IncidentBulkImportRequest) -> IncidentImportResult:
    imported = 0
    errors: list[str] = []
    with database_manager.transaction() as conn:
        if data.mode == "replace":
            conn.execute("DELETE FROM ro_incidents")
        for idx, item in enumerate(data.incidents):
            try:
                id_ = _new_id()
                now = utcnow_iso()
                ref = _seq_reference("INC", conn)
                conn.execute(
                    """INSERT INTO ro_incidents (id, reference, date_occurrence, description,
                       ligne_metier, type_evenement, cause_racine, perte_brute, perte_recuperee,
                       statut, cree_le, modifie_le) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (id_, ref, item.date_occurrence, item.description, item.ligne_metier,
                     item.type_evenement, item.cause_racine, item.perte_brute,
                     item.perte_recuperee, item.statut, now, now),
                )
                imported += 1
            except Exception as exc:
                errors.append(f"Ligne {idx + 2}: {exc}")
        if imported > 0:
            _log(conn, "Pertes", "IMPORT", f"{imported} incident(s) importé(s)")
    return IncidentImportResult(imported=imported, skipped=0, errors=errors)


# ─── KRI ──────────────────────────────────────────────────────────────────────

def _kri_statut(valeur: float | None, seuil_alerte: float, seuil_critique: float, sens: str) -> str:
    if valeur is None:
        return "non_renseigne"
    if sens == "superieur":
        if valeur >= seuil_critique:
            return "critique"
        if valeur >= seuil_alerte:
            return "alerte"
        return "normal"
    else:
        if valeur <= seuil_critique:
            return "critique"
        if valeur <= seuil_alerte:
            return "alerte"
        return "normal"


def get_kri_module() -> KriModuleData:
    with database_manager.transaction() as conn:
        defs = conn.execute("SELECT * FROM ro_kri_definitions ORDER BY id").fetchall()
        kri_list = []
        hors_seuil = 0
        for d in defs:
            vals = conn.execute(
                "SELECT * FROM ro_kri_valeurs WHERE kri_id = ? ORDER BY date_mesure DESC",
                (d["id"],),
            ).fetchall()
            historique = [
                KriValeurView(
                    id=v["id"], kri_id=v["kri_id"], date_mesure=v["date_mesure"],
                    valeur=v["valeur"], commentaire=v["commentaire"] or "", cree_le=v["cree_le"],
                )
                for v in vals
            ]
            derniere = vals[0] if vals else None
            derniere_valeur = derniere["valeur"] if derniere else None
            derniere_date = derniere["date_mesure"] if derniere else None
            statut = _kri_statut(derniere_valeur, d["seuil_alerte"], d["seuil_critique"], d["sens"])
            if statut in ("alerte", "critique"):
                hors_seuil += 1
            from .models import KriDefinition
            kri_list.append(KriView(
                definition=KriDefinition(
                    id=d["id"], nom=d["nom"], unite=d["unite"], formule=d["formule"],
                    seuil_alerte=d["seuil_alerte"], seuil_critique=d["seuil_critique"],
                    sens=d["sens"], frequence=d["frequence"],
                ),
                derniere_valeur=derniere_valeur,
                derniere_date=derniere_date,
                statut=statut,
                historique=historique,
            ))
    return KriModuleData(kri_list=kri_list, kri_hors_seuil=hors_seuil)


def add_kri_valeur(data: KriValeurCreate) -> KriValeurView:
    id_ = _new_id()
    now = utcnow_iso()
    with database_manager.transaction() as conn:
        conn.execute(
            """INSERT INTO ro_kri_valeurs (id, kri_id, date_mesure, valeur, commentaire, cree_le)
               VALUES (?,?,?,?,?,?)""",
            (id_, data.kri_id, data.date_mesure, data.valeur, data.commentaire, now),
        )
        _log(conn, "KRI", "CREATE", data.kri_id, "", str(data.valeur))
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT * FROM ro_kri_valeurs WHERE id = ?", (id_,)).fetchone()
    return KriValeurView(
        id=row["id"], kri_id=row["kri_id"], date_mesure=row["date_mesure"],
        valeur=row["valeur"], commentaire=row["commentaire"] or "", cree_le=row["cree_le"],
    )


def auto_calcul_kri() -> KriModuleData:
    """
    Calcule automatiquement les KRI dérivables des données existantes
    (incidents et contrôles) et insère de nouvelles valeurs datées d'aujourd'hui.
    Seuls les KRI dont les données sources sont disponibles sont mis à jour :
      kri-04 : Incidents IT  (ro_incidents, type_evenement)
      kri-05 : Délai moyen de résolution (ro_incidents, statut=Clôturé)
      kri-06 : Taux d'erreurs de traitement (% incidents processus / total mois)
      kri-08 : Contrôles non conformes (ro_controles)
    Les KRI RH (kri-01/02/03/07) restent en saisie manuelle.
    """
    from datetime import date, timedelta
    now      = utcnow_iso()
    today    = date.today().isoformat()
    start_m  = date.today().replace(day=1).isoformat()
    ago30    = (date.today() - timedelta(days=30)).isoformat()

    with database_manager.transaction() as conn:
        # kri-04 : incidents liés aux systèmes/IT ce mois
        kri04 = float(conn.execute(
            """SELECT COUNT(*) AS cnt FROM ro_incidents
               WHERE date_occurrence >= ?
                 AND (type_evenement LIKE '%IT%'
                   OR type_evenement LIKE '%nterruption%'
                   OR type_evenement LIKE '%système%'
                   OR type_evenement LIKE '%Système%'
                   OR type_evenement LIKE '%informatique%')""",
            (start_m,),
        ).fetchone()["cnt"])

        # kri-05 : délai moyen de résolution des incidents clôturés (30 derniers jours)
        row05 = conn.execute(
            """SELECT AVG(julianday(modifie_le) - julianday(date_occurrence)) AS avg_d
               FROM ro_incidents
               WHERE statut = 'Clôturé' AND date_occurrence >= ?""",
            (ago30,),
        ).fetchone()
        kri05 = round(row05["avg_d"] or 0.0, 1)

        # kri-06 : % incidents liés aux processus / total incidents du mois
        total_m = conn.execute(
            "SELECT COUNT(*) AS cnt FROM ro_incidents WHERE date_occurrence >= ?",
            (start_m,),
        ).fetchone()["cnt"] or 0
        proc_m  = conn.execute(
            """SELECT COUNT(*) AS cnt FROM ro_incidents
               WHERE date_occurrence >= ?
                 AND (type_evenement LIKE '%xécution%'
                   OR type_evenement LIKE '%processus%'
                   OR type_evenement LIKE '%raitement%')""",
            (start_m,),
        ).fetchone()["cnt"]
        kri06 = round((proc_m / total_m * 100), 2) if total_m > 0 else 0.0

        # kri-08 : contrôles non conformes (tous, pas de fenêtre temporelle)
        kri08 = float(conn.execute(
            "SELECT COUNT(*) AS cnt FROM ro_controles WHERE resultat = 'Non-conforme'"
        ).fetchone()["cnt"])

        auto_values = {
            "kri-04": kri04,
            "kri-05": kri05,
            "kri-06": kri06,
            "kri-08": kri08,
        }
        for kri_id, valeur in auto_values.items():
            conn.execute(
                """INSERT INTO ro_kri_valeurs (id, kri_id, date_mesure, valeur, commentaire, cree_le)
                   VALUES (?,?,?,?,?,?)""",
                (_new_id(), kri_id, today, valeur, "auto", now),
            )
            _log(conn, "KRI", "AUTO_CALCUL", kri_id, "", str(valeur))

    return get_kri_module()


# ─── Risques ──────────────────────────────────────────────────────────────────

def _row_to_risque(row) -> RisqueView:
    p = row["probabilite"] or 1
    i = row["impact"] or 1
    eff = row["efficacite_controle"] or 3
    brut = p * i
    residuel = round(brut / eff, 2)
    return RisqueView(
        id=row["id"], nom=row["nom"], categorie=row["categorie"],
        ligne_metier=row["ligne_metier"], probabilite=p, impact=i,
        niveau_brut=brut, controle_existant=row["controle_existant"] or "",
        efficacite_controle=eff, niveau_residuel=residuel,
        niveau_label=_niveau_label(brut), plan_action_ref=row["plan_action_ref"] or "",
        cree_le=row["cree_le"], modifie_le=row["modifie_le"],
    )


def list_risques() -> list[RisqueView]:
    with database_manager.transaction() as conn:
        rows = conn.execute("SELECT * FROM ro_risques ORDER BY probabilite * impact DESC").fetchall()
    return [_row_to_risque(r) for r in rows]


def create_risque(data: RisqueCreate) -> RisqueView:
    id_ = _new_id()
    now = utcnow_iso()
    with database_manager.transaction() as conn:
        conn.execute(
            """INSERT INTO ro_risques (id, nom, categorie, ligne_metier, probabilite, impact,
               controle_existant, efficacite_controle, plan_action_ref, cree_le, modifie_le)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (id_, data.nom, data.categorie, data.ligne_metier, data.probabilite, data.impact,
             data.controle_existant, data.efficacite_controle, data.plan_action_ref, now, now),
        )
        _log(conn, "Cartographie", "CREATE", data.nom)
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT * FROM ro_risques WHERE id = ?", (id_,)).fetchone()
    return _row_to_risque(row)


def update_risque(id_: str, data: RisqueUpdate) -> RisqueView:
    now = utcnow_iso()
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if updates:
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        with database_manager.transaction() as conn:
            old = conn.execute("SELECT nom FROM ro_risques WHERE id = ?", (id_,)).fetchone()
            conn.execute(
                f"UPDATE ro_risques SET {set_clause}, modifie_le = ? WHERE id = ?",
                (*updates.values(), now, id_),
            )
            _log(conn, "Cartographie", "UPDATE", old["nom"] if old else id_)
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT * FROM ro_risques WHERE id = ?", (id_,)).fetchone()
    return _row_to_risque(row)


def delete_risque(id_: str) -> None:
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT nom FROM ro_risques WHERE id = ?", (id_,)).fetchone()
        conn.execute("DELETE FROM ro_risques WHERE id = ?", (id_,))
        _log(conn, "Cartographie", "DELETE", row["nom"] if row else id_)


# ─── Contrôles ────────────────────────────────────────────────────────────────

def _row_to_controle(row) -> ControleView:
    pc = row["points_conformes"] or 0
    ptc = row["points_controles"] or 0
    taux = round((pc / ptc) * 100, 1) if ptc > 0 else 0.0
    return ControleView(
        id=row["id"], reference=row["reference"], perimetre=row["perimetre"],
        type_controle=row["type_controle"], frequence=row["frequence"],
        date_dernier_test=row["date_dernier_test"] or "",
        prochaine_date=_prochaine_date(row["date_dernier_test"] or "", row["frequence"]),
        resultat=row["resultat"], taux_conformite=taux,
        points_conformes=pc, points_controles=ptc,
        non_conformites=row["non_conformites"] or "", validateur=row["validateur"] or "",
        cree_le=row["cree_le"], modifie_le=row["modifie_le"],
    )


def list_controles() -> list[ControleView]:
    with database_manager.transaction() as conn:
        rows = conn.execute("SELECT * FROM ro_controles ORDER BY cree_le DESC").fetchall()
    return [_row_to_controle(r) for r in rows]


def create_controle(data: ControleCreate) -> ControleView:
    id_ = _new_id()
    now = utcnow_iso()
    with database_manager.transaction() as conn:
        ref = _seq_reference("CTRL", conn)
        conn.execute(
            """INSERT INTO ro_controles (id, reference, perimetre, type_controle, frequence,
               date_dernier_test, resultat, points_conformes, points_controles,
               non_conformites, validateur, cree_le, modifie_le)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (id_, ref, data.perimetre, data.type_controle, data.frequence,
             data.date_dernier_test, data.resultat, data.points_conformes,
             data.points_controles, data.non_conformites, data.validateur, now, now),
        )
        _log(conn, "Contrôles", "CREATE", ref)
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT * FROM ro_controles WHERE id = ?", (id_,)).fetchone()
    return _row_to_controle(row)


def update_controle(id_: str, data: ControleUpdate) -> ControleView:
    now = utcnow_iso()
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if updates:
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        with database_manager.transaction() as conn:
            old = conn.execute("SELECT reference FROM ro_controles WHERE id = ?", (id_,)).fetchone()
            conn.execute(
                f"UPDATE ro_controles SET {set_clause}, modifie_le = ? WHERE id = ?",
                (*updates.values(), now, id_),
            )
            _log(conn, "Contrôles", "UPDATE", old["reference"] if old else id_)
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT * FROM ro_controles WHERE id = ?", (id_,)).fetchone()
    return _row_to_controle(row)


def delete_controle(id_: str) -> None:
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT reference FROM ro_controles WHERE id = ?", (id_,)).fetchone()
        conn.execute("DELETE FROM ro_controles WHERE id = ?", (id_,))
        _log(conn, "Contrôles", "DELETE", row["reference"] if row else id_)


# ─── Plans d'actions ──────────────────────────────────────────────────────────

def _row_to_plan(row) -> PlanView:
    today = date.today()
    echeance_str = row["date_echeance"] or ""
    jours = 0
    en_retard = False
    if echeance_str:
        try:
            echeance = date.fromisoformat(echeance_str)
            diff = (echeance - today).days
            jours = max(0, diff)
            en_retard = diff < 0 and row["statut"] not in ("Terminé", "Abandonné")
        except ValueError:
            pass
    return PlanView(
        id=row["id"], reference=row["reference"], titre=row["titre"],
        description=row["description"] or "", type_action=row["type_action"],
        source=row["source"], source_ref=row["source_ref"] or "",
        responsable=row["responsable"] or "", date_debut=row["date_debut"] or "",
        date_echeance=echeance_str, jours_restants=jours, priorite=row["priorite"],
        statut=row["statut"], avancement=row["avancement"] or 0,
        en_retard=en_retard, cree_le=row["cree_le"], modifie_le=row["modifie_le"],
    )


def list_plans() -> list[PlanView]:
    with database_manager.transaction() as conn:
        rows = conn.execute("SELECT * FROM ro_plans_action ORDER BY cree_le DESC").fetchall()
    return [_row_to_plan(r) for r in rows]


def create_plan(data: PlanCreate) -> PlanView:
    id_ = _new_id()
    now = utcnow_iso()
    with database_manager.transaction() as conn:
        ref = _seq_reference("ACT", conn)
        conn.execute(
            """INSERT INTO ro_plans_action (id, reference, titre, description, type_action, source,
               source_ref, responsable, date_debut, date_echeance, priorite, statut, avancement,
               cree_le, modifie_le) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (id_, ref, data.titre, data.description, data.type_action, data.source,
             data.source_ref, data.responsable, data.date_debut, data.date_echeance,
             data.priorite, data.statut, data.avancement, now, now),
        )
        _log(conn, "Plans", "CREATE", ref, "", data.titre)
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT * FROM ro_plans_action WHERE id = ?", (id_,)).fetchone()
    return _row_to_plan(row)


def update_plan(id_: str, data: PlanUpdate) -> PlanView:
    now = utcnow_iso()
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if updates:
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        with database_manager.transaction() as conn:
            old = conn.execute("SELECT reference FROM ro_plans_action WHERE id = ?", (id_,)).fetchone()
            conn.execute(
                f"UPDATE ro_plans_action SET {set_clause}, modifie_le = ? WHERE id = ?",
                (*updates.values(), now, id_),
            )
            _log(conn, "Plans", "UPDATE", old["reference"] if old else id_)
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT * FROM ro_plans_action WHERE id = ?", (id_,)).fetchone()
    return _row_to_plan(row)


def delete_plan(id_: str) -> None:
    with database_manager.transaction() as conn:
        row = conn.execute("SELECT reference FROM ro_plans_action WHERE id = ?", (id_,)).fetchone()
        conn.execute("DELETE FROM ro_plans_action WHERE id = ?", (id_,))
        _log(conn, "Plans", "DELETE", row["reference"] if row else id_)


# ─── Historique ───────────────────────────────────────────────────────────────

def list_historique(limit: int = 200) -> list[HistoriqueView]:
    with database_manager.transaction() as conn:
        rows = conn.execute(
            "SELECT * FROM ro_historique ORDER BY date_evenement DESC LIMIT ?", (limit,)
        ).fetchall()
    return [
        HistoriqueView(
            id=r["id"], date_evenement=r["date_evenement"], utilisateur=r["utilisateur"],
            menu=r["menu"], type_action=r["type_action"], element=r["element"] or "",
            ancienne_valeur=r["ancienne_valeur"] or "", nouvelle_valeur=r["nouvelle_valeur"] or "",
            commentaire=r["commentaire"] or "",
        )
        for r in rows
    ]


# ─── Dashboard ────────────────────────────────────────────────────────────────

def get_dashboard() -> DashboardData:
    today = date.today()
    current_month = today.strftime("%Y-%m")
    last_month = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")

    with database_manager.transaction() as conn:
        # Widget 1 : on calcule K approché depuis les pertes (méthode indicateur de base simplifiée)
        total_pertes = conn.execute(
            "SELECT COALESCE(SUM(perte_nette_calc), 0) as total FROM "
            "(SELECT (perte_brute - perte_recuperee) as perte_nette_calc FROM ro_incidents)"
        ).fetchone()["total"] or 0
        # K_IB simplifié : 15% de la moyenne des pertes sur 3 ans (proxy)
        k_ib = total_pertes * 0.15
        apr = k_ib * 12.5

        # Widget 2
        inc_mois = conn.execute(
            "SELECT COUNT(*) as cnt FROM ro_incidents WHERE date_occurrence LIKE ?",
            (f"{current_month}%",),
        ).fetchone()["cnt"] or 0

        pertes_mois = conn.execute(
            "SELECT COALESCE(SUM(perte_brute - perte_recuperee), 0) as total "
            "FROM ro_incidents WHERE date_occurrence LIKE ?",
            (f"{current_month}%",),
        ).fetchone()["total"] or 0

        pertes_mois_prec = conn.execute(
            "SELECT COALESCE(SUM(perte_brute - perte_recuperee), 0) as total "
            "FROM ro_incidents WHERE date_occurrence LIKE ?",
            (f"{last_month}%",),
        ).fetchone()["total"] or 0

        non_clos = conn.execute(
            "SELECT COUNT(*) as cnt FROM ro_incidents WHERE statut != 'Clôturé'"
        ).fetchone()["cnt"] or 0

        evolution_pct: float | None = None
        if pertes_mois_prec and pertes_mois_prec != 0:
            evolution_pct = round(((pertes_mois - pertes_mois_prec) / abs(pertes_mois_prec)) * 100, 1)

        # Widget 3
        actions_retard = conn.execute(
            "SELECT COUNT(*) as cnt FROM ro_plans_action "
            "WHERE date_echeance != '' AND date_echeance < ? AND statut != 'Terminé'",
            (today.isoformat(),),
        ).fetchone()["cnt"] or 0

        ctrl_non_conf = conn.execute(
            "SELECT COUNT(*) as cnt FROM ro_controles WHERE resultat = 'Non-conforme'"
        ).fetchone()["cnt"] or 0

        # KRI hors seuil
        defs = conn.execute("SELECT * FROM ro_kri_definitions").fetchall()
        kri_hors = 0
        for d in defs:
            last_val = conn.execute(
                "SELECT valeur FROM ro_kri_valeurs WHERE kri_id = ? ORDER BY date_mesure DESC LIMIT 1",
                (d["id"],),
            ).fetchone()
            if last_val:
                st = _kri_statut(last_val["valeur"], d["seuil_alerte"], d["seuil_critique"], d["sens"])
                if st in ("alerte", "critique"):
                    kri_hors += 1

        # Évolution 12 mois
        evolution: list[EvolutionPoint] = []
        for i in range(11, -1, -1):
            d_ref = today.replace(day=1) - timedelta(days=i * 30)
            m_str = d_ref.strftime("%Y-%m")
            row_ev = conn.execute(
                "SELECT COALESCE(SUM(perte_brute), 0) as brute, "
                "COALESCE(SUM(perte_brute - perte_recuperee), 0) as nette "
                "FROM ro_incidents WHERE date_occurrence LIKE ?",
                (f"{m_str}%",),
            ).fetchone()
            evolution.append(EvolutionPoint(
                mois=d_ref.strftime("%b %Y"),
                perte_brute=row_ev["brute"] or 0,
                perte_nette=row_ev["nette"] or 0,
            ))

        # Répartitions
        lm_rows = conn.execute(
            "SELECT ligne_metier, COALESCE(SUM(perte_brute - perte_recuperee), 0) as valeur "
            "FROM ro_incidents GROUP BY ligne_metier"
        ).fetchall()
        repartition_lm = [RepartitionItem(label=r["ligne_metier"], valeur=int(r["valeur"])) for r in lm_rows]

        type_rows = conn.execute(
            "SELECT type_evenement, COUNT(*) as cnt FROM ro_incidents GROUP BY type_evenement"
        ).fetchall()
        repartition_type = [RepartitionItem(label=r["type_evenement"], valeur=r["cnt"]) for r in type_rows]

    return DashboardData(
        widget1=DashboardWidget1(
            exigence_fonds_propres=k_ib,
            apr_risque_op=apr,
            statut_reglementaire="Conforme" if k_ib >= 0 else "Attention",
        ),
        widget2=DashboardWidget2(
            total_incidents_mois=inc_mois,
            pertes_nettes_mois=pertes_mois,
            incidents_non_clos=non_clos,
            evolution_pertes_pct=evolution_pct,
        ),
        widget3=DashboardWidget3(
            actions_en_retard=actions_retard,
            controles_non_conformes=ctrl_non_conf,
            kri_hors_seuil=kri_hors,
        ),
        evolution_pertes=evolution,
        repartition_ligne_metier=repartition_lm,
        repartition_type=repartition_type,
    )


# ─── BIC — Approche Standard CRR3 ────────────────────────────────────────────

def _row_to_input(row) -> OpRiskInput:
    return OpRiskInput(
        annee=row["annee"],
        interets_percus=row["interets_percus"],
        interets_verses=row["interets_verses"],
        revenus_leasing=row["revenus_leasing"] if "revenus_leasing" in row.keys() else 0.0,
        dividendes_percus=row["dividendes_percus"],
        autres_produits_exploitation=row["autres_produits_exploitation"],
        autres_charges_exploitation=row["autres_charges_exploitation"],
        commissions_percues=row["commissions_percues"],
        commissions_versees=row["commissions_versees"],
        resultat_portefeuille_negociation=row["resultat_portefeuille_negociation"],
        resultat_portefeuille_bancaire=row["resultat_portefeuille_bancaire"],
        tresorerie_et_banques_centrales=row["tresorerie_et_banques_centrales"],
        creances_etablissements_credit=row["creances_etablissements_credit"],
        creances_clientele=row["creances_clientele"],
        provisions=row["provisions"],
        pnb=row["pnb"],
    )


def _row_to_params(row) -> OpRiskParametres:
    return OpRiskParametres(
        seuil_ildc=row["seuil_ildc"],
        coef_tranche1=row["coef_tranche1"],
        coef_tranche2=row["coef_tranche2"],
        coef_tranche3=row["coef_tranche3"],
        seuil1_fcfa=row["seuil1_fcfa"],
        seuil2_fcfa=row["seuil2_fcfa"],
        multiplicateur_rea=row["multiplicateur_rea"],
        taux_conversion_eur_fcfa=row["taux_conversion_eur_fcfa"],
    )


def get_op_risk_input(annee: int) -> OpRiskInput:
    with database_manager.read_connection() as conn:
        row = conn.execute(
            "SELECT * FROM op_risk_financial_inputs WHERE annee = ?", (annee,)
        ).fetchone()
        if row is None:
            return OpRiskInput(annee=annee)
        return _row_to_input(row)


def upsert_op_risk_input(annee: int, data: OpRiskInputUpdate) -> OpRiskInput:
    now = utcnow_iso()
    with database_manager.transaction() as conn:
        existing = conn.execute(
            "SELECT annee FROM op_risk_financial_inputs WHERE annee = ?", (annee,)
        ).fetchone()
        if existing is None:
            fields = {
                "annee": annee,
                "interets_percus": data.interets_percus or 0.0,
                "interets_verses": data.interets_verses or 0.0,
                "revenus_leasing": data.revenus_leasing or 0.0,
                "dividendes_percus": data.dividendes_percus or 0.0,
                "autres_produits_exploitation": data.autres_produits_exploitation or 0.0,
                "autres_charges_exploitation": data.autres_charges_exploitation or 0.0,
                "commissions_percues": data.commissions_percues or 0.0,
                "commissions_versees": data.commissions_versees or 0.0,
                "resultat_portefeuille_negociation": data.resultat_portefeuille_negociation or 0.0,
                "resultat_portefeuille_bancaire": data.resultat_portefeuille_bancaire or 0.0,
                "tresorerie_et_banques_centrales": data.tresorerie_et_banques_centrales or 0.0,
                "creances_etablissements_credit": data.creances_etablissements_credit or 0.0,
                "creances_clientele": data.creances_clientele or 0.0,
                "provisions": data.provisions or 0.0,
                "pnb": data.pnb or 0.0,
                "modifie_le": now,
            }
            cols = ", ".join(fields.keys())
            placeholders = ", ".join("?" * len(fields))
            conn.execute(
                f"INSERT INTO op_risk_financial_inputs ({cols}) VALUES ({placeholders})",
                list(fields.values()),
            )
        else:
            updates = {k: v for k, v in data.model_dump().items() if v is not None}
            updates["modifie_le"] = now
            set_clause = ", ".join(f"{k} = ?" for k in updates)
            conn.execute(
                f"UPDATE op_risk_financial_inputs SET {set_clause} WHERE annee = ?",
                [*updates.values(), annee],
            )
        row = conn.execute(
            "SELECT * FROM op_risk_financial_inputs WHERE annee = ?", (annee,)
        ).fetchone()
        return _row_to_input(row)


def get_op_risk_parametres() -> OpRiskParametres:
    with database_manager.read_connection() as conn:
        row = conn.execute(
            "SELECT * FROM op_risk_parametres WHERE id = 1"
        ).fetchone()
        return _row_to_params(row)


def update_op_risk_parametres(data: OpRiskParametresUpdate) -> OpRiskParametres:
    with database_manager.transaction() as conn:
        updates = {k: v for k, v in data.model_dump().items() if v is not None}
        if updates:
            set_clause = ", ".join(f"{k} = ?" for k in updates)
            conn.execute(
                f"UPDATE op_risk_parametres SET {set_clause} WHERE id = 1",
                list(updates.values()),
            )
        row = conn.execute(
            "SELECT * FROM op_risk_parametres WHERE id = 1"
        ).fetchone()
        return _row_to_params(row)


def calcul_bic(annee_n: int | None = None) -> OpRiskCalculResult:
    """
    Business Indicator Component (CRR3 Art. 315-321).
    ILM = 1 par hypothèse (non activé — cadre réglementaire UE actuel).
    annee_n : dernier exercice clos (défaut = année courante - 1).
    """
    year_n = annee_n if annee_n is not None else date.today().year - 1
    annees = [year_n - 2, year_n - 1, year_n]

    params = get_op_risk_parametres()

    with database_manager.read_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM op_risk_financial_inputs WHERE annee IN (?, ?, ?)",
            annees,
        ).fetchall()

    inputs_map = {row["annee"]: _row_to_input(row) for row in rows}
    inputs = [inputs_map.get(y, OpRiskInput(annee=y)) for y in annees]
    donnees_insuffisantes = len(rows) == 0

    def moy(attr: str) -> float:
        return sum(getattr(inp, attr) for inp in inputs) / 3

    # ── ILDC ──────────────────────────────────────────────────────────────────
    ic = moy("interets_percus") - moy("interets_verses")
    ac = (
        moy("tresorerie_et_banques_centrales")
        + moy("creances_etablissements_credit")
        + (moy("creances_clientele") - moy("provisions"))
    )
    plafond_ildc = ac * params.seuil_ildc
    plafond_actif = abs(ic) > plafond_ildc
    dividendes = moy("dividendes_percus")
    ildc = min(abs(ic), plafond_ildc) + dividendes

    # ── SC ────────────────────────────────────────────────────────────────────
    oi = moy("autres_produits_exploitation")
    oe = moy("autres_charges_exploitation")
    fi = moy("commissions_percues")
    fe = moy("commissions_versees")
    sc = max(oi, oe) + max(fi, fe)

    # ── FC ────────────────────────────────────────────────────────────────────
    tc = abs(moy("resultat_portefeuille_negociation"))
    bc = abs(moy("resultat_portefeuille_bancaire"))
    fc = tc + bc

    # ── BI et BIC (tranches marginales) ───────────────────────────────────────
    bi = ildc + sc + fc
    s1 = params.seuil1_fcfa
    s2 = params.seuil2_fcfa
    c1, c2, c3 = params.coef_tranche1, params.coef_tranche2, params.coef_tranche3

    if bi <= s1:
        bic = bi * c1
        tranche_active = 1
        marge: float | None = s1 - bi
    elif bi <= s2:
        bic = s1 * c1 + (bi - s1) * c2
        tranche_active = 2
        marge = s2 - bi
    else:
        bic = s1 * c1 + (s2 - s1) * c2 + (bi - s2) * c3
        tranche_active = 3
        marge = None

    # ── Résultats (ILM = 1) ───────────────────────────────────────────────────
    mult = params.multiplicateur_rea
    ofr_crr3 = bic
    rea_crr3 = ofr_crr3 * mult

    pnb_moy = moy("pnb")
    ofr_bia = pnb_moy * 0.15
    rea_bia = ofr_bia * mult
    ecart = ofr_crr3 - ofr_bia

    return OpRiskCalculResult(
        annees=annees,
        inputs=inputs,
        params=params,
        ildc_detail=OpRiskIldcDetail(
            ic=ic, ac=ac, plafond_ildc=plafond_ildc,
            plafond_actif=plafond_actif, dividendes=dividendes, ildc=ildc,
        ),
        sc_detail=OpRiskScDetail(oi=oi, oe=oe, fi=fi, fe=fe, sc=sc),
        fc_detail=OpRiskFcDetail(tc=tc, bc=bc, fc=fc),
        bi_detail=OpRiskBiDetail(
            bi=bi, tranche_active=tranche_active, bic=bic,
            marge_avant_tranche_suivante=marge,
        ),
        ofr_crr3=ofr_crr3,
        rea_crr3=rea_crr3,
        ofr_bia=ofr_bia,
        rea_bia=rea_bia,
        ecart=ecart,
        donnees_insuffisantes=donnees_insuffisantes,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# BLOC A1 — AIB (Approche Indicateur de Base) — art. 301
# ═══════════════════════════════════════════════════════════════════════════════

def _row_to_pnb_annuel(row) -> PnbAnnuelView:
    pbt = float(row["produit_brut_total"])
    return PnbAnnuelView(
        annee=int(row["annee"]),
        produit_brut_total=pbt,
        pnb_positif=pbt > 0,
        pnb_retenu_aib=pbt if pbt > 0 else 0.0,
        source_document=row["source_document"] or "",
        modifie_le=row["modifie_le"] or "",
    )


def list_pnb_annuel() -> list[PnbAnnuelView]:
    with database_manager.read_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM op_pnb_annuel ORDER BY annee"
        ).fetchall()
    return [_row_to_pnb_annuel(r) for r in rows]


def upsert_pnb_annuel(annee: int, data: PnbAnnuelCreate) -> PnbAnnuelView:
    now = utcnow_iso()
    with database_manager.transaction() as conn:
        existing = conn.execute(
            "SELECT annee FROM op_pnb_annuel WHERE annee = ?", (annee,)
        ).fetchone()
        if existing is None:
            conn.execute(
                "INSERT INTO op_pnb_annuel (annee, produit_brut_total, source_document, modifie_le)"
                " VALUES (?, ?, ?, ?)",
                (annee, data.produit_brut_total, data.source_document, now),
            )
        else:
            conn.execute(
                "UPDATE op_pnb_annuel SET produit_brut_total=?, source_document=?, modifie_le=?"
                " WHERE annee=?",
                (data.produit_brut_total, data.source_document, now, annee),
            )
        row = conn.execute(
            "SELECT * FROM op_pnb_annuel WHERE annee = ?", (annee,)
        ).fetchone()
    return _row_to_pnb_annuel(row)


def delete_pnb_annuel(annee: int) -> None:
    with database_manager.transaction() as conn:
        conn.execute("DELETE FROM op_pnb_annuel WHERE annee = ?", (annee,))


def get_aib_parametres() -> ParametresAib:
    with database_manager.read_connection() as conn:
        row = conn.execute("SELECT * FROM op_parametres_aib WHERE id = 1").fetchone()
    return ParametresAib(
        alpha=row["alpha"],
        multiplicateur_rwa=row["multiplicateur_rwa"],
        ratio_solvabilite_min=row["ratio_solvabilite_min"],
    )


def update_aib_parametres(data: ParametresAibUpdate) -> ParametresAib:
    with database_manager.transaction() as conn:
        updates = {k: v for k, v in data.model_dump().items() if v is not None}
        if updates:
            set_clause = ", ".join(f"{k} = ?" for k in updates)
            conn.execute(
                f"UPDATE op_parametres_aib SET {set_clause} WHERE id = 1",
                list(updates.values()),
            )
        row = conn.execute("SELECT * FROM op_parametres_aib WHERE id = 1").fetchone()
    return ParametresAib(
        alpha=row["alpha"],
        multiplicateur_rwa=row["multiplicateur_rwa"],
        ratio_solvabilite_min=row["ratio_solvabilite_min"],
    )


def calcul_aib() -> AibCalculResult:
    annees_saisies = list_pnb_annuel()
    params = get_aib_parametres()

    positifs = [a for a in annees_saisies if a.pnb_retenu_aib > 0]
    n = len(positifs)
    somme = sum(a.pnb_retenu_aib for a in positifs)
    pnb_moyen = somme / n if n > 0 else 0.0
    k_ib = pnb_moyen * params.alpha
    apr = k_ib * params.multiplicateur_rwa
    capital = apr * params.ratio_solvabilite_min

    return AibCalculResult(
        annees_saisies=annees_saisies,
        n=n,
        somme_pnb_positifs=somme,
        pnb_moyen=pnb_moyen,
        alpha=params.alpha,
        k_ib=k_ib,
        apr_aib=apr,
        capital_min_aib=capital,
        donnees_insuffisantes=n == 0,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# BLOC A2 — AS (Approche Standard) — art. 305-311
# ═══════════════════════════════════════════════════════════════════════════════

def list_beta_lignes() -> list[BetaLigneView]:
    with database_manager.read_connection() as conn:
        rows = conn.execute("SELECT * FROM op_beta_lignes ORDER BY beta DESC, ligne_metier").fetchall()
    return [BetaLigneView(ligne_metier=r["ligne_metier"], beta=r["beta"], description=r["description"]) for r in rows]


def update_beta_ligne(ligne_metier: str, data: BetaLigneUpdate) -> BetaLigneView:
    with database_manager.transaction() as conn:
        conn.execute(
            "UPDATE op_beta_lignes SET beta = ? WHERE ligne_metier = ?",
            (data.beta, ligne_metier),
        )
        row = conn.execute(
            "SELECT * FROM op_beta_lignes WHERE ligne_metier = ?", (ligne_metier,)
        ).fetchone()
    return BetaLigneView(ligne_metier=row["ligne_metier"], beta=row["beta"], description=row["description"])


def get_pnb_lignes(annee: int) -> list[PnbParLigneView]:
    with database_manager.read_connection() as conn:
        betas = {r["ligne_metier"]: r["beta"] for r in conn.execute("SELECT * FROM op_beta_lignes").fetchall()}
        rows = conn.execute(
            "SELECT * FROM op_pnb_par_ligne WHERE annee = ? ORDER BY ligne_metier", (annee,)
        ).fetchall()
    result = []
    for r in rows:
        beta = betas.get(r["ligne_metier"], 0.0)
        pbl = float(r["produit_brut_ligne"])
        result.append(PnbParLigneView(
            annee=annee, ligne_metier=r["ligne_metier"],
            produit_brut_ligne=pbl, beta=beta, k_ligne=pbl * beta,
        ))
    return result


def upsert_pnb_ligne(annee: int, ligne_metier: str, data: PnbParLigneCreate) -> PnbParLigneView:
    now = utcnow_iso()
    with database_manager.transaction() as conn:
        existing = conn.execute(
            "SELECT annee FROM op_pnb_par_ligne WHERE annee=? AND ligne_metier=?",
            (annee, ligne_metier),
        ).fetchone()
        if existing is None:
            conn.execute(
                "INSERT INTO op_pnb_par_ligne (annee, ligne_metier, produit_brut_ligne, modifie_le)"
                " VALUES (?, ?, ?, ?)",
                (annee, ligne_metier, data.produit_brut_ligne, now),
            )
        else:
            conn.execute(
                "UPDATE op_pnb_par_ligne SET produit_brut_ligne=?, modifie_le=?"
                " WHERE annee=? AND ligne_metier=?",
                (data.produit_brut_ligne, now, annee, ligne_metier),
            )
        betas = {r["ligne_metier"]: r["beta"] for r in conn.execute("SELECT * FROM op_beta_lignes").fetchall()}
        row = conn.execute(
            "SELECT * FROM op_pnb_par_ligne WHERE annee=? AND ligne_metier=?",
            (annee, ligne_metier),
        ).fetchone()
    beta = betas.get(ligne_metier, 0.0)
    pbl = float(row["produit_brut_ligne"])
    return PnbParLigneView(annee=annee, ligne_metier=ligne_metier, produit_brut_ligne=pbl, beta=beta, k_ligne=pbl * beta)


def get_as_parametres() -> ParametresAs:
    with database_manager.read_connection() as conn:
        row = conn.execute("SELECT * FROM op_parametres_as WHERE id = 1").fetchone()
    return ParametresAs(
        as_autorisee=bool(row["as_autorisee"]),
        date_autorisation=row["date_autorisation"] or "",
        reference_autorisation=row["reference_autorisation"] or "",
        multiplicateur_rwa=row["multiplicateur_rwa"],
        ratio_solvabilite_min=row["ratio_solvabilite_min"],
    )


def update_as_parametres(data: ParametresAsUpdate) -> ParametresAs:
    with database_manager.transaction() as conn:
        raw = data.model_dump()
        updates: dict = {}
        for k, v in raw.items():
            if v is not None:
                updates[k] = int(v) if k == "as_autorisee" else v
        if updates:
            set_clause = ", ".join(f"{k} = ?" for k in updates)
            conn.execute(
                f"UPDATE op_parametres_as SET {set_clause} WHERE id = 1",
                list(updates.values()),
            )
        row = conn.execute("SELECT * FROM op_parametres_as WHERE id = 1").fetchone()
    return ParametresAs(
        as_autorisee=bool(row["as_autorisee"]),
        date_autorisation=row["date_autorisation"] or "",
        reference_autorisation=row["reference_autorisation"] or "",
        multiplicateur_rwa=row["multiplicateur_rwa"],
        ratio_solvabilite_min=row["ratio_solvabilite_min"],
    )


def calcul_as() -> AsCalculResult:
    params = get_as_parametres()
    annees_saisies = sorted({r.annee for r in list_pnb_annuel()})

    detail_par_annee: list[AsAnneeDetail] = []
    with database_manager.read_connection() as conn:
        betas = {r["ligne_metier"]: float(r["beta"]) for r in conn.execute("SELECT * FROM op_beta_lignes").fetchall()}
        for annee in annees_saisies:
            rows = conn.execute(
                "SELECT * FROM op_pnb_par_ligne WHERE annee = ?", (annee,)
            ).fetchall()
            lignes: list[AsLigneDetail] = []
            for r in rows:
                beta = betas.get(r["ligne_metier"], 0.0)
                pbl = float(r["produit_brut_ligne"])
                lignes.append(AsLigneDetail(ligne_metier=r["ligne_metier"], pnb=pbl, beta=beta, k_ligne=pbl * beta))
            k_total = sum(l.k_ligne for l in lignes)
            k_retenu = max(k_total, 0.0)
            detail_par_annee.append(AsAnneeDetail(annee=annee, lignes=lignes, k_total=k_total, k_retenu=k_retenu))

    n = len(detail_par_annee)
    k_as = sum(d.k_retenu for d in detail_par_annee) / n if n > 0 else 0.0
    apr = k_as * params.multiplicateur_rwa
    capital = apr * params.ratio_solvabilite_min

    return AsCalculResult(
        as_autorisee=params.as_autorisee,
        detail_par_annee=detail_par_annee,
        k_as=k_as,
        apr_as=apr,
        capital_min_as=capital,
        donnees_insuffisantes=n == 0,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Seuils de reporting Pilier 2
# ═══════════════════════════════════════════════════════════════════════════════

def get_pertes_seuils() -> ParametresSeuils:
    with database_manager.read_connection() as conn:
        row = conn.execute("SELECT * FROM op_parametres_seuils WHERE id = 1").fetchone()
    return ParametresSeuils(
        seuil_reporting_interne=row["seuil_reporting_interne"],
        seuil_reporting_direction=row["seuil_reporting_direction"],
        seuil_reporting_conseil=row["seuil_reporting_conseil"],
    )


def update_pertes_seuils(data: ParametresSeuilsUpdate) -> ParametresSeuils:
    with database_manager.transaction() as conn:
        updates = {k: v for k, v in data.model_dump().items() if v is not None}
        if updates:
            set_clause = ", ".join(f"{k} = ?" for k in updates)
            conn.execute(
                f"UPDATE op_parametres_seuils SET {set_clause} WHERE id = 1",
                list(updates.values()),
            )
        row = conn.execute("SELECT * FROM op_parametres_seuils WHERE id = 1").fetchone()
    return ParametresSeuils(
        seuil_reporting_interne=row["seuil_reporting_interne"],
        seuil_reporting_direction=row["seuil_reporting_direction"],
        seuil_reporting_conseil=row["seuil_reporting_conseil"],
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Synthèse comparative AIB / AS / BIC
# ═══════════════════════════════════════════════════════════════════════════════

def get_synthese() -> SyntheseResult:
    lignes: list[SyntheseLigne] = []

    # AIB
    try:
        aib = calcul_aib()
        lignes.append(SyntheseLigne(
            methode="AIB — Indicateur de Base (art. 301)",
            k=aib.k_ib, apr=aib.apr_aib, capital_min=aib.capital_min_aib,
            disponible=not aib.donnees_insuffisantes,
        ))
        k_ib = aib.k_ib if not aib.donnees_insuffisantes else None
    except Exception:
        k_ib = None
        lignes.append(SyntheseLigne(methode="AIB — Indicateur de Base (art. 301)", k=0, apr=0, capital_min=0, disponible=False))

    # AS
    try:
        as_res = calcul_as()
        lignes.append(SyntheseLigne(
            methode="AS — Approche Standard (art. 305-311)",
            k=as_res.k_as, apr=as_res.apr_as, capital_min=as_res.capital_min_as,
            disponible=as_res.as_autorisee and not as_res.donnees_insuffisantes,
        ))
    except Exception:
        lignes.append(SyntheseLigne(methode="AS — Approche Standard (art. 305-311)", k=0, apr=0, capital_min=0, disponible=False))

    # BIC
    try:
        bic = calcul_bic()
        lignes.append(SyntheseLigne(
            methode="BIC — CRR3 Pilotage interne",
            k=bic.ofr_crr3, apr=bic.rea_crr3, capital_min=bic.rea_crr3 * 0.08,
            disponible=not bic.donnees_insuffisantes,
        ))
        ofr_bic = bic.ofr_crr3 if not bic.donnees_insuffisantes else None
    except Exception:
        ofr_bic = None
        lignes.append(SyntheseLigne(methode="BIC — CRR3 Pilotage interne", k=0, apr=0, capital_min=0, disponible=False))

    ecart = (ofr_bic - k_ib) if (ofr_bic is not None and k_ib is not None) else None

    # Ratio de couverture : K_IB / perte_nette_annuelle_moy
    ratio: float | None = None
    try:
        with database_manager.read_connection() as conn:
            row = conn.execute(
                "SELECT AVG(perte_brute - perte_recuperee) as pnm FROM ro_incidents"
            ).fetchone()
            pnm = float(row["pnm"]) if row and row["pnm"] else 0.0
        if pnm > 0 and k_ib is not None and k_ib > 0:
            ratio = k_ib / pnm
    except Exception:
        pass

    return SyntheseResult(lignes=lignes, ecart_bic_vs_aib=ecart, ratio_couverture=ratio)


# ═══════════════════════════════════════════════════════════════════════════════
# Décision de pilotage — Analyse et reporting (dispositif réglementaire UEMOA)
# ═══════════════════════════════════════════════════════════════════════════════

def get_decision_pilotage() -> DecisionPilotageResult:
    """
    Moteur de décision du pilotage interne du risque opérationnel.

    Agrège les signaux déjà mesurés par le module (BIC/CRR3, pertes
    opérationnelles, incidents, cartographie des risques, contrôles internes,
    KRI, plans d'action) pour produire :
      - un verdict global (Maîtrisé / Sous surveillance / Vigilance renforcée / Critique)
      - l'organe auquel le reporting doit être escaladé (seuils Pilier 2)
      - des recommandations argumentées, rattachées aux articles du dispositif
        réglementaire UEMOA déjà utilisés dans l'outil (Art. 89, 301/307, 313,
        313.b, 313.c, 314, 546).

    Chaque critère est noté 0 (conforme), 1 (attention) ou 2 (critique).
    Le score global (0 à 16) détermine le niveau de pilotage retenu.
    """
    criteres: list[DecisionCritere] = []

    # ── C1 — Disponibilité des données réglementaires (Art. 89) ───────────────
    try:
        bic = calcul_bic()
        bic_erreur = False
    except Exception:
        bic = None
        bic_erreur = True

    if bic_erreur or bic.donnees_insuffisantes:
        criteres.append(DecisionCritere(
            code="C1", libelle="Disponibilité des données de calcul (BIC/AIB)",
            reference_reglementaire="Art. 89",
            statut="critique", poids=2,
            valeur_observee="Données insuffisantes",
            seuil_reference="3 exercices de PNB requis",
            commentaire="Le calcul de l'exigence de fonds propres pour risque opérationnel "
                        "ne peut être finalisé faute de données suffisantes sur 3 exercices.",
        ))
    else:
        criteres.append(DecisionCritere(
            code="C1", libelle="Disponibilité des données de calcul (BIC/AIB)",
            reference_reglementaire="Art. 89",
            statut="conforme", poids=0,
            valeur_observee="Données complètes",
            seuil_reference="3 exercices de PNB requis",
            commentaire="Les données financières nécessaires au calcul réglementaire sont disponibles.",
        ))

    # ── C2 — Exposition en capital réglementaire (tranche BIC) ────────────────
    if not bic_erreur and not bic.donnees_insuffisantes:
        tranche = bic.bi_detail.tranche_active
        ecart_pct = (bic.ecart / bic.ofr_bia * 100) if bic.ofr_bia else 0.0
        if tranche == 3 or ecart_pct > 30:
            statut, poids = "critique", 2
        elif tranche == 2 or ecart_pct > 10:
            statut, poids = "attention", 1
        else:
            statut, poids = "conforme", 0
        criteres.append(DecisionCritere(
            code="C2", libelle="Exposition en capital réglementaire (BIC)",
            reference_reglementaire="Art. 315-321 / Art. 89",
            statut=statut, poids=poids,
            valeur_observee=f"Tranche {tranche} — écart CRR3/BIA {ecart_pct:+.1f} %",
            seuil_reference="Tranche 1 attendue, écart ≤ 10 %",
            commentaire=(
                f"Le niveau du Business Indicator place l'établissement en tranche {tranche} "
                "du dispositif ; un écart significatif avec la méthode indicateur de base doit "
                "être justifié auprès de la Commission Bancaire."
                if ecart_pct > 10 or tranche > 1 else
                f"Le niveau du Business Indicator place l'établissement en tranche {tranche} ; "
                "l'écart avec la méthode indicateur de base reste maîtrisé."
            ),
        ))
    else:
        criteres.append(DecisionCritere(
            code="C2", libelle="Exposition en capital réglementaire (BIC)",
            reference_reglementaire="Art. 315-321 / Art. 89",
            statut="critique", poids=2,
            valeur_observee="Non calculable",
            seuil_reference="Tranche 1 attendue",
            commentaire="Impossible d'évaluer l'exposition en capital faute de calcul BIC disponible.",
        ))

    # ── C3 — Pertes opérationnelles vs seuils de reporting Pilier 2 ───────────
    dashboard = get_dashboard()
    seuils = get_pertes_seuils()
    pertes_mois = dashboard.widget2.pertes_nettes_mois

    if pertes_mois >= seuils.seuil_reporting_conseil:
        statut, poids = "critique", 2
        organe = "Conseil d'administration"
        motif = (
            f"Pertes nettes du mois ({pertes_mois:,.0f} FCFA) supérieures au seuil de "
            f"reporting au Conseil ({seuils.seuil_reporting_conseil:,.0f} FCFA)."
        )
    elif pertes_mois >= seuils.seuil_reporting_direction:
        statut, poids = "attention", 1
        organe = "Direction Générale"
        motif = (
            f"Pertes nettes du mois ({pertes_mois:,.0f} FCFA) supérieures au seuil de "
            f"reporting à la Direction ({seuils.seuil_reporting_direction:,.0f} FCFA)."
        )
    elif pertes_mois >= seuils.seuil_reporting_interne:
        statut, poids = "attention", 1
        organe = "Comité des risques"
        motif = (
            f"Pertes nettes du mois ({pertes_mois:,.0f} FCFA) supérieures au seuil de "
            f"reporting interne ({seuils.seuil_reporting_interne:,.0f} FCFA)."
        )
    else:
        statut, poids = "conforme", 0
        organe = "Reporting interne standard"
        motif = "Les pertes nettes du mois restent inférieures aux seuils d'escalade Pilier 2."

    criteres.append(DecisionCritere(
        code="C3", libelle="Pertes opérationnelles vs seuils Pilier 2",
        reference_reglementaire="Art. 313.b",
        statut=statut, poids=poids,
        valeur_observee=f"{pertes_mois:,.0f} FCFA (mois en cours)",
        seuil_reference=(
            f"Interne {seuils.seuil_reporting_interne:,.0f} / "
            f"Direction {seuils.seuil_reporting_direction:,.0f} / "
            f"Conseil {seuils.seuil_reporting_conseil:,.0f} FCFA"
        ),
        commentaire=motif,
    ))

    # ── C4 — Incidents ouverts et tendance des pertes ─────────────────────────
    non_clos = dashboard.widget2.incidents_non_clos
    evolution = dashboard.widget2.evolution_pertes_pct or 0.0
    if non_clos > 5 or evolution > 50:
        statut, poids = "critique", 2
    elif non_clos >= 3 or evolution > 15:
        statut, poids = "attention", 1
    else:
        statut, poids = "conforme", 0
    criteres.append(DecisionCritere(
        code="C4", libelle="Incidents ouverts et tendance des pertes",
        reference_reglementaire="Art. 313.b",
        statut=statut, poids=poids,
        valeur_observee=f"{non_clos} incident(s) non clôturé(s), évolution {evolution:+.1f} %",
        seuil_reference="≤ 2 incidents non clôturés, évolution ≤ 15 %",
        commentaire=(
            "Le volume d'incidents ouverts et/ou la tendance des pertes nécessite un suivi "
            "renforcé et une accélération des clôtures."
            if statut != "conforme" else
            "Le flux d'incidents et son évolution restent maîtrisés."
        ),
    ))

    # ── C5 — Cartographie des risques ──────────────────────────────────────────
    risques = list_risques()
    nb_critique_risque = sum(1 for r in risques if r.niveau_label == "Critique")
    nb_eleve_risque = sum(1 for r in risques if r.niveau_label == "Élevé")
    if nb_critique_risque > 0:
        statut, poids = "critique", 2
    elif nb_eleve_risque > 0:
        statut, poids = "attention", 1
    else:
        statut, poids = "conforme", 0
    criteres.append(DecisionCritere(
        code="C5", libelle="Cartographie des risques",
        reference_reglementaire="Art. 313",
        statut=statut, poids=poids,
        valeur_observee=f"{nb_critique_risque} risque(s) critique(s), {nb_eleve_risque} élevé(s)",
        seuil_reference="Aucun risque en zone critique",
        commentaire=(
            "Des risques en zone critique ou élevée doivent être couverts par un plan "
            "d'action documenté et suivi."
            if statut != "conforme" else
            "Aucun risque en zone critique dans la cartographie actuelle."
        ),
    ))

    # ── C6 — Contrôles internes non conformes ─────────────────────────────────
    controles = list_controles()
    nb_non_conf = sum(1 for c in controles if c.resultat == "Non-conforme")
    if nb_non_conf > 2:
        statut, poids = "critique", 2
    elif nb_non_conf > 0:
        statut, poids = "attention", 1
    else:
        statut, poids = "conforme", 0
    criteres.append(DecisionCritere(
        code="C6", libelle="Efficacité du dispositif de contrôle interne",
        reference_reglementaire="Art. 314",
        statut=statut, poids=poids,
        valeur_observee=f"{nb_non_conf} contrôle(s) non conforme(s) sur {len(controles)}",
        seuil_reference="0 contrôle non conforme",
        commentaire=(
            "Des non-conformités de contrôle doivent être corrigées et documentées "
            "(conservation ≥ 7 ans, Art. 314)."
            if statut != "conforme" else
            "Le dispositif de contrôle interne ne présente pas de non-conformité ouverte."
        ),
    ))

    # ── C7 — Indicateurs clés de risque (KRI) hors seuil ──────────────────────
    kri = get_kri_module()
    if kri.kri_hors_seuil >= 2:
        statut, poids = "critique", 2
    elif kri.kri_hors_seuil == 1:
        statut, poids = "attention", 1
    else:
        statut, poids = "conforme", 0
    criteres.append(DecisionCritere(
        code="C7", libelle="Indicateurs clés de risque (KRI)",
        reference_reglementaire="Art. 313",
        statut=statut, poids=poids,
        valeur_observee=f"{kri.kri_hors_seuil} KRI hors seuil sur {len(kri.kri_list)}",
        seuil_reference="0 KRI en zone alerte/critique",
        commentaire=(
            "Un ou plusieurs indicateurs clés de risque dépassent leur seuil d'alerte et "
            "doivent faire l'objet d'un suivi rapproché."
            if statut != "conforme" else
            "L'ensemble des indicateurs clés de risque reste dans les seuils normaux."
        ),
    ))

    # ── C8 — Plans d'action en retard ──────────────────────────────────────────
    plans = list_plans()
    en_retard = sum(1 for p in plans if p.en_retard)
    if en_retard > 2:
        statut, poids = "critique", 2
    elif en_retard > 0:
        statut, poids = "attention", 1
    else:
        statut, poids = "conforme", 0
    criteres.append(DecisionCritere(
        code="C8", libelle="Avancement des plans d'action correctifs",
        reference_reglementaire="Art. 313.c",
        statut=statut, poids=poids,
        valeur_observee=f"{en_retard} plan(s) en retard sur {len(plans)}",
        seuil_reference="0 plan en retard",
        commentaire=(
            "Des plans d'action en retard exposent l'établissement à une récurrence des "
            "incidents ou risques déjà identifiés."
            if statut != "conforme" else
            "Les plans d'action correctifs et préventifs sont à jour."
        ),
    ))

    # ── Agrégation ──────────────────────────────────────────────────────────────
    score = sum(c.poids for c in criteres)
    score_max = 2 * len(criteres)

    if score <= 2:
        niveau = "Maîtrisé"
    elif score <= 6:
        niveau = "Sous surveillance"
    elif score <= 10:
        niveau = "Vigilance renforcée"
    else:
        niveau = "Critique"

    recommandations = [
        f"[{c.reference_reglementaire}] {c.commentaire}"
        for c in criteres if c.statut != "conforme"
    ]
    if not recommandations:
        recommandations = [
            "Aucune action corrective requise à ce stade — maintenir la surveillance "
            "courante du dispositif conformément à l'Art. 546 (reporting annuel à la "
            "Commission Bancaire de l'UMOA)."
        ]

    nb_attention = sum(1 for c in criteres if c.statut == "attention")
    nb_critique_total = sum(1 for c in criteres if c.statut == "critique")
    nb_conforme = len(criteres) - nb_attention - nb_critique_total

    synthese = (
        f"Sur {len(criteres)} critères analysés au titre du dispositif réglementaire UEMOA "
        f"de gestion et de surveillance du risque opérationnel, {nb_conforme} sont conformes, "
        f"{nb_attention} nécessitent une attention particulière et {nb_critique_total} "
        f"présentent un niveau critique. Le niveau de pilotage global retenu est « {niveau} » "
        f"(score {score}/{score_max}). Organe de reporting recommandé : {organe}."
    )

    return DecisionPilotageResult(
        date_analyse=utcnow_iso(),
        niveau_global=niveau,
        score_global=score,
        score_max=score_max,
        organe_reporting=organe,
        organe_reporting_motif=motif,
        criteres=criteres,
        recommandations=recommandations,
        synthese=synthese,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Critères transverses partagés — utilisés par les décisions AIB (A1) et AS (A2)
# ═══════════════════════════════════════════════════════════════════════════════
#
# Ces critères couvrent les signaux de risque opérationnel qui ne dépendent pas
# de la méthode de calcul du capital réglementaire (pertes, incidents,
# cartographie, contrôles, KRI, plans d'action). Ils sont dupliqués — et non
# extraits depuis get_decision_pilotage() — afin de ne pas modifier le moteur
# de décision BIC déjà validé.

def _critere_pertes_pilier2(code: str) -> tuple[DecisionCritere, str, str]:
    dashboard = get_dashboard()
    seuils = get_pertes_seuils()
    pertes_mois = dashboard.widget2.pertes_nettes_mois

    if pertes_mois >= seuils.seuil_reporting_conseil:
        statut, poids = "critique", 2
        organe = "Conseil d'administration"
        motif = (
            f"Pertes nettes du mois ({pertes_mois:,.0f} FCFA) supérieures au seuil de "
            f"reporting au Conseil ({seuils.seuil_reporting_conseil:,.0f} FCFA)."
        )
    elif pertes_mois >= seuils.seuil_reporting_direction:
        statut, poids = "attention", 1
        organe = "Direction Générale"
        motif = (
            f"Pertes nettes du mois ({pertes_mois:,.0f} FCFA) supérieures au seuil de "
            f"reporting à la Direction ({seuils.seuil_reporting_direction:,.0f} FCFA)."
        )
    elif pertes_mois >= seuils.seuil_reporting_interne:
        statut, poids = "attention", 1
        organe = "Comité des risques"
        motif = (
            f"Pertes nettes du mois ({pertes_mois:,.0f} FCFA) supérieures au seuil de "
            f"reporting interne ({seuils.seuil_reporting_interne:,.0f} FCFA)."
        )
    else:
        statut, poids = "conforme", 0
        organe = "Reporting interne standard"
        motif = "Les pertes nettes du mois restent inférieures aux seuils d'escalade Pilier 2."

    critere = DecisionCritere(
        code=code, libelle="Pertes opérationnelles vs seuils Pilier 2",
        reference_reglementaire="Art. 313.b",
        statut=statut, poids=poids,
        valeur_observee=f"{pertes_mois:,.0f} FCFA (mois en cours)",
        seuil_reference=(
            f"Interne {seuils.seuil_reporting_interne:,.0f} / "
            f"Direction {seuils.seuil_reporting_direction:,.0f} / "
            f"Conseil {seuils.seuil_reporting_conseil:,.0f} FCFA"
        ),
        commentaire=motif,
    )
    return critere, organe, motif


def _critere_incidents_tendance(code: str) -> DecisionCritere:
    dashboard = get_dashboard()
    non_clos = dashboard.widget2.incidents_non_clos
    evolution = dashboard.widget2.evolution_pertes_pct or 0.0
    if non_clos > 5 or evolution > 50:
        statut, poids = "critique", 2
    elif non_clos >= 3 or evolution > 15:
        statut, poids = "attention", 1
    else:
        statut, poids = "conforme", 0
    return DecisionCritere(
        code=code, libelle="Incidents ouverts et tendance des pertes",
        reference_reglementaire="Art. 313.b",
        statut=statut, poids=poids,
        valeur_observee=f"{non_clos} incident(s) non clôturé(s), évolution {evolution:+.1f} %",
        seuil_reference="≤ 2 incidents non clôturés, évolution ≤ 15 %",
        commentaire=(
            "Le volume d'incidents ouverts et/ou la tendance des pertes nécessite un suivi "
            "renforcé et une accélération des clôtures."
            if statut != "conforme" else
            "Le flux d'incidents et son évolution restent maîtrisés."
        ),
    )


def _critere_cartographie_risques(code: str) -> DecisionCritere:
    risques = list_risques()
    nb_critique_risque = sum(1 for r in risques if r.niveau_label == "Critique")
    nb_eleve_risque = sum(1 for r in risques if r.niveau_label == "Élevé")
    if nb_critique_risque > 0:
        statut, poids = "critique", 2
    elif nb_eleve_risque > 0:
        statut, poids = "attention", 1
    else:
        statut, poids = "conforme", 0
    return DecisionCritere(
        code=code, libelle="Cartographie des risques",
        reference_reglementaire="Art. 313",
        statut=statut, poids=poids,
        valeur_observee=f"{nb_critique_risque} risque(s) critique(s), {nb_eleve_risque} élevé(s)",
        seuil_reference="Aucun risque en zone critique",
        commentaire=(
            "Des risques en zone critique ou élevée doivent être couverts par un plan "
            "d'action documenté et suivi."
            if statut != "conforme" else
            "Aucun risque en zone critique dans la cartographie actuelle."
        ),
    )


def _critere_controles_internes(code: str) -> DecisionCritere:
    controles = list_controles()
    nb_non_conf = sum(1 for c in controles if c.resultat == "Non-conforme")
    if nb_non_conf > 2:
        statut, poids = "critique", 2
    elif nb_non_conf > 0:
        statut, poids = "attention", 1
    else:
        statut, poids = "conforme", 0
    return DecisionCritere(
        code=code, libelle="Efficacité du dispositif de contrôle interne",
        reference_reglementaire="Art. 314",
        statut=statut, poids=poids,
        valeur_observee=f"{nb_non_conf} contrôle(s) non conforme(s) sur {len(controles)}",
        seuil_reference="0 contrôle non conforme",
        commentaire=(
            "Des non-conformités de contrôle doivent être corrigées et documentées "
            "(conservation ≥ 7 ans, Art. 314)."
            if statut != "conforme" else
            "Le dispositif de contrôle interne ne présente pas de non-conformité ouverte."
        ),
    )


def _critere_kri_hors_seuil(code: str) -> DecisionCritere:
    kri = get_kri_module()
    if kri.kri_hors_seuil >= 2:
        statut, poids = "critique", 2
    elif kri.kri_hors_seuil == 1:
        statut, poids = "attention", 1
    else:
        statut, poids = "conforme", 0
    return DecisionCritere(
        code=code, libelle="Indicateurs clés de risque (KRI)",
        reference_reglementaire="Art. 313",
        statut=statut, poids=poids,
        valeur_observee=f"{kri.kri_hors_seuil} KRI hors seuil sur {len(kri.kri_list)}",
        seuil_reference="0 KRI en zone alerte/critique",
        commentaire=(
            "Un ou plusieurs indicateurs clés de risque dépassent leur seuil d'alerte et "
            "doivent faire l'objet d'un suivi rapproché."
            if statut != "conforme" else
            "L'ensemble des indicateurs clés de risque reste dans les seuils normaux."
        ),
    )


def _critere_plans_action(code: str) -> DecisionCritere:
    plans = list_plans()
    en_retard = sum(1 for p in plans if p.en_retard)
    if en_retard > 2:
        statut, poids = "critique", 2
    elif en_retard > 0:
        statut, poids = "attention", 1
    else:
        statut, poids = "conforme", 0
    return DecisionCritere(
        code=code, libelle="Avancement des plans d'action correctifs",
        reference_reglementaire="Art. 313.c",
        statut=statut, poids=poids,
        valeur_observee=f"{en_retard} plan(s) en retard sur {len(plans)}",
        seuil_reference="0 plan en retard",
        commentaire=(
            "Des plans d'action en retard exposent l'établissement à une récurrence des "
            "incidents ou risques déjà identifiés."
            if statut != "conforme" else
            "Les plans d'action correctifs et préventifs sont à jour."
        ),
    )


def _finaliser_decision(
    criteres: list[DecisionCritere], organe: str, motif: str,
) -> DecisionPilotageResult:
    score = sum(c.poids for c in criteres)
    score_max = 2 * len(criteres)

    if score <= 2:
        niveau = "Maîtrisé"
    elif score <= 6:
        niveau = "Sous surveillance"
    elif score <= 10:
        niveau = "Vigilance renforcée"
    else:
        niveau = "Critique"

    recommandations = [
        f"[{c.reference_reglementaire}] {c.commentaire}"
        for c in criteres if c.statut != "conforme"
    ]
    if not recommandations:
        recommandations = [
            "Aucune action corrective requise à ce stade — maintenir la surveillance "
            "courante du dispositif conformément à l'Art. 546 (reporting annuel à la "
            "Commission Bancaire de l'UMOA)."
        ]

    nb_attention = sum(1 for c in criteres if c.statut == "attention")
    nb_critique_total = sum(1 for c in criteres if c.statut == "critique")
    nb_conforme = len(criteres) - nb_attention - nb_critique_total

    synthese = (
        f"Sur {len(criteres)} critères analysés au titre du dispositif réglementaire UEMOA "
        f"de gestion et de surveillance du risque opérationnel, {nb_conforme} sont conformes, "
        f"{nb_attention} nécessitent une attention particulière et {nb_critique_total} "
        f"présentent un niveau critique. Le niveau de pilotage global retenu est « {niveau} » "
        f"(score {score}/{score_max}). Organe de reporting recommandé : {organe}."
    )

    return DecisionPilotageResult(
        date_analyse=utcnow_iso(),
        niveau_global=niveau,
        score_global=score,
        score_max=score_max,
        organe_reporting=organe,
        organe_reporting_motif=motif,
        criteres=criteres,
        recommandations=recommandations,
        synthese=synthese,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Décision de pilotage — AIB (Approche Indicateur de Base, Art. 301)
# ═══════════════════════════════════════════════════════════════════════════════

def get_decision_aib() -> DecisionPilotageResult:
    """
    Moteur de décision pour l'onglet A1 — AIB.
    Combine la robustesse des données PNB retenues pour le calcul déclaratoire
    avec les signaux transverses de risque opérationnel.
    """
    criteres: list[DecisionCritere] = []

    try:
        aib = calcul_aib()
        aib_erreur = False
    except Exception:
        aib = None
        aib_erreur = True

    if aib_erreur or aib.donnees_insuffisantes:
        criteres.append(DecisionCritere(
            code="C1", libelle="Disponibilité des données PNB (AIB)",
            reference_reglementaire="Art. 301",
            statut="critique", poids=2,
            valeur_observee="Aucun exercice de PNB positif exploitable",
            seuil_reference="≥ 1 exercice de PNB positif sur 3",
            commentaire="Le calcul déclaratoire de l'Indicateur de Base ne peut être établi "
                        "faute de PNB positif saisi sur les 3 derniers exercices.",
        ))
    elif aib.n < 3:
        criteres.append(DecisionCritere(
            code="C1", libelle="Disponibilité des données PNB (AIB)",
            reference_reglementaire="Art. 301",
            statut="attention", poids=1,
            valeur_observee=f"{aib.n} exercice(s) de PNB positif sur 3",
            seuil_reference="3 exercices de PNB positif",
            commentaire="La moyenne du PNB retenue repose sur moins de 3 exercices positifs, "
                        "ce qui peut fragiliser la robustesse de l'exigence de fonds propres.",
        ))
    else:
        criteres.append(DecisionCritere(
            code="C1", libelle="Disponibilité des données PNB (AIB)",
            reference_reglementaire="Art. 301",
            statut="conforme", poids=0,
            valeur_observee=f"{aib.n} exercices de PNB positif sur 3",
            seuil_reference="3 exercices de PNB positif",
            commentaire="La moyenne triennale du PNB repose sur des données complètes.",
        ))

    critere_pertes, organe, motif = _critere_pertes_pilier2("C2")
    criteres.append(critere_pertes)
    criteres.append(_critere_incidents_tendance("C3"))
    criteres.append(_critere_cartographie_risques("C4"))
    criteres.append(_critere_controles_internes("C5"))
    criteres.append(_critere_kri_hors_seuil("C6"))
    criteres.append(_critere_plans_action("C7"))

    return _finaliser_decision(criteres, organe, motif)


# ═══════════════════════════════════════════════════════════════════════════════
# Décision de pilotage — AS (Approche Standard, Art. 305-311)
# ═══════════════════════════════════════════════════════════════════════════════

def get_decision_as() -> DecisionPilotageResult:
    """
    Moteur de décision pour l'onglet A2 — AS.
    Vérifie en priorité l'autorisation réglementaire préalable de la Commission
    Bancaire (art. 300 UEMOA), puis applique les signaux transverses de risque
    opérationnel.
    """
    criteres: list[DecisionCritere] = []

    try:
        as_res = calcul_as()
        as_erreur = False
    except Exception:
        as_res = None
        as_erreur = True

    if as_erreur:
        criteres.append(DecisionCritere(
            code="C1", libelle="Autorisation et disponibilité des données (AS)",
            reference_reglementaire="Art. 300",
            statut="critique", poids=2,
            valeur_observee="Non calculable",
            seuil_reference="Autorisation CB + données par ligne de métier",
            commentaire="Impossible d'évaluer l'Approche Standard faute de calcul disponible.",
        ))
    elif not as_res.as_autorisee:
        criteres.append(DecisionCritere(
            code="C1", libelle="Autorisation et disponibilité des données (AS)",
            reference_reglementaire="Art. 300",
            statut="critique", poids=2,
            valeur_observee="Approche Standard non autorisée",
            seuil_reference="Autorisation préalable de la Commission Bancaire requise",
            commentaire="L'établissement n'est pas autorisé par la Commission Bancaire de "
                        "l'UMOA à utiliser l'Approche Standard ; les montants affichés sont "
                        "indicatifs et ne peuvent être retenus pour le reporting réglementaire.",
        ))
    elif as_res.donnees_insuffisantes:
        criteres.append(DecisionCritere(
            code="C1", libelle="Autorisation et disponibilité des données (AS)",
            reference_reglementaire="Art. 305-311",
            statut="attention", poids=1,
            valeur_observee="Autorisée — données par ligne de métier incomplètes",
            seuil_reference="PNB des 8 lignes de métier sur 3 exercices",
            commentaire="L'Approche Standard est autorisée mais les données de PNB par ligne "
                        "de métier restent incomplètes sur les exercices requis.",
        ))
    else:
        criteres.append(DecisionCritere(
            code="C1", libelle="Autorisation et disponibilité des données (AS)",
            reference_reglementaire="Art. 305-311",
            statut="conforme", poids=0,
            valeur_observee="Autorisée — données complètes",
            seuil_reference="PNB des 8 lignes de métier sur 3 exercices",
            commentaire="L'Approche Standard est autorisée et les données par ligne de "
                        "métier sont complètes.",
        ))

    critere_pertes, organe, motif = _critere_pertes_pilier2("C2")
    criteres.append(critere_pertes)
    criteres.append(_critere_incidents_tendance("C3"))
    criteres.append(_critere_cartographie_risques("C4"))
    criteres.append(_critere_controles_internes("C5"))
    criteres.append(_critere_kri_hors_seuil("C6"))
    criteres.append(_critere_plans_action("C7"))

    return _finaliser_decision(criteres, organe, motif)
