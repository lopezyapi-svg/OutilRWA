"""Acces centralise aux donnees Excel avec cache et overlays session."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
import json
from pathlib import Path
import re
from threading import Lock
from typing import Any
import unicodedata

from openpyxl import Workbook, load_workbook


def _resolve_excel_source_path() -> Path:
    desktop_candidates = (
        Path.home() / "OneDrive" / "Desktop",
        Path.home() / "Desktop",
    )
    workbook_candidates: list[Path] = []
    for desktop in desktop_candidates:
        if not desktop.exists():
            continue
        workbook_candidates.extend(
            path for path in desktop.glob("BASE_CALCUL_RWA*.xlsx") if path.is_file()
        )
    if workbook_candidates:
        return max(workbook_candidates, key=lambda path: path.stat().st_mtime)
    return Path(r"C:\Users\hp\OneDrive\Desktop\BASE_CALCUL_RWA_v2_pers-21-04-2026.xlsx")


EXCEL_SOURCE_PATH = _resolve_excel_source_path()
APP_DATA_PATH = Path(__file__).resolve().parents[2] / "data"
EXPOSURE_METADATA_PATH = APP_DATA_PATH / "exposure_metadata.json"
EXPOSURE_EXPORTS_PATH = APP_DATA_PATH / "exports"
EXPECTED_SHEETS = (
    "Template données",
    "LISTE",
    "Traitement_HB",
    "CRM_non_financee",
    "CRM_financée",
    "(a) souverains",
    "(b) organismes pub. hors Adm c",
    "(c) Expositions sur les BMD",
    "(d) institutions financières",
    "(e) entreprises",
    "(f) clientèle de détail",
    "(g) prêts garantis par l'immo R",
    "(h) prêts garantis par l'immo C",
    "(i) créances en souffrance",
    "(j) créances à risque élevé",
    "(k) autres actifs",
    "(l) Hors bilan",
    "Mapping des pondérations",
    "Ref_Ponderation",
)
EXPECTED_COLUMNS_BY_SHEET: dict[str, tuple[str, ...]] = {
    "Template données": (
        "Date d'analyse",
        "ID_Exposition",
        "Date d'octroi",
        "Date d'échéance",
        "Maturité de l'exposition",
        "Maturité résiduelle",
        "Contrepartie",
        "Notation_externe_contrepartie",
        "Pays_contrepartie",
        "Notation_externe_pays",
        "Pondération_pays",
        "Catégorie d'exposition",
        "Pondération (RW)",
        "PRÊT TOTAL",
        "Montant_exposition_but_au_bilan",
        "Montant d'exposition au HB",
        "Devise",
        "CRM_existe",
        "Type_CRM",
        "EAD_bilan",
        "EAD_HB",
        "EAD_HB_ccf",
        "EAD_Total",
        "RWA_EB",
        "RWA_HB",
        "RWA_crédit",
        "Capital_min_reg",
    ),
    "Traitement_HB": (
        "ID_Exposition",
        "Catégorie Hors bilan",
        "Facteur_conversion (CCF)",
        "EAD_HB_ccf",
    ),
    "CRM_non_financee": (
        "ID_Exposition",
        "Nom du garant",
        "Catégorie du garant",
        "Note_garant",
        "% Exp_couverte",
        "Part couverte",
        "Pondération du garant",
    ),
    "CRM_financée": (
        "ID_Exposition",
        "Valeur_Collatéral",
        "Type_emetteur",
        "Notation",
        "Bloc",
        "Maturite",
        "HE",
        "HC",
        "Hfx",
        "Eva_EB",
        "Eva_HB",
        "Cva",
    ),
}
OPTIONAL_COLUMNS_BY_SHEET: dict[str, tuple[str, ...]] = {
    "Template données": (
        "Commentaire",
    ),
}
EXPECTED_INDICATORS_BY_SHEET: dict[str, tuple[str, ...]] = {
    "Ref_Ponderation": (
        "Notation / RW Souverains",
        "Notation / RW Organismes publics",
        "Notation / RW Entreprises",
        "Notation / RW Banques <=3m",
        "Notation / RW Banques >3m",
        "Type_HB",
        "CCF_HB",
    ),
}
REF_REQUIRED_MARKERS = ("Type_HB", "CCF_HB")
REF_MIN_COLUMN_COUNT = 17


class ExcelImportValidationError(ValueError):
    """Erreur metier retournee si la structure du fichier Excel est invalide."""

    def __init__(self, payload: dict[str, Any]) -> None:
        super().__init__(payload.get("message", "Invalid Excel workbook structure."))
        self.payload = payload


def _normalize_apostrophes(value: str) -> str:
    return value.replace("’", "'").replace("`", "'")


def _normalize_for_key(value: str) -> str:
    normalized = _normalize_apostrophes(value.strip().lower())
    normalized = unicodedata.normalize("NFKD", normalized)
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    return " ".join(normalized.split())


def _as_clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _as_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    lowered = _normalize_for_key(text)
    if lowered in {"non eligible", "n/a", "na", "none"}:
        return 0.0
    compact = text.replace("\u202f", "").replace(" ", "").replace(",", ".")
    try:
        return float(compact)
    except ValueError:
        return 0.0


def _as_optional_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    lowered = _normalize_for_key(text)
    if lowered in {"nd", "non eligible", "n/a", "na", "none"}:
        return None
    compact = text.replace("\u202f", "").replace(" ", "").replace(",", ".")
    try:
        return float(compact)
    except ValueError:
        return None


def _as_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        candidate = value.strip().split(" ")[0]
        try:
            return date.fromisoformat(candidate)
        except ValueError:
            return date.today()
    return date.today()


def _as_optional_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        candidate = value.strip().split(" ")[0]
        for parser in (
            date.fromisoformat,
            lambda raw: datetime.strptime(raw, "%d/%m/%Y").date(),
            lambda raw: datetime.strptime(raw, "%d-%m-%Y").date(),
        ):
            try:
                return parser(candidate)
            except ValueError:
                continue
    return None


def _dashboard_category_from_raw(raw: str) -> str:
    cleaned = re.sub(r"^\([^)]+\)\s*", "", raw.strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or "Non renseigné"


def _standard_category_from_raw(raw: str) -> str:
    lowered = _normalize_for_key(raw)
    if lowered.startswith("(a)") or lowered.startswith("(b)"):
        return "Souverains"
    if lowered.startswith("(c)") or lowered.startswith("(d)"):
        return "Banques"
    if lowered.startswith("(e)") or lowered.startswith("(i)") or lowered.startswith("(j)") or lowered.startswith("(k)"):
        return "Entreprises"
    if lowered.startswith("(f)") or lowered.startswith("(g)") or lowered.startswith("(h)"):
        return "Particuliers"
    if lowered.startswith("(l)"):
        return "Entreprises"
    return "Entreprises"


def _normalize_country_display(country: str) -> str:
    lowered = _normalize_for_key(country)
    if lowered in {"cote d'ivoire", "cote d ivoire"}:
        return "Côte d'Ivoire"
    if lowered == "senegal":
        return "Sénégal"
    if lowered == "benin":
        return "Bénin"
    if lowered == "guinee-bissau":
        return "Guinée-Bissau"
    return country.strip()


def _crm_mode_from_type(crm_type: str) -> str:
    normalized = _normalize_for_key(crm_type)
    if not normalized or normalized in {"aucune", "sans crm", "non eligible"}:
        return "Aucune"
    if "non financ" in normalized or "garantie" in normalized or "assurance" in normalized:
        return "CRM non financee"
    if "cash" in normalized or "collateral" in normalized or "financee" in normalized:
        return "CRM financee"
    return "CRM non financee"


def _months_between(start: date | None, end: date | None) -> int | None:
    if start is None or end is None:
        return None
    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day < start.day:
        months -= 1
    return max(months, 0)


def _as_yes_no(value: bool) -> str:
    return "Oui" if value else "Non"


@dataclass(slots=True)
class ExcelWorkbookCache:
    """Cache parse du classeur Excel source."""

    template_rows: list[dict[str, Any]]
    hb_rows: list[dict[str, Any]]
    crm_non_fin_rows: list[dict[str, Any]]
    crm_fin_rows: list[dict[str, Any]]
    ref_rows: list[list[Any]]


@dataclass(slots=True)
class SessionOverlay:
    """Donnees ajoutees en session, sans ecriture sur le classeur."""

    exposures: list[dict[str, Any]] = field(default_factory=list)
    off_balance: list[dict[str, Any]] = field(default_factory=list)
    crm: list[dict[str, Any]] = field(default_factory=list)
    reports: list[dict[str, Any]] = field(default_factory=list)


class ExcelRepository:
    """Fournit les lignes metier reconstruites depuis le fichier Excel."""

    def __init__(
        self,
        path: Path = EXCEL_SOURCE_PATH,
        metadata_path: Path = EXPOSURE_METADATA_PATH,
        exports_path: Path = EXPOSURE_EXPORTS_PATH,
    ) -> None:
        self._path = path
        self._metadata_path = metadata_path
        self._exports_path = exports_path
        self._lock = Lock()
        self._cached_mtime: float | None = None
        self._cache: ExcelWorkbookCache | None = None
        self._overlay = SessionOverlay()

    @property
    def source_path(self) -> Path:
        return self._path

    def _sheet_headers(self, workbook, sheet_name: str) -> list[str]:
        sheet = workbook[sheet_name]
        headers: list[str] = []
        for index in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=index).value
            text = _as_clean_text(value)
            headers.append(text or "")
        return headers

    def _collect_sheet_markers(self, workbook, sheet_name: str, max_columns: int = 23, max_rows: int = 25) -> set[str]:
        sheet = workbook[sheet_name]
        markers: set[str] = set()
        for row_index in range(1, min(sheet.max_row, max_rows) + 1):
            for column_index in range(1, min(sheet.max_column, max_columns) + 1):
                value = sheet.cell(row=row_index, column=column_index).value
                text = _as_clean_text(value)
                if text:
                    markers.add(_normalize_for_key(text))
        return markers

    def _validate_workbook_structure(self, workbook) -> None:
        workbook_sheets = set(workbook.sheetnames)
        missing_sheets = [sheet for sheet in EXPECTED_SHEETS if sheet not in workbook_sheets]
        missing_columns_by_sheet: dict[str, list[str]] = {}
        available_columns_by_sheet: dict[str, list[str]] = {}
        missing_indicators_by_sheet: dict[str, list[str]] = {}
        sheet_diagnostics: dict[str, list[str]] = {}

        for sheet_name, required_columns in EXPECTED_COLUMNS_BY_SHEET.items():
            if sheet_name not in workbook_sheets:
                continue
            available_columns = [header for header in self._sheet_headers(workbook, sheet_name) if header]
            available_columns_by_sheet[sheet_name] = available_columns
            missing_columns = [column for column in required_columns if column not in available_columns]
            if missing_columns:
                missing_columns_by_sheet[sheet_name] = missing_columns

        if "Ref_Ponderation" in workbook_sheets:
            ref_sheet = workbook["Ref_Ponderation"]
            ref_markers = self._collect_sheet_markers(workbook, "Ref_Ponderation")
            missing_markers = [
                marker for marker in REF_REQUIRED_MARKERS if _normalize_for_key(marker) not in ref_markers
            ]
            if missing_markers:
                missing_indicators_by_sheet["Ref_Ponderation"] = missing_markers
            if ref_sheet.max_column < REF_MIN_COLUMN_COUNT:
                sheet_diagnostics["Ref_Ponderation"] = [
                    f"La feuille doit contenir au moins {REF_MIN_COLUMN_COUNT} colonnes utiles pour les RW et les CCF."
                ]

        if missing_sheets or missing_columns_by_sheet or missing_indicators_by_sheet or sheet_diagnostics:
            payload: dict[str, Any] = {
                "error_code": "excel_structure_invalid",
                "message": "Le fichier Excel importé ne correspond pas au format attendu.",
                "expected_sheets": list(EXPECTED_SHEETS),
                "available_sheets": list(workbook.sheetnames),
                "sheet_columns": {sheet: list(columns) for sheet, columns in EXPECTED_COLUMNS_BY_SHEET.items()},
                "optional_columns": {sheet: list(columns) for sheet, columns in OPTIONAL_COLUMNS_BY_SHEET.items()},
                "sheet_indicators": {sheet: list(indicators) for sheet, indicators in EXPECTED_INDICATORS_BY_SHEET.items()},
            }
            if missing_sheets:
                payload["missing_sheets"] = missing_sheets
            if missing_columns_by_sheet:
                payload["missing_columns_by_sheet"] = missing_columns_by_sheet
            if available_columns_by_sheet:
                payload["available_columns_by_sheet"] = available_columns_by_sheet
            if missing_indicators_by_sheet:
                payload["missing_indicators_by_sheet"] = missing_indicators_by_sheet
            if sheet_diagnostics:
                payload["sheet_diagnostics"] = sheet_diagnostics
            raise ExcelImportValidationError(payload)

    def _read_sheet_rows(self, workbook, sheet_name: str) -> list[dict[str, Any]]:
        sheet = workbook[sheet_name]
        header = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
        headers = [str(item).strip() if item is not None else "" for item in header]
        rows: list[dict[str, Any]] = []
        for row_index in range(2, sheet.max_row + 1):
            values = [sheet.cell(row=row_index, column=index).value for index in range(1, sheet.max_column + 1)]
            if not any(value not in (None, "") for value in values):
                continue
            row: dict[str, Any] = {}
            for index, value in enumerate(values):
                key = headers[index]
                if not key:
                    key = f"col_{index + 1}"
                row[key] = value
            rows.append(row)
        return rows

    def _read_ref_rows(self, workbook) -> list[list[Any]]:
        sheet = workbook["Ref_Ponderation"]
        rows: list[list[Any]] = []
        for row_index in range(1, sheet.max_row + 1):
            values = [sheet.cell(row=row_index, column=index).value for index in range(1, 24)]
            if any(value not in (None, "") for value in values):
                rows.append(values)
        return rows

    def _build_cache_from_workbook(self, workbook) -> ExcelWorkbookCache:
        self._validate_workbook_structure(workbook)
        return ExcelWorkbookCache(
            template_rows=self._read_sheet_rows(workbook, "Template données"),
            hb_rows=self._read_sheet_rows(workbook, "Traitement_HB"),
            crm_non_fin_rows=self._read_sheet_rows(workbook, "CRM_non_financee"),
            crm_fin_rows=self._read_sheet_rows(workbook, "CRM_financée"),
            ref_rows=self._read_ref_rows(workbook),
        )

    def _build_cache_for_path(self, path: Path) -> ExcelWorkbookCache:
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            return self._build_cache_from_workbook(workbook)
        finally:
            workbook.close()

    def _load_cache(self) -> ExcelWorkbookCache:
        with self._lock:
            mtime = self._path.stat().st_mtime
            if self._cache is not None and self._cached_mtime == mtime:
                return self._cache
            cache = self._build_cache_for_path(self._path)
            self._cached_mtime = mtime
            self._cache = cache
            return cache

    def _empty_metadata(self) -> dict[str, dict[str, dict[str, Any]]]:
        return {
            "exposures": {},
            "off_balance": {},
        }

    def _load_metadata_unlocked(self) -> dict[str, dict[str, dict[str, Any]]]:
        if not self._metadata_path.exists():
            return self._empty_metadata()
        try:
            payload = json.loads(self._metadata_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return self._empty_metadata()
        metadata = self._empty_metadata()
        exposures = payload.get("exposures")
        off_balance = payload.get("off_balance")
        if isinstance(exposures, dict):
            metadata["exposures"] = {
                str(key): dict(value) for key, value in exposures.items() if isinstance(value, dict)
            }
        if isinstance(off_balance, dict):
            metadata["off_balance"] = {
                str(key): dict(value) for key, value in off_balance.items() if isinstance(value, dict)
            }
        return metadata

    def _save_metadata_unlocked(self, payload: dict[str, dict[str, dict[str, Any]]]) -> None:
        self._metadata_path.parent.mkdir(parents=True, exist_ok=True)
        self._metadata_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )

    def _invalidate_cache_unlocked(self) -> None:
        self._cached_mtime = None
        self._cache = None

    def _sheet_header_index(self, sheet) -> dict[str, int]:
        headers: dict[str, int] = {}
        for column_index in range(1, sheet.max_column + 1):
            value = _as_clean_text(sheet.cell(row=1, column=column_index).value)
            if value:
                headers[value] = column_index
        return headers

    def _find_row_index_by_id(self, sheet, header_index: dict[str, int], exposure_id: str) -> int | None:
        id_column = header_index.get("ID_Exposition")
        if id_column is None:
            return None
        for row_index in range(2, sheet.max_row + 1):
            current_id = _as_clean_text(sheet.cell(row=row_index, column=id_column).value)
            if current_id == exposure_id:
                return row_index
        return None

    def _set_row_values(
        self,
        sheet,
        header_index: dict[str, int],
        row_index: int,
        values: dict[str, Any],
    ) -> None:
        for header, value in values.items():
            column_index = header_index.get(header)
            if column_index is None:
                continue
            sheet.cell(row=row_index, column=column_index, value=value)

    def _delete_rows_by_id(self, sheet, header_index: dict[str, int], exposure_id: str) -> None:
        id_column = header_index.get("ID_Exposition")
        if id_column is None:
            return
        for row_index in range(sheet.max_row, 1, -1):
            current_id = _as_clean_text(sheet.cell(row=row_index, column=id_column).value)
            if current_id == exposure_id:
                sheet.delete_rows(row_index, 1)

    def _append_row(self, sheet, header_index: dict[str, int], values: dict[str, Any]) -> int:
        row_index = sheet.max_row + 1
        self._set_row_values(sheet, header_index, row_index, values)
        return row_index

    def _delete_rows_by_ids(
        self,
        sheet,
        header_index: dict[str, int],
        exposure_ids: set[str],
    ) -> set[str]:
        id_column = header_index.get("ID_Exposition")
        if id_column is None or not exposure_ids:
            return set()
        deleted_ids: set[str] = set()
        for row_index in range(sheet.max_row, 1, -1):
            current_id = _as_clean_text(sheet.cell(row=row_index, column=id_column).value)
            if current_id in exposure_ids:
                deleted_ids.add(current_id)
                sheet.delete_rows(row_index, 1)
        return deleted_ids

    def _row_index_by_id(self, sheet, header_index: dict[str, int]) -> dict[str, int]:
        id_column = header_index.get("ID_Exposition")
        if id_column is None:
            return {}
        indexed: dict[str, int] = {}
        for row_index in range(2, sheet.max_row + 1):
            current_id = _as_clean_text(sheet.cell(row=row_index, column=id_column).value)
            if current_id:
                indexed[current_id] = row_index
        return indexed

    def _upsert_exposure_unlocked(
        self,
        item: dict[str, Any],
        *,
        metadata: dict[str, dict[str, dict[str, Any]]],
        template_sheet,
        crm_non_fin_sheet,
        crm_fin_sheet,
        template_headers: dict[str, int],
        crm_non_fin_headers: dict[str, int],
        crm_fin_headers: dict[str, int],
        template_row_index: dict[str, int] | None = None,
    ) -> None:
        exposure_id = str(item["id"])
        template_row = (
            template_row_index.get(exposure_id)
            if template_row_index is not None
            else self._find_row_index_by_id(template_sheet, template_headers, exposure_id)
        )
        row_values = {
            "Date d'analyse": item["analysis_date"],
            "ID_Exposition": exposure_id,
            "Date d'octroi": item.get("grant_date"),
            "Date d'échéance": item.get("maturity_date"),
            "Maturité de l'exposition": item.get("exposure_maturity_months")
            or _months_between(item.get("grant_date"), item.get("maturity_date")),
            "Maturité résiduelle": item.get("residual_maturity_months")
            or _months_between(item.get("analysis_date"), item.get("maturity_date")),
            "Contrepartie": item["counterparty_name"],
            "Notation_externe_contrepartie": item["rating"],
            "Pays_contrepartie": item["country"],
            "Notation_externe_pays": item.get("country_rating", "Non noté"),
            "Pondération_pays": item.get("country_risk_weight"),
            "Catégorie d'exposition": item["category_raw"],
            "Pondération (RW)": item["final_rw"],
            "PRÊT TOTAL": item.get("loan_total_amount", item["gross_amount"]),
            "Montant_exposition_but_au_bilan": item.get(
                "on_balance_exposure_amount",
                item["gross_amount"],
            ),
            "Montant d'exposition au HB": item.get("off_balance_exposure_amount", 0.0),
            "Devise": item.get("currency", "XOF"),
            "CRM_existe": "OUI" if item.get("crm_exists", False) else "NON",
            "Type_CRM": item["crm_type"],
            "EAD_bilan": item.get("ead_bilan_amount", item["ead"]),
            "EAD_HB": item.get("ead_hb_amount"),
            "EAD_HB_ccf": item.get("ead_hb_ccf_amount"),
            "EAD_Total": item.get("ead_total_amount", item["ead"]),
            "RWA_EB": item.get("rwa_eb_amount"),
            "RWA_HB": item.get("rwa_hb_amount"),
            "RWA_crédit": item["rwa"],
            "Capital_min_reg": item["capital"],
        }
        if template_row is None:
            template_row = self._append_row(template_sheet, template_headers, row_values)
            if template_row_index is not None:
                template_row_index[exposure_id] = template_row
        else:
            self._set_row_values(template_sheet, template_headers, template_row, row_values)

        self._delete_rows_by_id(crm_non_fin_sheet, crm_non_fin_headers, exposure_id)
        self._delete_rows_by_id(crm_fin_sheet, crm_fin_headers, exposure_id)

        crm_details = dict(item.get("crm_details", {}))
        crm_mode = _crm_mode_from_type(str(crm_details.get("mode") or item.get("crm_type") or "Aucune"))
        coverage = _as_float(item.get("crm_coverage_percent"))
        if crm_mode == "CRM non financee":
            guarantor_rw = _as_float(item.get("guarantor_rw"))
            covered_amount = round(_as_float(item["gross_amount"]) * coverage, 2)
            non_covered_ratio = max(0.0, round(1.0 - coverage, 4))
            self._append_row(
                crm_non_fin_sheet,
                crm_non_fin_headers,
                {
                    "ID_Exposition": exposure_id,
                    "Nom du garant": crm_details.get("guarantor_name") or "",
                    "Note_garant": crm_details.get("guarantor_rating") or "",
                    "Pays_garant": crm_details.get("guarantor_country") or "",
                    "Note_pays_garant": crm_details.get("guarantor_country_rating") or "",
                    "Pondération_pays_garant": crm_details.get("guarantor_country_rw") or "",
                    "Catégorie du garant": crm_details.get("guarantor_category") or "",
                    "Pondération du garant": guarantor_rw,
                    "% Exp_couverte": coverage,
                    "%Exp_Nn_couverte": non_covered_ratio,
                    "Part couverte": covered_amount,
                    "Part non couverte": round(_as_float(item["gross_amount"]) - covered_amount, 2),
                    "RWA_non_fin": item["rwa"],
                },
            )
        elif crm_mode == "CRM financee":
            self._append_row(
                crm_fin_sheet,
                crm_fin_headers,
                {
                    "ID_Exposition": exposure_id,
                    "Valeur_Collatéral": crm_details.get("collateral_value") or 0.0,
                    "Type_emetteur": crm_details.get("issuer_type") or "",
                    "Notation": crm_details.get("issuer_rating") or "",
                    "Bloc": crm_details.get("label") or item["crm_type"],
                    "Maturite": crm_details.get("maturity_bucket") or "<=1 an",
                    "HE": crm_details.get("haircut") or 0.0,
                    "HC": crm_details.get("haircut") or 0.0,
                    "Hfx": crm_details.get("fx_haircut") or 0.0,
                    "Eva": item["gross_amount"],
                    "Cva": item["ead"],
                },
            )

        metadata["exposures"][exposure_id] = {
            "currency": item.get("currency", "XOF"),
            "status": item.get("status", "Active"),
            "comment": item.get("comment") or "",
            "crm_coverage_percent": coverage,
            "crm_details": crm_details,
            "original_rw": item.get("original_rw", 0.0),
            "final_rw": item.get("final_rw", 0.0),
            "ead": item.get("ead", 0.0),
            "rwa": item.get("rwa", 0.0),
            "capital": item.get("capital", 0.0),
        }

    def _append_off_balance_unlocked(
        self,
        item: dict[str, Any],
        *,
        metadata: dict[str, dict[str, dict[str, Any]]],
        sheet,
        headers: dict[str, int],
    ) -> str:
        row_index = self._append_row(
            sheet,
            headers,
            {
                "ID_Exposition": item["counterparty_id"],
                "Catégorie Hors bilan": item["engagement_type"],
                "Facteur_conversion (CCF)": item["ccf"],
                "EAD_HB_ccf": item.get("ead"),
            },
        )
        commitment_id = f"HB{row_index - 1:03d}"
        metadata["off_balance"][commitment_id] = {
            "comment": item.get("comment") or "",
        }
        return commitment_id

    def _metadata_for_exposure(self, exposure_id: str) -> dict[str, Any]:
        with self._lock:
            metadata = self._load_metadata_unlocked()
            return dict(metadata["exposures"].get(exposure_id, {}))

    def _crm_non_fin_by_rows(self, rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        indexed: dict[str, dict[str, Any]] = {}
        for row in rows:
            exposure_id = _as_clean_text(row.get("ID_Exposition"))
            if exposure_id:
                indexed[exposure_id] = row
        return indexed

    def _crm_fin_by_rows(self, rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        indexed: dict[str, dict[str, Any]] = {}
        for row in rows:
            exposure_id = _as_clean_text(row.get("ID_Exposition"))
            if exposure_id:
                indexed[exposure_id] = row
        return indexed

    def _find_template_by_id(self, exposure_id: str) -> dict[str, Any] | None:
        for row in self._load_cache().template_rows:
            current_id = _as_clean_text(row.get("ID_Exposition"))
            if current_id == exposure_id:
                return row
        return None

    def _build_exposure_from_template(
        self,
        row: dict[str, Any],
        *,
        exposure_metadata: dict[str, Any] | None = None,
        crm_non_fin_row: dict[str, Any] | None = None,
        crm_fin_row: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        exposure_id = _as_clean_text(row.get("ID_Exposition")) or ""
        metadata = exposure_metadata or {}
        raw_category = _as_clean_text(row.get("Catégorie d'exposition")) or "Entreprises"
        country = _normalize_country_display(_as_clean_text(row.get("Pays_contrepartie")) or "")
        country_rating = (
            _as_clean_text(metadata.get("country_rating"))
            or _as_clean_text(row.get("Notation_externe_pays"))
            or "Non noté"
        )
        rating = _as_clean_text(row.get("Notation_externe_contrepartie")) or "Non noté"
        grant_date = _as_optional_date(metadata.get("grant_date")) or _as_optional_date(
            row.get("Date d'octroi")
        )
        maturity_date = _as_optional_date(metadata.get("maturity_date")) or _as_optional_date(
            row.get("Date d'échéance")
        )
        crm_type = _as_clean_text(row.get("Type_CRM")) or "Sans CRM"
        loan_total_amount = _as_optional_float(row.get("PRÊT TOTAL"))
        on_balance_exposure_amount = _as_optional_float(
            row.get("Montant_exposition_but_au_bilan")
        )
        off_balance_exposure_amount = _as_optional_float(
            row.get("Montant d'exposition au HB")
        )
        country_risk_weight = _as_optional_float(row.get("Pondération_pays"))
        ead_bilan_amount = _as_optional_float(row.get("EAD_bilan"))
        ead_hb_amount = _as_optional_float(row.get("EAD_HB"))
        ead_hb_ccf_amount = _as_optional_float(row.get("EAD_HB_ccf"))
        ead_total_amount = _as_optional_float(row.get("EAD_Total"))
        rwa_eb_amount = _as_optional_float(row.get("RWA_EB"))
        rwa_hb_amount = _as_optional_float(row.get("RWA_HB"))
        rwa_credit_amount = _as_optional_float(row.get("RWA_crédit"))
        capital_amount = _as_optional_float(row.get("Capital_min_reg"))
        stored_rw = _as_optional_float(row.get("Pondération (RW)")) or 0.0
        gross_amount = on_balance_exposure_amount or loan_total_amount or 0.0
        exposure_maturity_months = (
            int(raw_value)
            if (raw_value := _as_optional_float(row.get("Maturité de l'exposition"))) is not None
            else _months_between(grant_date, maturity_date)
        )
        residual_maturity_months = (
            int(raw_value)
            if (raw_value := _as_optional_float(row.get("Maturité résiduelle"))) is not None
            else _months_between(_as_date(row.get("Date d'analyse")), maturity_date)
        )
        is_off_balance = _normalize_for_key(raw_category).startswith("(l) hors bilan")
        computed_ead_total = (
            ead_total_amount
            if ead_total_amount is not None
            else (
                (ead_bilan_amount or 0.0) + (ead_hb_ccf_amount or 0.0)
                if ead_bilan_amount is not None or ead_hb_ccf_amount is not None
                else ead_bilan_amount
            )
        )
        ead = 0.0 if is_off_balance else (computed_ead_total or 0.0)
        rwa = 0.0 if is_off_balance else (
            rwa_credit_amount
            if rwa_credit_amount is not None
            else (
                ((rwa_eb_amount or 0.0) + (rwa_hb_amount or 0.0))
                if rwa_eb_amount is not None or rwa_hb_amount is not None
                else round((computed_ead_total or 0.0) * stored_rw, 2)
            )
        )
        capital = 0.0 if is_off_balance else (
            capital_amount if capital_amount is not None else round(rwa * 0.08, 2)
        )
        crm_coverage_percent = _as_float(metadata.get("crm_coverage_percent"))
        crm_mode = _crm_mode_from_type(crm_type)
        crm_details = metadata.get("crm_details")
        if not isinstance(crm_details, dict):
            crm_details = {
                "mode": crm_mode,
                "label": crm_type,
                "coverage_percent": crm_coverage_percent,
            }
            if crm_mode == "CRM non financee" and crm_non_fin_row is not None:
                coverage = _as_float(crm_non_fin_row.get("% Exp_couverte"))
                crm_coverage_percent = coverage
                crm_details.update(
                    {
                        "guarantor_name": _as_clean_text(crm_non_fin_row.get("Nom du garant")) or "",
                        "guarantor_category": _as_clean_text(crm_non_fin_row.get("Catégorie du garant")) or "",
                        "guarantor_rating": _as_clean_text(crm_non_fin_row.get("Note_garant")) or "",
                        "coverage_percent": coverage,
                    }
                )
            elif crm_mode == "CRM financee" and crm_fin_row is not None:
                coverage_amount = max(
                    (
                        _as_float(crm_fin_row.get("Eva_EB"))
                        + _as_float(crm_fin_row.get("Eva_HB"))
                    )
                    - _as_float(crm_fin_row.get("Cva")),
                    0.0,
                )
                coverage_base = loan_total_amount or gross_amount
                coverage = 0.0 if coverage_base == 0 else min(coverage_amount / coverage_base, 1.0)
                crm_coverage_percent = coverage
                crm_details.update(
                    {
                        "collateral_value": _as_float(crm_fin_row.get("Valeur_Collatéral")),
                        "issuer_type": _as_clean_text(crm_fin_row.get("Type_emetteur")) or "",
                        "issuer_rating": _as_clean_text(crm_fin_row.get("Notation")) or "",
                        "maturity_bucket": _as_clean_text(crm_fin_row.get("Maturite")) or "<=1 an",
                        "fx_haircut": _as_float(crm_fin_row.get("Hfx")),
                        "coverage_percent": coverage,
                    }
                )

        original_rw = _as_float(metadata.get("original_rw")) or stored_rw
        final_rw = _as_float(metadata.get("final_rw")) or (stored_rw if ead > 0 else 0.0)
        ead = _as_float(metadata.get("ead")) or ead
        rwa = _as_float(metadata.get("rwa")) or rwa
        capital = _as_float(metadata.get("capital")) or capital
        source_fields = {
            "loan_total_amount": loan_total_amount,
            "on_balance_exposure_amount": on_balance_exposure_amount,
            "off_balance_exposure_amount": off_balance_exposure_amount,
            "exposure_maturity_months": exposure_maturity_months,
            "residual_maturity_months": residual_maturity_months,
            "country_risk_weight": country_risk_weight,
            "ead_bilan_amount": ead_bilan_amount,
            "ead_hb_amount": ead_hb_amount,
            "ead_hb_ccf_amount": ead_hb_ccf_amount,
            "ead_total_amount": ead_total_amount,
            "rwa_eb_amount": rwa_eb_amount,
            "rwa_hb_amount": rwa_hb_amount,
        }
        return {
            "id": exposure_id,
            "analysis_date": _as_date(row.get("Date d'analyse")),
            "grant_date": grant_date,
            "maturity_date": maturity_date,
            "counterparty_name": _as_clean_text(row.get("Contrepartie")) or exposure_id,
            "country": country,
            "country_rating": country_rating,
            "category_raw": raw_category,
            "category_dashboard": _dashboard_category_from_raw(raw_category),
            "category_standard": _standard_category_from_raw(raw_category),
            "rating": rating,
            "gross_amount": gross_amount,
            **source_fields,
            "source_fields": source_fields,
            "currency": _as_clean_text(metadata.get("currency")) or _as_clean_text(row.get("Devise")) or "XOF",
            "status": _as_clean_text(metadata.get("status")) or "Active",
            "crm_exists": (_normalize_for_key(_as_clean_text(row.get("CRM_existe")) or "") == "oui"),
            "crm_type": crm_type,
            "crm_coverage_percent": crm_coverage_percent,
            "crm_details": crm_details,
            "original_rw": original_rw,
            "final_rw": final_rw,
            "ead": ead,
            "rwa": rwa,
            "capital": capital,
            "comment": _as_clean_text(metadata.get("comment")) or _as_clean_text(row.get("Commentaire")),
        }

    def exposure_template_fields_by_id(self) -> dict[str, dict[str, Any]]:
        cache = self._load_cache()
        indexed: dict[str, dict[str, Any]] = {}
        for row in cache.template_rows:
            exposure_id = _as_clean_text(row.get("ID_Exposition")) or ""
            if not exposure_id:
                continue
            indexed[exposure_id] = {
                "grant_date": _as_optional_date(row.get("Date d'octroi")),
                "maturity_date": _as_optional_date(row.get("Date d'échéance")),
                "country_rating": _as_clean_text(row.get("Notation_externe_pays")) or "Non noté",
            }
        return indexed

    def list_exposures(self) -> list[dict[str, Any]]:
        cache = self._load_cache()
        with self._lock:
            metadata = self._load_metadata_unlocked()["exposures"]
        crm_non_fin = self._crm_non_fin_by_rows(cache.crm_non_fin_rows)
        crm_fin = self._crm_fin_by_rows(cache.crm_fin_rows)
        base = [
            self._build_exposure_from_template(
                row,
                exposure_metadata=metadata.get(_as_clean_text(row.get("ID_Exposition")) or "", {}),
                crm_non_fin_row=crm_non_fin.get(_as_clean_text(row.get("ID_Exposition")) or ""),
                crm_fin_row=crm_fin.get(_as_clean_text(row.get("ID_Exposition")) or ""),
            )
            for row in cache.template_rows
            if not _normalize_for_key(_as_clean_text(row.get("Catégorie d'exposition")) or "").startswith("(l) hors bilan")
        ]
        return base + [item.copy() for item in self._overlay.exposures]

    def list_off_balance(self) -> list[dict[str, Any]]:
        cache = self._load_cache()
        with self._lock:
            metadata = self._load_metadata_unlocked()["off_balance"]
        rows: list[dict[str, Any]] = []
        for index, hb_row in enumerate(cache.hb_rows, start=1):
            exposure_id = _as_clean_text(hb_row.get("ID_Exposition")) or ""
            template = self._find_template_by_id(exposure_id)
            if template is None:
                continue
            exposure = self._build_exposure_from_template(
                template,
                exposure_metadata=self._metadata_for_exposure(exposure_id),
            )
            ccf = _as_float(hb_row.get("Facteur_conversion (CCF)"))
            hb_amount = (
                _as_optional_float(template.get("Montant d'exposition au HB"))
                or exposure.get("off_balance_exposure_amount")
                or 0.0
            )
            nominal = hb_amount if hb_amount > 0 else exposure["gross_amount"]
            ead = _as_optional_float(hb_row.get("EAD_HB_ccf"))
            if ead is None:
                ead = round(nominal * ccf, 2)
            rwa = (
                _as_optional_float(template.get("RWA_HB"))
                or exposure.get("rwa_hb_amount")
                or 0.0
            )
            risk_weight = (
                round(rwa / ead, 6)
                if ead > 0 and rwa > 0
                else (
                    exposure["original_rw"]
                    if exposure["original_rw"] > 0
                    else exposure["final_rw"]
                )
            )
            if rwa <= 0:
                rwa = round(ead * risk_weight, 2)
            capital = round(rwa * 0.08, 2)
            commitment_id = f"HB{index:03d}"
            rows.append(
                {
                    "id": commitment_id,
                    "analysis_date": exposure["analysis_date"],
                    "counterparty_id": exposure_id,
                    "counterparty_name": exposure["counterparty_name"],
                    "category": exposure["category_standard"],
                    "rating": exposure["rating"],
                    "engagement_type": _as_clean_text(hb_row.get("Catégorie Hors bilan")) or "Risque élevé",
                    "nominal_amount": nominal,
                    "ccf": ccf,
                    "ead": ead,
                    "risk_weight": risk_weight,
                    "rwa": rwa,
                    "capital": capital,
                    "comment": _as_clean_text(metadata.get(commitment_id, {}).get("comment")),
                }
            )
        return rows + [item.copy() for item in self._overlay.off_balance]

    def _crm_non_fin_by_id(self) -> dict[str, dict[str, Any]]:
        return self._crm_non_fin_by_rows(self._load_cache().crm_non_fin_rows)

    def _crm_fin_by_id(self) -> dict[str, dict[str, Any]]:
        return self._crm_fin_by_rows(self._load_cache().crm_fin_rows)

    def list_crm_items(self) -> list[dict[str, Any]]:
        non_fin = self._crm_non_fin_by_id()
        fin = self._crm_fin_by_id()
        items: list[dict[str, Any]] = []
        for exposure in self.list_exposures():
            exposure_id = exposure["id"]
            crm_type = _normalize_for_key(exposure["crm_type"])
            gross = exposure["gross_amount"]
            ead = exposure["ead"]
            borrower_rw = exposure["original_rw"] if exposure["original_rw"] > 0 else exposure["final_rw"]
            guarantor_name = "Sans garant"
            guarantor_category = exposure["category_standard"]
            guarantor_rating = exposure["rating"]
            coverage_amount = 0.0
            coverage_ratio = 0.0
            guarantee_type = "Sans CRM"
            guarantor_rw = borrower_rw

            if "non" in crm_type and "finance" in crm_type and exposure_id in non_fin:
                source = non_fin[exposure_id]
                guarantor_name = _as_clean_text(source.get("Nom du garant")) or "Garant non renseigné"
                guarantor_category = _standard_category_from_raw(_as_clean_text(source.get("Catégorie du garant")) or "")
                guarantor_rating = _as_clean_text(source.get("Note_garant")) or "Non noté"
                coverage_ratio = _as_float(source.get("% Exp_couverte"))
                coverage_amount = _as_float(source.get("Part couverte"))
                guarantee_type = "CRM non financée"
                guarantor_rw = _as_float(source.get("Pondération du garant"))
            elif "finance" in crm_type and exposure_id in fin:
                source = fin[exposure_id]
                guarantor_name = _as_clean_text(source.get("Type_emetteur")) or "Garant non renseigné"
                guarantor_category = "Souverains" if _normalize_for_key(guarantor_name) == "souverain" else "Entreprises"
                guarantor_rating = _as_clean_text(source.get("Notation")) or "Non noté"
                coverage_amount = max(_as_float(source.get("Eva")) - _as_float(source.get("Cva")), 0.0)
                coverage_ratio = 0.0 if gross == 0 else min(coverage_amount / gross, 1.0)
                guarantee_type = "CRM financée"
                guarantor_rw = borrower_rw

            final_rw = borrower_rw
            if coverage_ratio > 0:
                final_rw = round((coverage_ratio * guarantor_rw) + ((1 - coverage_ratio) * guarantor_rw), 4)
            rwa_before = round(ead * borrower_rw, 2)
            rwa_after = round(ead * final_rw, 2)
            capital_before = round(rwa_before * 0.08, 2)
            capital_after = round(rwa_after * 0.08, 2)

            items.append(
                {
                    "id": f"CRM{len(items) + 1:03d}",
                    "exposure_id": exposure_id,
                    "borrower_name": exposure["counterparty_name"],
                    "borrower_category": exposure["category_standard"],
                    "borrower_rating": exposure["rating"],
                    "gross_amount": gross,
                    "guarantor_name": guarantor_name,
                    "guarantor_category": guarantor_category,
                    "guarantor_rating": guarantor_rating,
                    "guarantee_type": guarantee_type,
                    "coverage_amount": coverage_amount,
                    "coverage_ratio": coverage_ratio,
                    "borrower_rw": borrower_rw,
                    "guarantor_rw": guarantor_rw,
                    "final_rw": final_rw,
                    "ead": ead,
                    "rwa_before": rwa_before,
                    "rwa_after": rwa_after,
                    "capital_before": capital_before,
                    "capital_after": capital_after,
                    "rwa_reduction": round(rwa_before - rwa_after, 2),
                    "capital_saving": round(capital_before - capital_after, 2),
                }
            )
        return items + [item.copy() for item in self._overlay.crm]

    def list_ref_rows(self) -> list[list[Any]]:
        return [row[:] for row in self._load_cache().ref_rows]

    def list_reports(self) -> list[dict[str, Any]]:
        return [item.copy() for item in self._overlay.reports]

    def _save_workbook(self, workbook) -> None:
        workbook.save(self._path)
        workbook.close()
        self._invalidate_cache_unlocked()

    def create_or_update_exposure(self, item: dict[str, Any]) -> dict[str, Any]:
        with self._lock:
            workbook = load_workbook(self._path)
            try:
                self._validate_workbook_structure(workbook)
                template_sheet = workbook["Template données"]
                crm_non_fin_sheet = workbook["CRM_non_financee"]
                crm_fin_sheet = workbook["CRM_financée"]
                template_headers = self._sheet_header_index(template_sheet)
                crm_non_fin_headers = self._sheet_header_index(crm_non_fin_sheet)
                crm_fin_headers = self._sheet_header_index(crm_fin_sheet)
                metadata = self._load_metadata_unlocked()
                self._upsert_exposure_unlocked(
                    item,
                    metadata=metadata,
                    template_sheet=template_sheet,
                    crm_non_fin_sheet=crm_non_fin_sheet,
                    crm_fin_sheet=crm_fin_sheet,
                    template_headers=template_headers,
                    crm_non_fin_headers=crm_non_fin_headers,
                    crm_fin_headers=crm_fin_headers,
                )
                self._save_metadata_unlocked(metadata)
                self._save_workbook(workbook)
            except Exception:
                workbook.close()
                raise
        return item.copy()

    def create_or_update_exposures(self, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        if not items:
            return []
        with self._lock:
            workbook = load_workbook(self._path)
            try:
                self._validate_workbook_structure(workbook)
                template_sheet = workbook["Template données"]
                crm_non_fin_sheet = workbook["CRM_non_financee"]
                crm_fin_sheet = workbook["CRM_financée"]
                template_headers = self._sheet_header_index(template_sheet)
                crm_non_fin_headers = self._sheet_header_index(crm_non_fin_sheet)
                crm_fin_headers = self._sheet_header_index(crm_fin_sheet)
                template_row_index = self._row_index_by_id(template_sheet, template_headers)
                metadata = self._load_metadata_unlocked()
                for item in items:
                    self._upsert_exposure_unlocked(
                        item,
                        metadata=metadata,
                        template_sheet=template_sheet,
                        crm_non_fin_sheet=crm_non_fin_sheet,
                        crm_fin_sheet=crm_fin_sheet,
                        template_headers=template_headers,
                        crm_non_fin_headers=crm_non_fin_headers,
                        crm_fin_headers=crm_fin_headers,
                        template_row_index=template_row_index,
                    )
                self._save_metadata_unlocked(metadata)
                self._save_workbook(workbook)
            except Exception:
                workbook.close()
                raise
        return [item.copy() for item in items]

    def delete_exposure(self, exposure_id: str) -> bool:
        target_ids = {str(exposure_id).strip()}
        if not target_ids or "" in target_ids:
            return False
        with self._lock:
            workbook = load_workbook(self._path)
            try:
                self._validate_workbook_structure(workbook)
                deleted_ids: set[str] = set()
                for sheet_name in ("Template données", "CRM_non_financee", "CRM_financée", "Traitement_HB"):
                    sheet = workbook[sheet_name]
                    headers = self._sheet_header_index(sheet)
                    deleted_ids.update(self._delete_rows_by_ids(sheet, headers, target_ids))
                metadata = self._load_metadata_unlocked()
                exposure_id = next(iter(target_ids))
                if exposure_id in metadata["exposures"]:
                    deleted_ids.add(exposure_id)
                    metadata["exposures"].pop(exposure_id, None)
                self._save_metadata_unlocked(metadata)
                self._save_workbook(workbook)
            except Exception:
                workbook.close()
                raise
        return bool(deleted_ids)

    def delete_exposures(self, exposure_ids: list[str]) -> dict[str, int | list[str]]:
        requested_ids: list[str] = []
        seen_ids: set[str] = set()
        for raw_id in exposure_ids:
            exposure_id = str(raw_id or "").strip()
            if not exposure_id or exposure_id in seen_ids:
                continue
            seen_ids.add(exposure_id)
            requested_ids.append(exposure_id)

        if not requested_ids:
            return {
                "requested_ids": [],
                "deleted_ids": [],
                "missing_ids": [],
                "deleted_count": 0,
                "missing_count": 0,
            }

        target_ids = set(requested_ids)
        deleted_ids: set[str] = set()
        with self._lock:
            workbook = load_workbook(self._path)
            try:
                self._validate_workbook_structure(workbook)
                for sheet_name in ("Template données", "CRM_non_financee", "CRM_financée", "Traitement_HB"):
                    sheet = workbook[sheet_name]
                    headers = self._sheet_header_index(sheet)
                    deleted_ids.update(self._delete_rows_by_ids(sheet, headers, target_ids))
                metadata = self._load_metadata_unlocked()
                for exposure_id in list(target_ids):
                    if exposure_id in metadata["exposures"]:
                        deleted_ids.add(exposure_id)
                        metadata["exposures"].pop(exposure_id, None)
                self._save_metadata_unlocked(metadata)
                self._save_workbook(workbook)
            except Exception:
                workbook.close()
                raise

        deleted_list = [exposure_id for exposure_id in requested_ids if exposure_id in deleted_ids]
        missing_list = [exposure_id for exposure_id in requested_ids if exposure_id not in deleted_ids]
        return {
            "requested_ids": requested_ids,
            "deleted_ids": deleted_list,
            "missing_ids": missing_list,
            "deleted_count": len(deleted_list),
            "missing_count": len(missing_list),
        }

    def create_off_balance(self, item: dict[str, Any]) -> dict[str, Any]:
        with self._lock:
            workbook = load_workbook(self._path)
            try:
                self._validate_workbook_structure(workbook)
                sheet = workbook["Traitement_HB"]
                headers = self._sheet_header_index(sheet)
                metadata = self._load_metadata_unlocked()
                self._append_off_balance_unlocked(
                    item,
                    metadata=metadata,
                    sheet=sheet,
                    headers=headers,
                )
                self._save_metadata_unlocked(metadata)
                self._save_workbook(workbook)
            except Exception:
                workbook.close()
                raise
        return item.copy()

    def create_off_balances(self, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        if not items:
            return []
        with self._lock:
            workbook = load_workbook(self._path)
            try:
                self._validate_workbook_structure(workbook)
                sheet = workbook["Traitement_HB"]
                headers = self._sheet_header_index(sheet)
                metadata = self._load_metadata_unlocked()
                for item in items:
                    self._append_off_balance_unlocked(
                        item,
                        metadata=metadata,
                        sheet=sheet,
                        headers=headers,
                    )
                self._save_metadata_unlocked(metadata)
                self._save_workbook(workbook)
            except Exception:
                workbook.close()
                raise
        return [item.copy() for item in items]

    def export_exposures_workbook(self, rows: list[dict[str, Any]]) -> Path:
        self._exports_path.mkdir(parents=True, exist_ok=True)
        export_path = self._exports_path / f"expositions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Expositions"
        headers = [
            "ID",
            "Date Analyse",
            "Contrepartie",
            "Pays",
            "Categorie",
            "Notation",
            "Montant Brut",
            "Devise",
            "Statut",
            "Type CRM",
            "Couverture CRM",
            "RW Initial",
            "RW Final",
            "EAD",
            "RWA",
            "Capital",
            "Commentaire",
        ]
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=column_index, value=header)
        for row_index, row in enumerate(rows, start=2):
            values = [
                row["id"],
                row["analysis_date"].isoformat() if isinstance(row["analysis_date"], date) else row["analysis_date"],
                row["counterparty_name"],
                row["country"],
                row["category_raw"],
                row["rating"],
                row["gross_amount"],
                row.get("currency", "XOF"),
                row.get("status", "Active"),
                row.get("crm_type", "Aucune"),
                row.get("crm_coverage_percent", 0.0),
                row.get("original_rw", 0.0),
                row.get("final_rw", 0.0),
                row.get("ead", 0.0),
                row.get("rwa", 0.0),
                row.get("capital", 0.0),
                row.get("comment") or "",
            ]
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)
        workbook.save(export_path)
        workbook.close()
        return export_path

    def add_exposure(self, item: dict[str, Any]) -> None:
        self.create_or_update_exposure(item)

    def add_off_balance(self, item: dict[str, Any]) -> None:
        self.create_off_balance(item)

    def add_crm(self, item: dict[str, Any]) -> None:
        self._overlay.crm.append(item.copy())

    def add_report(self, item: dict[str, Any]) -> None:
        self._overlay.reports.append(item.copy())


excel_repository = ExcelRepository()
