"""Export / import Excel des déclarations FODEP basé strictement sur la matrice
officielle BCEAO (`Matrice_FODEP_Officielle.xlsx`).
"""

from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any

from app.fodep.dispru import FONDS_PROPRES_CODES

_CHEMINS_MATRICE_OFFICIELLE = (
    Path(__file__).parent / "templates" / "Matrice_FODEP_Officielle.xlsx",
    Path(r"C:\RisqueManagement\backend\app\fodep\templates\Matrice_FODEP_Officielle.xlsx"),
    Path(r"C:\RisqueManagement\Matrice_FODEP_Officielle.xlsx"),
)


def get_matrice_officielle_path() -> Path | None:
    """Retourne le chemin absolu vers le classeur officiel FODEP BCEAO."""
    for candidat in _CHEMINS_MATRICE_OFFICIELLE:
        if candidat.exists():
            return candidat
    return None


def get_matrice_officielle_template_bytes() -> bytes:
    """Retourne les octets bruts du modèle officiel FODEP BCEAO, sans modification."""
    chemin = get_matrice_officielle_path()
    if chemin is None:
        raise FileNotFoundError(
            "La matrice FODEP officielle (Matrice_FODEP_Officielle.xlsx) est introuvable."
        )
    return chemin.read_bytes()


def _ecrire(ws: Any, ligne: int, colonne: int, valeur: Any) -> None:
    """Ecrit une valeur en tenant compte des cellules fusionnees.

    openpyxl refuse toute ecriture sur une cellule recouverte par une fusion :
    seule la cellule haut-gauche porte la valeur. Viser directement une case
    fusionnee (ainsi la denomination de l'etablissement, dans la fusion C5:K5
    de l'onglet ADPE) faisait echouer l'export entier.
    """

    for fusion in ws.merged_cells.ranges:
        if fusion.min_row <= ligne <= fusion.max_row and fusion.min_col <= colonne <= fusion.max_col:
            ligne, colonne = fusion.min_row, fusion.min_col
            break
    ws.cell(ligne, colonne, valeur)


# Categorie prudentielle (code rwa_credit) -> ligne EP09 (exposition brute) et
# onglet/ligne de synthese ou ecrire le total de la categorie (exposition
# apres ARC, RWA). Le detail par tranche de ponderation reglementaire n'est PAS
# rempli : rwa_credit calcule une ponderation continue par exposition (effet
# des suretes partielles, plafonnements pays...) alors que ces onglets
# attendent des tranches discretes fixes (0%, 20%, 35%, 50%, 75%, 100%, 150%).
# Forcer une ponderation continue dans la tranche la plus proche romprait la
# formule propre de la ligne (RWA = montant x ponderation figee de la ligne) et
# fausserait silencieusement le detail. Le total par categorie, lui, est fiable
# et vaut mieux qu'une case vide.
_VENTILATION_CREDIT: dict[str, tuple[int, str, int]] = {
    "a": (12, "EP12", 43),  # Souverains
    "b": (13, "EP13", 43),  # Organismes publics hors administration centrale
    "c": (14, "EP14", 43),  # Banques multilaterales de developpement
    "d": (15, "EP15", 43),  # Institutions financieres
    "e": (18, "EP16", 47),  # Entreprises (ligne "Autres entreprises", pas de suivi PME distinct cote app)
    "f": (21, "EP17", 38),  # Clientele de detail (ligne "Autres clientele de detail")
    "g": (22, "EP18", 35),  # Prets garantis par l'immobilier residentiel
    "h": (25, "EP19", 38),  # Prets garantis par l'immobilier commercial (ligne "Autres")
    "k": (26, "EP20", 27),  # Autres actifs
}

# Lignes de detail (ponderation par tranche) a vider dans EP12-EP19 : ce sont
# les memes categories que ci-dessus, colonnes C (avant ARC) a H (apres ARC).
# On ne les remplit pas (cf. note ci-dessus) mais on retire les montants d'un
# autre etablissement laisses par le modele, pour ne pas les laisser
# contredire silencieusement le total desormais correct de la categorie.
_LIGNES_DETAIL_CREDIT: dict[str, list[int]] = {
    "EP12": [12, 13, 14, 15, 16, 20, 21, 22, 23, 24, 28, 29, 30, 31, 32, 36, 37, 38, 39, 40],
    "EP13": [12, 13, 14, 15, 16, 20, 21, 22, 23, 24, 28, 29, 30, 31, 32, 36, 37, 38, 39, 40],
    "EP14": [12, 13, 14, 15, 16, 20, 21, 22, 23, 24, 28, 29, 30, 31, 32, 36, 37, 38, 39, 40],
    "EP15": [12, 13, 14, 15, 16, 20, 21, 22, 23, 24, 28, 29, 30, 31, 32, 36, 37, 38, 39, 40],
    "EP16": [12, 13, 14, 15, 16, 17, 21, 22, 23, 24, 25, 26, 30, 31, 32, 33, 34, 35, 39, 40, 41, 42, 43, 44],
    "EP17": [12, 13, 14, 15, 16, 17, 21, 22, 23, 24, 25, 26, 30, 31, 32, 33, 34, 35],
    "EP18": [12, 13, 14, 15, 16, 20, 21, 22, 23, 24, 28, 29, 30, 31, 32],
    "EP19": [12, 13, 14, 15, 16, 17, 21, 22, 23, 24, 25, 26, 30, 31, 32, 33, 34, 35],
}
_COLONNES_DETAIL_CREDIT = range(3, 9)  # C a H

# EP20 (autres actifs) : structure differente (poste par poste, pas de tranche
# de ponderation partagee) -> lignes de detail a vider, colonne C uniquement.
_LIGNES_DETAIL_EP20 = list(range(10, 27))

# EP09 : ligne -> categorie, pour les 4 categories qui se subdivisent en
# PME/Autres dans le formulaire officiel sans que rwa_credit ne fasse cette
# distinction (l'app ne suit pas separement les PME) : le total va sur la
# ligne "Autres", la ligne PME reste a 0 (deja le cas dans le modele).


def _ecrire_ventilation_credit(wb: Any, analyse: Any) -> None:
    """Renseigne EP09 (expositions brutes par categorie) et les totaux par
    categorie d'EP12 a EP20, a partir de l'analyse RWA credit reelle (module
    rwa_credit, la meme que celle affichee a l'ecran dans l'application).
    """

    agents = {agent.code: agent for agent in getattr(analyse, "agents", [])}
    ws09 = wb["EP09"] if "EP09" in wb.sheetnames else None

    for code, (ligne_ep09, onglet_total, ligne_total) in _VENTILATION_CREDIT.items():
        agent = agents.get(code)
        if agent is None:
            continue

        if ws09 is not None:
            _ecrire(ws09, ligne_ep09, 4, agent.gross_exposure)

        if onglet_total not in wb.sheetnames:
            continue
        ws = wb[onglet_total]

        if onglet_total == "EP20":
            _ecrire(ws, ligne_total, 3, agent.ead)
            _ecrire(ws, ligne_total, 5, agent.rwa)
            for r in _LIGNES_DETAIL_EP20:
                _ecrire(ws, r, 3, 0)
            continue

        _ecrire(ws, ligne_total, 9, agent.ead)
        _ecrire(ws, ligne_total, 10, agent.rwa)
        for r in _LIGNES_DETAIL_CREDIT.get(onglet_total, ()):
            for c in _COLONNES_DETAIL_CREDIT:
                _ecrire(ws, r, c, 0)


# Onglets ou le modele officiel contient des donnees de portefeuille d'un
# autre etablissement (grands risques, participations) qu'aucun module de
# l'application ne calcule aujourd'hui. Plutot que de les laisser dans l'export
# - un vrai risque de confidentialite et de conformite, la Commission Bancaire
# recevrait le portefeuille d'un tiers - on les vide : colonnes de donnees
# remises a blanc, code DISPRU de la ligne (colonne A) conserve car il fait
# partie de la structure officielle du formulaire.
_ONGLETS_A_VIDER: dict[str, tuple[int, int, int]] = {
    # onglet: (premiere_ligne, derniere_ligne, derniere_colonne)
    # EP09 : ventilation du bilan par categorie de risque de credit. Le modele
    # officiel la livre pre-remplie avec le portefeuille d'un autre
    # etablissement (montants negatifs, en milliers) : on la vide avant d'y
    # reinjecter, le cas echeant, l'analyse RWA credit reelle de l'application.
    "EP09": (12, 28, 9),
    "EP10": (12, 45, 9),
    "EP29": (16, 91, 30),
    "EP30": (16, 116, 30),
    "EP31": (11, 31, 30),
    "EP32": (18, 68, 15),
    "EP34": (13, 127, 6),
}


def _vider_donnees_etrangeres(wb: Any) -> None:
    """Efface les montants et noms de contreparties laisses par le modele
    officiel dans les onglets que l'application ne renseigne pas encore.
    """

    for nom, (premiere, derniere, derniere_colonne) in _ONGLETS_A_VIDER.items():
        if nom not in wb.sheetnames:
            continue
        ws = wb[nom]
        for r in range(premiere, derniere + 1):
            for c in range(2, derniere_colonne + 1):
                cellule = ws.cell(r, c)
                if isinstance(cellule.value, str) and cellule.value.startswith("="):
                    # Une formule de simple agregation (=D17+D18) reste correcte
                    # une fois ses entrees a 0. Mais le modele officiel code
                    # aussi des montants du specimen EN DUR dans des formules
                    # (=-203717+173) : on efface celles qui portent un nombre de
                    # 3 chiffres ou plus.
                    if re.search(r"\d{3,}", cellule.value):
                        cellule.value = None
                    continue
                if cellule.value not in (None, ""):
                    # ws.cell(r, c, None) n'efface rien : pour openpyxl, une
                    # valeur None passee a cell() signifie « ne rien changer »,
                    # pas « vider ». Seule l'affectation directe .value = None
                    # efface reellement la cellule.
                    cible = ws.cell(r, c)
                    for fusion in ws.merged_cells.ranges:
                        if fusion.min_row <= r <= fusion.max_row and fusion.min_col <= c <= fusion.max_col:
                            cible = ws.cell(fusion.min_row, fusion.min_col)
                            break
                    cible.value = None


def _effacer(ws: Any, ligne: int, colonne: int) -> None:
    """Vide reellement une cellule (openpyxl ignore un None passe a cell())."""

    for fusion in ws.merged_cells.ranges:
        if fusion.min_row <= ligne <= fusion.max_row and fusion.min_col <= colonne <= fusion.max_col:
            ligne, colonne = fusion.min_row, fusion.min_col
            break
    ws.cell(ligne, colonne).value = None


def build_fonds_propres_export(
    periode: str | None,
    postes: dict[str, float],
    *,
    etablissement: Any = None,
    participations: list[Any] | None = None,
    apr: Any = None,
    totaux: dict[str, float] | None = None,
    analyse_credit: Any = None,
    attestation: Any = None,
) -> bytes:
    """Génère un export basé strictement sur la matrice officielle BCEAO,
    en renseignant les postes dans les états réglementaires (EP03, EP21, EP33, EP35, etc.)
    tout en préservant intacte la structure et les formules officielles.
    """
    import openpyxl

    chemin = get_matrice_officielle_path()
    if chemin is None:
        raise FileNotFoundError("Matrice FODEP officielle introuvable.")

    wb = openpyxl.load_workbook(chemin)

    # 0. Renseigner l'en-tête et l'établissement dans ADPE
    if "ADPE" in wb.sheetnames:
        ws = wb["ADPE"]
        if etablissement:
            nom = getattr(etablissement, "denomination", None) or (
                etablissement.get("denomination") if isinstance(etablissement, dict) else ""
            )
            code_b = getattr(etablissement, "code_bceao", None) or (
                etablissement.get("code_bceao") if isinstance(etablissement, dict) else ""
            )
            if nom:
                _ecrire(ws, 5, 4, str(nom).strip())
            if code_b:
                _ecrire(ws, 6, 4, str(code_b).strip())
        if periode:
            try:
                parts = periode.split("-")
                if len(parts) == 3:
                    _ecrire(ws, 9, 4, f"{parts[2]}/{parts[1]}/{parts[0]}")
            except Exception:
                _ecrire(ws, 9, 4, str(periode))

    totaux = totaux or {}

    def _codes_ligne(valeur_a: Any) -> set[str]:
        """Codes DISPRU d'une cellule de colonne A, en minuscules. Gere la
        forme « FPI41 / FPC41 » des lignes de rappel (EP02, EP36-EP38)."""

        if not valeur_a:
            return set()
        return {
            morceau.strip().lower()
            for morceau in str(valeur_a).replace("\\", "/").split("/")
            if morceau.strip()
        }

    def _poste_ou_total(code_min: str) -> float | None:
        """Valeur a ecrire pour un code DISPRU : saisie si c'est un poste,
        sinon total deja calcule par ``calculations.py``. ``None`` si le code
        n'est ni l'un ni l'autre (la cellule garde alors sa formule)."""

        if code_min in postes:
            return postes[code_min]
        if code_min in totaux:
            return totaux[code_min]
        return None

    # 1. Renseigner EP03 (Fonds Propres Base Individuelle)
    #
    # On ecrit AUSSI les totaux (FPI08, FPI14, FPI22, FPI29, FPI41...) par
    # dessus les formules du modele : celles du classeur officiel livre sont
    # partielles (FPI14 ne deduit que IM012, FPI26/FPI28/FPI39/FPI40 sont des
    # zeros en dur, FPI29 = FPI22, FPI41 = FPI29). ``calculations.py`` reste la
    # seule source de verite - sinon le PDF/Excel diverge de l'ecran des qu'il
    # y a de l'AT1, du T2 ou une deduction CET1 autre que IM012.
    if "EP03" in wb.sheetnames:
        ws = wb["EP03"]
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if not c1:
                continue
            code = str(c1).strip().lower()
            valeur = _poste_ou_total(code)
            if valeur is not None:
                _ecrire(ws, r, 3, valeur)

    # 2. Renseigner EP21 (Produit brut) - postes RO001..RO008 + total RO009.
    if "EP21" in wb.sheetnames:
        ws = wb["EP21"]
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if not c1:
                continue
            code = str(c1).strip().lower()
            valeur = _poste_ou_total(code)
            if valeur is not None:
                _ecrire(ws, r, 4, valeur)
        # Serie du produit brut sur 3 ans (C26=annee-3, D26=annee-2,
        # E26=annee-1) : le modele officiel la livre avec les chiffres du
        # specimen. L'application ne suit que l'arrete courant -> annee-1 =
        # RO009, annees anterieures effacees (a saisir a la main si besoin).
        if "ro009" in totaux:
            _ecrire(ws, 26, 3, 0)
            _ecrire(ws, 26, 4, 0)
            _ecrire(ws, 26, 5, totaux["ro009"])

    # 3. Renseigner EP33 (Ratio de levier) - postes RL001..RL012, totaux
    # RL004/RL007/RL010/RL013/RL015, et le rappel RL014 = fonds propres T1.
    if "EP33" in wb.sheetnames:
        ws = wb["EP33"]
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if not c1:
                continue
            code = str(c1).strip().lower()
            valeur = _poste_ou_total(code)
            if valeur is not None:
                _ecrire(ws, r, 3, valeur)
            elif code == "rl014" and "fpi29" in totaux:
                _ecrire(ws, r, 3, totaux["fpi29"])

    # 4. Renseigner EP35 (Participations dans des entités commerciales).
    # On efface d'abord toutes les lignes (le modele officiel les livre avec
    # les participations d'un autre etablissement : SCIE, GIM UEMOA...).
    if "EP35" in wb.sheetnames:
        ws = wb["EP35"]
        for r in range(12, 66):
            _effacer(ws, r, 2)             # denomination
            for c in (4, 5, 6, 7, 8, 9):   # capital, brut, net, % (formules du specimen)
                _ecrire(ws, r, c, 0)
        for idx, part in enumerate(participations or []):
            row_num = 12 + idx
            if row_num > 65:
                break
            denom = getattr(part, "denomination_emettrice", "") or (
                part.get("denomination_emettrice") if isinstance(part, dict) else ""
            )
            cap = getattr(part, "capital_emettrice", 0.0) or (
                part.get("capital_emettrice") if isinstance(part, dict) else 0.0
            )
            net = getattr(part, "montant_net", 0.0) or (
                part.get("montant_net") if isinstance(part, dict) else 0.0
            )
            if denom:
                _ecrire(ws, row_num, 2, denom)
            if cap > 0:
                _ecrire(ws, row_num, 4, cap)
            if net > 0:
                _ecrire(ws, row_num, 5, net)
                _ecrire(ws, row_num, 6, net)

    # 5. Renseigner EP36 (Immobilisations hors exploitation).
    if "EP36" in wb.sheetnames:
        ws = wb["EP36"]
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if c1 and str(c1).strip().lower() in postes:
                _ecrire(ws, r, 3, postes[str(c1).strip().lower()])
        # IM004 = IM001 + IM002 + IM003 (le modele le laisse en dur), et le
        # rappel de fonds propres de base T1 en bas d'onglet.
        im004 = sum(float(postes.get(k, 0.0) or 0.0) for k in ("im001", "im002", "im003"))
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if not c1:
                continue
            code = str(c1).strip().lower()
            if code == "im004":
                _ecrire(ws, r, 3, im004)
                _ecrire(ws, r, 4, im004)
            elif _codes_ligne(c1) & {"fpi29", "fpc29"} and "fpi29" in totaux:
                _ecrire(ws, r, 3, totaux["fpi29"])

    # 6. Renseigner EP37 (Immobilisations d'exploitation et participations).
    if "EP37" in wb.sheetnames:
        ws = wb["EP37"]
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if c1 and str(c1).strip().lower() in postes:
                _ecrire(ws, r, 3, postes[str(c1).strip().lower()])
        im004 = sum(float(postes.get(k, 0.0) or 0.0) for k in ("im001", "im002", "im003"))
        im008 = im004 + float(postes.get("im007", 0.0) or 0.0)
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if not c1:
                continue
            code = str(c1).strip().lower()
            if code == "im004":
                _ecrire(ws, r, 3, im004)
                _ecrire(ws, r, 4, im004)
            elif code == "im008":
                _ecrire(ws, r, 3, im008)
                _ecrire(ws, r, 4, im008)
            elif _codes_ligne(c1) & {"fpi41", "fpc41"} and "fpi41" in totaux:
                _ecrire(ws, r, 3, totaux["fpi41"])

    # 7. Renseigner EP38 (Prêts dirigeants/actionnaires).
    if "EP38" in wb.sheetnames:
        ws = wb["EP38"]
        categories = ("a", "b", "c", "d", "e", "f", "g", "h")
        pr001 = [float(postes.get(f"pr001{cat}", 0.0) or 0.0) for cat in categories]
        pr002 = [float(postes.get(f"pr002{cat}", 0.0) or 0.0) for cat in categories]
        for idx in range(8):
            col = 3 + idx
            _ecrire(ws, 11, col, pr001[idx])          # PR001 - concours
            _ecrire(ws, 12, col, pr002[idx])          # PR002 - engagements signature
            _ecrire(ws, 13, col, pr001[idx] + pr002[idx])  # PR003 - total
        _ecrire(ws, 11, 11, sum(pr001))               # K11 - total concours
        _ecrire(ws, 12, 11, sum(pr002))               # K12 - total engagements
        _ecrire(ws, 13, 11, sum(pr001) + sum(pr002))  # K13 - total general
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if c1 and _codes_ligne(c1) & {"fpi41", "fpc41"} and "fpi41" in totaux:
                _ecrire(ws, r, 3, totaux["fpi41"])

    # 8. Renseigner EP08 (Actifs pondérés des risques)
    if "EP08" in wb.sheetnames and apr:
        ws = wb["EP08"]
        rwa_credit = getattr(apr, "rwa_credit", 0.0) or (apr.get("rwa_credit") if isinstance(apr, dict) else 0.0)
        rwa_marche = getattr(apr, "rwa_marche", 0.0) or (apr.get("rwa_marche") if isinstance(apr, dict) else 0.0)
        rwa_op = getattr(apr, "rwa_operationnel", 0.0) or (apr.get("rwa_operationnel") if isinstance(apr, dict) else 0.0)
        _ecrire(ws, 19, 5, rwa_credit)
        _ecrire(ws, 26, 5, rwa_marche)
        _ecrire(ws, 32, 5, rwa_op)

    # 9. Renseigner EP02 (Solvabilité)
    if "EP02" in wb.sheetnames and totaux:
        ws = wb["EP02"]
        if "fpi22" in totaux:
            _ecrire(ws, 14, 6, totaux["fpi22"])
        if "fpi29" in totaux:
            _ecrire(ws, 15, 6, totaux["fpi29"])
        if "fpi41" in totaux:
            _ecrire(ws, 16, 6, totaux["fpi41"])

    # 10. Retirer les donnees de portefeuille d'un autre etablissement que le
    # modele officiel contient (grands risques, ventilation credit,
    # participations). A faire AVANT la ventilation credit reelle ci-dessous,
    # sans quoi le nettoyage d'EP09 effacerait les montants qu'on vient d'y
    # ecrire.
    _vider_donnees_etrangeres(wb)

    # 11. Renseigner EP09 et les totaux EP12-EP20 (ventilation credit reelle)
    if analyse_credit is not None:
        _ecrire_ventilation_credit(wb, analyse_credit)

    # 12. Onglet d'attestation de declaration prudentielle (signe par la
    # direction et le commissaire aux comptes) : present a la fois dans le
    # fichier Excel et, par conversion fidèle, dans le PDF.
    _ajouter_feuille_attestation(wb, attestation, etablissement, periode)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _ajouter_feuille_attestation(wb: Any, attestation: Any, etablissement: Any, periode: str | None) -> None:
    """Ajoute (ou remplace) l'onglet « ATTESTATION » selon le modèle BCEAO officiel."""

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # Récupérer les valeurs (objet ou dictionnaire) de manière tolérante.
    def _champ(obj: Any, *cles: str, defaut: str = "") -> str:
        if obj is None:
            return defaut
        for cle in cles:
            if isinstance(obj, dict):
                val = obj.get(cle)
            else:
                val = getattr(obj, cle, None)
            if val:
                return str(val)
        return defaut

    # Données établissement
    nom = _champ(etablissement, "denomination") or "L'ÉTABLISSEMENT ASSUJETTI"
    code_b = _champ(etablissement, "code_bceao")

    # Données attestation BCEAO
    rens_nom = _champ(attestation, "rens_prenoms_nom")
    rens_fonction = _champ(attestation, "rens_fonction")
    rens_telephone = _champ(attestation, "rens_telephone")
    rens_poste = _champ(attestation, "rens_poste")
    rens_email = _champ(attestation, "rens_email")

    trans_nom = _champ(attestation, "trans_prenoms_nom")
    trans_fonction = _champ(attestation, "trans_fonction")
    trans_telephone = _champ(attestation, "trans_telephone")
    trans_poste = _champ(attestation, "trans_poste")
    trans_email = _champ(attestation, "trans_email")

    certif_1 = _champ(attestation, "certif_nous_1")
    certif_2 = _champ(attestation, "certif_nous_2")

    sign1_code = _champ(attestation, "sign1_code")
    sign1_fonction = _champ(attestation, "sign1_fonction")
    sign1_date = _champ(attestation, "sign1_date")

    sign2_code = _champ(attestation, "sign2_code")
    sign2_fonction = _champ(attestation, "sign2_fonction")
    sign2_date = _champ(attestation, "sign2_date")

    if periode:
        try:
            parts = periode.split("-")
            if len(parts) == 3:
                date_arrete = f"{parts[2]}/{parts[1]}/{parts[0]}"
            else:
                date_arrete = periode
        except Exception:
            date_arrete = str(periode)
    else:
        date_arrete = "—"

    # ── Nettoyage préalable ──────────────────────────────────────────────────
    if "ATTESTATION" in wb.sheetnames:
        del wb["ATTESTATION"]
    ws = wb.create_sheet("ATTESTATION")

    # ── Largeurs de colonnes (8 colonnes) ────────────────────────────────────
    largeurs = [12.0, 18.0, 18.0, 18.0, 14.0, 10.0, 10.0, 14.0]
    for i, larg in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    # ── Palette et styles ────────────────────────────────────────────────────
    NAVY = "FF172554"
    GRIS_BORD = "FFCBD5E1"
    JAUNE = "FFFFF7DB"

    bord_gris = Side(style="thin", color=GRIS_BORD)
    fill_jaune = PatternFill("solid", fgColor=JAUNE)
    bordure_fine = Border(left=bord_gris, right=bord_gris, top=bord_gris, bottom=bord_gris)

    police_titre = Font(name="Arial", size=13, bold=True, color=NAVY)
    police_section = Font(name="Arial", size=10, bold=True, color=NAVY)
    police_label = Font(name="Arial", size=9, color="FF1F2937")
    police_valeur = Font(name="Arial", size=9, bold=True, color=NAVY)
    police_texte = Font(name="Arial", size=9, color="FF1F2937")
    police_note = Font(name="Arial", size=8, italic=True, color="FF64748B")

    align_centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_gauche = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _fusion(l1: int, c1: int, l2: int, c2: int) -> None:
        ws.merge_cells(start_row=l1, start_column=c1, end_row=l2, end_column=c2)

    def _case(ligne: int, colonne: int, valeur: Any = None, *, police=None,
              alignement=None, remplissage=None, bordure_obj=None) -> None:
        cellule = ws.cell(ligne, colonne, valeur)
        if police is not None:
            cellule.font = police
        if alignement is not None:
            cellule.alignment = alignement
        if remplissage is not None:
            cellule.fill = remplissage
        if bordure_obj is not None:
            cellule.border = bordure_obj

    def _ligne_champ(ligne: int, label: str, valeur: str) -> None:
        _fusion(ligne, 1, ligne, 2)
        _case(ligne, 1, label, police=police_label, alignement=align_gauche)
        _fusion(ligne, 3, ligne, 8)
        _case(ligne, 3, valeur, police=police_valeur,
              alignement=align_gauche, remplissage=fill_jaune, bordure_obj=bordure_fine)
        ws.row_dimensions[ligne].height = 16

    def _ligne_tel_poste(ligne: int, tel: str, poste: str) -> None:
        _fusion(ligne, 1, ligne, 2)
        _case(ligne, 1, "Téléphone :", police=police_label, alignement=align_gauche)
        _fusion(ligne, 3, ligne, 6)
        _case(ligne, 3, tel, police=police_valeur,
              alignement=align_gauche, remplissage=fill_jaune, bordure_obj=bordure_fine)
        _case(ligne, 7, "Poste :", police=police_label, alignement=align_gauche)
        _case(ligne, 8, poste, police=police_valeur,
              alignement=align_gauche, remplissage=fill_jaune, bordure_obj=bordure_fine)
        ws.row_dimensions[ligne].height = 16

    r = 1

    # ── Titre ────────────────────────────────────────────────────────────────
    _fusion(r, 1, r, 8)
    _case(r, 1, "ATTESTATION DE DECLARATION PRUDENTIELLE",
          police=police_titre, alignement=align_centre)
    ws.row_dimensions[r].height = 28
    r += 2

    # ── Identification ───────────────────────────────────────────────────────
    _fusion(r, 1, r, 4)
    _case(r, 1, f"ETAT :", police=police_label, alignement=align_gauche)
    _fusion(r, 5, r, 6)
    _case(r, 5, "ETABLISSEMENT :", police=police_label, alignement=align_gauche)
    _fusion(r, 7, r, 8)
    _case(r, 7, nom, police=police_valeur, alignement=align_gauche,
          remplissage=fill_jaune, bordure_obj=bordure_fine)
    ws.row_dimensions[r].height = 18
    r += 1
    _fusion(r, 1, r, 4)
    _case(r, 1, f"  Date d'arrêté : {date_arrete}", police=police_label, alignement=align_gauche)
    _fusion(r, 5, r, 8)
    _case(r, 5, f"Code : {code_b}" if code_b else "Code :", police=police_label, alignement=align_gauche)
    ws.row_dimensions[r].height = 18
    r += 2

    # ── Personne-responsable du renseignement ────────────────────────────────
    _fusion(r, 1, r, 8)
    _case(r, 1, "Personne-responsable du renseignement du FODEP :",
          police=police_section, alignement=align_gauche)
    ws.row_dimensions[r].height = 18
    r += 1
    _ligne_champ(r, "Prénoms et Nom :", rens_nom); r += 1
    _ligne_champ(r, "Fonction :", rens_fonction); r += 1
    _ligne_tel_poste(r, rens_telephone, rens_poste); r += 1
    _ligne_champ(r, "E-mail :", rens_email); r += 1
    r += 1

    # ── Personne-responsable de la transmission ──────────────────────────────
    _fusion(r, 1, r, 8)
    _case(r, 1, "Personne-responsable de la transmission du FODEP à la plateforme de reporting BCEAO",
          police=police_section, alignement=align_gauche)
    ws.row_dimensions[r].height = 18
    r += 1
    _ligne_champ(r, "Prénoms et Nom :", trans_nom); r += 1
    _ligne_champ(r, "Fonction :", trans_fonction); r += 1
    _ligne_tel_poste(r, trans_telephone, trans_poste); r += 1
    _ligne_champ(r, "E-mail :", trans_email); r += 1
    r += 1

    # ── CERTIFICATION ────────────────────────────────────────────────────────
    _fusion(r, 1, r, 8)
    _case(r, 1, "CERTIFICATION", police=police_section, alignement=align_centre)
    ws.row_dimensions[r].height = 18
    r += 1

    _case(r, 1, "Nous,", police=police_label, alignement=align_gauche)
    _fusion(r, 2, r, 5)
    _case(r, 2, certif_1, police=police_valeur, alignement=align_gauche,
          remplissage=fill_jaune, bordure_obj=bordure_fine)
    _case(r, 6, "et", police=police_label, alignement=align_centre)
    _fusion(r, 7, r, 8)
    _case(r, 7, certif_2, police=police_valeur, alignement=align_gauche,
          remplissage=fill_jaune, bordure_obj=bordure_fine)
    ws.row_dimensions[r].height = 16
    r += 1

    certif_texte = (
        "certifions que le présent formulaire a été rempli conformément aux exigences du dispositif "
        "prudentiel applicable aux établissements de crédit et aux compagnies financières de l'Union "
        "Monétaire Ouest Africaine.\n\n"
        "En outre, nous attestons qu'au meilleur de notre connaissance, les données contenues dans le "
        "présent formulaire sont fiables, intègres et exhaustives."
    )
    _fusion(r, 1, r + 3, 8)
    _case(r, 1, certif_texte, police=police_texte, alignement=align_gauche)
    for i in range(4):
        ws.row_dimensions[r + i].height = 16
    r += 5

    # ── Signature 1 ──────────────────────────────────────────────────────────
    row_s1 = r
    _fusion(r, 1, r, 2)
    _case(r, 1, "Code Signature :", police=police_label, alignement=align_gauche)
    _fusion(r, 3, r, 5)
    _case(r, 3, sign1_code, police=police_valeur, alignement=align_gauche,
          remplissage=fill_jaune, bordure_obj=bordure_fine)
    ws.row_dimensions[r].height = 16
    r += 1
    _case(r, 1, "Fonction :", police=police_label, alignement=align_gauche)
    _fusion(r, 2, r, 5)
    _case(r, 2, sign1_fonction, police=police_valeur, alignement=align_gauche,
          remplissage=fill_jaune, bordure_obj=bordure_fine)
    _case(r, 6, "Date :", police=police_label, alignement=align_gauche)
    _fusion(r, 7, r, 8)
    _case(r, 7, sign1_date, police=police_valeur, alignement=align_gauche,
          remplissage=fill_jaune, bordure_obj=bordure_fine)
    ws.row_dimensions[r].height = 16
    r += 1
    
    # Espace pour l'image 1
    _fusion(r, 1, r+1, 8)
    sign1_b64 = _champ(attestation, "sign1_image")
    if sign1_b64:
        try:
            import base64
            from io import BytesIO
            sig_img = openpyxl.drawing.image.Image(BytesIO(base64.b64decode(sign1_b64)))
            sig_img.width = 150
            sig_img.height = 50
            ws.add_image(sig_img, f"D{r}")
        except Exception:
            pass
    ws.row_dimensions[r].height = 25
    ws.row_dimensions[r+1].height = 25
    r += 2
    r += 1

    # ── Signature 2 ──────────────────────────────────────────────────────────
    row_s2 = r
    _fusion(r, 1, r, 2)
    _case(r, 1, "Code Signature :", police=police_label, alignement=align_gauche)
    _fusion(r, 3, r, 5)
    _case(r, 3, sign2_code, police=police_valeur, alignement=align_gauche,
          remplissage=fill_jaune, bordure_obj=bordure_fine)
    ws.row_dimensions[r].height = 16
    r += 1
    _case(r, 1, "Fonction :", police=police_label, alignement=align_gauche)
    _fusion(r, 2, r, 5)
    _case(r, 2, sign2_fonction, police=police_valeur, alignement=align_gauche,
          remplissage=fill_jaune, bordure_obj=bordure_fine)
    _case(r, 6, "Date :", police=police_label, alignement=align_gauche)
    _fusion(r, 7, r, 8)
    _case(r, 7, sign2_date, police=police_valeur, alignement=align_gauche,
          remplissage=fill_jaune, bordure_obj=bordure_fine)
    ws.row_dimensions[r].height = 16
    r += 1
    
    # Espace pour l'image 2
    _fusion(r, 1, r+1, 8)
    sign2_b64 = _champ(attestation, "sign2_image")
    if sign2_b64:
        try:
            import base64
            from io import BytesIO
            sig_img = openpyxl.drawing.image.Image(BytesIO(base64.b64decode(sign2_b64)))
            sig_img.width = 150
            sig_img.height = 50
            ws.add_image(sig_img, f"D{r}")
        except Exception:
            pass
    ws.row_dimensions[r].height = 25
    ws.row_dimensions[r+1].height = 25
    r += 2
    r += 1

    # ── Note de confidentialité ──────────────────────────────────────────────
    _fusion(r, 1, r, 8)
    _case(r, 1,
          "Document réglementaire officiel — Confidentiel. La présente attestation, revêtue des "
          "signatures et du cachet de l'établissement, accompagne la déclaration transmise à la "
          "BCEAO et à la Commission Bancaire de l'UMOA.",
          police=police_note, alignement=align_centre)
    ws.row_dimensions[r].height = 26

    # ── Mise en page : paysage, zone d'impression explicite ─────────────────
    from openpyxl.worksheet.properties import PageSetupProperties

    ws.page_setup.orientation = "landscape"
    ws.print_area = "A1:H26"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False


DISPRU_ALIASES: dict[str, str] = {
    "IM12": "im012",
    "ID09": "id009",
    "IM06": "im006",
    "IM10": "im010",
    "PR04": "pr004",
    "PA07": "pa149",
    "PA14": "pa156",
    "PA32": "pa173",
    "EP21_PB_N1": "ro001",
    "EP21_PB_N2": "ro001_n2",
    "EP21_PB_N3": "ro001_n3",
    "EP35_CHANGE_POSITION_NETTE": "rm067",
}


def parse_fonds_propres_import(contenu: bytes) -> dict[str, float]:
    """Lit un classeur FODEP et extrait les postes DISPRU détectés.

    Prend en charge :
    1. Le classeur officiel BCEAO multi-onglets (EP03/EP05, EP21, EP33, EP36, EP37, EP38...)
       tel que la matrice réglementaire officielle (`Matrice_FODEP_Officielle.xlsx`).
    2. Les classeurs de saisie multi-onglets (colonnes Catégorie / Code / Libellé / Valeur).
    3. Le format tabulaire simple ou export à 5 colonnes.
    """
    from openpyxl import load_workbook

    codes_connus: dict[str, str] = {c.code.upper(): c.code.lower() for c in FONDS_PROPRES_CODES}
    codes_connus.update(DISPRU_ALIASES)

    wb = load_workbook(io.BytesIO(contenu), data_only=True)
    postes: dict[str, float] = {}

    sheet_names_upper = {name.strip().upper(): name for name in wb.sheetnames}

    # ── Cas 1 : Classeur officiel BCEAO multi-onglets ────────────────────────
    # 1. EP03 (Fonds Propres Base Individuelle) ou EP05 (Base Consolidée)
    ep03_name = sheet_names_upper.get("EP03") or sheet_names_upper.get("EP05")
    if ep03_name:
        ws = wb[ep03_name]
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if c1:
                code = str(c1).strip().upper()
                if code in codes_connus:
                    val = ws.cell(r, 3).value
                    if val is not None and str(val).strip() != "":
                        try:
                            postes[codes_connus[code]] = float(val)
                        except (ValueError, TypeError):
                            pass

    # 2. EP21 (Produit brut risque opérationnel - RO001 à RO008)
    if "EP21" in sheet_names_upper:
        ws = wb[sheet_names_upper["EP21"]]
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if c1:
                code = str(c1).strip().upper()
                if code in codes_connus:
                    for col in (4, 3, 5):
                        val = ws.cell(r, col).value
                        if val is not None and str(val).strip() != "":
                            try:
                                postes[codes_connus[code]] = float(val)
                                break
                            except (ValueError, TypeError):
                                pass

    # 3. EP33 (Ratio de levier - RL001 à RL012)
    if "EP33" in sheet_names_upper:
        ws = wb[sheet_names_upper["EP33"]]
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if c1:
                code = str(c1).strip().upper()
                if code in codes_connus:
                    val = ws.cell(r, 3).value
                    if val is not None and str(val).strip() != "":
                        try:
                            postes[codes_connus[code]] = float(val)
                        except (ValueError, TypeError):
                            pass

    # 4. EP36 (Immobilisations hors exploitation - IM001, IM002, IM003, PA084)
    if "EP36" in sheet_names_upper:
        ws = wb[sheet_names_upper["EP36"]]
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if c1:
                code = str(c1).strip().upper()
                if code in codes_connus and codes_connus[code] not in postes:
                    for col in (4, 3, 5):
                        val = ws.cell(r, col).value
                        if val is not None and str(val).strip() != "":
                            try:
                                postes[codes_connus[code]] = float(val)
                                break
                            except (ValueError, TypeError):
                                pass

    # 5. EP37 (Immobilisations d'exploitation et participations - IM007, PA106)
    if "EP37" in sheet_names_upper:
        ws = wb[sheet_names_upper["EP37"]]
        for r in range(1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if c1:
                code = str(c1).strip().upper()
                if code in codes_connus and codes_connus[code] not in postes:
                    for col in (4, 3, 5):
                        val = ws.cell(r, col).value
                        if val is not None and str(val).strip() != "":
                            try:
                                postes[codes_connus[code]] = float(val)
                                break
                            except (ValueError, TypeError):
                                pass

    # 6. EP38 (Prêts aux actionnaires, dirigeants et personnel - PR001A-H, PR002A-H)
    if "EP38" in sheet_names_upper:
        ws = wb[sheet_names_upper["EP38"]]
        categories = ("A", "B", "C", "D", "E", "F", "G", "H")
        for idx, cat in enumerate(categories):
            col = 3 + idx
            val_pr1 = ws.cell(11, col).value
            if val_pr1 is not None and str(val_pr1).strip() != "":
                try:
                    postes[f"pr001{cat.lower()}"] = float(val_pr1)
                except (ValueError, TypeError):
                    pass
            val_pr2 = ws.cell(12, col).value
            if val_pr2 is not None and str(val_pr2).strip() != "":
                try:
                    postes[f"pr002{cat.lower()}"] = float(val_pr2)
                except (ValueError, TypeError):
                    pass

    # ── Cas 2 : Balayage universel (toutes feuilles, colonnes de codes 1 ou 2)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(1, ws.max_row + 1):
            for c_idx in (1, 2):
                c_val = ws.cell(r, c_idx).value
                if c_val:
                    code_raw = str(c_val).strip().upper()
                    if code_raw in codes_connus:
                        tgt = codes_connus[code_raw]
                        if tgt not in postes:
                            for col in range(c_idx + 1, min(ws.max_column + 1, 10)):
                                v = ws.cell(r, col).value
                                if isinstance(v, (int, float)) and v != "":
                                    postes[tgt] = float(v)
                                    break

    return postes
