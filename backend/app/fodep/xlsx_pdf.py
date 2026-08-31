"""Rendu PDF d'un classeur Excel, en Python pur.

Ce module reproduit le classeur FODEP officiel en PDF sans Excel, sans
LibreOffice et sans aucun logiciel a installer : il ne depend que d'openpyxl et
de reportlab, deux paquets pip livres avec l'application (image Docker comme
installeur Windows).

Le rendu vise la fidelite au classeur tel qu'il s'ouvre dans Excel : meme
decoupage en onglets, memes largeurs de colonnes, memes fusions, memes polices,
remplissages, bordures, alignements, formats de nombre et mises en forme
conditionnelles. Les formules sont recalculees en amont par `formula_engine`,
sans quoi les totaux seraient vides.

Perimetre mesure sur le classeur officiel (44 onglets, 23294 cellules utiles) :
6 couleurs de fond, 2 styles de bordure, 4 alignements horizontaux, 8 formats
de nombre numeriques, 0 rotation, 0 indent, 0 volet fige, 0 ligne masquee.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import (
    BaseDocTemplate,
    Flowable,
    Frame,
    Image,
    NextPageTemplate,
    PageBreak,
    PageTemplate,
    Table,
    TableStyle,
)

from app.fodep.formula_engine import ErreurExcel, MoteurFormules

# ── Conversions d'unites Excel ──────────────────────────────────────────────

_LARGEUR_COLONNE_DEFAUT = 8.43  # unites Excel
_HAUTEUR_LIGNE_DEFAUT = 12.75  # points, pour de l'Arial 10
_POINTS_PAR_PIXEL = 0.75  # 96 DPI : 72/96

# Rapport hauteur de ligne / corps de police, proche du rendu d'Excel.
_INTERLIGNE = 1.18

# Garde-fou : au-dela, la plage declaree par le classeur est jugee aberrante.
# Necessaire car openpyxl annonce parfois max_column=1027 (colonnes AMB/AMD)
# des qu'une ligne entiere porte un style, sans aucun contenu reel derriere.
_LIGNES_MAX = 5000
_COLONNES_MAX = 256

_EPAISSEUR_BORDURE = {
    "hair": 0.25,
    "thin": 0.4,
    "dotted": 0.4,
    "dashed": 0.4,
    "dashDot": 0.4,
    "dashDotDot": 0.4,
    "medium": 0.9,
    "mediumDashed": 0.9,
    "mediumDashDot": 0.9,
    "mediumDashDotDot": 0.9,
    "slantDashDot": 0.9,
    "thick": 1.4,
    "double": 1.0,
}


def _largeur_colonne_points(feuille: Any, index: int) -> float:
    """Largeur d'une colonne en points, selon la formule exacte d'Excel.

    Une unite de largeur Excel est calibree sur le chiffre le plus large de la
    police par defaut du classeur (7 px en Arial 10). Excel convertit ensuite
    cette unite en pixels par ``px = round(((256*w + round(128/7)) / 256) * 7)``
    avant de passer a l'affichage a 96 DPI (0,75 pt/px). Une approximation plus
    grossiere derive sur les onglets les plus larges (jusqu'a 2552 pt de
    contenu pour 740 pt utiles en paysage), ou l'erreur cumulee sur des
    dizaines de colonnes suffit a faire deborder la mise en page.
    """

    dimension = feuille.column_dimensions.get(get_column_letter(index))
    if dimension is not None and dimension.hidden:
        return 0.0
    largeur = None
    if dimension is not None and dimension.width:
        largeur = float(dimension.width)
    if largeur is None:
        defaut = getattr(feuille.sheet_format, "defaultColWidth", None)
        largeur = float(defaut) if defaut else _LARGEUR_COLONNE_DEFAUT

    pixels = round(((256 * largeur + round(128 / 7)) / 256) * 7)
    return pixels * _POINTS_PAR_PIXEL


def _hauteur_ligne_points(feuille: Any, index: int) -> float:
    """Hauteur d'une ligne en points, telle que stockee par Excel.

    Ces hauteurs sont deja calibrees par Excel pour son propre retour a la
    ligne (109 valeurs distinctes mesurees sur le classeur). Les recalculer a
    partir des metriques reportlab ferait deriver la pagination des onglets
    longs (jusqu'a 130 lignes).
    """

    dimension = feuille.row_dimensions.get(index)
    if dimension is not None and dimension.hidden:
        return 0.0
    if dimension is not None and dimension.height:
        return float(dimension.height)
    defaut = getattr(feuille.sheet_format, "defaultRowHeight", None)
    return float(defaut) if defaut else _HAUTEUR_LIGNE_DEFAUT


# ── Couleurs ────────────────────────────────────────────────────────────────

# Ordre de declaration des couleurs du theme (ECMA-376), puis correspondance
# avec l'index que portent les cellules : Excel intervertit fond et texte pour
# les deux premieres paires, d'ou ces deux tableaux distincts.
_ORDRE_THEME = (
    "dk1", "lt1", "dk2", "lt2",
    "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
    "hlink", "folHlink",
)
_INDEX_THEME = (
    "lt1", "dk1", "lt2", "dk2",
    "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
    "hlink", "folHlink",
)

# Palette indexee historique d'Excel, restreinte aux entrees courantes.
_PALETTE_INDEXEE = {
    0: "000000", 1: "FFFFFF", 2: "FF0000", 3: "00FF00", 4: "0000FF",
    5: "FFFF00", 6: "FF00FF", 7: "00FFFF", 8: "000000", 9: "FFFFFF",
    10: "FF0000", 11: "00FF00", 12: "0000FF", 13: "FFFF00", 14: "FF00FF",
    15: "00FFFF", 22: "C0C0C0", 23: "808080", 64: "000000", 65: "FFFFFF",
}


def _extraire_theme(classeur: Any) -> dict[str, str]:
    """Lit la palette du theme du classeur (couleurs `theme=` des cellules)."""

    brut = getattr(classeur, "loaded_theme", None)
    if not brut:
        return {}
    if isinstance(brut, bytes):
        brut = brut.decode("utf-8", "ignore")

    bloc = re.search(r"<a:clrScheme\b.*?</a:clrScheme>", brut, re.DOTALL)
    if not bloc:
        return {}

    palette: dict[str, str] = {}
    for index, element in enumerate(
        re.finditer(r"<a:(\w+)>\s*<a:(srgbClr|sysClr)([^>]*)/>", bloc.group(0))
    ):
        if index >= len(_ORDRE_THEME):
            break
        attributs = element.group(3)
        valeur = re.search(r'val="([0-9A-Fa-f]{6})"', attributs)
        dernier = re.search(r'lastClr="([0-9A-Fa-f]{6})"', attributs)
        source = valeur or dernier
        if source:
            palette[_ORDRE_THEME[index]] = source.group(1).upper()
    return palette


def _appliquer_teinte(hexa: str, teinte: float) -> str:
    """Applique la teinte Excel (`tint`) : eclaircit si positive, assombrit sinon."""

    if not teinte:
        return hexa
    composantes = []
    for decalage in (0, 2, 4):
        canal = int(hexa[decalage: decalage + 2], 16)
        if teinte > 0:
            canal = canal * (1 - teinte) + 255 * teinte
        else:
            canal = canal * (1 + teinte)
        composantes.append(max(0, min(255, int(round(canal)))))
    return "%02X%02X%02X" % tuple(composantes)


def _couleur(valeur: Any, theme: dict[str, str]) -> colors.Color | None:
    """Traduit une couleur openpyxl en couleur reportlab.

    Renvoie None pour une couleur transparente/automatique : c'est ainsi que
    les remplissages differentiels (mise en forme conditionnelle) codent
    « pas de cette couleur-la », avec un alpha nul dans le canal ARGB.
    """

    if valeur is None:
        return None

    type_couleur = getattr(valeur, "type", None)
    hexa: str | None = None

    if type_couleur == "rgb":
        brut = getattr(valeur, "rgb", None)
        if isinstance(brut, str) and len(brut) >= 6:
            if len(brut) == 8 and brut[:2].upper() == "00":
                return None
            hexa = brut[-6:].upper()
    elif type_couleur == "theme":
        index = getattr(valeur, "theme", None)
        if isinstance(index, int) and 0 <= index < len(_INDEX_THEME):
            hexa = theme.get(_INDEX_THEME[index])
    elif type_couleur == "indexed":
        hexa = _PALETTE_INDEXEE.get(getattr(valeur, "indexed", None))

    if not hexa or len(hexa) != 6:
        return None

    try:
        hexa = _appliquer_teinte(hexa, float(getattr(valeur, "tint", 0.0) or 0.0))
        return colors.HexColor("#" + hexa)
    except ValueError:
        return None


# ── Formats de nombre ───────────────────────────────────────────────────────

_SEPARATEUR_MILLIERS = " "  # espace fine insecable, usage francais


def _nettoyer_format(code: str) -> str:
    """Retire les prefixes de locale et de couleur d'un format Excel."""

    code = re.sub(r"\[\$?-?[0-9A-Fa-f]+\]", "", code)
    code = re.sub(r"\[[A-Za-z]+\]", "", code)
    return code


def _formater_general(nombre: float) -> str:
    """Format « Standard » d'Excel : ni separateur de milliers, ni zero inutile."""

    if nombre == int(nombre) and abs(nombre) < 1e15:
        return str(int(nombre))
    texte = repr(round(nombre, 10))
    if "." in texte:
        texte = texte.rstrip("0").rstrip(".")
    # Le classeur est en locale francaise ([$-40C]) : la virgule y separe les
    # decimales, comme dans les formats explicites du meme document.
    return texte.replace(".", ",")


def _appliquer_motif(nombre: float, motif: str) -> str:
    """Formate un nombre selon un motif Excel (sous-ensemble reellement utilise)."""

    if "%" in motif:
        nombre *= 100.0

    if re.search(r"[Ee][+-]?0+", motif):
        decimales = 0
        mantisse = re.match(r"[#0]*(?:\.(0+))?", motif)
        if mantisse and mantisse.group(1):
            decimales = len(mantisse.group(1))
        return ("%." + str(decimales) + "e") % nombre

    milliers = "," in re.sub(r'"[^"]*"', "", motif)
    partie_decimale = re.search(r"\.([0#]+)", motif)
    decimales = len(partie_decimale.group(1)) if partie_decimale else 0

    negatif = nombre < 0
    valeur = round(abs(nombre), decimales)
    entier = int(valeur)
    texte = f"{entier:,}".replace(",", _SEPARATEUR_MILLIERS) if milliers else str(entier)
    if decimales:
        texte = f"{texte},{f'{valeur - entier:.{decimales}f}'[2:]}"
    if negatif and valeur != 0:
        texte = "-" + texte

    # Le reste du motif (espaces echappes, symboles, %) est restitue tel quel :
    # « 0\ % » rend ainsi « 12 % ». Le souligne (« _€ ») ne fait que reserver
    # la largeur d'un caractere sans jamais s'afficher : il n'est pas capture.
    suffixe = ""
    for correspondance in re.finditer(r'\\(.)|"([^"]*)"|(%)', motif):
        suffixe += correspondance.group(1) or correspondance.group(2) or correspondance.group(3)
    return texte + suffixe


def _formater(valeur: Any, code_format: str) -> str:
    """Rend une valeur de cellule comme Excel l'afficherait."""

    if valeur is None:
        return ""
    if isinstance(valeur, ErreurExcel):
        return str(valeur)
    if isinstance(valeur, bool):
        return "VRAI" if valeur else "FAUX"
    if isinstance(valeur, str):
        return valeur
    if not isinstance(valeur, (int, float)):
        return str(valeur)

    motif = _nettoyer_format(code_format or "General").strip()
    if not motif or motif.upper() == "GENERAL" or motif == "@":
        return _formater_general(float(valeur))

    # Un format peut porter des sections positif;negatif;zero.
    sections = re.split(r"(?<!\\);", motif)
    nombre = float(valeur)
    if len(sections) > 1 and nombre < 0:
        motif, nombre = sections[1], abs(nombre)
    elif len(sections) > 2 and nombre == 0:
        motif = sections[2]
    else:
        motif = sections[0]

    if not re.search(r"[#0]", motif):
        return _formater_general(nombre)
    return _appliquer_motif(nombre, motif)


# ── Glyphes hors Latin-1 ─────────────────────────────────────────────────────

# Les polices Type1 integrees a reportlab (Helvetica, Times, Courier) sont
# limitees a WinAnsiEncoding : l'apostrophe typographique et le tiret
# demi-cadratin y figurent (rendus tels quels), mais pas <=, >= ni le signe
# somme. Sans police Unicode embarquee, on substitue une equivalence ASCII
# plutot que de laisser un rectangle vide s'afficher a la place du glyphe.
_SUBSTITUTIONS_GLYPHES = {
    "≤": "<=",
    "≥": ">=",
    "∑": "Somme ",
}


def _substituer_glyphes_non_supportes(texte: str) -> str:
    for cible, remplacement in _SUBSTITUTIONS_GLYPHES.items():
        if cible in texte:
            texte = texte.replace(cible, remplacement)
    return texte


# ── Polices ─────────────────────────────────────────────────────────────────


def _police(nom: str | None, gras: bool, italique: bool) -> str:
    """Choisit la police PDF correspondant a celle du classeur.

    La matrice officielle est en Arial (le nom « Arial1 » qui y apparait aussi
    est un artefact de deduplication de LibreOffice, pas une police reelle :
    il designe la meme Arial). Helvetica, presente dans tout lecteur PDF, en
    partage les metriques : les largeurs de texte du PDF correspondent donc a
    celles d'Excel, sans embarquer le moindre fichier de police.
    """

    reference = (nom or "").lower()
    if "courier" in reference or "consol" in reference or "mono" in reference:
        famille = ("Courier", "Courier-Bold", "Courier-Oblique", "Courier-BoldOblique")
    elif any(cle in reference for cle in ("times", "serif", "cambria", "georgia", "garamond")):
        famille = ("Times-Roman", "Times-Bold", "Times-Italic", "Times-BoldItalic")
    else:
        famille = ("Helvetica", "Helvetica-Bold", "Helvetica-Oblique", "Helvetica-BoldOblique")

    if gras and italique:
        return famille[3]
    if gras:
        return famille[1]
    if italique:
        return famille[2]
    return famille[0]


_ALIGNEMENT_HORIZONTAL = {
    "left": "LEFT",
    "right": "RIGHT",
    "center": "CENTER",
    "centerContinuous": "CENTER",
    "justify": "LEFT",
    "distributed": "CENTER",
    "fill": "LEFT",
}
_ALIGNEMENT_VERTICAL = {
    "top": "TOP",
    "center": "MIDDLE",
    "bottom": "BOTTOM",
    "justify": "MIDDLE",
    "distributed": "MIDDLE",
}


# ── Mise en page du texte ───────────────────────────────────────────────────


def _replier(texte: str, police: str, taille: float, largeur: float) -> list[str]:
    """Replie un texte sur plusieurs lignes, comme le renvoi automatique d'Excel."""

    if largeur <= 0:
        return [texte]

    lignes: list[str] = []
    for paragraphe in texte.split("\n"):
        courante = ""
        for mot in paragraphe.split(" "):
            candidate = f"{courante} {mot}" if courante else mot
            if pdfmetrics.stringWidth(candidate, police, taille) <= largeur or not courante:
                courante = candidate
            else:
                lignes.append(courante)
                courante = mot
            # Un mot seul plus large que la cellule est coupe en force,
            # faute de quoi il deborderait sur les colonnes voisines.
            while pdfmetrics.stringWidth(courante, police, taille) > largeur and len(courante) > 1:
                coupe = len(courante) - 1
                while coupe > 1 and pdfmetrics.stringWidth(courante[:coupe], police, taille) > largeur:
                    coupe -= 1
                lignes.append(courante[:coupe])
                courante = courante[coupe:]
        lignes.append(courante)
    return lignes or [""]


def _tronquer(texte: str, police: str, taille: float, largeur: float) -> str:
    """Coupe un texte trop long, comme Excel le fait contre une cellule occupee."""

    if largeur <= 0:
        return ""
    if pdfmetrics.stringWidth(texte, police, taille) <= largeur:
        return texte
    court = texte
    while court and pdfmetrics.stringWidth(court, police, taille) > largeur:
        court = court[:-1]
    return court


def _dieses_pour_largeur(police: str, taille: float, largeur: float) -> str:
    """Repere « ### » d'Excel pour un nombre qui ne tient pas dans sa colonne.

    Un nombre ne se deverse JAMAIS dans les cellules voisines, contrairement au
    texte : Excel montrerait une valeur fausse (des chiffres coupes en cours de
    nombre) plutot que d'admettre le depassement. Le sens exact du chiffre
    tronque etant irrecuperable, ### signale le depassement sans mentir sur la
    valeur.
    """

    if largeur <= 0:
        return ""
    largeur_diese = pdfmetrics.stringWidth("#", police, taille)
    n = max(1, int(largeur // largeur_diese)) if largeur_diese > 0 else 1
    texte = "#" * n
    while texte and pdfmetrics.stringWidth(texte, police, taille) > largeur:
        texte = texte[:-1]
    return texte or "#"


# ── Mise en forme conditionnelle ─────────────────────────────────────────────

# Perimetre mesure sur le classeur : 6 regles cellIs porteuses de sens
# reglementaire (CONFORME/INFRACTION sur EP01, depassements >15 sur
# EP36/EP37) et 4 regles iconSet sur Liste_EP_a_renseigner qui masquent la
# valeur numerique (showValue=false) pour n'afficher qu'une icone. Ignorer ces
# regles ferait apparaitre au lecteur des nombres qu'Excel ne montre jamais
# (les placeholders 68/32 de la liste des etats) ou des indicateurs de
# conformite sans leur code couleur.


@dataclass
class _RegleMFC:
    min_ligne: int
    min_col: int
    max_ligne: int
    max_col: int
    genre: str  # "egal_texte" | "superieur_nombre" | "icone"
    cible: Any
    couleur_police: colors.Color | None
    gras: bool
    couleur_fond: colors.Color | None


def _couleur_dxf_fill(dxf: Any, theme: dict[str, str]) -> colors.Color | None:
    """Couleur effective d'un remplissage differentiel.

    Un dxf de mise en forme conditionnelle code souvent son fond dans
    `bgColor` avec `fgColor` transparent (00000000) : c'est une convention
    OOXML pour les remplissages « unis » definis par difference, pas une
    inversion a corriger.
    """

    remplissage = getattr(dxf, "fill", None)
    if remplissage is None:
        return None
    return _couleur(getattr(remplissage, "fgColor", None), theme) or _couleur(
        getattr(remplissage, "bgColor", None), theme
    )


def _extraire_regles_mfc(feuille: Any, theme: dict[str, str]) -> list[_RegleMFC]:
    regles: list[_RegleMFC] = []
    for plage_cf, jeu_regles in feuille.conditional_formatting._cf_rules.items():
        boites: list[tuple[int, int, int, int]] = []
        for morceau in str(plage_cf.sqref).split():
            try:
                min_col, min_ligne, max_col, max_ligne = range_boundaries(morceau)
            except ValueError:
                continue
            boites.append((min_ligne, min_col, max_ligne, max_col))

        for regle in jeu_regles:
            for min_ligne, min_col, max_ligne, max_col in boites:
                if regle.type == "cellIs" and regle.formula:
                    brut = regle.formula[0]
                    dxf = regle.dxf
                    if regle.operator == "equal":
                        police_dxf = getattr(dxf, "font", None)
                        couleur_police = (
                            _couleur(police_dxf.color, theme)
                            if police_dxf is not None and police_dxf.color is not None
                            else None
                        )
                        regles.append(
                            _RegleMFC(
                                min_ligne, min_col, max_ligne, max_col,
                                "egal_texte", brut.strip('"'),
                                couleur_police,
                                bool(police_dxf.b) if police_dxf is not None else False,
                                None,
                            )
                        )
                    elif regle.operator == "greaterThan":
                        try:
                            seuil = float(brut)
                        except (TypeError, ValueError):
                            continue
                        regles.append(
                            _RegleMFC(
                                min_ligne, min_col, max_ligne, max_col,
                                "superieur_nombre", seuil, None, False,
                                _couleur_dxf_fill(dxf, theme),
                            )
                        )
                elif regle.type == "iconSet" and not getattr(regle.iconSet, "showValue", True):
                    regles.append(
                        _RegleMFC(min_ligne, min_col, max_ligne, max_col, "icone", None, None, False, None)
                    )
    return regles


def _regle_applicable(regles: list[_RegleMFC], ligne: int, colonne: int, valeur: Any) -> _RegleMFC | None:
    for regle in regles:
        if not (regle.min_ligne <= ligne <= regle.max_ligne and regle.min_col <= colonne <= regle.max_col):
            continue
        if regle.genre == "egal_texte":
            if isinstance(valeur, str) and valeur == regle.cible:
                return regle
        elif regle.genre == "superieur_nombre":
            if isinstance(valeur, (int, float)) and not isinstance(valeur, bool) and valeur > regle.cible:
                return regle
        elif regle.genre == "icone":
            return regle
    return None


# ── Rendu d'un onglet ───────────────────────────────────────────────────────


@dataclass
class _Plage:
    lignes: int
    colonnes: int


def _zone_impression(feuille: Any) -> tuple[int, int] | None:
    """Derniere ligne/colonne de la zone d'impression definie dans Excel.

    Honorer print_area n'est pas une nuance esthetique : EP07 porte 40
    cellules de contenu (colonnes O a BB) volontairement exclues de
    l'impression, et les ignorer ferait passer l'onglet de 14 a 54 colonnes,
    faussant a la fois la largeur et la mise a l'echelle de la page entiere.
    """

    brut = feuille.print_area
    if not brut:
        return None

    max_ligne = max_colonne = 0
    for morceau in str(brut).split(","):
        morceau = morceau.strip()
        if "!" in morceau:
            morceau = morceau.split("!", 1)[1]
        try:
            _, _, colonne_fin, ligne_fin = range_boundaries(morceau.replace("$", ""))
        except ValueError:
            continue
        max_ligne = max(max_ligne, ligne_fin)
        max_colonne = max(max_colonne, colonne_fin)

    if max_ligne == 0 or max_colonne == 0:
        return None
    return (min(max_ligne, _LIGNES_MAX), min(max_colonne, _COLONNES_MAX))


def _plage_utilisee(feuille: Any) -> _Plage:
    """Derniere ligne et derniere colonne a rendre pour cet onglet."""

    zone = _zone_impression(feuille)
    if zone is not None:
        return _Plage(*zone)

    # Pas de zone d'impression definie : reconstruire une bbox de contenu, sans
    # jamais se fier a ws.max_row/max_column (une ligne entiere stylee suffit a
    # les faire grimper a plus de 1000, pour zero contenu reel derriere).
    lignes_max = min(feuille.max_row or 0, _LIGNES_MAX)
    colonnes_max = min(feuille.max_column or 0, _COLONNES_MAX)
    if lignes_max <= 0 or colonnes_max <= 0:
        return _Plage(0, 0)

    derniere_ligne = 0
    derniere_colonne = 0
    for ligne in feuille.iter_rows(min_row=1, max_row=lignes_max, max_col=colonnes_max):
        for cellule in ligne:
            valeur = cellule.value
            if valeur is None or (isinstance(valeur, str) and not valeur.strip()):
                continue
            derniere_ligne = max(derniere_ligne, cellule.row)
            derniere_colonne = max(derniere_colonne, cellule.column)

    if derniere_ligne == 0:
        return _Plage(0, 0)

    # Une fusion qui commence dans la zone renseignee en fait partie, meme si
    # sa derniere cellule est vide : c'est le cas des titres etales sur
    # plusieurs colonnes.
    for fusion in feuille.merged_cells.ranges:
        if fusion.min_row <= derniere_ligne and fusion.min_col <= derniere_colonne:
            derniere_ligne = max(derniere_ligne, min(fusion.max_row, lignes_max))
            derniere_colonne = max(derniere_colonne, min(fusion.max_col, colonnes_max))

    return _Plage(derniere_ligne, derniere_colonne)


def _carte_fusions(feuille: Any, plage: _Plage) -> tuple[dict, set]:
    """Cellules pilotes des fusions, et cellules qu'elles recouvrent."""

    ancres: dict[tuple[int, int], tuple[int, int]] = {}
    couvertes: set[tuple[int, int]] = set()

    for fusion in feuille.merged_cells.ranges:
        if fusion.min_row > plage.lignes or fusion.min_col > plage.colonnes:
            continue
        ligne_fin = min(fusion.max_row, plage.lignes)
        colonne_fin = min(fusion.max_col, plage.colonnes)
        ancres[(fusion.min_row, fusion.min_col)] = (ligne_fin, colonne_fin)
        for ligne in range(fusion.min_row, ligne_fin + 1):
            for colonne in range(fusion.min_col, colonne_fin + 1):
                if (ligne, colonne) != (fusion.min_row, fusion.min_col):
                    couvertes.add((ligne, colonne))

    return ancres, couvertes


def _bordure_arete(feuille: Any, l1: int, c1: int, l2: int, c2: int, cote: str) -> Any:
    """Cherche une bordure sur une arete de rectangle (fusion ou cellule simple).

    388 fusions du classeur portent leur trait sur leur DERNIERE cellule (bas
    ou droite) plutot que sur la cellule d'ancrage : ne lire que l'ancre en
    perd une grande partie. On balaie donc toute l'arete concernee.
    """

    if cote in ("left", "right"):
        colonne = c1 if cote == "left" else c2
        plage = range(l1, l2 + 1)
        for ligne in plage:
            bordure = getattr(feuille.cell(ligne, colonne).border, cote, None)
            if bordure is not None and bordure.style:
                return bordure
    else:
        ligne = l1 if cote == "top" else l2
        for colonne in range(c1, c2 + 1):
            bordure = getattr(feuille.cell(ligne, colonne).border, cote, None)
            if bordure is not None and bordure.style:
                return bordure
    return None


def _construire_table(
    feuille: Any,
    plage: _Plage,
    moteur: MoteurFormules,
    theme: dict[str, str],
    echelle: float,
) -> Table:
    """Assemble la table reportlab correspondant a un onglet."""

    largeurs = [
        _largeur_colonne_points(feuille, index) * echelle
        for index in range(1, plage.colonnes + 1)
    ]
    hauteurs = [
        _hauteur_ligne_points(feuille, index) * echelle
        for index in range(1, plage.lignes + 1)
    ]

    ancres, couvertes = _carte_fusions(feuille, plage)
    regles_mfc = _extraire_regles_mfc(feuille, theme)

    donnees: list[list[str]] = []
    commandes: list[tuple] = []

    # Valeurs brutes de toutes les cellules : elles servent aussi a savoir si
    # une voisine est occupee, ce qui determine si un texte peut deborder.
    valeurs: dict[tuple[int, int], Any] = {}
    for index_ligne in range(1, plage.lignes + 1):
        for index_colonne in range(1, plage.colonnes + 1):
            cellule = feuille.cell(index_ligne, index_colonne)
            valeur = cellule.value
            if isinstance(valeur, str) and valeur.startswith("="):
                valeur = moteur.valeur(feuille.title, cellule.coordinate)
            if isinstance(valeur, str):
                valeur = _substituer_glyphes_non_supportes(valeur)
            valeurs[(index_ligne, index_colonne)] = valeur

    for index_ligne in range(1, plage.lignes + 1):
        ligne_rendue: list[str] = []

        for index_colonne in range(1, plage.colonnes + 1):
            position = (index_colonne - 1, index_ligne - 1)

            if (index_ligne, index_colonne) in couvertes:
                ligne_rendue.append("")
                continue

            cellule = feuille.cell(index_ligne, index_colonne)
            valeur = valeurs[(index_ligne, index_colonne)]
            regle = _regle_applicable(regles_mfc, index_ligne, index_colonne, valeur)

            fin_ligne, fin_colonne = ancres.get(
                (index_ligne, index_colonne), (index_ligne, index_colonne)
            )
            largeur_cellule = sum(largeurs[index_colonne - 1: fin_colonne])

            police_cellule = cellule.font
            taille = max(float(police_cellule.sz or 10) * echelle, 3.2)
            gras = bool(police_cellule.b) or (regle is not None and regle.genre == "egal_texte" and regle.gras)
            nom_police = _police(police_cellule.name, gras, bool(police_cellule.i))
            interligne = taille * _INTERLIGNE

            # L'icone masque la valeur (showValue=false cote Excel) : afficher
            # le nombre brut ("68") montrerait au lecteur ce qu'Excel ne
            # montre jamais.
            if regle is not None and regle.genre == "icone":
                texte = ""
            else:
                texte = _formater(valeur, cellule.number_format)

            alignement = cellule.alignment
            est_nombre = isinstance(valeur, (int, float)) and not isinstance(valeur, bool)
            horizontal = _ALIGNEMENT_HORIZONTAL.get(
                alignement.horizontal or "", "RIGHT" if est_nombre else "LEFT"
            )

            if texte:
                marges = 2 * 1.2 * echelle
                if alignement.wrap_text:
                    lignes_texte = _replier(
                        texte, nom_police, taille, max(largeur_cellule - marges, 1.0)
                    )
                    texte = "\n".join(lignes_texte)
                    besoin = len(lignes_texte) * interligne
                    if fin_ligne == index_ligne and besoin > hauteurs[index_ligne - 1]:
                        hauteurs[index_ligne - 1] = besoin
                elif est_nombre:
                    # Un nombre ne se deverse JAMAIS dans les cellules voisines,
                    # contrairement au texte : Excel afficherait ### plutot que
                    # de couper des chiffres au milieu d'une valeur, ce qui la
                    # rendrait fausse. Deux totaux dans des cellules adjacentes
                    # qui debordaient chacun vers l'autre se melangeaient
                    # visuellement avant ce correctif.
                    disponible = max(largeur_cellule - marges, 1.0)
                    if pdfmetrics.stringWidth(texte, nom_police, taille) > disponible:
                        texte = _dieses_pour_largeur(nom_police, taille, disponible)
                else:
                    # Sans renvoi automatique, Excel laisse le TEXTE deborder sur
                    # les cellules vides voisines, puis le coupe des qu'une
                    # cellule est occupee.
                    disponible = largeur_cellule
                    if horizontal in ("LEFT", "CENTER"):
                        suivante = fin_colonne + 1
                        while suivante <= plage.colonnes and not valeurs.get((index_ligne, suivante)):
                            disponible += largeurs[suivante - 1]
                            suivante += 1
                    if horizontal in ("RIGHT", "CENTER"):
                        precedente = index_colonne - 1
                        while precedente >= 1 and not valeurs.get((index_ligne, precedente)):
                            disponible += largeurs[precedente - 1]
                            precedente -= 1
                    texte = _tronquer(texte, nom_police, taille, max(disponible - marges, 1.0))

            ligne_rendue.append(texte)

            commandes.append(("FONT", position, position, nom_police, taille, interligne))

            couleur_texte = _couleur(police_cellule.color, theme)
            if regle is not None and regle.genre == "egal_texte" and regle.couleur_police is not None:
                couleur_texte = regle.couleur_police
            if couleur_texte is not None:
                commandes.append(("TEXTCOLOR", position, position, couleur_texte))

            fond = None
            remplissage = cellule.fill
            if remplissage is not None and remplissage.fill_type == "solid":
                fond = _couleur(remplissage.fgColor, theme)
            if regle is not None and regle.genre == "superieur_nombre" and regle.couleur_fond is not None:
                fond = regle.couleur_fond
            if fond is not None:
                commandes.append(("BACKGROUND", position, (fin_colonne - 1, fin_ligne - 1), fond))

            commandes.append(("ALIGN", position, position, horizontal))
            commandes.append(
                (
                    "VALIGN",
                    position,
                    position,
                    _ALIGNEMENT_VERTICAL.get(alignement.vertical or "bottom", "BOTTOM"),
                )
            )

            for cote, commande in (
                ("left", "LINEBEFORE"),
                ("top", "LINEABOVE"),
                ("right", "LINEAFTER"),
                ("bottom", "LINEBELOW"),
            ):
                bordure = _bordure_arete(feuille, index_ligne, index_colonne, fin_ligne, fin_colonne, cote)
                if bordure is None:
                    continue
                epaisseur = max(_EPAISSEUR_BORDURE.get(bordure.style, 0.4) * echelle, 0.15)
                trait = _couleur(bordure.color, theme) or colors.black
                commandes.append(
                    (commande, position, (fin_colonne - 1, fin_ligne - 1), epaisseur, trait)
                )

        donnees.append(ligne_rendue)

    for (ligne_debut, colonne_debut), (ligne_fin, colonne_fin) in ancres.items():
        if (ligne_fin, colonne_fin) != (ligne_debut, colonne_debut):
            commandes.append(
                (
                    "SPAN",
                    (colonne_debut - 1, ligne_debut - 1),
                    (colonne_fin - 1, ligne_fin - 1),
                )
            )

    marge = 1.2 * echelle
    commandes.extend(
        [
            ("LEFTPADDING", (0, 0), (-1, -1), marge),
            ("RIGHTPADDING", (0, 0), (-1, -1), marge),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]
    )

    table = Table(donnees, colWidths=largeurs, rowHeights=hauteurs, repeatRows=0)
    table.setStyle(TableStyle(commandes))
    return table


class _ImageAncree(Flowable):
    """Image d'un onglet, dessinee a sa position d'ancrage reelle (et non
    ajoutee en flux apres le tableau, ce qui la ferait deborder sur une page
    suivante). Ne consomme aucun espace de flot (wrap -> 0, 0)."""

    def __init__(self, donnees: bytes, x: float, y: float, largeur: float, hauteur: float):
        super().__init__()
        self.donnees = donnees
        self.x = x
        self.y = y
        self.largeur = largeur
        self.hauteur = hauteur

    def wrap(self, *args):
        return (0, 0)

    def draw(self):
        if self.donnees:
            self.canv.drawImage(
                ImageReader(io.BytesIO(self.donnees)),
                self.x,
                self.y,
                self.largeur,
                self.hauteur,
                mask="auto",
            )


def _images_feuille(
    feuille: Any, echelle: float, marge: float, format_page: tuple[float, float]
) -> list[Flowable]:
    """Images d'un onglet, dessinees a leur ancre d'origine (logo de page de
    garde, signature de l'attestation, etc.).

    Dans le classeur, chaque image est ancree sur des cellules (souvent
    laissees vides) qui reservent deja sa place dans le tableau. On les
    restitue en surimpression, a l'echelle de la feuille, exactement la ou
    Excel les place, plutot qu'en flux apres le tableau (ce qui les ferait
    basculer sur une page supplementaire).
    """

    images = getattr(feuille, "_images", None)
    if not images:
        return []

    resultats: list[Flowable] = []
    for brut in images:
        ancrage = getattr(brut, "anchor", None)
        if not hasattr(ancrage, "_from"):
            continue
        try:
            donnees = brut._data()
            largeur_native = float(brut.width) * _POINTS_PAR_PIXEL
            hauteur_native = float(brut.height) * _POINTS_PAR_PIXEL
        except Exception:
            continue
        if largeur_native <= 0 or hauteur_native <= 0:
            continue

        depart = ancrage._from
        col_off = (depart.colOff or 0) / 12700.0
        row_off = (depart.rowOff or 0) / 12700.0
        gauche_pt = (
            sum(_largeur_colonne_points(feuille, c + 1) for c in range(depart.col)) + col_off
        )
        haut_pt = (
            sum(_hauteur_ligne_points(feuille, r + 1) for r in range(depart.row)) + row_off
        )

        largeur = largeur_native * echelle
        hauteur = hauteur_native * echelle
        x = marge + gauche_pt * echelle
        y_haut = (format_page[1] - marge) - haut_pt * echelle
        y = y_haut - hauteur

        resultats.append(_ImageAncree(donnees, x, y, largeur, hauteur))

    return resultats


# ── Assemblage du document ──────────────────────────────────────────────────


def convertir_classeur_en_pdf(contenu_xlsx: bytes) -> bytes:
    """Convertit un classeur Excel (octets) en PDF fidele (octets)."""

    import openpyxl

    classeur = openpyxl.load_workbook(io.BytesIO(contenu_xlsx))
    moteur = MoteurFormules(classeur)
    theme = _extraire_theme(classeur)

    tampon = io.BytesIO()
    marge = 8 * mm

    document = BaseDocTemplate(
        tampon,
        pagesize=A4,
        leftMargin=marge,
        rightMargin=marge,
        topMargin=marge,
        bottomMargin=marge,
        title="Declaration prudentielle FODEP",
        author="Risk management",
    )

    modeles: list[PageTemplate] = []
    elements: list[Any] = []

    for nom in classeur.sheetnames:
        feuille = classeur[nom]
        if feuille.sheet_state != "visible":
            continue

        plage = _plage_utilisee(feuille)
        if plage.lignes == 0 or plage.colonnes == 0:
            continue

        paysage = (feuille.page_setup.orientation or "portrait") == "landscape"
        marge = 8 * mm if paysage else 4 * mm
        format_page = landscape(A4) if paysage else A4
        largeur_utile = format_page[0] - 2 * marge

        largeur_brute = sum(
            _largeur_colonne_points(feuille, index) for index in range(1, plage.colonnes + 1)
        )
        # Echelle : ajustement a la largeur de la page (equivalent Excel
        # « ajuster a 1 page de large »). Le contenu remplit toute la largeur
        # utile, sans marge laterale ni centrage flottant. Les onglets dont la
        # hauteur depasse la page sont fractionnes naturellement (fidele au
        # classeur). Le grossissement est autorise pour les onglets plus
        # etroits que la page, afin d'eviter les blocs etroits centres au
        # milieu de feuilles blanches.
        echelle = (largeur_utile / largeur_brute) if largeur_brute > 0 else 1.0

        # Pour les feuilles dont le contenu tient sur une seule page (largeur
        # ajustee), on comble aussi la hauteur utile afin d'eviter les grandes
        # marges blanches verticales (cas typique des pages portrait courtes,
        # ex. l'attestation). Les onglets volontairement multiplicatifs (etats
        # longs) ne sont pas concernes : s'ils depassent nettement la page,
        # ils conservent le fractionnement naturel.
        hauteur_utile = format_page[1] - 2 * marge
        hauteur_brute = sum(
            _hauteur_ligne_points(feuille, r + 1) for r in range(1, plage.lignes + 1)
        )
        if largeur_brute > 0 and hauteur_brute > 0:
            hauteur_contenu = hauteur_brute * echelle
            if hauteur_contenu <= hauteur_utile * 1.25:
                echelle = min(echelle, hauteur_utile / hauteur_brute)

        identifiant = f"onglet-{len(modeles)}"
        modeles.append(
            PageTemplate(
                id=identifiant,
                pagesize=format_page,
                frames=[
                    Frame(
                        marge,
                        marge,
                        format_page[0] - 2 * marge,
                        format_page[1] - 2 * marge,
                        leftPadding=0,
                        rightPadding=0,
                        topPadding=0,
                        bottomPadding=0,
                        id=f"cadre-{identifiant}",
                    )
                ],
            )
        )

        # Le changement de modele ne prend effet qu'a la page suivante : il est
        # donc annonce avant le saut, et le premier onglet ouvre le document.
        elements.append(NextPageTemplate(identifiant))
        if len(modeles) > 1:
            elements.append(PageBreak())
        elements.append(_construire_table(feuille, plage, moteur, theme, echelle))

        for image in _images_feuille(feuille, echelle, marge, format_page):
            elements.append(image)

    if not modeles:
        raise ValueError("Le classeur ne contient aucun onglet imprimable.")

    document.addPageTemplates(modeles)
    document.build(elements)
    return tampon.getvalue()
