"""Services metier du module dashboard."""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
import math
import unicodedata

from app.core.calculations import (
    calculate_capital,
    convert_currency_amount,
    safe_ratio,
)
from app.core.bceao_calculations import (
    calculate_fonds_propres,
    evaluate_ratios
)
from app.market.services import resolve_market_capital
from app.risque_operationnel.services import calcul_aib as _calcul_aib_uemoa
from app.dashboard.models import (
    DashboardMetric,
    DashboardProjectionPoint,
    DashboardSnapshot,
    DistributionEntry,
    PortfolioRow,
    TopExposure,
    FondsPropresDetail,
    FondsPropresUpdate
)
from app.expositions.models import ExposureView
from app.expositions.services import list_expositions
from database.connection import database_manager
from database.services.rwa_calculation_service import (
    normalize_exposure_category_label,
    normalize_exposure_crm_mode,
    normalize_exposure_rating_label,
)

_MONTH_LABELS = [
    "Janv.", "Fevr.", "Mars", "Avr.", "Mai", "Juin",
    "Juil.", "Aout", "Sept.", "Oct.", "Nov.", "Dec.",
]

_TOP5_FALLBACK_COUNTRIES = [
    "Côte d'Ivoire", "Sénégal", "Bénin", "Togo",
    "Burkina Faso", "Mali", "Niger", "Guinée-Bissau",
]

_CRM_BUCKET_ORDER = ("CRM financee", "CRM non financee", "Aucune")
_DISPLAY_CURRENCY = "XOF"

def _build_metric(key: str, label: str, value: float) -> DashboardMetric:
    trend = [value for _ in range(7)]
    return DashboardMetric(
        key=key,
        label=label,
        value=value,
        variation="+0.0% M/M",
        trend=trend,
    )

def _build_distribution_from_buckets(
    buckets: dict[str, float],
    ordered_labels: tuple[str, ...] | None = None,
) -> list[DistributionEntry]:
    total = sum(buckets.values())
    entries = [
        DistributionEntry(
            label=label,
            amount=round(amount, 2),
            percentage=safe_ratio(amount, total),
        )
        for label, amount in buckets.items()
    ]
    entries.sort(key=lambda item: item.amount, reverse=True)

    if ordered_labels is None:
        return entries

    ordered_lookup = {entry.label: entry for entry in entries}
    return [
        ordered_lookup.get(
            label,
            DistributionEntry(label=label, amount=0.0, percentage=0.0),
        )
        for label in ordered_labels
    ]

def _normalize_row(row: ExposureView) -> dict[str, object]:
    on_balance_source = (
        row.on_balance_exposure_amount
        if row.on_balance_exposure_amount is not None
        else row.gross_amount
    )
    off_balance_source = row.off_balance_exposure_amount or 0.0
    has_exposure_breakdown = (
        row.on_balance_exposure_amount is not None
        or row.off_balance_exposure_amount is not None
    )
    if has_exposure_breakdown:
        gross_source = on_balance_source + off_balance_source
    elif row.loan_total_amount is not None and row.loan_total_amount > 0:
        gross_source = row.loan_total_amount
    else:
        gross_source = row.gross_amount
    gross_amount = convert_currency_amount(
        gross_source,
        from_currency=row.currency,
        to_currency=_DISPLAY_CURRENCY,
    )
    on_balance_amount = convert_currency_amount(
        on_balance_source,
        from_currency=row.currency,
        to_currency=_DISPLAY_CURRENCY,
    )
    off_balance_amount = convert_currency_amount(
        off_balance_source,
        from_currency=row.currency,
        to_currency=_DISPLAY_CURRENCY,
    )
    ead = convert_currency_amount(
        row.ead,
        from_currency=row.currency,
        to_currency=_DISPLAY_CURRENCY,
    )
    rwa = convert_currency_amount(
        row.rwa,
        from_currency=row.currency,
        to_currency=_DISPLAY_CURRENCY,
    )
    capital = convert_currency_amount(
        row.capital,
        from_currency=row.currency,
        to_currency=_DISPLAY_CURRENCY,
    )
    return {
        "row": row,
        "category": normalize_exposure_category_label(row.counterparty.category),
        # La note de la contrepartie est une donnee de la ligne : l'echelon
        # prudentiel qui en decoule est un regroupement de lecture, pas une
        # notation. Confondre les deux revient a dire qu'une contrepartie
        # notee BBB- « n'a pas de note precise ».
        "rating": _rating_display(row.counterparty.rating),
        "rating_band": normalize_exposure_rating_label(row.counterparty.rating),
        "crm_mode": normalize_exposure_crm_mode(row.crm_type, row.crm_details),
        "gross_amount": gross_amount,
        "on_balance_exposure_amount": on_balance_amount,
        "off_balance_exposure_amount": off_balance_amount,
        "ead": ead,
        "rwa": rwa,
        "capital": capital,
    }

def _rating_display(rating: str) -> str:
    """Notation telle qu'elle a ete saisie, sans regroupement."""

    return str(rating or "").strip() or "Non noté"

def _crm_snapshot(row: ExposureView, *, ead_xof: float) -> dict[str, object]:
    """Detaille la technique d'attenuation d'une ligne, en XOF.

    Le RWA « avant CRM » est reconstitue a partir du gain d'exposition mesure
    par le moteur (`crm_gain`, nul hors CRM financee) : EAD avant CRM =
    EAD + gain, pondere au poids du debiteur. La comparaison avec le RWA
    retenu rend visible une garantie sans effet, voire defavorable lorsque le
    garant est pondere plus lourdement que le debiteur.
    """

    details = row.crm_details
    collateral_currency = details.collateral_currency or row.currency
    collateral_value = convert_currency_amount(
        details.collateral_value,
        from_currency=collateral_currency,
        to_currency=_DISPLAY_CURRENCY,
    )
    collateral_after_haircut = convert_currency_amount(
        details.cva,
        from_currency=collateral_currency,
        to_currency=_DISPLAY_CURRENCY,
    )
    crm_gain = convert_currency_amount(
        details.crm_gain,
        from_currency=row.currency,
        to_currency=_DISPLAY_CURRENCY,
    )
    return {
        "crm_label": details.label or row.crm_type,
        "crm_coverage_percent": round(float(details.coverage_percent or 0.0), 4),
        "collateral_type": details.collateral_type if details.collateral_value else "",
        "collateral_value": round(collateral_value, 2),
        "collateral_value_after_haircut": round(collateral_after_haircut, 2),
        "collateral_haircut": round(float(details.hc or 0.0) + float(details.hfx or 0.0), 4),
        "crm_eligible": bool(details.eligible),
        "crm_ineligibility_reason": details.eligibility_reason,
        "guarantor_name": details.guarantor_name,
        "guarantor_category": details.guarantor_category,
        "guarantor_rating": details.guarantor_rating,
        # L'echelon accompagne la note du garant comme celle du debiteur :
        # c'est lui qui designe la ligne de grille, donc la ponderation.
        "guarantor_rating_band": normalize_exposure_rating_label(
            details.guarantor_rating
        )
        if details.guarantor_rating
        else "",
        "guarantor_risk_weight": round(float(details.guarantor_risk_weight or 0.0), 4),
        "original_risk_weight": round(float(row.original_rw or 0.0), 4),
        "final_risk_weight": round(float(row.final_rw or 0.0), 4),
        "rwa_before_crm": round((ead_xof + crm_gain) * float(row.original_rw or 0.0), 2),
    }

def _text_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return normalized.encode("ascii", "ignore").decode("ascii").lower()

def _country_key(value: str) -> str:
    return _text_key(value).replace("'", " ").replace("-", " ").strip()

def _canonical_country_label(value: str) -> str:
    aliases = {
        "cote d ivoire": "Côte d'Ivoire",
        "senegal": "Sénégal",
        "benin": "Bénin",
        "guinee bissau": "Guinée-Bissau",
        "guinee equatoriale": "Guinée équatoriale",
    }
    normalized = " ".join(_country_key(value).split())
    return aliases.get(normalized, value.strip())

def _is_defaulted_exposure(row: ExposureView, category: str) -> bool:
    # Le statut prudentiel fait foi : une ligne déclassée par ses impayés
    # (>= 90 jours) ou par décision motivée est en défaut, quel que soit le
    # statut de gestion hérité de l'import, resté « Sain ». L'ignorer
    # sous-estimerait le taux de créances en souffrance du portefeuille.
    if row.statut_prudentiel == "douteuse":
        return True
    status_key = _text_key(row.status)
    category_key = _text_key(category)
    raw_category_key = _text_key(row.counterparty.category)
    return (
        "defaut" in status_key
        or "souffrance" in status_key
        or "defaut" in category_key
        or "souffrance" in category_key
        or "defaut" in raw_category_key
        or "souffrance" in raw_category_key
    )

def _build_country_distribution(buckets: dict[str, float]) -> list[DistributionEntry]:
    total = sum(buckets.values())
    entries = sorted(
        [
            DistributionEntry(
                label=country,
                amount=round(amount, 2),
                percentage=safe_ratio(amount, total),
            )
            for country, amount in buckets.items()
        ],
        key=lambda item: item.amount,
        reverse=True,
    )
    top = entries[:5]
    existing = {item.label for item in top}
    for candidate in _TOP5_FALLBACK_COUNTRIES:
        if len(top) >= 5:
            break
        if candidate in existing:
            continue
        top.append(DistributionEntry(label=candidate, amount=0.0, percentage=0.0))
    return top

def _build_projection(valuation_date: date, base_rwa: float) -> list[DashboardProjectionPoint]:
    points: list[DashboardProjectionPoint] = []
    for month_index in range(12):
        point_month = valuation_date.month + month_index + 1
        year_shift = (point_month - 1) // 12
        month = ((point_month - 1) % 12) + 1
        _ = valuation_date.year + year_shift
        ratio = max(0.0, 1 - (month_index * 0.045))
        projected = round(base_rwa * ratio, 2)
        points.append(
            DashboardProjectionPoint(
                label=_MONTH_LABELS[month - 1],
                value=projected,
            )
        )
    return points

def _max_grouped_share(
    exposure_rows: list[dict[str, object]],
    total: float,
    *,
    key_name: str,
    amount_name: str,
) -> float:
    if total <= 0:
        return 0.0

    buckets: dict[str, float] = defaultdict(float)
    for item in exposure_rows:
        key_value = item.get(key_name)
        if key_name == "counterparty":
            row = item["row"]
            key_value = row.counterparty.name
        elif key_name == "category":
            key_value = item["category"]
        buckets[str(key_value or "Non renseigne")] += float(item[amount_name])

    return max(buckets.values(), default=0.0) / total

def _portfolio_value_at_risk(
    exposure_rows: list[dict[str, object]],
    *,
    gross_total: float,
    rwa_total: float,
) -> float:
    if not exposure_rows or rwa_total <= 0:
        return 0.0

    densities = [
        float(row["rwa"]) / float(row["ead"])
        for row in exposure_rows
        if float(row["ead"]) > 0
    ]
    average_density = safe_ratio(rwa_total, gross_total)
    density_mean = (
        sum(densities) / len(densities)
        if densities
        else average_density
    )
    density_variance = (
        sum((density - density_mean) ** 2 for density in densities) / len(densities)
        if densities
        else 0.0
    )
    concentration_share = _max_grouped_share(
        exposure_rows,
        rwa_total,
        key_name="counterparty",
        amount_name="rwa",
    )
    volatility_proxy = min(
        max(
            math.sqrt(density_variance) * 0.45
            + average_density * 0.035
            + concentration_share * 0.09,
            0.006,
        ),
        0.16,
    )
    return rwa_total * volatility_proxy * 1.65

def _critical_incident_count(
    exposure_rows: list[dict[str, object]],
    *,
    default_rate: float,
    gross_total: float,
    rwa_total: float,
) -> int:
    if not exposure_rows:
        return 0

    incidents = 0
    counterparty_share = _max_grouped_share(
        exposure_rows,
        gross_total,
        key_name="counterparty",
        amount_name="gross_amount",
    )
    category_share = _max_grouped_share(
        exposure_rows,
        rwa_total,
        key_name="category",
        amount_name="rwa",
    )
    has_very_high_density = any(
        float(row["ead"]) > 0 and float(row["rwa"]) / float(row["ead"]) >= 1.5
        for row in exposure_rows
    )
    has_uncovered_large_risk = any(
        rwa_total > 0
        and float(row["rwa"]) / rwa_total >= 0.18
        and str(row["crm_mode"]).strip().lower() in {"", "aucune", "none"}
        for row in exposure_rows
    )

    if default_rate >= 0.05:
        incidents += 1
    if counterparty_share >= 0.35:
        incidents += 1
    if category_share >= 0.50:
        incidents += 1
    if has_very_high_density:
        incidents += 1
    if has_uncovered_large_risk:
        incidents += 1

    return incidents

def get_dashboard_snapshot() -> DashboardSnapshot:
    """Construit le contenu complet du tableau de bord."""
    # FETCH REAL DATA
    with database_manager.read_connection() as conn:
        cursor = conn.cursor()
        
        # Fonds propres
        cursor.execute("SELECT * FROM fonds_propres ORDER BY date_analyse DESC LIMIT 1")
        fp_row = cursor.fetchone()
        fp_data = dict(fp_row) if fp_row else {}
        
        # Risque Marché
        cursor.execute("SELECT * FROM risque_marche ORDER BY date_analyse DESC LIMIT 1")
        rm_row = cursor.fetchone()
        rm_data = dict(rm_row) if rm_row else {}

    # CALCULATIONS BCEAO
    fp_calc = calculate_fonds_propres(fp_data)
    rm_calc = resolve_market_capital(rm_data)
    # RWA Opérationnel = APR de l'Approche Indicateur de Base (AIB, art. 301
    # du dispositif prudentiel BCEAO) - c'est la méthode réglementaire UEMOA
    # effectivement applicable, à la différence du BIC/CRR3 qui n'est qu'un
    # outil de pilotage interne calé sur le référentiel bâlois européen (cf.
    # libellé « BIC - CRR3 Pilotage interne » dans la synthèse du module
    # Risque Opérationnel). `apr_aib` est déjà exprimé en équivalent RWA
    # (K_IB × 12,5), comme le risque de marché (rwa_marche = capital requis
    # × 12,5) : pas de multiplicateur à réappliquer ici.
    rwa_operationnel = _calcul_aib_uemoa().apr_aib

    fp_detail = FondsPropresDetail(
        capital_ordinaire=fp_data.get("capital_ordinaire", 0.0),
        reserves=fp_data.get("reserves", 0.0),
        resultats_report=fp_data.get("resultats_report", 0.0),
        resultat_eligible=fp_data.get("resultat_eligible", 0.0),
        deductions_prud_cet1=fp_data.get("deductions_prud_cet1", 0.0),
        cet1=fp_calc["cet1"],
        instruments_at1=fp_data.get("instruments_at1", 0.0),
        primes_emission_at1=fp_data.get("primes_emission_at1", 0.0),
        deductions_prud_at1=fp_data.get("deductions_prud_at1", 0.0),
        at1=fp_calc["at1"],
        tier1=fp_calc["t1"],
        dettes_subordonnees_t2=fp_data.get("dettes_subordonnees_t2", 0.0),
        provisions_generales_t2=fp_data.get("provisions_generales_t2", 0.0),
        deductions_prud_t2=fp_data.get("deductions_prud_t2", 0.0),
        tier2=fp_calc["t2"],
        total_fp=fp_calc["total_capital"],
    )

    exposure_rows = [_normalize_row(item) for item in list_expositions()]

    gross_total = sum(float(row["gross_amount"]) for row in exposure_rows)
    ead_total = sum(float(row["ead"]) for row in exposure_rows)
    rwa_credit = sum(float(row["rwa"]) for row in exposure_rows)
    
    # RWA Total = Crédit + Marché + Opérationnel
    rwa_total = rwa_credit + rwa_operationnel + rm_calc["rwa_marche"]

    # Exigence de fonds propres du portefeuille : elle porte sur le RWA total
    # (crédit + marché + opérationnel), pas sur le seul risque de crédit, et
    # se distingue des fonds propres effectivement détenus.
    capital_requis = calculate_capital(rwa_total)
    # Exigence du seul risque de crédit : la tuile du dashboard l'affiche
    # séparément de l'exigence totale (clé "capital_requis").
    capital_credit = calculate_capital(rwa_credit)
    risk_ratio = safe_ratio(rwa_total, gross_total if gross_total > 0 else ead_total)
    
    default_gross_total = sum(
        float(row["gross_amount"])
        for row in exposure_rows
        if _is_defaulted_exposure(row["row"], str(row["category"]))
    )
    default_rate = safe_ratio(default_gross_total, gross_total)
    
    crm_gross = sum(
        float(row["gross_amount"]) * row["row"].crm_coverage_percent
        for row in exposure_rows
        if row["crm_mode"] != "Aucune"
    )
    residual_risk = max(gross_total - crm_gross, 0.0)
    covered_ratio = safe_ratio(crm_gross, gross_total)

    # Calculate Official Ratios (Solvency, Leverage, etc.)
    ratios = evaluate_ratios(rwa_total, fp_calc, gross_total)

    portfolio_rows = [
        PortfolioRow(
            id=row["row"].id,
            analysis_date=row["row"].analysis_date.isoformat(),
            counterparty=row["row"].counterparty.name,
            country=row["row"].counterparty.country,
            category=row["category"],
            rating=row["rating"],
            rating_band=row["rating_band"],
            crm_type=row["crm_mode"],
            gross_amount=round(float(row["gross_amount"]), 2),
            on_balance_exposure_amount=round(
                float(row["on_balance_exposure_amount"]),
                2,
            ),
            off_balance_exposure_amount=round(
                float(row["off_balance_exposure_amount"]),
                2,
            ),
            ead=round(float(row["ead"]), 2),
            rwa=round(float(row["rwa"]), 2),
            capital=round(float(row["capital"]), 2),
            **_crm_snapshot(row["row"], ead_xof=float(row["ead"])),
        )
        for row in exposure_rows
    ]

    category_buckets: dict[str, float] = defaultdict(float)
    rwa_category_buckets: dict[str, float] = defaultdict(float)
    country_buckets: dict[str, float] = defaultdict(float)
    crm_buckets: dict[str, float] = defaultdict(float)
    rating_buckets: dict[str, float] = defaultdict(float)
    for item in exposure_rows:
        row = item["row"]
        category = str(item["category"])
        # La repartition se lit par echelon : seize notations distinctes
        # feraient un graphique illisible, sans rapport avec la grille.
        rating = str(item["rating_band"])
        crm_mode = str(item["crm_mode"])
        category_buckets[category] += float(item["gross_amount"])
        rwa_category_buckets[category] += float(item["rwa"])
        country_buckets[_canonical_country_label(row.counterparty.country)] += float(item["rwa"])
        crm_buckets[crm_mode] += float(item["gross_amount"])
        rating_buckets[rating] += float(item["gross_amount"])

    # Create RWA Type distribution
    total_all_risks = rwa_credit + rm_calc["rwa_marche"] + rwa_operationnel
    rwa_type_distribution = [
        DistributionEntry(label="Crédit", amount=rwa_credit, percentage=(rwa_credit/total_all_risks*100) if total_all_risks > 0 else 0.0),
        DistributionEntry(label="Opérationnel", amount=rwa_operationnel, percentage=(rwa_operationnel/total_all_risks*100) if total_all_risks > 0 else 0.0),
        DistributionEntry(label="Marché", amount=rm_calc["rwa_marche"], percentage=(rm_calc["rwa_marche"]/total_all_risks*100) if total_all_risks > 0 else 0.0),
    ]

    category_distribution = _build_distribution_from_buckets(category_buckets)
    rwa_category_distribution = _build_distribution_from_buckets(rwa_category_buckets)
    crm_distribution = _build_distribution_from_buckets(crm_buckets, ordered_labels=_CRM_BUCKET_ORDER)
    rating_distribution = _build_distribution_from_buckets(rating_buckets)

    valuation_date = max((row["row"].analysis_date for row in exposure_rows), default=date.today())
    projection_base = rwa_total
    value_at_risk = _portfolio_value_at_risk(
        exposure_rows,
        gross_total=gross_total,
        rwa_total=rwa_total,
    )
    critical_incidents = _critical_incident_count(
        exposure_rows,
        default_rate=default_rate,
        gross_total=gross_total,
        rwa_total=rwa_total,
    )
    concentration_max = _max_grouped_share(
        exposure_rows,
        gross_total,
        key_name="counterparty",
        amount_name="gross_amount",
    )
    
    # Update metrics with the real ratios
    metrics = [
        _build_metric("encours", "Exposition totale brute", gross_total),
        _build_metric("risque_residuel", "Risque residuel", residual_risk),
        _build_metric("rwa", "RWA Total", rwa_total),
        _build_metric("capital", "Capital Total (FPE)", fp_calc["total_capital"]),
        _build_metric("capital_requis", "Exigence de fonds propres", capital_requis),
        _build_metric("capital_credit", "Exigence de fonds propres (crédit)", capital_credit),
        _build_metric("taux_risque", "Taux de risque", risk_ratio),
        _build_metric("taux_defaut", "Taux de defaut", default_rate),
        _build_metric("solvabilite", "Ratio Solvabilite Globale", ratios["solvency"]["value"] / 100.0),
        _build_metric("cet1_ratio", "Ratio CET1", ratios["cet1"]["value"] / 100.0),
        _build_metric("tier1_ratio", "Ratio Tier 1", ratios["tier1"]["value"] / 100.0),
        _build_metric("ratio_levier", "Ratio de Levier", ratios["leverage"]["value"] / 100.0),
        _build_metric("rwa_credit", "RWA Crédit", rwa_credit),
        # Exigence portee par le SEUL risque de credit. Sans elle, la tuile
        # « Capital requis (credit) » retombait sur l'exigence totale, marche
        # et operationnel compris : elle annoncait un besoin de fonds propres
        # de credit superieur a la realite.
        _build_metric("capital_credit", "Exigence de fonds propres (crédit)", calculate_capital(rwa_credit)),
        _build_metric("rwa_market", "RWA Marché", rm_calc["rwa_marche"]),
        _build_metric("rwa_op", "RWA Opérationnel", rwa_operationnel),
        _build_metric("crm", "Couverture CRM", covered_ratio),
        _build_metric("value_at_risk", "VaR globale", value_at_risk),
        _build_metric(
            "incidents_critiques",
            "Incidents critiques",
            float(critical_incidents),
        ),
        _build_metric(
            "concentration_max",
            "Concentration max",
            concentration_max,
        ),
    ]

    grouped_risk = {}
    for item in exposure_rows:
        row = item["row"]
        group_name = row.counterparty.name
        if group_name not in grouped_risk:
            grouped_risk[group_name] = {
                "country": row.counterparty.country if row.counterparty.country else "Non spécifié",
                "rating": item["rating"],
                "category": item["category"],
                "gross_amount": 0.0,
                "net_exposure": 0.0,
                "rwa_amount": 0.0,
            }
        grouped_risk[group_name]["gross_amount"] += float(item["gross_amount"])
        grouped_risk[group_name]["net_exposure"] += float(item["ead"])
        grouped_risk[group_name]["rwa_amount"] += float(item["rwa"])

    actual_top10 = []
    own_funds = fp_calc["total_capital"]
    for group_name, agg in grouped_risk.items():
        gross = agg["gross_amount"]
        net = agg["net_exposure"]
        rwa_agg = agg["rwa_amount"]
        ratio_fp = (net / own_funds * 100.0) if own_funds > 0 else 0.0

        # Les expositions souveraines ne sont pas considérées comme des grands risques
        category_norm = str(agg["category"]).lower().replace('é', 'e').replace('è', 'e')
        if "souverain" in category_norm:
            continue

        if ratio_fp >= 10.0:
            status = "Dans la norme"
            if ratio_fp > 25.0:
                status = "Dépassement"

            actual_top10.append(TopExposure(
                counterparty=group_name,
                sector=str(agg["category"]),
                country=str(agg["country"]),
                rating=str(agg["rating"]),
                exposure_amount=gross,
                net_exposure=net,
                rwa_amount=rwa_agg,
                fp_ratio=ratio_fp,
                status=status,
            ))
            
    actual_top10.sort(key=lambda x: x.net_exposure, reverse=True)
    top10_exposures = actual_top10[:10]
    grands_risques = actual_top10

    return DashboardSnapshot(
        metrics=metrics,
        fonds_propres=fp_detail,
        valuation_date=valuation_date.isoformat(),
        category_distribution=category_distribution,
        rwa_category_distribution=rwa_category_distribution,
        rwa_type_distribution=rwa_type_distribution,
        country_distribution=_build_country_distribution(country_buckets),
        crm_distribution=crm_distribution,
        rating_distribution=rating_distribution,
        rwa_projection=_build_projection(valuation_date, projection_base),
        portfolio_overview=portfolio_rows,
        top10_exposures=top10_exposures,
        grands_risques=grands_risques,
    )

def update_fonds_propres(update_data: FondsPropresUpdate) -> DashboardSnapshot:
    with database_manager.transaction() as conn:
        cursor = conn.cursor()
        
        # Obtenir l'ID existant ou en generer un nouveau
        cursor.execute("SELECT id, date_analyse FROM fonds_propres ORDER BY date_analyse DESC LIMIT 1")
        row = cursor.fetchone()
        
        now = datetime.now().isoformat()
        if row:
            # Mettre a jour l'existant
            cursor.execute('''
                UPDATE fonds_propres SET
                    capital_ordinaire = ?, reserves = ?, resultats_report = ?, resultat_eligible = ?, deductions_prud_cet1 = ?,
                    instruments_at1 = ?, primes_emission_at1 = ?, deductions_prud_at1 = ?,
                    dettes_subordonnees_t2 = ?, provisions_generales_t2 = ?, deductions_prud_t2 = ?,
                    modifie_le = ?
                WHERE id = ?
            ''', (
                update_data.capital_ordinaire, update_data.reserves, update_data.resultats_report,
                update_data.resultat_eligible, update_data.deductions_prud_cet1,
                update_data.instruments_at1, update_data.primes_emission_at1, update_data.deductions_prud_at1,
                update_data.dettes_subordonnees_t2, update_data.provisions_generales_t2, update_data.deductions_prud_t2,
                now, row['id']
            ))
        else:
            # Creer une nouvelle ligne
            import uuid
            new_id = str(uuid.uuid4())
            cursor.execute('''
                INSERT INTO fonds_propres (
                    id, date_analyse, capital_ordinaire, reserves, resultats_report, resultat_eligible, deductions_prud_cet1,
                    instruments_at1, primes_emission_at1, deductions_prud_at1,
                    dettes_subordonnees_t2, provisions_generales_t2, deductions_prud_t2,
                    cree_le, modifie_le
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                new_id, now, update_data.capital_ordinaire, update_data.reserves, update_data.resultats_report,
                update_data.resultat_eligible, update_data.deductions_prud_cet1,
                update_data.instruments_at1, update_data.primes_emission_at1, update_data.deductions_prud_at1,
                update_data.dettes_subordonnees_t2, update_data.provisions_generales_t2, update_data.deductions_prud_t2,
                now, now
            ))
        
        conn.commit()

    return get_dashboard_snapshot()


# Libellés des 11 postes des Fonds Propres Réglementaires, dans l'ordre
# attendu par le modèle Excel d'import. Doit rester synchronisé avec
# `_fpFields`/`_fpLabels` côté frontend
# (dashboard/widgets/dashboard_fonds_propres_import_dialog.dart) - chaque
# libellé précise son groupe (CET1/AT1/Tier 2) car "Réduction prudentielle"
# apparaît dans les trois groupes et doit rester non-ambigu.
FONDS_PROPRES_INPUT_FIELDS: tuple[tuple[str, str, str], ...] = (
    ("capital_ordinaire", "CET1", "Capital ordinaire"),
    ("reserves", "CET1", "Réserves"),
    ("resultats_report", "CET1", "Résultats en report"),
    ("resultat_eligible", "CET1", "Résultat éligible"),
    ("deductions_prud_cet1", "CET1", "Réduction prudentielle (CET1)"),
    ("instruments_at1", "AT1", "Instruments additionnels (AT1)"),
    ("primes_emission_at1", "AT1", "Primes d'émission (AT1)"),
    ("deductions_prud_at1", "AT1", "Réduction prudentielle (AT1)"),
    ("dettes_subordonnees_t2", "Tier 2", "Dettes subordonnées (Tier 2)"),
    ("provisions_generales_t2", "Tier 2", "Provisions générales (Tier 2)"),
    ("deductions_prud_t2", "Tier 2", "Réduction prudentielle (Tier 2)"),
)


def build_fonds_propres_import_template() -> bytes:
    """Génère un classeur Excel de modèle pour l'import des Fonds Propres
    Réglementaires (CET1 / AT1 / Tier 2).

    Une seule "photo" à la fois (pas de dimension année) : le modèle contient
    une ligne par poste, avec son groupe (CET1/AT1/Tier 2) et sa valeur -
    l'import remplace entièrement les fonds propres actuellement enregistrés,
    exactement comme le fait le formulaire "Mettre à jour".
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()

    BLUE_DARK = "1D4ED8"
    BLUE_LIGHT = "DBEAFE"
    BORDER_CLR = "CBD5E1"
    thin = Side(style="thin", color=BORDER_CLR)
    thin_b = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def center(wrap: bool = False) -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def left(wrap: bool = False) -> Alignment:
        return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

    ws = wb.active
    ws.title = "Fonds propres"
    ws.freeze_panes = "A3"

    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = "Modèle d'import - Fonds Propres Réglementaires (CET1 / AT1 / Tier 2)"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = fill(BLUE_DARK)
    title_cell.alignment = center()

    headers = [
        ("A", "Groupe", 12),
        ("B", "Poste", 36),
        ("C", "Valeur (FCFA)", 18),
    ]
    ws.row_dimensions[2].height = 22
    for col_letter, label, width in headers:
        c = ws[f"{col_letter}2"]
        c.value = label
        c.font = Font(bold=True, size=10, color=BLUE_DARK)
        c.fill = fill(BLUE_LIGHT)
        c.border = thin_b
        c.alignment = center()
        ws.column_dimensions[col_letter].width = width

    row_idx = 3
    for _, groupe, label in FONDS_PROPRES_INPUT_FIELDS:
        row_fill = fill("FFFFFF") if row_idx % 2 == 1 else fill("F8FAFC")
        c_groupe = ws.cell(row=row_idx, column=1, value=groupe)
        c_groupe.font = Font(size=10, bold=True, color=BLUE_DARK)
        c_groupe.fill = row_fill
        c_groupe.border = thin_b
        c_groupe.alignment = center()

        c_label = ws.cell(row=row_idx, column=2, value=label)
        c_label.font = Font(size=10)
        c_label.fill = row_fill
        c_label.border = thin_b
        c_label.alignment = left(wrap=True)

        c_val = ws.cell(row=row_idx, column=3, value=0)
        c_val.font = Font(size=10)
        c_val.fill = row_fill
        c_val.border = thin_b
        c_val.alignment = center()

        row_idx += 1

    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 78
    ws2.merge_cells("A1:A1")
    ws2["A1"].value = "Instructions - Import des Fonds Propres Réglementaires"
    ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = fill(BLUE_DARK)
    ws2["A1"].alignment = center()

    instructions = [
        "Ce modèle représente UNE photo des fonds propres (pas de notion "
        "d'année) : l'import remplace entièrement les valeurs actuellement "
        "enregistrées, comme le ferait le formulaire \"Mettre à jour\".",
        "Ne pas modifier les colonnes « Groupe » et « Poste ».",
        "Renseigner les montants (en FCFA) dans la colonne « Valeur ». Les "
        "postes de réduction prudentielle doivent être saisis en valeur "
        "positive (ils sont soustraits automatiquement).",
        "Les cases vides ou non numériques sont importées comme 0.",
    ]
    for ri, text in enumerate(instructions, start=2):
        c = ws2.cell(row=ri, column=1, value=text)
        c.font = Font(size=10)
        c.alignment = left(wrap=True)
        ws2.row_dimensions[ri].height = 30

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
