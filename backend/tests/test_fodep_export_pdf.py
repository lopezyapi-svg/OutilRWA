"""L'export PDF FODEP doit reproduire le classeur officiel, pas un autre document.

Deux bugs corriges par ce fichier :
1. build_fonds_propres_export plantait (AttributeError: 'MergedCell' object
   attribute 'value' is read-only) des qu'une denomination d'etablissement
   etait renseignee, car ADPE!D5 est dans la fusion C5:K5 et openpyxl refuse
   d'ecrire ailleurs que sur l'ancre d'une fusion.
2. openpyxl ne recalcule jamais les formules : apres load_workbook + save, les
   588 formules du classeur (dont les totaux EP03 et les ratios EP01/EP02)
   perdent leur valeur en cache et valent None. Sans un recalcul explicite
   (formula_engine.MoteurFormules), le PDF affiche des totaux vides.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from app.fodep import excel
from app.fodep.formula_engine import MoteurFormules
from app.fodep.xlsx_pdf import convertir_classeur_en_pdf


class _Etablissement:
    denomination = "BANQUE ATLANTIQUE CI"
    code_bceao = "CI0042"


def test_moteur_formules_reproduit_les_valeurs_calculees_par_excel():
    """Non-regression du risque n1 identifie a la conception : sans moteur de
    recalcul, un simple load_workbook + save d'openpyxl remet a None les 588
    formules du classeur (totaux EP03, ratios EP01/EP02 compris). Le moteur
    doit reproduire exactement les valeurs qu'Excel avait mises en cache dans
    le modele officiel, avant toute modification."""

    chemin = excel.get_matrice_officielle_path()
    assert chemin is not None, "Matrice FODEP officielle introuvable pour le test."

    wb_formules = openpyxl.load_workbook(chemin)
    wb_valeurs = openpyxl.load_workbook(chemin, data_only=True)

    moteur = MoteurFormules(wb_formules)
    ecarts: list[tuple[str, str, object, object]] = []

    for nom in wb_formules.sheetnames:
        for ligne in wb_formules[nom].iter_rows():
            for cellule in ligne:
                if not (isinstance(cellule.value, str) and cellule.value.startswith("=")):
                    continue
                attendu = wb_valeurs[nom][cellule.coordinate].value
                obtenu = moteur.valeur(nom, cellule.coordinate)
                if attendu is None:
                    continue
                if isinstance(attendu, (int, float)) and isinstance(obtenu, (int, float)):
                    if abs(float(attendu) - float(obtenu)) > max(1e-6, abs(float(attendu)) * 1e-9):
                        ecarts.append((nom, cellule.coordinate, attendu, obtenu))
                elif str(attendu) != str(obtenu):
                    ecarts.append((nom, cellule.coordinate, attendu, obtenu))

    assert not ecarts, f"Le moteur diverge d'Excel sur {len(ecarts)} cellule(s) : {ecarts[:10]}"
    assert not moteur.formules_ignorees, (
        f"Formules hors grammaire supportee : {moteur.formules_ignorees[:5]}"
    )


def test_export_avec_denomination_etablissement_ne_plante_pas():
    """ADPE!D5 (denomination) est dans la fusion C5:K5 : y ecrire directement
    levait AttributeError avant l'introduction de excel._ecrire."""

    contenu = excel.build_fonds_propres_export(
        "2026-06-30",
        {"fpi01": 1_000_000.0},
        etablissement=_Etablissement(),
        participations=[],
    )

    wb = openpyxl.load_workbook(__import__("io").BytesIO(contenu))
    assert wb["ADPE"]["C5"].value == "BANQUE ATLANTIQUE CI"


def test_convertir_classeur_en_pdf_produit_un_pdf_valide():
    """Le PDF genere doit etre un document PDF exploitable, pas un fichier vide
    ou tronque : c'est la garantie minimale avant tout controle de contenu."""

    contenu_xlsx = excel.build_fonds_propres_export(
        "2026-06-30",
        {"fpi01": 15_000_000_000.0, "fpi02": 2_500_000_000.0},
        etablissement=_Etablissement(),
        participations=[],
    )

    pdf = convertir_classeur_en_pdf(contenu_xlsx)

    assert pdf[:5] == b"%PDF-"
    assert len(pdf) > 10_000, "PDF anormalement petit pour un classeur de 44 onglets."


def test_conversion_pdf_echoue_proprement_sur_un_classeur_vide():
    """Un classeur sans le moindre onglet imprimable doit lever une erreur
    explicite plutot que produire un PDF vide silencieux."""

    from io import BytesIO

    wb = openpyxl.Workbook()  # une seule feuille visible, entierement vide

    tampon = BytesIO()
    wb.save(tampon)

    with pytest.raises(ValueError):
        convertir_classeur_en_pdf(tampon.getvalue())


def test_feuille_attestation_presente_et_rendue_dans_le_pdf():
    """L'onglet « ATTESTATION » (modele BCEAO) doit etre injecte dans le
    classeur avec les donnees etablissement + declarant, et figurer dans le
    PDF genere sans casser la conversion."""

    attestation = {
        "rens_prenoms_nom": "Amadou Diallo",
        "rens_fonction": "Directeur General",
        "rens_telephone": "+225 01 02 03",
        "rens_poste": "101",
        "rens_email": "a.diallo@banque.ci",
        "trans_prenoms_nom": "Awa Kone",
        "trans_fonction": "Responsable Reporting",
        "trans_telephone": "+225 04 05 06",
        "trans_poste": "202",
        "trans_email": "a.kone@banque.ci",
        "certif_nous_1": "Amadou Diallo",
        "certif_nous_2": "Awa Kone",
        "sign1_code": "SIG-001",
        "sign1_fonction": "Directeur General",
        "sign1_date": "30/06/2026",
        "sign1_image": "",
        "sign2_code": "SIG-002",
        "sign2_fonction": "Commissaire aux comptes",
        "sign2_date": "30/06/2026",
        "sign2_image": "",
    }

    contenu = excel.build_fonds_propres_export(
        "2026-06-30",
        {"fpi01": 1_000_000.0},
        etablissement=_Etablissement(),
        participations=[],
        attestation=attestation,
    )

    wb = openpyxl.load_workbook(BytesIO(contenu))
    assert "ATTESTATION" in wb.sheetnames
    ws = wb["ATTESTATION"]
    cellules = [str(c.value) for ligne in ws.iter_rows() for c in ligne if c.value]
    texte = " ".join(cellules)
    assert "ATTESTATION DE DECLARATION PRUDENTIELLE" in texte
    assert "BANQUE ATLANTIQUE CI" in texte
    assert "Amadou Diallo" in texte and "Awa Kone" in texte
    assert "SIG-001" in texte and "Commissaire aux comptes" in texte

    pdf = convertir_classeur_en_pdf(contenu)
    assert pdf[:5] == b"%PDF-"
    assert len(pdf) > 10_000


def test_les_totaux_du_classeur_suivent_calculations_pas_les_formules_du_modele():
    """Regression du bug « le PDF ne montre pas les valeurs de l'outil » : le
    modele officiel livre des formules de total partielles (FPI14 ne deduit que
    IM012, FPI26/FPI28/FPI39/FPI40 = 0 en dur, FPI29 = FPI22, FPI41 = FPI29).
    L'export doit ecrire par-dessus les totaux calcules par calculations.py."""

    from app.fodep.calculations import (
        calculer_exposition_levier,
        calculer_fonds_propres_detailles,
        calculer_produit_brut,
    )
    from app.fodep.dispru import FONDS_PROPRES_CODES

    postes = {c.code.lower(): 0.0 for c in FONDS_PROPRES_CODES}
    postes.update(
        {
            "fpi01": 20_000.0, "fpi03": 5_623.0, "fpi06": 14_211.0,
            "im012": -376.0, "pa163": -50.0,          # 2e deduction CET1 -> hors formule modele
            "fpi23": 1_000.0,                          # AT1 -> ignore par le modele
            "fpi30": 2_000.0,                          # T2  -> ignore par le modele
            "ro001": 49_350.0, "ro003": -12_761.0,
            "rl001": 537_474.0, "rl002": -376.0, "rl011": 35.5, "rl012": 21_527.0,
        }
    )
    totaux = calculer_fonds_propres_detailles(postes)
    totaux.update(calculer_produit_brut(postes))
    totaux.update(calculer_exposition_levier(postes))

    contenu = excel.build_fonds_propres_export(
        "2026-06-30", postes, etablissement=_Etablissement(),
        participations=[], totaux=totaux,
    )
    wb = openpyxl.load_workbook(BytesIO(contenu))
    moteur = MoteurFormules(wb)

    def v(feuille: str, cellule: str) -> float:
        return float(moteur.valeur(feuille, cellule))

    assert v("EP03", "C40") == pytest.approx(totaux["fpi22"])   # TOTAL CET1
    assert v("EP03", "C52") == pytest.approx(totaux["fpi29"])   # TOTAL T1 (avec AT1)
    assert v("EP03", "C69") == pytest.approx(totaux["fpi41"])   # FONDS PROPRES EFFECTIFS (avec T2)
    assert totaux["fpi29"] > totaux["fpi22"]                     # l'AT1 compte vraiment
    assert totaux["fpi41"] > totaux["fpi29"]                     # le T2 compte vraiment
    assert v("EP21", "D18") == pytest.approx(totaux["ro009"])   # produit brut total
    assert v("EP33", "C34") == pytest.approx(totaux["rl015"])   # exposition levier totale
    assert v("EP33", "C33") == pytest.approx(totaux["fpi29"])   # rappel T1
    assert v("EP02", "F14") == pytest.approx(totaux["fpi22"])
    assert v("EP02", "F16") == pytest.approx(totaux["fpi41"])


def test_export_ne_laisse_aucune_participation_du_specimen_dans_ep35():
    """Le modele officiel livre EP35 pre-rempli avec les participations d'un
    autre etablissement (SCIE, GIM UEMOA...). Sans registre pour l'arrete,
    l'onglet doit ressortir vide."""

    contenu = excel.build_fonds_propres_export(
        "2026-06-30", {"fpi01": 1_000_000.0},
        etablissement=_Etablissement(), participations=[],
    )
    ws = openpyxl.load_workbook(BytesIO(contenu))["EP35"]
    for r in range(12, 66):
        assert ws.cell(r, 2).value in (None, "")
        for c in (4, 5, 6):
            assert (ws.cell(r, c).value or 0) == 0
