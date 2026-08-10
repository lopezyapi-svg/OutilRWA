"""Tests de l'import Excel de l'historique VaR (POST /market/upload-var-history).

Le modèle Excel (2 onglets Obligations/Actions, format long : une ligne =
un titre à une date) doit produire des CSV lisibles par la couche de
données VaR (app.var_marche.portefeuille_data) sans intervention manuelle —
c'est le seul moyen, pour un utilisateur, d'alimenter la VaR Historique.
"""

from __future__ import annotations

import sys
from pathlib import Path
from unittest.mock import patch

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from fastapi.testclient import TestClient  # noqa: E402

from app.main import app  # noqa: E402
from app.market.routes import _build_var_history_template  # noqa: E402
from app.var_marche import portefeuille_data  # noqa: E402

client = TestClient(app)


def _uploader_modele(tmp_path):
    """Génère le modèle officiel et l'importe, isolé dans tmp_path (aucun
    portefeuille importé via l'application ne doit interférer)."""

    with patch.object(portefeuille_data, "app_data_root", return_value=tmp_path), patch.object(
        portefeuille_data, "_lire_valeur_marche_portefeuilles_sqlite", return_value=None
    ):
        xlsx = _build_var_history_template()
        reponse = client.post(
            "/market/upload-var-history",
            files={
                "file": (
                    "modele.xlsx",
                    xlsx,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    return reponse


def test_modele_officiel_structure_des_onglets():
    """Le classeur généré contient exactement les onglets attendus :
    Obligations, Actions (positions + historique au format long) et la
    feuille optionnelle Courbe des taux (facteur de risque, voie
    recommandée pour la VaR historique obligataire)."""

    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(_build_var_history_template()))
    assert wb.sheetnames == ["Obligations", "Actions", "Courbe des taux"]


def test_upload_modele_officiel_alimente_la_var(tmp_path):
    """Le modèle généré par l'application doit être ré-importable tel quel
    et rendre la VaR calculable — non-régression du bug où (1) les notes
    d'aide étaient relues comme des positions malformées et (2) les
    fréquences de coupon en toutes lettres (« Semestrielle ») faisaient
    échouer la lecture des positions."""

    reponse = _uploader_modele(tmp_path)
    assert reponse.status_code == 200
    corps = reponse.json()
    assert corps["obligations_positions"] == 2
    assert corps["obligations_historique"] == 6
    assert corps["actions_positions"] == 2
    assert corps["actions_historique"] == 6

    with patch.object(portefeuille_data, "app_data_root", return_value=tmp_path), patch.object(
        portefeuille_data, "_lire_valeur_marche_portefeuilles_sqlite", return_value=None
    ):
        portefeuille_data.invalider_cache_series()
        try:
            positions = portefeuille_data.charger_positions_obligations()
            assert len(positions) == 2
            # Fréquences en toutes lettres correctement normalisées (2, 1).
            frequences = {p.isin: p.frequence_coupon for p in positions}
            assert frequences["OAT-CI-2023-01"] == 2  # Semestrielle
            assert frequences["OAT-SN-2024-02"] == 1  # Annuelle

            serie_obligations = portefeuille_data.get_serie_pnl("obligations", 2, 1)
            assert serie_obligations.valeur_portefeuille > 0
            assert len(serie_obligations.pertes) == 2

            serie_actions = portefeuille_data.get_serie_pnl("actions", 2, 1)
            assert serie_actions.valeur_portefeuille > 0
            assert len(serie_actions.pertes) == 2
        finally:
            portefeuille_data.invalider_cache_series()


def test_upload_position_sans_historique_est_conservee(tmp_path):
    """Une ligne sans Date ni Prix (position connue, pas encore de série de
    prix) doit être conservée comme position mais ne pas alimenter
    l'historique — permet l'import progressif d'un portefeuille."""

    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Obligations"
    ws.append(
        [
            "ID Titre", "Emetteur", "Devise", "Valeur nominale unitaire",
            "Coupon (%)", "Fréquence de paiement des intérêts",
            "Date d'émission", "Date d'échéance", "quantités",
            "Date", "Prix de marché (%)",
        ]
    )
    ws.append(["OBSEUL", "Etat Test", "XOF", 10000, 5.0, "Annuelle", "2020-01-01", "2030-01-01", 1000, None, None])
    ws2 = wb.create_sheet("Actions")
    ws2.append(["ID Instrument", "Émetteur / Société", "Secteur", "Quantité", "Date", "Cours de clôture"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    with patch.object(portefeuille_data, "app_data_root", return_value=tmp_path), patch.object(
        portefeuille_data, "_lire_valeur_marche_portefeuilles_sqlite", return_value=None
    ):
        reponse = client.post(
            "/market/upload-var-history",
            files={
                "file": (
                    "partiel.xlsx",
                    buf.read(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert reponse.status_code == 200
        corps = reponse.json()
        assert corps["obligations_positions"] == 1
        assert corps["obligations_historique"] == 0

        portefeuille_data.invalider_cache_series()
        try:
            positions = portefeuille_data.charger_positions_obligations()
            assert [p.isin for p in positions] == ["OBSEUL"]
            historique = portefeuille_data.charger_historique_prix_obligations()
            assert historique == {}
        finally:
            portefeuille_data.invalider_cache_series()


def test_import_courbe_des_taux_remplace_historique(tmp_path):
    """La feuille optionnelle « Courbe des taux » alimente historique_taux.csv
    (format long date/maturité/taux lu par la couche VaR) et remplace
    l'historique existant."""

    reponse = _uploader_modele(tmp_path)
    assert reponse.status_code == 200
    corps = reponse.json()
    # Le modèle officiel contient 2 dates d'exemple sur la feuille courbe.
    assert corps["courbe_taux_jours"] == 2

    with patch.object(portefeuille_data, "app_data_root", return_value=tmp_path), patch.object(
        portefeuille_data, "_lire_valeur_marche_portefeuilles_sqlite", return_value=None
    ):
        portefeuille_data.invalider_cache_series()
        try:
            courbes = portefeuille_data.charger_historique_taux()
            assert len(courbes) == 2
            for points in courbes.values():
                # 4 maturités par date dans l'exemple, triées croissantes.
                assert len(points) == 4
                assert points == sorted(points)
        finally:
            portefeuille_data.invalider_cache_series()


def test_actualisation_courbe_accumule_les_dates(tmp_path, monkeypatch):
    """_ecrire_historique_taux EMPILE les dates (une par actualisation) au
    lieu d'écraser : réactualiser le même jour reste idempotent, un jour
    nouveau augmente la profondeur — c'est ce qui permet à la VaR historique
    de se débloquer avec le temps."""

    from app.var_marche.routes import _ecrire_historique_taux

    monkeypatch.setattr(portefeuille_data, "app_data_root", lambda: tmp_path)

    points_j1 = [
        {"maturite_annees": 1, "taux_pct": 5.5},
        {"maturite_annees": 5, "taux_pct": 6.0},
    ]
    points_j2 = [
        {"maturite_annees": 1, "taux_pct": 5.6},
        {"maturite_annees": 5, "taux_pct": 6.1},
    ]

    _ecrire_historique_taux(points_j1, "2026-07-16")
    _ecrire_historique_taux(points_j2, "2026-07-17")
    # Réactualisation du même jour avec des taux corrigés : remplace le jour,
    # sans dupliquer ni toucher aux autres dates.
    _ecrire_historique_taux(
        [{"maturite_annees": 1, "taux_pct": 5.65}], "2026-07-17"
    )

    contenu = (tmp_path / "historique_taux.csv").read_text(encoding="utf-8")
    lignes = [l for l in contenu.splitlines() if l and not l.startswith("date;")]
    assert len(lignes) == 3  # 2 points du 16 + 1 point corrigé du 17
    assert sum(1 for l in lignes if l.startswith("2026-07-16;")) == 2
    assert sum(1 for l in lignes if l.startswith("2026-07-17;")) == 1
    assert any("5.65" in l for l in lignes if l.startswith("2026-07-17;"))

    with patch.object(
        portefeuille_data, "_lire_valeur_marche_portefeuilles_sqlite", return_value=None
    ):
        portefeuille_data.invalider_cache_series()
        try:
            courbes = portefeuille_data.charger_historique_taux()
            assert len(courbes) == 2  # 2 dates distinctes accumulées
        finally:
            portefeuille_data.invalider_cache_series()
