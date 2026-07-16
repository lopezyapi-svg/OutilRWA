"""Nouveau pipeline d'import Excel optimisé pour SQLite."""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
import json
import logging
from time import perf_counter
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from app.core.excel_repository import (
    EXPECTED_COLUMNS_BY_SHEET,
    EXPECTED_SHEETS,
    OPTIONAL_COLUMNS_BY_SHEET,
    excel_repository,
)
from app.core.runtime_paths import seed_data_path
from app.validators.excel_import_validator import (
    IMPORT_SHEET_SPECS,
    ExcelImportValidationError,
    build_excel_import_spec,
    inspect_workbook_structure,
    read_sheet_headers,
)
from database.connection import database_manager, utcnow_iso
from database.repositories.import_repository import import_repository

# Options figées réutilisées pour les listes déroulantes du modèle Excel.
# Importées directement du moteur de calcul partagé (build_exposure_record /
# resolve_category / coerce_bank_institution_case...) pour garantir qu'une
# valeur choisie dans le classeur est TOUJOURS reconnue par le calcul —
# exactement comme les options proposées par le formulaire "Ajouter une
# exposition" (frontend/lib/modules/expositions/models/exposition_models.dart).
from database.services.rwa_calculation_service import (
    BANK_INSTITUTION_ELIGIBLE_CATEGORIES_CASE,
    BANK_INSTITUTION_EQUIVALENT_RULES_CASE,
    BANK_INSTITUTION_WEAK_PRUDENTIAL_CASE,
    CATEGORY_OPTIONS,
    OFF_BALANCE_RISK_LEVEL_OPTIONS,
    OTHER_ASSET_TYPE_OPTIONS,
    SOVEREIGN_NO_SPECIAL_CASE,
    SOVEREIGN_ZERO_WEIGHT_SPECIAL_CASES,
)

logger = logging.getLogger(__name__)

# ── Listes d'options pour les colonnes à choix du modèle Excel ───────────────

# Toutes les catégories d'exposition (libellé prudentiel), dans l'ordre (a)..(k).
CATEGORY_PRUDENTIAL_LABELS: tuple[str, ...] = tuple(
    option["prudential"] for option in CATEGORY_OPTIONS
)
# Sous-ensemble de catégories acceptées comme "Catégorie du garant" en CRM non
# financée (un garant ne peut pas être de catégorie immobilier/défaut) — mêmes
# codes que guarantorEligibleCategoryCodes côté frontend.
_GUARANTOR_ELIGIBLE_CATEGORY_CODES = ("a", "b", "c", "d", "e", "f", "k")
GUARANTOR_CATEGORY_OPTIONS: tuple[str, ...] = tuple(
    option["prudential"]
    for option in CATEGORY_OPTIONS
    if option["code"] in _GUARANTOR_ELIGIBLE_CATEGORY_CODES
)

# Libellé prudentiel associé à chaque code de catégorie (a, b, c...) — sert à
# construire les listes déroulantes en cascade ci-dessous : le choix de la
# catégorie détermine les options valides des colonnes qui n'ont de sens que
# pour certaines catégories, exactement comme le formulaire "Ajouter une
# exposition" masque ces mêmes champs selon la catégorie choisie (voir
# frontend/lib/modules/expositions/widgets/exposure_form_card.dart,
# _handleCategoryChanged et les getters _is*Category).
CATEGORY_PRUDENTIAL_BY_CODE: dict[str, str] = {
    option["code"]: option["prudential"] for option in CATEGORY_OPTIONS
}

CRM_TYPE_OPTIONS: tuple[str, ...] = ("Aucune", "CRM financee", "CRM non financee")

# Devises supportées par le convertisseur partagé (voir _CURRENCY_RATES_IN_XAF
# dans app/core/calculations.py et currencyRatesInXaf côté frontend).
CURRENCY_OPTIONS: tuple[str, ...] = ("XOF", "EUR", "USD")

RATING_OPTIONS: tuple[str, ...] = (
    "AAA", "AA+", "AA", "AA-", "A+", "A", "A-",
    "BBB+", "BBB", "BBB-", "BB+", "BB", "BB-",
    "B+", "B", "B-", "< B-", "Non noté",
)

# Notation du collatéral/émetteur en CRM financée : mêmes valeurs que
# financedCrmDebtRatings côté frontend (pas de B+/B/B- isolés).
FINANCED_CRM_DEBT_RATING_OPTIONS: tuple[str, ...] = (
    "AAA", "AA+", "AA", "AA-", "A+", "A", "A-",
    "BBB+", "BBB", "BBB-", "BB+", "BB", "BB-",
    "< B-", "Non noté",
)

FINANCED_CRM_ISSUER_ROLE_OPTIONS: tuple[str, ...] = (
    "emprunteur souverain",
    "autre émetteur",
)

FINANCED_CRM_MATURITY_BUCKET_OPTIONS: tuple[str, ...] = (
    "<=1 an", "1-3 ans", "3-5 ans", "5-10 ans", ">10 ans",
)

FINANCED_CRM_COLLATERAL_TYPE_OPTIONS: tuple[str, ...] = (
    "Liquidités dans la même devise",
    "Liquidités dans une devise différente",
    "Or",
    "Titre de dette souverain",
    "Titre non noté émis par un État de l UMOA",
    "Titre de dette émis par un autre émetteur",
    "Titre garanti par un agent agréé par la BRVM",
    "Titre bancaire non noté",
    "Action de l indice BRVM 10",
    "Action d un indice principal reconnu",
    "Autre action cotée à la BRVM ou sur une bourse reconnue",
    "Obligation convertible en action",
    "OPCVM / FI",
    "Panier d actifs",
    "Autre sûreté non éligible",
)

FINANCED_CRM_OPCVM_HAIRCUT_OPTIONS: tuple[float, ...] = (
    0.0, 0.005, 0.01, 0.02, 0.03, 0.04, 0.06, 0.09, 0.12, 0.15, 0.20, 0.30,
)

SOVEREIGN_SPECIAL_CASE_OPTIONS: tuple[str, ...] = (
    *SOVEREIGN_ZERO_WEIGHT_SPECIAL_CASES,
    SOVEREIGN_NO_SPECIAL_CASE,
)

SOVEREIGN_OCE_NOTE_OPTIONS: tuple[str, ...] = ("0", "1", "2", "3", "4", "5", "6", "7")

BANK_INSTITUTION_CASE_OPTIONS: tuple[str, ...] = (
    BANK_INSTITUTION_EQUIVALENT_RULES_CASE,
    BANK_INSTITUTION_WEAK_PRUDENTIAL_CASE,
    BANK_INSTITUTION_ELIGIBLE_CATEGORIES_CASE,
)

DEFAULTED_EXPOSURE_INITIAL_RISK_WEIGHT_OPTIONS: tuple[float, ...] = (
    0.2, 0.35, 0.5, 0.75, 1.0, 1.5, 2.5,
)

# Explication humaine des clés techniques Cas_institution_bancaire (utilisée
# uniquement pour l'affichage dans l'onglet "Listes de référence" — la
# valeur importée reste la clé technique elle-même, seule forme reconnue
# sans ambiguïté par coerce_bank_institution_case côté calcul).
BANK_INSTITUTION_CASE_LABELS: dict[str, str] = {
    BANK_INSTITUTION_EQUIVALENT_RULES_CASE: (
        "Banque hors UMOA soumise à des règles équivalentes à celles de l'UMOA"
    ),
    BANK_INSTITUTION_WEAK_PRUDENTIAL_CASE: "Banque en difficulté prudentielle",
    BANK_INSTITUTION_ELIGIBLE_CATEGORIES_CASE: (
        "Institution faisant partie des catégories éligibles (voir notation)"
    ),
}

OUI_NON_OPTIONS: tuple[str, ...] = ("Oui", "Non")

# Valeur affichée dans les listes déroulantes en cascade lorsque la catégorie
# choisie sur la ligne ne concerne pas la colonne (équivalent au masquage du
# champ dans le formulaire "Ajouter une exposition"). Cette valeur n'est
# jamais interprétée par le calcul : une cellule laissée sur cette valeur est
# traitée exactement comme une cellule vide.
NOT_APPLICABLE_OPTION: tuple[str, ...] = ("(Sans objet pour cette catégorie)",)

# Référentiel pays (mêmes libellés que worldCountries côté frontend).
WORLD_COUNTRY_OPTIONS: tuple[str, ...] = (
    "Afghanistan", "Afrique du Sud", "Albanie", "Algerie", "Allemagne", "Andorre",
    "Angola", "Antigua-et-Barbuda", "Arabie saoudite", "Argentine", "Armenie",
    "Australie", "Autriche", "Azerbaidjan", "Bahamas", "Bahrein", "Bangladesh",
    "Barbade", "Belgique", "Belize", "Benin", "Bhoutan", "Bielorussie", "Birmanie",
    "Bolivie", "Bosnie-Herzegovine", "Botswana", "Bresil", "Brunei", "Bulgarie",
    "Burkina Faso", "Burundi", "Cambodge", "Cameroun", "Canada", "Cap-Vert",
    "Centrafrique", "Chili", "Chine", "Chypre", "Colombie", "Comores", "Congo",
    "Coree du Nord", "Coree du Sud", "Costa Rica", "Cote d'Ivoire", "Croatie",
    "Cuba", "Danemark", "Djibouti", "Dominique", "Egypte", "Emirats arabes unis",
    "Equateur", "Erythree", "Espagne", "Estonie", "Eswatini", "Etats-Unis",
    "Ethiopie", "Fidji", "Finlande", "France", "Gabon", "Gambie", "Georgie",
    "Ghana", "Grece", "Grenade", "Guatemala", "Guinee", "Guinee-Bissau",
    "Guinee equatoriale", "Guyana", "Haiti", "Honduras", "Hongrie",
    "Iles Marshall", "Iles Salomon", "Inde", "Indonesie", "Irak", "Iran",
    "Irlande", "Islande", "Israel", "Italie", "Jamaïque", "Japon", "Jordanie",
    "Kazakhstan", "Kenya", "Kirghizistan", "Kiribati", "Kosovo", "Koweit",
    "Laos", "Lesotho", "Lettonie", "Liban", "Liberia", "Libye", "Liechtenstein",
    "Lituanie", "Luxembourg", "Macedoine du Nord", "Madagascar", "Malaisie",
    "Malawi", "Maldives", "Mali", "Malte", "Maroc", "Maurice", "Mauritanie",
    "Mexique", "Micronesie", "Moldavie", "Monaco", "Mongolie", "Montenegro",
    "Mozambique", "Namibie", "Nauru", "Nepal", "Nicaragua", "Niger", "Nigeria",
    "Norvege", "Nouvelle-Zelande", "Oman", "Ouganda", "Ouzbekistan", "Pakistan",
    "Palaos", "Palestine", "Panama", "Papouasie-Nouvelle-Guinee", "Paraguay",
    "Pays-Bas", "Perou", "Philippines", "Pologne", "Portugal", "Qatar",
    "Republique centrafricaine", "Republique democratique du Congo",
    "Republique dominicaine", "Republique tcheque", "Roumanie", "Royaume-Uni",
    "Russie", "Rwanda", "Saint-Christophe-et-Nieves", "Sainte-Lucie",
    "Saint-Marin", "Saint-Vincent-et-les-Grenadines", "Salvador", "Samoa",
    "Sao Tome-et-Principe", "Senegal", "Serbie", "Seychelles", "Sierra Leone",
    "Singapour", "Slovaquie", "Slovenie", "Somalie", "Soudan", "Soudan du Sud",
    "Sri Lanka", "Suede", "Suisse", "Suriname", "Syrie", "Tadjikistan",
    "Taiwan", "Tanzanie", "Tchad", "Thailande", "Timor oriental", "Togo",
    "Tonga", "Trinite-et-Tobago", "Tunisie", "Turkmenistan", "Turquie",
    "Tuvalu", "Ukraine", "Uruguay", "Vanuatu", "Vatican", "Venezuela",
    "Vietnam", "Yemen", "Zambie", "Zimbabwe",
)

# Colonnes à choix fixes -> liste d'options associée. Les colonnes de type
# "oui/non" (cf. IMPORT_SHEET_SPECS) sont traitées séparément et n'ont pas
# besoin d'être répétées ici.
FIXED_OPTIONS_BY_COLUMN: dict[str, tuple] = {
    "Catégorie d'exposition": CATEGORY_PRUDENTIAL_LABELS,
    "Type_CRM": CRM_TYPE_OPTIONS,
    "Devise": CURRENCY_OPTIONS,
    "Devise_Collatéral": CURRENCY_OPTIONS,
    "Notation_externe_contrepartie": RATING_OPTIONS,
    "Notation_externe_pays": RATING_OPTIONS,
    "Pays_contrepartie": WORLD_COUNTRY_OPTIONS,
    "Niveau de risque HB": OFF_BALANCE_RISK_LEVEL_OPTIONS,
    "Cas_particulier_souverain": SOVEREIGN_SPECIAL_CASE_OPTIONS,
    "Souverain_note_OCE": SOVEREIGN_OCE_NOTE_OPTIONS,
    "Cas_institution_bancaire": BANK_INSTITUTION_CASE_OPTIONS,
    "Type_autre_actif": OTHER_ASSET_TYPE_OPTIONS,
    "Ponderation_initiale_avant_defaut": DEFAULTED_EXPOSURE_INITIAL_RISK_WEIGHT_OPTIONS,
    "Catégorie du garant": GUARANTOR_CATEGORY_OPTIONS,
    "Note_garant": RATING_OPTIONS,
    "Pays_garant": WORLD_COUNTRY_OPTIONS,
    "Note_pays_garant": RATING_OPTIONS,
    "Type_emetteur": FINANCED_CRM_ISSUER_ROLE_OPTIONS,
    "Notation": FINANCED_CRM_DEBT_RATING_OPTIONS,
    "Maturite": FINANCED_CRM_MATURITY_BUCKET_OPTIONS,
    "Type_Collatéral": FINANCED_CRM_COLLATERAL_TYPE_OPTIONS,
    "Decote_OPCVM_max": FINANCED_CRM_OPCVM_HAIRCUT_OPTIONS,
}

# Colonnes de "Template données" dont les options valides dépendent de la
# catégorie choisie sur la même ligne (colonne "Catégorie d'exposition") :
# pour chaque colonne, les codes de catégorie qui l'activent et la liste
# d'options à proposer dans ce cas. Sur les autres catégories, la colonne ne
# propose que "(Sans objet pour cette catégorie)" — comme le formulaire
# "Ajouter une exposition" masque ces mêmes champs selon la catégorie.
# Remplace, pour ces colonnes, l'entrée correspondante de
# FIXED_OPTIONS_BY_COLUMN (conservée telle quelle pour l'onglet "Listes de
# référence", qui reste un simple rappel indépendant de la catégorie).
CATEGORY_DEPENDENT_COLUMNS: dict[str, tuple[tuple[str, ...], tuple]] = {
    "Cas_particulier_souverain": (("a",), SOVEREIGN_SPECIAL_CASE_OPTIONS),
    "Souverain_ponderation_pref_nulle": (("a",), OUI_NON_OPTIONS),
    "Souverain_OCE_etabli": (("a",), OUI_NON_OPTIONS),
    "Souverain_note_OCE": (("a",), SOVEREIGN_OCE_NOTE_OPTIONS),
    "Organisme_public_cas_UEMOA_FCFA": (("b",), OUI_NON_OPTIONS),
    "Organisme_public_activite_non_publique": (("b",), OUI_NON_OPTIONS),
    "BMD_cas_haute_qualite": (("c",), OUI_NON_OPTIONS),
    "BMD_cas_UEMOA_FCFA": (("c",), OUI_NON_OPTIONS),
    "BMD_criteres_UEMOA_respectes": (("c",), OUI_NON_OPTIONS),
    "BMD_institution_listee_FCFA": (("c",), OUI_NON_OPTIONS),
    "Cas_institution_bancaire": (("d",), BANK_INSTITUTION_CASE_OPTIONS),
    "Type_autre_actif": (("k",), OTHER_ASSET_TYPE_OPTIONS),
    "Clientele_detail_criteres_respectes": (("f",), OUI_NON_OPTIONS),
    "Immobilier_residentiel_eligible": (("g",), OUI_NON_OPTIONS),
    "Immobilier_commercial_eligible": (("h",), OUI_NON_OPTIONS),
    "Ponderation_initiale_avant_defaut": (
        ("i",),
        DEFAULTED_EXPOSURE_INITIAL_RISK_WEIGHT_OPTIONS,
    ),
    "Defaut_pret_immo_residentiel": (("i",), OUI_NON_OPTIONS),
    "Defaut_provision_min_20pct": (("i",), OUI_NON_OPTIONS),
    "Entreprise_depasse_seuil_degradation_BCEAO": (
        ("e", "f", "h"),
        OUI_NON_OPTIONS,
    ),
    "Entreprise_procedure_prudentielle": (("e", "f", "h"), OUI_NON_OPTIONS),
    "Entreprise_investissement_hors_loi_bancaire": (
        ("e", "f", "h"),
        OUI_NON_OPTIONS,
    ),
}

# Nombre de lignes de saisie couvertes par les listes déroulantes (au-delà de
# ce nombre de lignes, l'utilisateur peut toujours copier la validation avec
# Excel — poignée de recopie — donc une valeur généreuse suffit).
_VALIDATION_ROW_COUNT = 2000


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _create_import_run(
    connection,
    *,
    source_type: str,
    source_name: str,
    total_rows: int,
    details: dict[str, Any] | None = None,
) -> int:
    cursor = connection.execute(
        """
        INSERT INTO import_runs(source_type, source_name, imported_at, total_rows, imported_rows, rejected_rows, status, details_json)
        VALUES (?, ?, ?, ?, 0, 0, 'running', ?)
        """,
        (
            source_type,
            source_name,
            utcnow_iso(),
            total_rows,
            json.dumps(details or {}, ensure_ascii=False),
        ),
    )
    return int(cursor.lastrowid)


def _finish_import_run(
    connection,
    run_id: int,
    *,
    imported_rows: int,
    rejected_rows: int,
    details: dict[str, Any],
) -> None:
    connection.execute(
        """
        UPDATE import_runs
        SET imported_rows = ?, rejected_rows = ?, status = 'success', details_json = ?
        WHERE id = ?
        """,
        (
            imported_rows,
            rejected_rows,
            json.dumps(details, ensure_ascii=False),
            run_id,
        ),
    )


def _write_metadata(connection, key: str, value: str) -> None:
    connection.execute(
        """
        INSERT INTO metadonnees_app(cle, valeur)
        VALUES(?, ?)
        ON CONFLICT(cle) DO UPDATE SET valeur = excluded.valeur
        """,
        (key, value),
    )


@dataclass(slots=True)
class ParsedImportBundle:
    exposure_records: list[dict[str, Any]]
    off_balance_records: list[dict[str, Any]]
    rows_read: int
    valid_rows: int
    rejected_rows: int
    errors: list[dict[str, Any]] = field(default_factory=list)
    rows_read_by_sheet: dict[str, int] = field(default_factory=dict)


class ImportProfiler:
    def __init__(self, *, source_name: str) -> None:
        self.source_name = source_name
        self._started_at = perf_counter()
        self.timings_ms: dict[str, float] = {}

    def measure(self, step_name: str):
        start = perf_counter()

        class _StepContext:
            def __enter__(inner_self):
                return inner_self

            def __exit__(inner_self, exc_type, exc, tb):
                elapsed = round((perf_counter() - start) * 1000, 2)
                self.timings_ms[step_name] = elapsed
                logger.info(
                    "Import Excel [%s] - %s: %.2f ms",
                    self.source_name,
                    step_name,
                    elapsed,
                )
                return False

        return _StepContext()

    @property
    def total_ms(self) -> float:
        return round((perf_counter() - self._started_at) * 1000, 2)


class ExcelImportService:
    """Service d'import Excel moderne, rapide et batch."""

    def get_import_spec(self) -> dict[str, Any]:
        return build_excel_import_spec()

    def build_template_workbook(self) -> bytes:
        template_candidates = (seed_data_path("modele_import_rwa.xlsx"),)
        for candidate in template_candidates:
            if candidate.is_file():
                return candidate.read_bytes()

        return self._build_fallback_template_workbook()

    def _build_fallback_template_workbook(self) -> bytes:
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        # ── Palette (identique aux autres modèles Pertes Opérationnelles /
        # BIC / Fonds Propres, pour une expérience visuelle cohérente) ──────
        BLUE_DARK = "1D4ED8"
        BLUE_LIGHT = "DBEAFE"
        GREY_HEADER = "F1F5F9"
        BORDER_CLR = "CBD5E1"

        thin = Side(style="thin", color=BORDER_CLR)
        thin_b = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hdr_fill(hex_color: str) -> PatternFill:
            return PatternFill("solid", fgColor=hex_color)

        def center(wrap: bool = False) -> Alignment:
            return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

        def left(wrap: bool = False) -> Alignment:
            return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

        column_specs_by_sheet = {spec.name: spec for spec in IMPORT_SHEET_SPECS}
        value_type_by_column_by_sheet = {
            spec.name: {
                column.name: column.value_type
                for column in (spec.required_columns + spec.optional_columns)
            }
            for spec in IMPORT_SHEET_SPECS
        }

        # ── Feuille technique cachée : source des listes déroulantes ────────
        # Une colonne par liste d'options (dédupliquée par identité d'objet,
        # donc partagée entre toutes les colonnes qui utilisent la même
        # liste, ex: Notation_externe_contrepartie / Notation_externe_pays /
        # Note_garant / Note_pays_garant partagent RATING_OPTIONS). Cachée
        # (pas supprimée) pour ne pas encombrer l'utilisateur tout en restant
        # inspectable si besoin.
        validation_lists_sheet = workbook.create_sheet("Listes (validation)")
        _seen_option_ids: dict[int, str] = {}
        _next_validation_col = 1

        def _validation_range_for(options: tuple) -> str:
            """Écrit `options` dans une colonne de la feuille technique (si pas
            déjà fait) et renvoie la référence de plage Excel correspondante."""

            nonlocal _next_validation_col
            key = id(options)
            existing = _seen_option_ids.get(key)
            if existing is not None:
                return existing

            col_index = _next_validation_col
            _next_validation_col += 1
            col_letter = validation_lists_sheet.cell(row=1, column=col_index).column_letter
            for row_offset, value in enumerate(options, start=2):
                validation_lists_sheet.cell(row=row_offset, column=col_index, value=value)
            last_row = 1 + len(options)
            cell_range = (
                f"'Listes (validation)'!${col_letter}$2:${col_letter}${last_row}"
            )
            _seen_option_ids[key] = cell_range
            return cell_range

        oui_non_range = _validation_range_for(OUI_NON_OPTIONS)
        # Pré-calcule la plage de chaque liste utilisée par au moins une
        # colonne, pour que l'ordre des colonnes dans la feuille technique
        # reste stable (référentiels d'abord, oui/non en premier).
        for options in FIXED_OPTIONS_BY_COLUMN.values():
            _validation_range_for(options)

        # ── Noms courts (plages nommées) pour les listes en cascade ────────
        # Les formules de validation en cascade (INDIRECT + IF selon la
        # catégorie) doivent rester courtes : Excel limite la longueur des
        # formules de validation. On donne donc un nom court à chaque plage
        # utilisée par une colonne dépendant de la catégorie, plutôt que d'y
        # répéter la référence complète "'Listes (validation)'!$X$2:$X$9".
        _short_name_by_range: dict[str, str] = {}
        _next_short_name_index = 1

        def _short_name_for(cell_range: str) -> str:
            nonlocal _next_short_name_index
            existing = _short_name_by_range.get(cell_range)
            if existing is not None:
                return existing
            name = f"ListeCascade{_next_short_name_index}"
            _next_short_name_index += 1
            workbook.defined_names[name] = DefinedName(name, attr_text=cell_range)
            _short_name_by_range[cell_range] = name
            return name

        not_applicable_range = _validation_range_for(NOT_APPLICABLE_OPTION)
        not_applicable_name = _short_name_for(not_applicable_range)
        for _codes, _cascade_options in CATEGORY_DEPENDENT_COLUMNS.values():
            _short_name_for(_validation_range_for(_cascade_options))

        validation_lists_sheet.sheet_state = "hidden"

        # ── Feuille "Instructions" ──────────────────────────────────────────
        instructions = workbook.create_sheet("Instructions")
        instructions.column_dimensions["A"].width = 26
        instructions.column_dimensions["B"].width = 90
        instructions.merge_cells("A1:B1")
        instructions["A1"].value = "Instructions — Import des expositions (Risque de crédit)"
        instructions["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        instructions["A1"].fill = hdr_fill(BLUE_DARK)
        instructions["A1"].alignment = center()
        instructions.row_dimensions[1].height = 32

        instruction_rows = [
            (
                "Principe",
                "Ce fichier reprend uniquement les données brutes que vous saisiriez dans "
                "le formulaire « Ajouter une exposition ». Les colonnes calculées "
                "(pondération finale, EAD, RWA, capital, maturités déduites...) ne "
                "figurent pas ici : elles sont recalculées automatiquement à l'import, "
                "avec le même moteur de calcul prudentiel que la saisie manuelle.",
            ),
            (
                "Couleurs des en-têtes",
                "Bleu = colonne obligatoire. Gris = colonne optionnelle (peut rester vide "
                "si elle ne concerne pas la ligne). Survolez un en-tête pour voir son aide.",
            ),
            (
                "Lignes d'exemple",
                "Les feuilles de saisie contiennent quelques lignes d'exemple (IDs "
                "« EXP-EX-00x ») illustrant différentes catégories et cas de CRM. "
                "Supprimez-les avant d'importer vos propres données.",
            ),
            ("Template données", "Une ligne par exposition."),
            (
                "CRM_non_financee",
                "Une ligne par exposition couverte par une garantie "
                "(Type_CRM = CRM non financee sur la ligne correspondante de Template données).",
            ),
            (
                "CRM_financée",
                "Une ligne par exposition couverte par un collatéral financier "
                "(Type_CRM = CRM financee sur la ligne correspondante de Template données).",
            ),
            (
                "Listes déroulantes",
                "Chaque colonne à choix (catégorie, notation, devise, pays, "
                "type de CRM, cas particuliers, champs Oui/Non...) propose une "
                "liste déroulante : cliquez sur la cellule puis sur la flèche "
                "qui apparaît pour choisir une valeur reconnue par le calcul. "
                "Le détail des valeurs figure aussi dans l'onglet « Listes de "
                "référence ».",
            ),
            (
                "Listes en cascade",
                "Comme dans le formulaire « Ajouter une exposition », certaines "
                "listes dépendent de la « Catégorie d'exposition » choisie sur la "
                "même ligne (ex. Cas_institution_bancaire ne propose des options "
                "que si la catégorie est Institutions financières, Type_autre_actif "
                "seulement pour Autres actifs, Cas_particulier_souverain et les "
                "champs Souverain_* seulement pour Souverains, etc.). Choisissez "
                "d'abord la catégorie : les colonnes qui ne la concernent pas "
                "n'affichent alors que « (Sans objet pour cette catégorie) », "
                "et se réactivent automatiquement si vous changez la catégorie.",
            ),
            (
                "Champs Oui/Non",
                "Choisir « Oui » ou « Non » dans la liste déroulante (laisser "
                "vide si le cas ne s'applique pas à la ligne).",
            ),
            (
                "Calcul à l'import",
                "Chaque ligne (identifiée par ID_Exposition) est recalculée "
                "individuellement par le même moteur de calcul prudentiel que "
                "le formulaire « Ajouter une exposition » : pondération, EAD, "
                "RWA et capital requis seront donc identiques entre une saisie "
                "manuelle et un import de cette ligne.",
            ),
        ]
        for row_index, (label, text) in enumerate(instruction_rows, start=2):
            c_label = instructions.cell(row=row_index, column=1, value=label)
            c_text = instructions.cell(row=row_index, column=2, value=text)
            c_label.font = Font(bold=True, size=10, color=BLUE_DARK)
            c_label.fill = hdr_fill(BLUE_LIGHT)
            c_label.border = thin_b
            c_label.alignment = left(wrap=True)
            c_text.font = Font(size=10)
            c_text.fill = hdr_fill("FFFFFF")
            c_text.border = thin_b
            c_text.alignment = left(wrap=True)
            instructions.row_dimensions[row_index].height = 34

        # ── Feuilles de saisie : en-têtes stylés + commentaires d'aide +
        # lignes d'exemple. Le TEXTE des en-têtes reste strictement identique
        # à EXPECTED_COLUMNS_BY_SHEET / IMPORT_SHEET_SPECS (aucun préfixe
        # décoratif type "★") car le contrôle de structure à l'import fait
        # une comparaison exacte des noms de colonnes.
        example_rows_by_sheet: dict[str, list[dict[str, Any]]] = {
            "Template données": [
                {
                    "Date d'analyse": "2026-06-30",
                    "ID_Exposition": "EXP-EX-001",
                    "Contrepartie": "État XYZ",
                    "Notation_externe_contrepartie": "AAA",
                    "Pays_contrepartie": "Côte d'Ivoire",
                    "Notation_externe_pays": "AAA",
                    "Catégorie d'exposition": "Souverains",
                    "Montant_exposition_but_au_bilan": 500000000,
                    "Devise": "XOF",
                    "Type_CRM": "Aucune",
                },
                {
                    "Date d'analyse": "2026-06-30",
                    "ID_Exposition": "EXP-EX-002",
                    "Date d'octroi": "2024-01-10",
                    "Date d'échéance": "2027-01-10",
                    "Contrepartie": "Société ABC SA",
                    "Notation_externe_contrepartie": "BB",
                    "Pays_contrepartie": "Sénégal",
                    "Notation_externe_pays": "A",
                    "Catégorie d'exposition": "Entreprises",
                    "PRÊT TOTAL": 160000000,
                    "Montant_exposition_but_au_bilan": 150000000,
                    "Devise": "XOF",
                    "Type_CRM": "CRM financee",
                },
                {
                    "Date d'analyse": "2026-06-30",
                    "ID_Exposition": "EXP-EX-003",
                    "Contrepartie": "Particulier — M. Koffi",
                    "Notation_externe_contrepartie": "Non noté",
                    "Pays_contrepartie": "Côte d'Ivoire",
                    "Notation_externe_pays": "AAA",
                    "Catégorie d'exposition": "Clientèle de détail",
                    "Montant_exposition_but_au_bilan": 8000000,
                    "Montant d'exposition au HB": 2000000,
                    "Niveau de risque HB": "Risque moyen",
                    "Devise": "XOF",
                    "Type_CRM": "CRM non financee",
                    "Clientele_detail_criteres_respectes": "Oui",
                },
                {
                    "Date d'analyse": "2026-06-30",
                    "ID_Exposition": "EXP-EX-004",
                    "Contrepartie": "Banque Régionale XYZ",
                    "Notation_externe_contrepartie": "A",
                    "Pays_contrepartie": "Bénin",
                    "Notation_externe_pays": "A",
                    "Catégorie d'exposition": "Institutions financières",
                    "Montant_exposition_but_au_bilan": 75000000,
                    "Devise": "XOF",
                    "Type_CRM": "Aucune",
                },
            ],
            "CRM_financée": [
                {
                    "ID_Exposition": "EXP-EX-002",
                    "Valeur_Collatéral": 50000000,
                    "Type_emetteur": "emprunteur souverain",
                    "Notation": "AAA",
                    "Bloc": "Garantie espèces",
                    "Maturite": "<=1 an",
                },
            ],
            "CRM_non_financee": [
                {
                    "ID_Exposition": "EXP-EX-003",
                    "Nom du garant": "Fonds de Garantie XYZ",
                    "Catégorie du garant": "Souverains",
                    "Note_garant": "AAA",
                    "Pays_garant": "Côte d'Ivoire",
                    "Note_pays_garant": "AAA",
                    "Part couverte": 1500000,
                },
            ],
        }

        for sheet_name in EXPECTED_SHEETS:
            sheet = workbook.create_sheet(sheet_name)
            required_columns = list(EXPECTED_COLUMNS_BY_SHEET.get(sheet_name, ()))
            optional_columns = list(OPTIONAL_COLUMNS_BY_SHEET.get(sheet_name, ()))
            headers = required_columns + optional_columns
            if not headers:
                continue

            spec = column_specs_by_sheet.get(sheet_name)
            description_by_column = {
                column.name: column.description
                for column in ((spec.required_columns + spec.optional_columns) if spec else ())
            }

            value_type_by_column = value_type_by_column_by_sheet.get(sheet_name, {})
            # Une DataValidation par plage source (et par feuille) : plusieurs
            # colonnes de la même feuille partageant la même liste (ex: les
            # deux colonnes de notation) réutilisent le même objet plutôt que
            # d'en recréer un par colonne.
            data_validation_by_range: dict[str, DataValidation] = {}

            def _data_validation_for(cell_range: str) -> DataValidation:
                existing = data_validation_by_range.get(cell_range)
                if existing is not None:
                    return existing
                dv = DataValidation(
                    type="list",
                    formula1=cell_range,
                    allow_blank=True,
                    showDropDown=False,
                    showErrorMessage=True,
                    errorTitle="Valeur non reconnue",
                    error=(
                        "Choisissez une valeur dans la liste déroulante pour que "
                        "le calcul prudentiel s'applique correctement à cette ligne."
                    ),
                )
                sheet.add_data_validation(dv)
                data_validation_by_range[cell_range] = dv
                return dv

            # Colonne "Catégorie d'exposition" de cette feuille (si présente) :
            # sert de référence aux listes déroulantes en cascade ci-dessous.
            category_col_letter = (
                get_column_letter(headers.index("Catégorie d'exposition") + 1)
                if "Catégorie d'exposition" in headers
                else None
            )

            sheet.freeze_panes = "A2"
            sheet.row_dimensions[1].height = 30
            for col_index, header in enumerate(headers, start=1):
                is_required = header in required_columns
                cell = sheet.cell(row=1, column=col_index, value=header)
                cell.font = Font(bold=True, size=10, color=BLUE_DARK if is_required else "475569")
                cell.fill = hdr_fill(BLUE_LIGHT if is_required else GREY_HEADER)
                cell.border = thin_b
                cell.alignment = center(wrap=True)
                width = max(16, min(38, len(header) + 4))
                col_letter = cell.column_letter
                sheet.column_dimensions[col_letter].width = width

                description = description_by_column.get(header)
                if description:
                    comment_text = f"{'Obligatoire' if is_required else 'Optionnel'} — {description}"
                    cell.comment = Comment(comment_text, "OutilRWA")

                # Liste déroulante en cascade : les colonnes qui ne concernent
                # qu'une (ou plusieurs) catégorie(s) précise(s) ne proposent
                # les vraies options que si la catégorie de la ligne
                # correspond ; sinon seule "(Sans objet pour cette catégorie)"
                # est proposée — reproduit le masquage conditionnel du
                # formulaire "Ajouter une exposition" selon la catégorie.
                cascade_spec = (
                    CATEGORY_DEPENDENT_COLUMNS.get(header)
                    if category_col_letter
                    else None
                )
                if cascade_spec is not None:
                    active_codes, active_options = cascade_spec
                    active_range = _validation_range_for(active_options)
                    active_name = _short_name_for(active_range)
                    conditions = [
                        f'${category_col_letter}2="{CATEGORY_PRUDENTIAL_BY_CODE[code]}"'
                        for code in active_codes
                    ]
                    condition = (
                        conditions[0]
                        if len(conditions) == 1
                        else f"OR({','.join(conditions)})"
                    )
                    formula = (
                        f'INDIRECT(IF({condition},"{active_name}",'
                        f'"{not_applicable_name}"))'
                    )
                    data_range = f"{col_letter}2:{col_letter}{1 + _VALIDATION_ROW_COUNT}"
                    _data_validation_for(formula).add(data_range)
                else:
                    # Liste déroulante fixe : colonnes à choix fixes
                    # explicites, sinon Oui/Non pour toute colonne de ce type
                    # dans la spec du validateur
                    # (app/validators/excel_import_validator.py) — donc tout
                    # futur champ oui/non hérite automatiquement de son menu
                    # déroulant sans modification à faire ici.
                    options = FIXED_OPTIONS_BY_COLUMN.get(header)
                    if options is not None:
                        cell_range = _validation_range_for(options)
                    elif value_type_by_column.get(header) == "oui/non":
                        cell_range = oui_non_range
                    else:
                        cell_range = None
                    if cell_range is not None:
                        data_range = f"{col_letter}2:{col_letter}{1 + _VALIDATION_ROW_COUNT}"
                        _data_validation_for(cell_range).add(data_range)

            for row_offset, example in enumerate(example_rows_by_sheet.get(sheet_name, ()), start=2):
                row_fill = hdr_fill("FFFFFF") if row_offset % 2 == 0 else hdr_fill("F8FAFC")
                for col_index, header in enumerate(headers, start=1):
                    value = example.get(header)
                    cell = sheet.cell(row=row_offset, column=col_index, value=value)
                    cell.border = thin_b
                    cell.fill = row_fill
                    cell.font = Font(size=10)
                    cell.alignment = left()

        # ── Feuille informative "Listes de référence" (non requise à
        # l'import — sert uniquement de rappel des valeurs valides) ────────
        reference = workbook.create_sheet("Listes de référence")
        reference.column_dimensions["A"].width = 40
        reference.column_dimensions["B"].width = 40
        reference.merge_cells("A1:B1")
        reference["A1"].value = "Listes de référence (rappel — feuille informative)"
        reference["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        reference["A1"].fill = hdr_fill(BLUE_DARK)
        reference["A1"].alignment = center()
        reference.row_dimensions[1].height = 26

        # Reprend les constantes partagées ci-dessus (mêmes listes que celles
        # utilisées pour les menus déroulants) pour qu'il n'existe qu'une
        # seule source de vérité pour ces valeurs dans ce fichier.
        reference_blocks = [
            ("Catégorie d'exposition", CATEGORY_PRUDENTIAL_LABELS),
            ("Type_CRM", CRM_TYPE_OPTIONS),
            ("Niveau de risque HB", OFF_BALANCE_RISK_LEVEL_OPTIONS),
            ("Catégorie du garant (CRM non financée)", GUARANTOR_CATEGORY_OPTIONS),
            ("Notation (contrepartie / pays / garant)", RATING_OPTIONS),
            ("Notation (collatéral CRM financée)", FINANCED_CRM_DEBT_RATING_OPTIONS),
            ("Type_emetteur (CRM financée)", FINANCED_CRM_ISSUER_ROLE_OPTIONS),
            ("Maturite (CRM financée)", FINANCED_CRM_MATURITY_BUCKET_OPTIONS),
            ("Type_Collatéral (CRM financée)", FINANCED_CRM_COLLATERAL_TYPE_OPTIONS),
            (
                "Cas_particulier_souverain (cascade : catégorie = Souverains)",
                SOVEREIGN_SPECIAL_CASE_OPTIONS,
            ),
            (
                "Souverain_note_OCE (cascade : catégorie = Souverains)",
                SOVEREIGN_OCE_NOTE_OPTIONS,
            ),
            (
                "Cas_institution_bancaire (cascade : catégorie = Institutions financières)",
                BANK_INSTITUTION_CASE_OPTIONS,
            ),
            (
                "Type_autre_actif (cascade : catégorie = Autres actifs)",
                OTHER_ASSET_TYPE_OPTIONS,
            ),
            (
                "Ponderation_initiale_avant_defaut (cascade : catégorie = Créances en souffrance)",
                DEFAULTED_EXPOSURE_INITIAL_RISK_WEIGHT_OPTIONS,
            ),
            ("Devise / Devise_Collatéral", CURRENCY_OPTIONS),
            (
                "Champs Oui/Non (tous les Souverain_*, BMD_*, Organisme_public_*, "
                "Clientele_detail_*, Immobilier_*, Defaut_*, Entreprise_*, "
                "Obligation_convertible_indice_principal) — la plupart sont en "
                "cascade selon Catégorie d'exposition, voir l'onglet "
                "Instructions",
                OUI_NON_OPTIONS,
            ),
        ]
        row_index = 3
        for title, options in reference_blocks:
            title_cell = reference.cell(row=row_index, column=1, value=title)
            title_cell.font = Font(bold=True, size=10, color=BLUE_DARK)
            title_cell.fill = hdr_fill(BLUE_LIGHT)
            title_cell.border = thin_b
            reference.cell(row=row_index, column=2, value="").border = thin_b
            row_index += 1
            for option in options:
                c = reference.cell(row=row_index, column=1, value=option)
                c.border = thin_b
                c.font = Font(size=10)
                c.fill = hdr_fill("FFFFFF") if row_index % 2 == 0 else hdr_fill("F8FAFC")
                explanation = BANK_INSTITUTION_CASE_LABELS.get(str(option), "")
                b_cell = reference.cell(row=row_index, column=2, value=explanation)
                b_cell.border = thin_b
                b_cell.font = Font(size=9, italic=True)
                row_index += 1
            row_index += 1

        output = BytesIO()
        workbook.save(output)
        workbook.close()
        output.seek(0)
        return output.getvalue()

    def inspect_uploaded_workbook(self, workbook_bytes: bytes, filename: str | None = None) -> dict[str, Any]:
        source_name = filename or "upload.xlsx"
        profiler = ImportProfiler(source_name=source_name)

        with profiler.measure("lecture_fichier_excel"):
            workbook = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)

        try:
            inspection = self._inspect_structure(workbook, profiler=profiler)

            with profiler.measure("lecture_des_lignes_utiles"):
                rows_read_by_sheet = self._collect_rows_read_by_sheet(workbook)
        finally:
            workbook.close()

        return {
            "file": source_name,
            "valid": inspection["valid"],
            "sheet_count": inspection["sheet_count"],
            "detected_sheets": inspection["detected_sheets"],
            "sheets": inspection["sheets"],
            "errors": inspection["errors"],
            "rows_read_by_sheet": rows_read_by_sheet,
            "duration_ms": profiler.total_ms,
            "steps_ms": profiler.timings_ms,
        }

    def import_uploaded_workbook(
        self,
        workbook_bytes: bytes,
        filename: str | None = None,
        *,
        mode: str = "merge",
    ) -> dict[str, Any]:
        source_name = filename or "upload.xlsx"
        normalized_mode = mode.strip().lower()
        if normalized_mode not in {"merge", "replace"}:
            raise ValueError("Le mode d'import doit être 'merge' ou 'replace'.")

        profiler = ImportProfiler(source_name=source_name)

        with profiler.measure("lecture_fichier_excel"):
            workbook = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)

        try:
            inspection = self._inspect_structure(workbook, profiler=profiler)
            self._ensure_valid_inspection(inspection)

            parsed = self._parse_workbook(workbook, profiler=profiler)
        finally:
            workbook.close()

        backup_path = None
        if normalized_mode == "replace":
            with profiler.measure("sauvegarde_avant_ecrasement"):
                backup_path = database_manager.create_backup("before_excel_replace_import")

        with profiler.measure("insertion_sqlite"):
            with database_manager.transaction() as connection:
                run_id = _create_import_run(
                    connection,
                    source_type="excel_upload",
                    source_name=source_name,
                    total_rows=parsed.rows_read,
                    details={
                        "mode": normalized_mode,
                        "rows_read_by_sheet": parsed.rows_read_by_sheet,
                    },
                )
                persistence_stats = import_repository.persist_import(
                    exposure_records=parsed.exposure_records,
                    off_balance_records=parsed.off_balance_records,
                    mode=normalized_mode,
                    connection=connection,
                )
                _write_metadata(connection, "last_upload_import_at", utcnow_iso())
                _write_metadata(connection, "storage_backend", "sqlite")
                if backup_path is not None:
                    _write_metadata(connection, "last_backup_path", str(backup_path))

                report_details = {
                    "mode": normalized_mode,
                    "rows_read_by_sheet": parsed.rows_read_by_sheet,
                    "valid_rows": parsed.valid_rows,
                    "inserted_exposures": persistence_stats["inserted_exposures"],
                    "updated_exposures": persistence_stats["updated_exposures"],
                    "inserted_off_balance": persistence_stats["inserted_off_balance"],
                    "updated_off_balance": persistence_stats["updated_off_balance"],
                    "errors": parsed.errors,
                    "steps_ms": profiler.timings_ms,
                    "duration_ms": profiler.total_ms,
                    "backup_path": str(backup_path) if backup_path else None,
                }
                _finish_import_run(
                    connection,
                    run_id,
                    imported_rows=(
                        len(parsed.exposure_records) + len(parsed.off_balance_records)
                    ),
                    rejected_rows=parsed.rejected_rows,
                    details=report_details,
                )

        total_imported = (
            persistence_stats["inserted_exposures"] + persistence_stats["inserted_off_balance"]
        )
        total_updated = (
            persistence_stats["updated_exposures"] + persistence_stats["updated_off_balance"]
        )

        report = {
            "status": "success",
            "file": source_name,
            "mode": normalized_mode,
            "rows_read": parsed.rows_read,
            "valid_rows": parsed.valid_rows,
            "imported_rows": total_imported,
            "updated_rows": total_updated,
            "rejected_rows": parsed.rejected_rows,
            "rows_read_by_sheet": parsed.rows_read_by_sheet,
            "errors": parsed.errors,
            "backup_path": str(backup_path) if backup_path else None,
            "duration_ms": profiler.total_ms,
            "steps_ms": profiler.timings_ms,
        }
        logger.info(
            "Import Excel [%s] terminé en %.2f ms | lues=%s importées=%s mises_a_jour=%s rejetées=%s",
            source_name,
            profiler.total_ms,
            parsed.rows_read,
            total_imported,
            total_updated,
            parsed.rejected_rows,
        )
        return report

    def _inspect_structure(self, workbook, *, profiler: ImportProfiler) -> dict[str, Any]:
        with profiler.measure("detection_des_feuilles"):
            detected_sheets = list(workbook.sheetnames)

        with profiler.measure("validation_des_colonnes"):
            inspection = inspect_workbook_structure(workbook)

        inspection["detected_sheets"] = detected_sheets
        inspection["sheet_count"] = len(detected_sheets)
        return inspection

    def _ensure_valid_inspection(self, inspection: dict[str, Any]) -> None:
        if inspection.get("valid") is True:
            return
        raise ExcelImportValidationError(
            {
                "message": "Le fichier Excel ne respecte pas le format d'import attendu.",
                **inspection,
            }
        )

    def _collect_rows_read_by_sheet(self, workbook) -> dict[str, int]:
        counts: dict[str, int] = {}
        for spec in IMPORT_SHEET_SPECS:
            if spec.name not in workbook.sheetnames:
                counts[spec.name] = 0
                continue
            rows = self._read_sheet_rows(workbook[spec.name], spec)
            counts[spec.name] = len(rows)
        return counts

    def _read_sheet_rows(self, sheet, spec) -> list[tuple[int, dict[str, Any]]]:
        headers = read_sheet_headers(sheet)
        selected_indexes = {
            header: index
            for index, header in enumerate(headers)
            if header and header in spec.import_columns
        }
        rows: list[tuple[int, dict[str, Any]]] = []
        for excel_row_index, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if not selected_indexes:
                continue
            if not any(
                index < len(values) and values[index] not in (None, "")
                for index in selected_indexes.values()
            ):
                continue
            row = {
                column_name: values[index] if index < len(values) else None
                for column_name, index in selected_indexes.items()
            }
            rows.append((excel_row_index, row))
        return rows

    def _parse_workbook(self, workbook, *, profiler: ImportProfiler) -> ParsedImportBundle:
        spec_by_name = {spec.name: spec for spec in IMPORT_SHEET_SPECS}
        with profiler.measure("lecture_feuille_template_donnees"):
            template_rows = self._read_sheet_rows(
                workbook["Template données"],
                spec_by_name["Template données"],
            )
        with profiler.measure("lecture_feuille_crm_financee"):
            crm_fin_rows = (
                self._read_sheet_rows(
                    workbook["CRM_financée"],
                    spec_by_name["CRM_financée"],
                )
                if "CRM_financée" in workbook.sheetnames
                else []
            )
        with profiler.measure("lecture_feuille_crm_non_financee"):
            crm_non_fin_rows = (
                self._read_sheet_rows(
                    workbook["CRM_non_financee"],
                    spec_by_name["CRM_non_financee"],
                )
                if "CRM_non_financee" in workbook.sheetnames
                else []
            )
        rows_read_by_sheet = {
            "Template données": len(template_rows),
            "CRM_financée": len(crm_fin_rows),
            "CRM_non_financee": len(crm_non_fin_rows),
        }

        with profiler.measure("normalisation_des_donnees"):
            crm_non_fin = excel_repository._crm_non_fin_by_rows(
                [row for _, row in crm_non_fin_rows]
            )
            crm_fin = excel_repository._crm_fin_by_rows([row for _, row in crm_fin_rows])

        exposure_records_by_id: dict[str, dict[str, Any]] = {}
        errors: list[dict[str, Any]] = []

        # Chaque ligne "Template données" devient un enregistrement d'exposition
        # complet via le même moteur de calcul prudentiel que la saisie
        # manuelle (build_exposure_record / compute_metrics). La composante
        # hors bilan éventuelle ("Montant d'exposition au HB" +
        # "Niveau de risque HB") est déjà intégrée dans cette même exposition
        # par le moteur de calcul : il n'y a plus de feuille séparée à
        # traiter pour le hors bilan.
        with profiler.measure("calcul_des_expositions"):
            for excel_row_index, row in template_rows:
                exposure_id = _clean_text(row.get("ID_Exposition"))
                if not exposure_id:
                    errors.append(
                        {
                            "sheet": "Template données",
                            "row": excel_row_index,
                            "column": "ID_Exposition",
                            "message": "ID_Exposition manquant.",
                        }
                    )
                    continue
                if exposure_id in exposure_records_by_id:
                    errors.append(
                        {
                            "sheet": "Template données",
                            "row": excel_row_index,
                            "column": "ID_Exposition",
                            "message": f"ID dupliqué dans le fichier: {exposure_id}.",
                        }
                    )
                    continue
                try:
                    built = excel_repository._build_exposure_from_template(
                        row,
                        crm_non_fin_row=crm_non_fin.get(exposure_id),
                        crm_fin_row=crm_fin.get(exposure_id),
                    )
                except Exception as exc:
                    errors.append(
                        {
                            "sheet": "Template données",
                            "row": excel_row_index,
                            "column": None,
                            "message": str(exc),
                        }
                    )
                    continue

                exposure_records_by_id[exposure_id] = dict(built)

        exposure_records = list(exposure_records_by_id.values())
        off_balance_records: list[dict[str, Any]] = []
        rows_read = sum(rows_read_by_sheet.values())
        valid_rows = len(exposure_records)
        return ParsedImportBundle(
            exposure_records=exposure_records,
            off_balance_records=off_balance_records,
            rows_read=rows_read,
            valid_rows=valid_rows,
            rejected_rows=len(errors),
            errors=errors,
            rows_read_by_sheet=rows_read_by_sheet,
        )


excel_import_service = ExcelImportService()
