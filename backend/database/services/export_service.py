"""Services d'export Excel depuis SQLite."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.core.excel_repository import excel_repository
from app.core.runtime_paths import exports_dir
from database.repositories.exposure_repository import exposure_repository
from database.repositories.off_balance_repository import off_balance_repository


EXPORTS_DIR = exports_dir()


class ExportService:
    """Construit les exports Excel depuis la base SQLite."""

    _TEMPLATE_SHEET = "Template données"
    _CRM_FINANCED_SHEET = "CRM_financée"
    _CRM_NON_FINANCED_SHEET = "CRM_non_financee"
    _OFF_BALANCE_SHEET = "Traitement_HB"

    def export_excel_workbook_bytes(self) -> bytes:
        """Construit un classeur Excel complet basé sur le vrai modèle RWA."""

        source_path = excel_repository.source_path
        exposures = exposure_repository.list_exposures()
        off_balance_rows = off_balance_repository.list_commitments()
        workbook = self._load_export_workbook(source_path)
        try:
            self._fill_template_sheet(workbook[self._TEMPLATE_SHEET], exposures)
            self._fill_crm_financed_sheet(
                workbook[self._CRM_FINANCED_SHEET],
                exposures,
            )
            self._fill_crm_non_financed_sheet(
                workbook[self._CRM_NON_FINANCED_SHEET],
                exposures,
            )
            self._fill_off_balance_sheet(
                workbook[self._OFF_BALANCE_SHEET],
                off_balance_rows,
            )

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
        finally:
            workbook.close()

    def export_excel_workbook(self) -> Path:
        """Sauvegarde un export Excel sur disque et retourne son chemin."""

        EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        export_path = (
            EXPORTS_DIR
            / f"expositions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        export_path.write_bytes(self.export_excel_workbook_bytes())
        return export_path

    def _load_export_workbook(self, source_path: Path):
        if source_path.exists():
            return load_workbook(source_path)
        return self._build_fallback_workbook()

    def _build_fallback_workbook(self):
        workbook = Workbook()
        template_sheet = workbook.active
        template_sheet.title = self._TEMPLATE_SHEET
        template_sheet.append(
            [
                "Date d'analyse",
                "ID_Exposition",
                "Date d'octroi",
                "Date d'échéance",
                "Maturité de l'exposition",
                "Maturité résiduelle",
                "Contrepartie",
                "Notation_externe_contrepartie",
                "Pays_contrepartie",
                "Notation_externe_pays",
                "Pondération_pays",
                "Catégorie d'exposition",
                "Pondération (RW)",
                "PRÊT TOTAL",
                "Montant_exposition_but_au_bilan",
                "Montant d'exposition au HB",
                "Devise",
                "CRM_existe",
                "Type_CRM",
                "EAD_bilan",
                "EAD_HB",
                "EAD_HB_ccf",
                "EAD_Total",
                "RWA_EB",
                "RWA_HB",
                "RWA_crédit",
                "Capital_min_reg",
            ]
        )
        workbook.create_sheet(self._CRM_FINANCED_SHEET).append(
            [
                "ID_Exposition",
                "Valeur_Collatéral",
                "Type_emetteur",
                "Notation",
                "Bloc",
                "Maturite",
                "HE",
                "HC",
                "Hfx",
                "Eva_EB",
                "Eva_HB",
                "Cva",
            ]
        )
        workbook.create_sheet(self._CRM_NON_FINANCED_SHEET).append(
            [
                "ID_Exposition",
                "Nom du garant",
                "Note_garant",
                "Pays_garant",
                "Notation_externe_pays_garant",
                "Pondération_pays_garant",
                "Catégorie du garant",
                "Pondération du garant",
                "% Exp_couverte",
                "% Exp_non_couverte",
                "Part couverte",
                "Part non couverte",
                "RWA_crédit",
            ]
        )
        workbook.create_sheet(self._OFF_BALANCE_SHEET).append(
            [
                "ID_Exposition",
                "Catégorie Hors bilan",
                "Facteur_conversion (CCF)",
                "EAD_HB_ccf",
            ]
        )
        for sheet_name in (
            self._TEMPLATE_SHEET,
            self._CRM_FINANCED_SHEET,
            self._CRM_NON_FINANCED_SHEET,
            self._OFF_BALANCE_SHEET,
        ):
            workbook[sheet_name].freeze_panes = "A2"
        return workbook

    def _clear_sheet_values(self, sheet, *, max_columns: int) -> None:
        """Efface les valeurs d'une feuille en conservant sa structure."""

        if sheet.max_row < 2:
            return
        final_column = max(sheet.max_column, max_columns)
        for row in sheet.iter_rows(
            min_row=2,
            max_row=sheet.max_row,
            min_col=1,
            max_col=final_column,
        ):
            for cell in row:
                cell.value = None

    def _write_row(self, sheet, row_index: int, values: list[Any]) -> None:
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)

    def _coerce_excel_date(self, value: Any) -> date | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        if not text:
            return None
        for parser in (date.fromisoformat,):
            try:
                return parser(text.split("T")[0])
            except ValueError:
                continue
        return None

    def _coerce_float(self, value: Any, default: float = 0.0) -> float:
        if value in (None, ""):
            return default
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(" ", "").replace(",", ".")
        if not text:
            return default
        try:
            return float(text)
        except ValueError:
            return default

    def _coerce_optional_float(self, value: Any) -> float | None:
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(" ", "").replace(",", ".")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None

    def _months_between(self, start: date | None, end: date | None) -> int | None:
        if start is None or end is None:
            return None
        months = (end.year - start.year) * 12 + (end.month - start.month)
        if end.day < start.day:
            months -= 1
        return max(months, 0)

    def _coverage_ratio(self, exposure: dict[str, Any]) -> float:
        return min(
            max(self._coerce_float(exposure.get("crm_coverage_percent")), 0.0),
            1.0,
        )

    def _crm_type_label(self, exposure: dict[str, Any]) -> str:
        return str(exposure.get("crm_type") or "")

    def _fill_template_sheet(
        self,
        sheet,
        exposures: list[dict[str, Any]],
    ) -> None:
        self._clear_sheet_values(sheet, max_columns=27)

        for row_index, exposure in enumerate(exposures, start=2):
            analysis_date = self._coerce_excel_date(exposure.get("analysis_date"))
            grant_date = self._coerce_excel_date(exposure.get("grant_date"))
            maturity_date = self._coerce_excel_date(exposure.get("maturity_date"))
            identifier = str(exposure["id"])
            crm_exists = bool(exposure.get("crm_exists"))
            exposure_maturity = exposure.get("exposure_maturity_months")
            residual_maturity = exposure.get("residual_maturity_months")
            loan_total_amount = exposure.get("loan_total_amount", exposure.get("gross_amount"))
            on_balance_amount = exposure.get("on_balance_exposure_amount", exposure.get("gross_amount"))
            off_balance_amount = exposure.get("off_balance_exposure_amount")
            country_risk_weight = self._coerce_optional_float(
                exposure.get("country_risk_weight")
            )
            self._write_row(
                sheet,
                row_index,
                [
                    analysis_date,
                    identifier,
                    grant_date,
                    maturity_date,
                    exposure_maturity
                    if exposure_maturity is not None
                    else self._months_between(grant_date, maturity_date),
                    residual_maturity
                    if residual_maturity is not None
                    else self._months_between(analysis_date, maturity_date),
                    str(exposure.get("counterparty_name") or ""),
                    str(exposure.get("rating") or ""),
                    str(exposure.get("country") or ""),
                    str(exposure.get("country_rating") or "Non noté"),
                    country_risk_weight,
                    str(exposure.get("category_raw") or ""),
                    self._coerce_float(exposure.get("final_rw")),
                    self._coerce_optional_float(loan_total_amount),
                    self._coerce_optional_float(on_balance_amount),
                    self._coerce_optional_float(off_balance_amount),
                    str(exposure.get("currency") or "XOF"),
                    "OUI" if crm_exists else "NON",
                    self._crm_type_label(exposure),
                    self._coerce_optional_float(
                        exposure.get("ead_bilan_amount", exposure.get("ead"))
                    ),
                    self._coerce_optional_float(exposure.get("ead_hb_amount")),
                    self._coerce_optional_float(exposure.get("ead_hb_ccf_amount")),
                    self._coerce_optional_float(
                        exposure.get("ead_total_amount", exposure.get("ead"))
                    ),
                    self._coerce_optional_float(exposure.get("rwa_eb_amount")),
                    self._coerce_optional_float(exposure.get("rwa_hb_amount")),
                    self._coerce_float(exposure.get("rwa")),
                    self._coerce_float(exposure.get("capital")),
                ],
            )

    def _fill_crm_financed_sheet(
        self,
        sheet,
        exposures: list[dict[str, Any]],
    ) -> None:
        self._clear_sheet_values(sheet, max_columns=12)

        row_index = 2
        for exposure in exposures:
            crm_details = exposure.get("crm_details", {})
            if crm_details.get("mode") != "CRM financee":
                continue

            collateral_value = self._coerce_float(crm_details.get("collateral_value"))
            eva_eb = self._coerce_optional_float(crm_details.get("eva_eb"))
            eva_hb = self._coerce_optional_float(crm_details.get("eva_hb"))
            self._write_row(
                sheet,
                row_index,
                [
                    str(exposure["id"]),
                    collateral_value,
                    str(crm_details.get("issuer_type") or ""),
                    str(crm_details.get("issuer_rating") or ""),
                    str(crm_details.get("label") or exposure.get("crm_type") or ""),
                    str(crm_details.get("maturity_bucket") or ""),
                    self._coerce_float(crm_details.get("he")),
                    self._coerce_float(crm_details.get("hc")),
                    self._coerce_float(crm_details.get("hfx")),
                    eva_eb if eva_eb is not None else self._coerce_optional_float(
                        exposure.get("on_balance_exposure_amount")
                    ),
                    eva_hb if eva_hb is not None else self._coerce_optional_float(
                        exposure.get("off_balance_exposure_amount")
                    ),
                    self._coerce_float(crm_details.get("cva")),
                ],
            )
            row_index += 1

    def _fill_crm_non_financed_sheet(
        self,
        sheet,
        exposures: list[dict[str, Any]],
    ) -> None:
        self._clear_sheet_values(sheet, max_columns=13)

        row_index = 2
        for exposure in exposures:
            crm_details = exposure.get("crm_details", {})
            if crm_details.get("mode") != "CRM non financee":
                continue

            gross_amount = self._coerce_float(exposure.get("gross_amount"))
            coverage = self._coverage_ratio(exposure)
            covered_amount = round(gross_amount * coverage, 2)
            uncovered_amount = round(gross_amount - covered_amount, 2)
            self._write_row(
                sheet,
                row_index,
                [
                    str(exposure["id"]),
                    str(crm_details.get("guarantor_name") or ""),
                    str(crm_details.get("guarantor_rating") or ""),
                    str(crm_details.get("guarantor_country") or ""),
                    str(
                        crm_details.get("guarantor_country_rating") or ""
                    ),
                    self._coerce_float(crm_details.get("guarantor_country_rw")),
                    str(crm_details.get("guarantor_category") or ""),
                    self._coerce_float(exposure.get("guarantor_rw")),
                    coverage,
                    max(0.0, round(1.0 - coverage, 4)),
                    covered_amount,
                    uncovered_amount,
                    self._coerce_float(exposure.get("rwa")),
                ],
            )
            row_index += 1

    def _fill_off_balance_sheet(
        self,
        sheet,
        off_balance_rows: list[dict[str, Any]],
    ) -> None:
        self._clear_sheet_values(sheet, max_columns=4)

        for row_index, row in enumerate(off_balance_rows, start=2):
            self._write_row(
                sheet,
                row_index,
                [
                    str(row.get("counterparty_id") or ""),
                    str(row.get("engagement_type") or ""),
                    self._coerce_float(row.get("ccf")),
                    self._coerce_float(row.get("ead")),
                ],
            )


export_service = ExportService()
