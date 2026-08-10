import openpyxl

wb = openpyxl.Workbook()

# Remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Obligations
ws_bonds = wb.create_sheet("Obligations")
bond_headers = [
    "ID Titre",
    "Date d'analyse",
    "Pays émetteur",
    "Zone",
    "Type d'instrument",
    "Code type d'instrument",
    "Emetteur",
    "Mode de Placement",
    "Notation externe_S&P",
    "Notation externe_Moody's",
    "Notation externe_Fitch",
    "La pire notation externe",
    "Notation Interne",
    "Intention comptable",
    "Classification des titres",
    "Date d'émission",
    "Devise",
    "Valeur nominale unitaire",
    "quantités",
    "Capital initial",
    "Prix d'émission",
    "Prime d'émission",
    "Prix de remboursement",
    "Prime de remboursement",
    "Date d'échéance",
    "Coupon (%)",
    "Profil d'amortissement",
    "Fréquence de paiement des intérêts",
    "Maturité (mois)",
    "Maturité résiduelle (mois)"
]
ws_bonds.append(bond_headers)

# Actions
ws_equities = wb.create_sheet("Actions")
equity_headers = [
    "ID Instrument",
    "Type d'instrument",
    "Émetteur / Société",
    "Secteur",
    "Pays / marché",
    "Bourse",
    "Devise",
    "Intention comptable",
    "Quantité",
    "Prix d'acquisition",
    "Cours actuel",
    "Bêta",
    "Volatilité annualisée (%)",
    "Rendement dividende (%)",
    "Liquide et diversifié (Oui/Non)"
]
ws_equities.append(equity_headers)

# Historique Prix Obligations
ws_hist_bonds = wb.create_sheet("Historique Prix Obligations")
hist_bond_headers = [
    "Date",
    "ID Titre",
    "Prix de marché"
]
ws_hist_bonds.append(hist_bond_headers)

# Historique Cours Actions
ws_hist_equities = wb.create_sheet("Historique Cours Actions")
hist_equity_headers = [
    "Date",
    "ID Instrument",
    "Cours de clôture"
]
ws_hist_equities.append(hist_equity_headers)

wb.save("Modele_Import_Risque_Marche.xlsx")
